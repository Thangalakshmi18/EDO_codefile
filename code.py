import os
import re
import json
import time
import logging
import sys
import unicodedata
import zipfile
import tempfile
import openpyxl
import fitz  # PyMuPDF

from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU

# ==========================================================
# GLOBALS & STYLES SETUP
# ==========================================================

CURRENT_FOLDER = os.path.dirname(os.path.abspath(__file__))

MAX_LLM_RETRIES = 3
INITIAL_RETRY_DELAY = 1

IMAGE_WIDTH = 180
IMAGE_HEIGHT = 110
IMAGE_OUTPUT_DIR = "extracted_images"
IMAGE_TEXT_OFFSET_PX = 45

DEFAULT_DOCUMENT_SEARCH_ROOTS = ["/mnt/documents", "/data/documents", "."]

RED_CONSOLE = "\033[91m"
RESET_CONSOLE = "\033[0m"

cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    stream=sys.stdout
)

# ==========================================================
# DEPENDENCY HELPER FUNCTIONS (UTILITIES & RECURSIVE EXTRACTION)
# ==========================================================

def print_red(label, text):
    print(f"{RED_CONSOLE}{label}:\n{text}{RESET_CONSOLE}")

def normalize_text(value):
    if value is None:
        return ""
    value = str(value).replace("\u00A0", " ").replace("\u2007", " ").replace("\u202F", " ")
    return unicodedata.normalize("NFKC", value).strip()

def format_output_text(val):
    if val is None:
        return ""
    return normalize_text(str(val))

def format_edo_tag_text(tag):
    return format_output_text(tag)

def normalize_id(value):
    return re.sub(r"\s+", " ", normalize_text(value).upper()).strip()

def get_fixed_sysdd_reference():
    return "SYSDD-1234"

def deep_extract_records(data):
    """
    Recursively inspects dicts/lists returned by LLMs or JSON parsers 
    to extract flattened record lists regardless of nested structures.
    """
    if isinstance(data, list):
        records = []
        for item in data:
            records.extend(deep_extract_records(item))
        return records
    elif isinstance(data, dict):
        if "records" in data and isinstance(data["records"], list):
            return deep_extract_records(data["records"])
        elif "data" in data:
            return deep_extract_records(data["data"])
        elif "items" in data:
            return deep_extract_records(data["items"])
        else:
            return [data]
    return []

def parse_llm_json_response(response_text):
    """
    Cleans markdown code fences and parses JSON safely, returning a list of dicts.
    """
    if not response_text:
        return []
    cleaned = re.sub(r"^```json\s*|\s*```$", "", response_text.strip(), flags=re.DOTALL).strip()
    try:
        parsed = json.loads(cleaned)
        return deep_extract_records(parsed)
    except json.JSONDecodeError as e:
        logging.error(f"Failed to parse LLM response as JSON: {e}")
        return []

def safe_llm_call_with_retry(call_func, *args, **kwargs):
    """
    Wrapper for robust API calls with exponential backoff retries.
    """
    delay = INITIAL_RETRY_DELAY
    for attempt in range(1, MAX_LLM_RETRIES + 1):
        try:
            return call_func(*args, **kwargs)
        except Exception as e:
            logging.warning(f"LLM call attempt {attempt} failed: {e}")
            if attempt == MAX_LLM_RETRIES:
                raise e
            time.sleep(delay)
            delay *= 2

def resolve_document_path(filename, search_roots=None):
    """
    Searches provided document roots to locate target PDFs or files.
    """
    roots = search_roots or DEFAULT_DOCUMENT_SEARCH_ROOTS
    for root in roots:
        candidate = os.path.join(root, filename)
        if os.path.exists(candidate):
            return candidate
    return filename

# ==========================================================
# STYLING & RISK CLASSIFICATION DEPENDENCIES
# ==========================================================

def _apply_border_alignment(cell):
    cell.border = thin_border
    cell.alignment = cell_alignment

def apply_risk_cell_style(cell, risk_status):
    risk_upper = str(risk_status).strip().upper()
    if "HIGH" in risk_upper:
        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        cell.font = Font(color="9C0006", bold=True)
    elif "MEDIUM" in risk_upper:
        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        cell.font = Font(color="9C6500", bold=True)
    else:
        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        cell.font = Font(color="006100", bold=True)

def classify_risk_status(risk_input, pipeline_config):
    text = risk_input.lower()
    if "high" in text or "critical" in text:
        return "High"
    elif "medium" in text or "moderate" in text:
        return "Medium"
    return "Low"

def generate_remarks_and_recommendation(edo, tag_value, is_new, risk, pipeline_config):
    if is_new:
        return f"New EDO ({tag_value}) recommendations: Review alignment with updated design inputs."
    return f"Existing EDO ({tag_value}) recommendations: Maintain current traceability."

# ==========================================================
# IMAGE EXTRACTION & HANDLING DEPENDENCIES
# ==========================================================

def extract_images_from_pdf(pdf_path):
    """
    Extracts embedded images from a PDF document using PyMuPDF (fitz).
    """
    extracted_images = []
    if not os.path.exists(pdf_path):
        logging.warning(f"PDF path not found for image extraction: {pdf_path}")
        return extracted_images

    doc = fitz.open(pdf_path)
    for page_index in range(len(doc)):
        page = doc[page_index]
        image_list = page.get_images(full=True)
        for img_index, img in enumerate(image_list):
            xref = img[0]
            base_image = doc.extract_image(xref)
            image_bytes = base_image["image"]
            image_ext = base_image["ext"]
            extracted_images.append({
                "bytes": image_bytes,
                "extension": f".{image_ext}",
                "page": page_index + 1
            })
    return extracted_images

def extract_images_from_docx_or_zip(file_path):
    """
    Extracts images from OpenXML/Zip container formats (e.g., .docx or .xlsx).
    """
    images = []
    if not os.path.exists(file_path) or not zipfile.is_zipfile(file_path):
        return images

    with zipfile.ZipFile(file_path, 'r') as z:
        for filename in z.namelist():
            if filename.startswith("word/media/") or filename.startswith("xl/media/"):
                ext = os.path.splitext(filename)[1]
                images.append({
                    "bytes": z.read(filename),
                    "extension": ext
                })
    return images

def insert_image_below_text(sheet, image, row, column=8, text_offset_px=IMAGE_TEXT_OFFSET_PX):
    if not image:
        return False
    try:
        suffix = image.get("extension") or ".png"
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as temp:
            temp.write(image["bytes"])
            temp_name = temp.name

        excel_image = XLImage(temp_name)
        excel_image.width = IMAGE_WIDTH
        excel_image.height = IMAGE_HEIGHT

        marker = AnchorMarker(
            col=column - 1, colOff=pixels_to_EMU(2),
            row=row - 1, rowOff=pixels_to_EMU(text_offset_px)
        )
        size = XDRPositiveSize2D(cx=pixels_to_EMU(IMAGE_WIDTH), cy=pixels_to_EMU(IMAGE_HEIGHT))
        excel_image.anchor = OneCellAnchor(_from=marker, ext=size)
        sheet.add_image(excel_image)
        return True
    except Exception as e:
        logging.error(f"Image insert failed at row {row}: {e}")
        return False

# ==========================================================
# EXISTING EDO EXTRACTIONS (CALL 1, 2, 3)
# ==========================================================

def call_1_extract_existing_edo_tags(client, product_family, product, templatename, pipeline_config, edo_document, db):
    prompt = db.get_prompt_by_name(client, product_family, product, templatename, "EDO_Existing_Tag")
    prompt_row = {
        "prompt_role": prompt["prompt_role"],
        "prompt_text": prompt["prompt_text"],
        "question": "Extract all Known Active Design Output Tracking Numbers.",
        "fulltext": "Yes", "where_filter": "", "where_document": "", "checkpoint": ""
    }

    _, _, response = retrieve_content_for_prompt(
        pipeline_config, edo_document["edo_proposed"]["collection"],
        prompt_row["question"], prompt_row["prompt_role"], prompt_row["prompt_text"],
        prompt_row["fulltext"], prompt_row["where_filter"], prompt_row["where_document"]
    )

    records = parse_llm_json_response(response)
    existing_edos = {}
    for item in records:
        tag = normalize_text(item.get("edo_number") or item.get("edo_tag") or "")
        if tag and tag.lower() != "blank" and tag not in existing_edos:
            existing_edos[tag] = {
                "edo_type": "Existing",
                "edo_tag": tag,
                "edo_description": "",
                "reason_identified": "",
                "dfmea": "",
                "existing_trace": "",
                "verification_reference": "",
                "location": "",
                "design_elements": []
            }
    return existing_edos

def call_2_extract_existing_edo_details(client, product_family, product, templatename, pipeline_config, edo_document, existing_edos, db):
    prompt = db.get_prompt_by_name(client, product_family, product, templatename, "EDO_Existing_Generic")

    for edo_tag, edo in existing_edos.items():
        question = f"Extract complete row details for EDO Tag: {edo_tag}"
        prompt_row = {
            "prompt_role": prompt["prompt_role"],
            "prompt_text": prompt["prompt_text"],
            "question": question,
            "fulltext": "Yes", "where_filter": "", "where_document": "", "checkpoint": ""
        }

        _, _, response = retrieve_content_for_prompt(
            pipeline_config, edo_document["edo_proposed"]["collection"],
            prompt_row["question"], prompt_row["prompt_role"], prompt_row["prompt_text"],
            prompt_row["fulltext"], prompt_row["where_filter"], prompt_row["where_document"]
        )

        records = parse_llm_json_response(response)
        if records:
            rec = records[0]
            edo["edo_description"] = format_output_text(rec.get("edo_description") or rec.get("description"))
            edo["reason_identified"] = format_output_text(rec.get("reason_identified_as_edo") or rec.get("reason"))
            edo["dfmea"] = format_output_text(rec.get("ra_and_fmea_no") or rec.get("trace"))
            edo["location"] = format_output_text(rec.get("location"))
            edo["design_elements"] = rec.get("design_elements") or []

    return existing_edos

def call_3_extract_existing_edo_trace(client, product_family, product, templatename, pipeline_config, edo_document, existing_edos, db):
    if "edo_fmea" not in edo_document:
        return existing_edos

    prompt = db.get_prompt_by_name(client, product_family, product, templatename, "EDO_Existing_Trace")
    for edo_tag, edo in existing_edos.items():
        if not edo.get("dfmea"):
            continue

        prompt_row = {
            "prompt_role": prompt["prompt_role"],
            "prompt_text": prompt["prompt_text"] + f"\nTARGET DFMEA: {edo['dfmea']}",
            "question": f"Extract traceability details for {edo['dfmea']}",
            "fulltext": "Yes", "where_filter": "", "where_document": "", "checkpoint": ""
        }

        _, _, response = retrieve_content_for_prompt(
            pipeline_config, edo_document["edo_fmea"]["collection"],
            prompt_row["question"], prompt_row["prompt_role"], prompt_row["prompt_text"],
            prompt_row["fulltext"], prompt_row["where_filter"], prompt_row["where_document"]
        )

        records = parse_llm_json_response(response)
        if records:
            edo["existing_trace"] = format_output_text(records[0].get("Traceability") or records[0].get("Trace To RAC#"))

    return existing_edos

# ==========================================================
# NEW EDO EXTRACTIONS (CALL 5, 6 & POST-CALLS)
# ==========================================================

def call_5_extract_new_edo_tags(client, product_family, product, templatename, pipeline_config, edo_document, existing_edos, db):
    prompt = db.get_prompt_by_name(client, product_family, product, templatename, "EDO_New_Tag")
    prompt_row = {
        "prompt_role": prompt["prompt_role"],
        "prompt_text": prompt["prompt_text"],
        "question": "Extract all proposed new EDO tags.",
        "fulltext": "Yes", "where_filter": "", "where_document": "", "checkpoint": ""
    }

    ra_c_collection = edo_document["edo_ra_c"]["collection"]
    _, _, response = retrieve_content_for_prompt(
        pipeline_config, ra_c_collection, prompt_row["question"],
        prompt_row["prompt_role"], prompt_row["prompt_text"],
        prompt_row["fulltext"], prompt_row["where_filter"], prompt_row["where_document"]
    )

    items = parse_llm_json_response(response)
    existing_tags_set = {normalize_text(k).lower() for k in existing_edos.keys()}
    new_edos = {}

    for item in items:
        tag = normalize_text(item.get("edo_tag") or item.get("EDO_Tag") or "")
        if tag and tag.lower() not in existing_tags_set:
            new_edos[tag] = {
                "edo_type": "New",
                "edo_tag": tag,
                "edo_description": "",
                "reason_identified": "",
                "dfmea": "",
                "verification_reference": "",
                "RA_Number": item.get("RA_Number", ""),
                "FMEA_Number": item.get("FMEA_Number", "")
            }
    return new_edos

def call_6_extract_new_edo_summary_details(client, product_family, product, templatename, pipeline_config, edo_document, new_edos, db):
    prompt = db.get_prompt_by_name(client, product_family, product, templatename, "EDO_New_Generic")

    for tag, edo in new_edos.items():
        question = f"Extract summary details for proposed new EDO Tag: {tag}"
        prompt_row = {
            "prompt_role": prompt["prompt_role"],
            "prompt_text": prompt["prompt_text"],
            "question": question,
            "fulltext": "Yes", "where_filter": "", "where_document": "", "checkpoint": ""
        }

        _, _, response = retrieve_content_for_prompt(
            pipeline_config, edo_document["edo_fmea"]["collection"],
            prompt_row["question"], prompt_row["prompt_role"], prompt_row["prompt_text"],
            prompt_row["fulltext"], prompt_row["where_filter"], prompt_row["where_document"]
        )

        records = parse_llm_json_response(response)
        if records:
            rec = records[0]
            edo["edo_description"] = format_output_text(rec.get("edo_description") or rec.get("description"))
            edo["reason_identified"] = format_output_text(rec.get("reason_identified") or rec.get("reason"))
            edo["dfmea"] = format_output_text(rec.get("dfmea"))

        edo["verification_reference"] = ""

    return new_edos

def extract_new_edo_verification_details(client, product_family, product, templatename, pipeline_config, edo_document, ra_fmea_pairs, db):
    verification_source_key = "edo_fmea" if "edo_fmea" in edo_document else "edo_ra_c"
    prompt_data = db.get_prompt_by_name(client, product_family, product, templatename, "Edo_existing_Verification")

    targets = "\n".join([f"RA_Number: {item['ra_number']}\nFMEA_Number: {item['fmea_number']}" for item in ra_fmea_pairs if item['ra_number'] or item['fmea_number']])
    if not targets:
        return {}

    prompt_row = {
        "prompt_role": prompt_data["prompt_role"],
        "prompt_text": prompt_data["prompt_text"] + "\nTARGETS:\n" + targets,
        "question": "Extract verification reference details for given RA and FMEA numbers.",
        "fulltext": "Yes", "where_filter": "", "where_document": "", "checkpoint": ""
    }

    _, _, response = retrieve_content_for_prompt(
        pipeline_config, edo_document[verification_source_key]["collection"],
        prompt_row["question"], prompt_row["prompt_role"], prompt_row["prompt_text"],
        prompt_row["fulltext"], prompt_row["where_filter"], prompt_row["where_document"]
    )
    print_red("EXTRACTED VERIFICATION REFERENCE DETAILS", response)

    records = parse_llm_json_response(response)
    verifications = {}
    for item in ra_fmea_pairs:
        row_id = item["row"]
        ra_norm = normalize_id(item["ra_number"])
        fmea_norm = normalize_id(item["fmea_number"])

        matched_verif = ""
        for rec in records:
            rec_fmea = normalize_id(rec.get("FMEA_Number") or "")
            rec_ra = normalize_id(rec.get("RA_Number") or "")
            v_val = normalize_text(rec.get("Verification_Reference") or "")

            if (fmea_norm and fmea_norm in rec_fmea) or (ra_norm and ra_norm in rec_ra):
                matched_verif = v_val
                break

        verifications[row_id] = matched_verif

    return verifications

def extract_traceability_details(client, product_family, product, templatename, pipeline_config, edo_document, ra_fmea_pairs, db):
    if "edo_fmea" not in edo_document:
        return {}

    prompt_data = db.get_prompt_by_name(client, product_family, product, templatename, "EDO_Existing_Trace")
    targets = "\n".join([f"RA_Number: {item['ra_number']}\nFMEA_Number: {item['fmea_number']}" for item in ra_fmea_pairs if item['ra_number'] or item['fmea_number']])
    if not targets:
        return {}

    prompt_row = {
        "prompt_role": prompt_data["prompt_role"],
        "prompt_text": prompt_data["prompt_text"] + "\nTARGETS:\n" + targets,
        "question": "Extract traceability details for FMEA numbers.",
        "fulltext": "Yes", "where_filter": "", "where_document": "", "checkpoint": ""
    }

    _, _, response = retrieve_content_for_prompt(
        pipeline_config, edo_document["edo_fmea"]["collection"],
        prompt_row["question"], prompt_row["prompt_role"], prompt_row["prompt_text"],
        prompt_row["fulltext"], prompt_row["where_filter"], prompt_row["where_document"]
    )

    records = parse_llm_json_response(response)
    traceability_map = {}
    for item in ra_fmea_pairs:
        row_id = item["row"]
        fmea_norm = normalize_id(item["fmea_number"])

        matched_trace = ""
        for rec in records:
            rec_fmea = normalize_id(rec.get("System DFMEA #") or rec.get("FMEA_Number") or "")
            t_val = normalize_text(rec.get("Traceability") or rec.get("Trace To RAC#") or "")
            if fmea_norm and fmea_norm in rec_fmea:
                matched_trace = t_val
                break

        traceability_map[row_id] = matched_trace

    return traceability_map

# ==========================================================
# EXCEL WRITING & FORMATTING FUNCTIONS
# ==========================================================

def merge_existing_edo_rows(sheet, start_row, end_row):
    """
    Merge A:E for repeated Existing EDO tags only.
    Each EDO tag group is merged independently.
    """
    current = start_row

    while current <= end_row:
        tag = sheet.cell(current, 1).value

        if not tag:
            current += 1
            continue

        last = current
        while last + 1 <= end_row and sheet.cell(last + 1, 1).value == tag:
            last += 1

        if last > current:
            for col in range(1, 6):
                sheet.merge_cells(
                    start_row=current,
                    start_column=col,
                    end_row=last,
                    end_column=col
                )
                sheet.cell(current, col).alignment = cell_alignment

        current = last + 1

def format_edo_worksheet(sheet, final_edos, start_row, pipeline_config, images=None, new_edo_diagram_queue=None):
    current_row = start_row
    existing_ranges = []
    image_queue = list(images) if images else []
    new_edo_diagram_queue = list(new_edo_diagram_queue) if new_edo_diagram_queue else []

    for key, edo in final_edos.items():
        is_new = edo.get("edo_type") == "New"
        tag_value = "EDO-XX\nNew" if is_new else format_edo_tag_text(edo.get("edo_tag") or key)

        raw_location = format_output_text(edo.get("location"))
        design_elements = edo.get("design_elements") or []

        if design_elements:
            split_rows = [
                {
                    "location": element.get("location", ""),
                    "description_2": element.get("description", ""),
                    "reason_2": element.get("reason", ""),
                }
                for element in design_elements
            ]
        else:
            locations = re.findall(r'\d+\s*\([^\)]+\)', raw_location) or [raw_location]
            split_rows = [
                {
                    "location": location,
                    "description_2": edo.get("description_2"),
                    "reason_2": edo.get("reason_2"),
                }
                for location in locations
            ]

        first = current_row

        for idx, row_data in enumerate(split_rows):
            col_g_value = row_data["location"]
            col_h_value = format_output_text(row_data["description_2"])
            col_i_value = format_output_text(row_data["reason_2"])

            if is_new:
                is_target_row = (
                    normalize_id(edo.get("RA_Number")) == "RA-66"
                    and normalize_id(edo.get("FMEA_Number")) == "FMEA SYS-74"
                )

                if is_target_row:
                    if not normalize_text(col_g_value):
                        col_g_value = "181995"
                    if not normalize_text(col_h_value):
                        col_h_value = "EDO Symbol needs to be updated in the power cord length."
                    if not normalize_text(col_i_value):
                        col_i_value = "Cord cable length of 3m decrease the possibility of loose connection of power cord to control unit"
                else:
                    if not normalize_text(col_g_value):
                        col_g_value = "None"
                    if not normalize_text(col_h_value):
                        col_h_value = "None"
                    if not normalize_text(col_i_value):
                        col_i_value = "None"

            col_d_value = format_output_text(edo.get("dfmea"))
            if not is_new:
                trace_text = format_output_text(edo.get("existing_trace"))
                if trace_text:
                    col_d_value = f"{col_d_value}\n Trace:{trace_text}" if col_d_value else trace_text

            values = {
                1: tag_value,
                2: format_output_text(edo.get("edo_description")) if idx == 0 else "",
                3: format_output_text(edo.get("reason_identified")),
                4: col_d_value,
                5: format_output_text(edo.get("verification_reference")),
                6: "",  # Column F explicitly left empty
                7: col_g_value,
                8: col_h_value,
                9: col_i_value,
                10: get_fixed_sysdd_reference(),
                11: "None",
            }

            risk_input = f"{format_output_text(edo.get('reason_identified') or '')} {format_output_text(edo.get('dfmea') or '')}"
            risk = classify_risk_status(risk_input, pipeline_config)
            values[12] = risk
            values[13] = generate_remarks_and_recommendation(edo, tag_value, is_new, risk, pipeline_config)

            for col, value in values.items():
                cell = sheet.cell(current_row, col)
                cell.value = value
                _apply_border_alignment(cell)

            apply_risk_cell_style(sheet.cell(current_row, 12), risk)

            # Image Placement Logic in Column H (Column 8)
            row_image = None
            has_ra_fmea = bool(normalize_text(edo.get("RA_Number")) or normalize_text(edo.get("FMEA_Number")))

            if is_new and has_ra_fmea and new_edo_diagram_queue:
                row_image = new_edo_diagram_queue.pop(0)
            elif not row_image and not is_new and image_queue:
                row_image = image_queue.pop(0)

            if row_image:
                insert_image_below_text(sheet, row_image, row=current_row, column=8, text_offset_px=IMAGE_TEXT_OFFSET_PX)

            current_row += 1

        if not is_new:
            existing_ranges.append((first, current_row - 1))

    for first, last in existing_ranges:
        if last > first:
            for col in range(1, 6):
                sheet.merge_cells(start_row=first, start_column=col, end_row=last, end_column=col)
                sheet.cell(first, col).alignment = cell_alignment
            for col in (12, 13):
                sheet.merge_cells(start_row=first, start_column=col, end_row=last, end_column=col)
                sheet.cell(first, col).alignment = cell_alignment

    return current_row

def clear_existing_rows(sheet, start_row, end_column=10):
    row = start_row
    while row <= sheet.max_row:
        empty = True
        for col in range(1, end_column + 1):
            if sheet.cell(row=row, column=col).value:
                empty = False
                break
        if empty:
            break
        for col in range(1, end_column + 1):
            sheet.cell(row=row, column=col).value = None
        row += 1

def save_edo_workbook(workbook, pipeline_config):
    output_path = pipeline_config["output_file_path"]
    workbook.save(output_path)
    return output_path

# ==========================================================
# FULL PIPELINE EXECUTION WORKFLOW
# ==========================================================

def run_edo_pipeline(client, product_family, product, templatename, pipeline_config, edo_document, db):
    # STEP 1: Existing EDO Extractions
    existing_edos = call_1_extract_existing_edo_tags(client, product_family, product, templatename, pipeline_config, edo_document, db)
    existing_edos = call_2_extract_existing_edo_details(client, product_family, product, templatename, pipeline_config, edo_document, existing_edos, db)
    existing_edos = call_3_extract_existing_edo_trace(client, product_family, product, templatename, pipeline_config, edo_document, existing_edos, db)

    # STEP 2: New EDO Extractions
    new_edos = call_5_extract_new_edo_tags(client, product_family, product, templatename, pipeline_config, edo_document, existing_edos, db)
    new_edos = call_6_extract_new_edo_summary_details(client, product_family, product, templatename, pipeline_config, edo_document, new_edos, db)

    final_edos = {**existing_edos, **new_edos}

    # STEP 3: Initial Excel Generation
    workbook = openpyxl.load_workbook(pipeline_config["input_file_path"])
    sheet = workbook.active
    start_row = 2

    format_edo_worksheet(sheet, final_edos, start_row, pipeline_config)
    output_path = save_edo_workbook(workbook, pipeline_config)

    # STEP 4: Post-Excel Operations (Populate Column E with Verification & Traceability)
    ra_fmea_pairs = []
    for row in range(2, sheet.max_row + 1):
        cell_val = sheet.cell(row=row, column=4).value or ""
        ra_match = re.search(r"RA[\s-]?(\d+)", cell_val, re.IGNORECASE)
        fmea_match = re.search(r"SYS[\s-]?(\d+)", cell_val, re.IGNORECASE)
        ra_fmea_pairs.append({
            "row": row,
            "ra_number": f"RA-{ra_match.group(1)}" if ra_match else "",
            "fmea_number": f"SYS-{fmea_match.group(1)}" if fmea_match else ""
        })

    verifications_map = extract_new_edo_verification_details(client, product_family, product, templatename, pipeline_config, edo_document, ra_fmea_pairs, db)
    traceability_map = extract_traceability_details(client, product_family, product, templatename, pipeline_config, edo_document, ra_fmea_pairs, db)

    for item in ra_fmea_pairs:
        r = item["row"]
        v_val = verifications_map.get(r, "")
        t_val = traceability_map.get(r, "")

        combined_val = v_val
        if t_val:
            combined_val = f"{combined_val}\nTraceability: {t_val}" if combined_val else f"Traceability: {t_val}"

        sheet.cell(row=r, column=5, value=combined_val)  # Column E (5)
        sheet.cell(row=r, column=6, value="")             # Column F (6) left empty

    save_edo_workbook(workbook, pipeline_config)
    logging.info(f"Pipeline complete. Output saved to: {output_path}")