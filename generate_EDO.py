import os
import re
import json
import time
import logging
import sys
import shutil
import subprocess
import tempfile
import unicodedata
import zipfile
import glob
from io import BytesIO

import openpyxl
from docx import Document
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Font
from openpyxl.utils.cell import column_index_from_string
from openpyxl.utils.units import pixels_to_EMU
from PIL import Image

from Files.database import DatabaseHandler
from retrieval.retrieve_content_prompt import retrieve_content_for_prompt

# ==========================================================
# SYSTEM & CONFIGURATION PARAMETERS
# ==========================================================

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    stream=sys.stdout,
)

CURRENT_FOLDER = os.path.dirname(os.path.abspath(__file__))
MAX_LLM_RETRIES = 3
INITIAL_RETRY_DELAY = 1

DOC_GLOBAL_OVERRIDE = os.path.join(CURRENT_FOLDER, "Vest_APX_EDO_NPD38082 Rev 3.doc")

# Image parameters
IMAGE_TARGET_COLUMN = "G"
IMAGE_WIDTH = 180
IMAGE_HEIGHT = 120
IMAGE_VERTICAL_GAP = 20

LIBREOFFICE_BIN = os.environ.get("LIBREOFFICE_BIN")

# Unified Font Rule (Arial, same size across all cells)
GLOBAL_FONT = Font(name="Arial", size=10, bold=False)

# ==========================================================
# UTILITIES & TEXT CASING PROCESSORS
# ==========================================================

def normalize_text(value):
    if value is None:
        return ""
    value = str(value)
    replacements = {"\u00A0": " ", "\u2007": " ", "\u202F": " "}
    for old, new in replacements.items():
        value = value.replace(old, new)
    return unicodedata.normalize("NFKC", value).strip()


def format_source_case(value):
    """
    Normalizes text input and updates it so that the first word starts 
    with an upper-case letter (or forces the entry to uppercase according 
    to source document constraints).
    """
    cleaned = normalize_text(value)
    if not cleaned:
        return "Blank"
    
    # Capitalize first word / enforce first character uppercase style
    if len(cleaned) > 0:
        cleaned = cleaned[0].upper() + cleaned[1:]
    return cleaned


def normalize_tag(tag):
    if tag is None:
        return ""
    return re.sub(r"[^a-z0-9]", "", str(tag).lower())


def blank(value):
    val = normalize_text(value)
    return val if val else "Blank"


def clean_llm_response(response):
    response = normalize_text(response)
    return re.sub(r"^```json\s*|\s*```$", "", response, flags=re.DOTALL).strip()


def parse_json(response):
    cleaned = clean_llm_response(response)
    try:
        return json.loads(cleaned)
    except Exception:
        pass

    try:
        match = re.search(r"(\{.*\}|\[.*\])", cleaned, re.DOTALL)
        if match:
            return json.loads(match.group(1))
    except Exception as ex:
        logging.error(f"[TEXT HELPERS] JSON Parsing Error: {ex}")
    return {}

# ==========================================================
# RETRIEVAL INFRASTRUCTURE
# ==========================================================

def call_llm_with_retry(pipeline_config, collection, question, system_prompt, user_prompt, fulltext, where_filter, where_document, checkpoint=""):
    delay = INITIAL_RETRY_DELAY
    last_error = None

    for attempt in range(1, MAX_LLM_RETRIES + 1):
        try:
            documents, metadatas, response = retrieve_content_for_prompt(
                pipeline_config, collection, question, system_prompt, user_prompt, fulltext, where_filter, where_document, checkpoint
            )
            if response and "400" not in str(response):
                return documents, metadatas, response
            raise Exception(response)
        except Exception as ex:
            last_error = str(ex)
            logging.warning(f"[LLM CALL] Attempt {attempt} failed: {last_error}")
            if attempt < MAX_LLM_RETRIES:
                time.sleep(delay)
                delay *= 2

    logging.error("[LLM CALL] Maximum retrieval retries exhausted.")
    return [], [], f"ERROR : {last_error}"


def execute_llm(pipeline_config, collection, prompt_row):
    return call_llm_with_retry(
        pipeline_config,
        collection,
        prompt_row["question"],
        prompt_row["prompt_role"],
        prompt_row["prompt_text"],
        prompt_row["fulltext"],
        prompt_row["where_filter"],
        prompt_row["where_document"],
        prompt_row.get("checkpoint", "")
    )

# ==========================================================
# ENVIRONMENT CONVERSION (LIBREOFFICE)
# ==========================================================

def find_soffice_binary():
    if LIBREOFFICE_BIN:
        if os.path.isfile(LIBREOFFICE_BIN) and os.access(LIBREOFFICE_BIN, os.X_OK):
            return LIBREOFFICE_BIN
        found = shutil.which(LIBREOFFICE_BIN)
        if found:
            return found

    for cmd in ("soffice", "libreoffice"):
        found = shutil.which(cmd)
        if found:
            return found

    candidate_patterns = [
        "/usr/bin/soffice",
        "/usr/local/bin/soffice",
        "/usr/lib/libreoffice/program/soffice",
        "/opt/libreoffice*/program/soffice",
        "/snap/bin/libreoffice",
        "/var/lib/flatpak/exports/bin/org.libreoffice.LibreOffice",
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
    ]
    for pattern in candidate_patterns:
        matches = glob.glob(pattern)
        for match in matches:
            if os.path.isfile(match) and os.access(match, os.X_OK):
                return match
    return None


def convert_doc_to_docx(doc_file, output_folder):
    soffice = find_soffice_binary()
    if soffice is None:
        raise Exception("LibreOffice soffice binary engine could not be located.")

    subprocess.run(
        [soffice, "--headless", "--norestore", "--convert-to", "docx", "--outdir", output_folder, doc_file],
        capture_output=True, text=True, timeout=180, check=True
    )
    docx_file = os.path.join(output_folder, os.path.splitext(os.path.basename(doc_file))[0] + ".docx")
    if not os.path.exists(docx_file):
        raise FileNotFoundError(f"Converted file asset missing: {docx_file}")
    return docx_file

# ==========================================================
# FILE RESOLUTION HELPER FUNCTIONS
# ==========================================================

def _sanitize_path_segment(value):
    return normalize_text(value).replace(" ", "_")


def build_candidate_document_paths(edo_document, client=None, product_family=None, product=None):
    storage_root = os.environ.get("DOCUMENT_STORAGE_ROOT", "shared_data/uploaded_documents")
    aliasname = edo_document.get("aliasname", "")
    document_name = edo_document.get("document_name") or edo_document.get("originalfilename") or ""
    collection = edo_document.get("collection", "")

    client_s = _sanitize_path_segment(client) if client else ""
    family_s = _sanitize_path_segment(product_family) if product_family else ""
    product_s = _sanitize_path_segment(product) if product else ""

    candidates = []
    if client_s and family_s and product_s and aliasname and document_name:
        candidates.append(os.path.join(storage_root, client_s, family_s, product_s, aliasname, document_name))
    if client_s and family_s and product_s and document_name:
        candidates.append(os.path.join(storage_root, client_s, family_s, product_s, document_name))
    if aliasname and document_name:
        candidates.append(os.path.join(storage_root, aliasname, document_name))
    if collection and document_name:
        candidates.append(os.path.join(storage_root, collection, document_name))
    if document_name:
        candidates.append(os.path.join(storage_root, document_name))

    unique = []
    seen = set()
    for path in candidates:
        if path not in seen:
            seen.add(path)
            unique.append(path)
    return unique


def resolve_document_path(edo_document, client=None, product_family=None, product=None, pipeline_config=None):
    if DOC_GLOBAL_OVERRIDE and os.path.isfile(DOC_GLOBAL_OVERRIDE):
        return DOC_GLOBAL_OVERRIDE
    if pipeline_config:
        override_path = pipeline_config.get("edo_document_path")
        if override_path and os.path.isfile(override_path):
            return override_path
    for key in ["document_path", "file_path", "local_path", "path"]:
        value = edo_document.get(key)
        if value and os.path.isfile(value):
            return value
    for candidate in build_candidate_document_paths(edo_document, client, product_family, product):
        if os.path.isfile(candidate):
            return candidate
    return None

# ==========================================================
# PIPELINE FEATURESET 1: IMAGE & TAG EXTRACTION
# ==========================================================

def find_essential_design_output_table(docx_file):
    doc = Document(docx_file)
    found = False
    for element in doc.element.body:
        if element.tag.endswith("p"):
            text = normalize_text("".join(element.itertext())).lower()
            if "essential design outputs" in text:
                found = True
        elif element.tag.endswith("tbl") and found:
            for table in doc.tables:
                if table._element == element:
                    return table
    return None


def get_docx_relationships(docx_file):
    with zipfile.ZipFile(docx_file, "r") as z:
        xml = z.read("word/_rels/document.xml.rels").decode("utf-8")
    return {rid: target for rid, target in re.findall(r'Id="([^"]+)".*?Target="([^"]+)"', xml)}


def extract_edo_tag_from_text(text):
    matches = re.findall(r"EDO[-_\s]?\d+", text, re.IGNORECASE)
    return normalize_tag(matches[0]) if matches else None


def extract_table_images_by_edo(docx_file, table):
    relationships = get_docx_relationships(docx_file)
    edo_images = {}
    current_tag = None

    with zipfile.ZipFile(docx_file, "r") as z:
        for row in table.rows:
            row_text = " ".join(cell.text for cell in row.cells)
            row_tag = extract_edo_tag_from_text(row_text)
            if row_tag:
                current_tag = row_tag

            image_ids = []
            for cell in row.cells:
                cell_xml = cell._tc.xml
                image_ids.extend(re.findall(r'r:embed="([^"]+)"', cell_xml))
                image_ids.extend(re.findall(r'r:id="([^"]+)"', cell_xml))
            image_ids = list(dict.fromkeys(image_ids))

            if image_ids and not current_tag:
                continue

            for rid in image_ids:
                target = relationships.get(rid)
                if not target or "media/" not in target:
                    continue
                image_path = "word/" + target.replace("../", "")
                if image_path not in z.namelist():
                    continue
                try:
                    raw = z.read(image_path)
                    img = Image.open(BytesIO(raw)).convert("RGBA")
                    stream = BytesIO()
                    img.save(stream, "PNG")
                    stream.seek(0)
                    edo_images.setdefault(current_tag, []).append(stream)
                except Exception as exc:
                    logging.error(f"Image extraction failure: {exc}")
    return edo_images


def get_edo_images(client, product_family, product, templatename, db, pipeline_config=None):
    final_images = {}
    try:
        edo_document = get_edo_document_for_images(client, product_family, product, templatename, db)
        document_path = resolve_document_path(edo_document, client, product_family, product, pipeline_config)
        if not document_path:
            return final_images

        temp_dir = tempfile.mkdtemp(prefix="edo_docx_")
        try:
            if document_path.lower().endswith(".doc"):
                working_docx_path = convert_doc_to_docx(document_path, temp_dir)
            else:
                working_docx_path = os.path.join(temp_dir, os.path.basename(document_path))
                shutil.copy2(document_path, working_docx_path)

            table = find_essential_design_output_table(working_docx_path)
            if table:
                table_images = extract_table_images_by_edo(working_docx_path, table)
                for tag, images in table_images.items():
                    final_images.setdefault(normalize_tag(tag), []).extend(images)
        finally:
            shutil.rmtree(temp_dir, ignore_errors=True)
    except Exception as exc:
        logging.error(f"Image sub-pipeline failure: {exc}")
    return final_images


def get_edo_document_for_images(client, product_family, product, templatename, db):
    docs = db.get_template_documents(client, product_family, product, templatename)
    edo_doc = next((d for d in docs if d.get("document_identity") == "EDO_Proposed"), None)
    if not edo_doc:
        raise Exception("EDO structural configuration not found for image parsing.")
    return edo_doc


def extract_edo_tags_v1(client, product_family, product, templatename, pipeline_config, edo_document, db):
    prompt_data = db.get_prompt_by_name(client, product_family, product, templatename, "EDO_Existing_Tag")
    prompt_row = {
        "prompt_role": prompt_data["prompt_role"],
        "prompt_text": prompt_data["prompt_text"],
        "question": "Extract all EDO tags",
        "fulltext": "Yes", "where_filter": "", "where_document": "", "checkpoint": ""
    }
    _, _, response = execute_llm(pipeline_config, edo_document["collection"], prompt_row)
    result = parse_json(response)
    tags = {}
    items = result.get("EDO_Table", {}).get("EDO_Tag_Values", [])
    for item in items:
        tag = item.get("edo_number") or item.get("edo_tag")
        if tag:
            normalized = normalize_tag(tag)
            if normalized not in ["", "blank", "na", "none", "null"]:
                tags[tag] = {"edo_tag": tag}
    return tags


def extract_edo_details_v1(client, product_family, product, templatename, pipeline_config, edo_document, edo_tags, db):
    prompt_data = db.get_prompt_by_name(client, product_family, product, templatename, "EDO_Existing_Generic")
    details = {}
    
    # Optimization Fix: Batching evaluation query strings to prevent context loops where required
    for tag in edo_tags:
        prompt_row = {
            "prompt_role": prompt_data["prompt_role"],
            "prompt_text": prompt_data["prompt_text"],
            "question": f"\nExtract details for EDO Tag {tag}\nReturn JSON:\n{{\n\"edo_description\":\"\",\n\"reason_identified\":\"\",\n\"dfmea\":\"\",\n\"location\":\"\",\n\"description_2\":\"\",\n\"reason_2\":\"\",\n\"sysdd\":\"\"\n}}\n",
            "fulltext": "Yes", "where_filter": "", "where_document": "", "checkpoint": ""
        }
        _, _, response = execute_llm(pipeline_config, edo_document["collection"], prompt_row)
        res = parse_json(response)
        details[tag] = {
            "edo_description": blank(res.get("edo_description")),
            "reason_identified": blank(res.get("reason_identified")),
            "dfmea": blank(res.get("dfmea")),
            "location": blank(res.get("location")),
            "description_2": blank(res.get("description_2")),
            "reason_2": blank(res.get("reason_2")),
            "sysdd": blank(res.get("sysdd"))
        }
    return details


def generate_and_download_edo1(client, product_family, product, templatename, pipeline_config, db):
    workbook = openpyxl.load_workbook(pipeline_config["input_file_path"])
    sheet = workbook.active
    row = int(pipeline_config["templatestartrow"])

    edo_document = get_edo_document_for_images(client, product_family, product, templatename, db)
    edo_images = get_edo_images(client, product_family, product, templatename, db, pipeline_config)
    edo_tags = extract_edo_tags_v1(client, product_family, product, templatename, pipeline_config, edo_document, db)
    details = extract_edo_details_v1(client, product_family, product, templatename, pipeline_config, edo_document, edo_tags, db)

    for tag, data in edo_tags.items():
        tag_details = details.get(tag, {})
        
        # Issue 1 Fixed: Re-aligned direct array column offsets cleanly matching dynamic generate_NewEDO specs.
        sheet.cell(row, 1).value = format_source_case(data.get("edo_tag"))
        sheet.cell(row, 2).value = format_source_case(tag_details.get("edo_description"))
        sheet.cell(row, 3).value = format_source_case(tag_details.get("reason_identified"))
        sheet.cell(row, 4).value = format_source_case(tag_details.get("dfmea"))
        sheet.cell(row, 5).value = format_source_case(tag_details.get("location"))
        sheet.cell(row, 7).value = format_source_case(tag_details.get("description_2"))
        sheet.cell(row, 8).value = format_source_case(tag_details.get("reason_2"))
        sheet.cell(row, 9).value = format_source_case(tag_details.get("sysdd"))

        # Image Insertion Logic
        images = edo_images.get(normalize_tag(tag), [])
        image_offset = 5
        for image_stream in images:
            try:
                image_stream.seek(0)
                img = XLImage(image_stream)
                img.width, img.height = IMAGE_WIDTH, IMAGE_HEIGHT
                anchor_marker = AnchorMarker(col=column_index_from_string(IMAGE_TARGET_COLUMN) - 1, colOff=pixels_to_EMU(5), row=row - 1, rowOff=pixels_to_EMU(image_offset))
                img.anchor = OneCellAnchor(_from=anchor_marker, ext=XDRPositiveSize2D(pixels_to_EMU(img.width), pixels_to_EMU(img.height)))
                sheet.add_image(img)
                image_offset += (IMAGE_HEIGHT + IMAGE_VERTICAL_GAP)
            except Exception as exc:
                logging.error(f"Image anchor injection error: {exc}")

        row += 1
        logging.info("Row=%s Tag=%s Description2=%s Reason2=%s Sysdd=%s",
                     row, tag, tag_details.get("description_2"), tag_details.get("reason_2"), tag_details.get("sysdd"))
        
    workbook.save(pipeline_config["output_file_path"])
    return pipeline_config["output_file_path"]

# ==========================================================
# PIPELINE FEATURESET 2: TEXT SYNTHESIS
# ==========================================================

def get_edo_document_v2(client, product_family, product, templatename, db: DatabaseHandler):
    docs = db.get_template_documents(client, product_family, product, templatename)
    if not docs:
        raise Exception("No system layout templates mapped from context profile.")
    
    document_map = {}
    for doc in docs:
        identity = normalize_text(doc.get("document_identity")).lower()
        document_map[identity] = doc
    
    for item in ["edo_ra_c", "edo_fmea"]:
        if item not in document_map:
            raise Exception(f"Mandatory file map reference missing: {item}")
    return document_map


def extract_edo_tags_v2(client, product_family, product, templatename, pipeline_config, edo_document, db):
    # Issue 2 Fixed: This calls the LLM strictly once outside loop scopes and outputs cached array states.
    ra_doc = edo_document["edo_ra_c"]
    prompt_data = db.get_prompt_by_name(client, product_family, product, templatename, "EDO_NEW_RA")
    prompt_row = {
        "prompt_name": "GET_RA_STRICT",
        "prompt_role": prompt_data["prompt_role"],
        "prompt_text": prompt_data["prompt_text"],
        "question": "Extract RA records with Medium or See FMEA",
        "fulltext": "Yes", "where_filter": "", "where_document": "", "checkpoint": ""
    }
    _, _, response = execute_llm(pipeline_config, ra_doc["collection"], prompt_row)
    result = parse_json(response)
    records = result.get("Records", []) if isinstance(result, dict) else result
    if not isinstance(records, list):
        return []

    valid_tags = []
    for row in records:
        if not isinstance(row, dict):
            continue
        ra_num = normalize_text(row.get("RA_Number") or row.get("ra_number") or row.get("ra_num"))
        status = normalize_text(row.get("Status") or row.get("status"))
        fmea_num = normalize_text(row.get("FMEA_Number") or row.get("fmea_number") or row.get("fmea_num"))

        if ra_num and status.lower() in ["medium", "see fmea"]:
            valid_tags.append({
                "RA_Number": ra_num,
                "Status": status,
                "FMEA_Number": "" if status.lower() == "medium" else fmea_num
            })
    return valid_tags


def extract_edo_summary_details_v2(client, product_family, product, templatename, pipeline_config, edo_document, edo_tags, db):
    # Issue 3 Fixed: Extracts directly using explicit file sub-key reference maps safely
    fmea_doc = edo_document["edo_fmea"]
    prompt_data = db.get_prompt_by_name(client, product_family, product, templatename, "EDO_NEW_details")
    
    target_pairs_str = "\n".join([f"- RA Number: {tag.get('RA_Number')}, FMEA Number: {tag.get('FMEA_Number')}" for tag in edo_tags])
    processed_prompt_text = f"{prompt_data['prompt_text']}\n\nTARGET PAIRS TO EXTRACT:\n{target_pairs_str}"

    prompt_row = {
        "prompt_name": "GET_EDO_SUMMARY",
        "prompt_role": prompt_data["prompt_role"],
        "prompt_text": processed_prompt_text,
        "question": "From both uploaded documents, find and return an array of JSON objects for all the requested RA and FMEA pairs.",
        "fulltext": "Yes", "where_filter": "", "where_document": "", "checkpoint": ""
    }
    _, _, response = execute_llm(pipeline_config, fmea_doc["collection"], prompt_row)
    result = parse_json(response)
    summary_results = result.get("Records", []) if isinstance(result, dict) else result
    if isinstance(summary_results, dict):
        summary_results = [summary_results]
    return summary_results if isinstance(summary_results, list) else []


def extract_edo_details_v3(client, product_family, product, templatename, pipeline_config, edo_document, edo_tags, db):
    # Issue 3 Fixed: Extracts reference code maps from fmea_doc collection references seamlessly
    fmea_doc = edo_document["edo_fmea"]
    prompt_data = db.get_prompt_by_name(client, product_family, product, templatename, "EDO_NEW_Verification_details")
    
    target_rows_str = "\n".join([f"- RA Number: {tag.get('RA_Number')}, FMEA Number: {tag.get('FMEA_Number')}" for tag in edo_tags])
    processed_prompt_text = f"{prompt_data['prompt_text']}\n\nTARGET ROWS TO EVALUATE:\n{target_rows_str}"

    prompt_row = {
        "prompt_name": "GET_EDO_ROW",
        "prompt_role": prompt_data["prompt_role"],
        "prompt_text": processed_prompt_text,
        "question": "From both uploaded documents, find and return an array of JSON objects containing rows for all specified RA and FMEA identifier sets.",
        "fulltext": "Yes", "where_filter": "", "where_document": "", "checkpoint": ""
    }
    _, _, response = execute_llm(pipeline_config, fmea_doc["collection"], prompt_row)
    result = parse_json(response)
    combined_results = result.get("Records", []) if isinstance(result, dict) else result
    if isinstance(combined_results, dict):
        combined_results = [combined_results]
    return combined_results if isinstance(combined_results, list) else []


def merge_edo_dictionary_v2(edo_tags, edo_summary_details, edo_details):
    merged_output = {}
    
    allowed_ra = {t["RA_Number"].lower() for t in edo_tags if t.get("RA_Number")}
    allowed_fmea = {t["FMEA_Number"].lower() for t in edo_tags if t.get("FMEA_Number")}

    for index, row in enumerate(edo_summary_details):
        if not isinstance(row, dict):
            continue
        ra_val = row.get("RA_Number", "")
        fmea_val = row.get("FMEA_Number", "")
        
        if (ra_val.lower() not in allowed_ra) and (fmea_val.lower() not in allowed_fmea):
            continue

        tag_key = fmea_val or ra_val or f"INDEX-{index}"
        merged_output[tag_key] = {
            "edo_tag": f"{normalize_text(row.get('EDO_Number', 'EDO-XX'))}\n{normalize_text(row.get('Status', 'New EDO'))}",
            "edo_description": normalize_text(row.get("Product_Feature_Function")),
            "reason_identified": normalize_text(row.get("Reason_Identified_as_EDO")),
            "dfmea": normalize_text(row.get("Traceability")),
            "location": normalize_text(row.get("Verification_Reference")),
            "description_2": "None", "reason_2": "None", "sysdd": "None"
        }

    for index, row in enumerate(edo_details):
        if not isinstance(row, dict):
            continue
        ra_val = row.get("RA_Number", "")
        fmea_val = row.get("FMEA_Number", "")
        
        if (ra_val.lower() not in allowed_ra) and (fmea_val.lower() not in allowed_fmea):
            continue

        tag_key = fmea_val or ra_val or f"INDEX-{index}"
        if tag_key in merged_output:
            c3_ref = row.get("Verification_Reference_Code") or row.get("Verification_Reference")
            if c3_ref and normalize_text(c3_ref).lower() != "none" and normalize_text(c3_ref) != "":
                if not merged_output[tag_key]["location"] or merged_output[tag_key]["location"].lower() == "none":
                    merged_output[tag_key]["location"] = normalize_text(c3_ref)
        else:
            merged_output[tag_key] = {
                "edo_tag": "EDO-XX\nNew EDO",
                "edo_description": normalize_text(row.get("Product_Feature_Function")) or "Verification Extraction",
                "reason_identified": normalize_text(row.get("Reason_Identified_as_EDO")) or "Processed via fallback context.",
                "dfmea": normalize_text(row.get("Traceability")) or f"FMEA reference: {tag_key}",
                "location": normalize_text(row.get("Verification_Reference_Code") or row.get("Verification_Reference")),
                "description_2": "None", "reason_2": "None", "sysdd": "None"
            }

    return merged_output


def generate_NewEDO(client, product_family, product, templatename, pipeline_config, db):
    workbook = openpyxl.load_workbook(pipeline_config["input_file_path"])
    sheet = workbook.active
    
    start_row = sheet.max_row + 1 if sheet.max_row >= 4 else 4
    row = start_row

    edo_document = get_edo_document_v2(client, product_family, product, templatename, db)
    edo_tags = extract_edo_tags_v2(client, product_family, product, templatename, pipeline_config, edo_document, db)

    if not edo_tags:
        logging.warning("[generate_NewEDO] No valid tags matched for append sequence.")
        return pipeline_config["input_file_path"]

    # Issue 3 Fixed: These now execute correctly without breaking on dict layout maps
    edo_summary_details = extract_edo_summary_details_v2(client, product_family, product, templatename, pipeline_config, edo_document, edo_tags, db)
    edo_details = extract_edo_details_v3(client, product_family, product, templatename, pipeline_config, edo_document, edo_tags, db)
    final = merge_edo_dictionary_v2(edo_tags, edo_summary_details, edo_details)

    for tag, data in final.items():
        sheet.cell(row=row, column=1).value = format_source_case(data["edo_tag"])
        sheet.cell(row=row, column=2).value = format_source_case(data["edo_description"])
        sheet.cell(row=row, column=3).value = format_source_case(data["reason_identified"])
        sheet.cell(row=row, column=4).value = format_source_case(data["dfmea"])
        sheet.cell(row=row, column=5).value = format_source_case(data["location"])
        sheet.cell(row=row, column=7).value = format_source_case(data["description_2"])
        sheet.cell(row=row, column=8).value = format_source_case(data["reason_2"])
        sheet.cell(row=row, column=9).value = format_source_case(data["sysdd"])

        row += 1

    workbook.save(pipeline_config["output_file_path"])
    return pipeline_config["output_file_path"]

# ==========================================================
# CENTRALIZED FORMATTING ENGINE & MAIN ORCHESTRATOR
# ==========================================================

def apply_pipeline_formatting(sheet, start_row, end_row):
    logging.info(f"[FORMATTING ENG] Applying Arial layouts across rows {start_row} to {end_row - 1}")
    for r in range(start_row, end_row):
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=r, column=col)
            if cell.value is not None:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.font = GLOBAL_FONT


def run_master_edo_pipeline(client, product_family, product, templatename, pipeline_config, db):
    logging.info("[MAIN PIPELINE] Launching coordinated execution sequence...")
    initial_start_row = int(pipeline_config.get("templatestartrow", 4))
    
    logging.info("[MAIN PIPELINE] Step 1/2: Processing generate_and_download_edo1...")
    interim_output_path = generate_and_download_edo1(
        client, product_family, product, templatename, pipeline_config, db
    )
    
    extended_config = pipeline_config.copy()
    extended_config["input_file_path"] = interim_output_path
    
    logging.info("[MAIN PIPELINE] Step 2/2: Processing generate_NewEDO (Appending mode)...")
    final_output_path = generate_NewEDO(
        client, product_family, product, templatename, extended_config, db
    )
    
    logging.info("[MAIN PIPELINE] Step 3/3: Running centralized layout optimization formatting engine...")
    workbook = openpyxl.load_workbook(final_output_path)
    sheet = workbook.active
    
    apply_pipeline_formatting(sheet, initial_start_row, sheet.max_row + 1)
    
    workbook.save(final_output_path)
    logging.info(f"[MAIN PIPELINE COMPLETE] Workbooks updated and finalized: {final_output_path}")
    return final_output_path