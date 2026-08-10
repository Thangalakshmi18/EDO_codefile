import os
import re
import json
import time
import logging
from typing import List, Dict, Any
import sys
import unicodedata
import zipfile
import tempfile
import hashlib
import math
import openpyxl
import fitz

from openpyxl.styles import (
    Alignment,
    Border,
    Side,
    Font
)
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.utils import get_column_letter

from Files.database import DatabaseHandler
from retrieval.retrieve_content_prompt import retrieve_content_for_prompt


# ==========================================================
# GLOBAL CONSTANTS / CONFIGURATION
# ==========================================================

CURRENT_FOLDER = os.path.dirname(os.path.abspath(__file__))

MAX_LLM_RETRIES = 3
INITIAL_RETRY_DELAY = 1

# ---- Image pipeline (from edo_image.py) ----
IMAGE_WIDTH = 180
IMAGE_HEIGHT = 110

# ---- PDF image extraction (ported from generate_ICU_Template_final_working_code.py) ----
PDF_GLOBAL_OVERRIDE = os.path.join(CURRENT_FOLDER, "80369-6.pdf")
IMAGE_OUTPUT_DIR = "extracted_images"
IMAGE_TEXT_OFFSET_PX = 45

# ---- New EDO diagram guaranteed-placement target ----
# The New EDO diagram queue (see extract_new_edo_diagram_queue()) is
# otherwise pure FIFO / content-blind. Per requirement, the row whose
# RA_Number/FMEA_Number match this specific pair is guaranteed to get
# the next available diagram from that queue, reserved for it ahead of
# every other New EDO row - see format_edo_worksheet().
TARGET_IMAGE_RA_NUMBER = "RA-141"
TARGET_IMAGE_FMEA_NUMBER = "FMEA Sys-152"


# ==========================================================
# DOCUMENT RETRIEVAL
# ==========================================================
# Loads EDO Proposed / RA&C / FMEA / PDF documents and the Excel template once, up front, for the whole pipeline to reuse.

DEFAULT_DOCUMENT_SEARCH_ROOTS = [
    "/mnt/documents",
    "/data/documents",
    ".",
]

def get_document_search_roots(pipeline_config=None):
    """
    Builds the ordered list of directories to search for the actual
    source .docx file, per the priority described above. The
    EDO_DOCUMENTS_DIR environment variable is read here (at call time),
    not baked into a module-level constant, so it's picked up correctly
    regardless of when it was set relative to module import.
    """
    roots = []

    if pipeline_config:
        configured_root = pipeline_config.get("documents_root")
        if configured_root:
            roots.append(configured_root)

    env_root = os.environ.get("EDO_DOCUMENTS_DIR")
    if env_root:
        roots.append(env_root)

    for root in DEFAULT_DOCUMENT_SEARCH_ROOTS:
        if root and root not in roots:
            roots.append(root)

    return roots


def find_file_by_name(filename, search_dir="."):
    if not search_dir or not os.path.isdir(search_dir):
        return None
    for root, _, files in os.walk(search_dir):
        if filename in files:
            return os.path.join(root, filename)
    return None


def get_template_documents(
    client,
    product_family,
    product,
    templatename,
    db: DatabaseHandler
):
    docs = db.get_template_documents(
        client,
        product_family,
        product,
        templatename
    )

    if not docs:
        raise Exception("No template documents were found.")

    logging.info("=" * 80)
    logging.info("AVAILABLE TEMPLATE DOCUMENTS")
    logging.info("=" * 80)

    for doc in docs:
        logging.info(
            f"{doc.get('document_identity')}  -->  "
            f"{doc.get('document_name')}"
        )

    logging.info("=" * 80)

    return docs


def get_edo_document(
    client,
    product_family,
    product,
    templatename,
    db: DatabaseHandler
):
    """
    CANONICAL version - updated to resolve EDO_Proposed, EDO_RA_C, EDO_FMEA,
    EDO_PDF_New, and 4 Excel documents from Chroma DB.
    Raises if EDO_Proposed is missing.
    """
    documents = get_template_documents(
        client,
        product_family,
        product,
        templatename,
        db
    )

    if not documents:
        raise Exception("No EDO template documents found.")

    edo_documents = {}
    excel_documents = []

    for document in documents:
        identity = normalize_text(document.get("document_identity")).lower()
        doc_name = normalize_text(document.get("document_name")).lower()
        logging.info(f"AVAILABLE DOCUMENT : {identity}")

        if identity == "edo_proposed":
            edo_documents["edo_proposed"] = document
        elif identity in ["edo_ra_c", "edo_ra&c", "edo_rac", "edo_ra"]:
            edo_documents["edo_ra_c"] = document
        elif identity in ["edo_fmea", "fmea", "system_fmea"]:
            edo_documents["edo_fmea"] = document
        elif identity in ["edo_pdf_new", "edo_pdf-new", "edo pdf new"]:
            edo_documents["edo_pdf_new"] = document
        # Explicit matching for 4 Excel document identities
        elif identity in ["edo_excel_1", "excel_1", "edo_excel1", "excel1"]:
            edo_documents["edo_excel_1"] = document
            excel_documents.append(document)
        elif identity in ["edo_excel_2", "excel_2", "edo_excel2", "excel2"]:
            edo_documents["edo_excel_2"] = document
            excel_documents.append(document)
        elif identity in ["edo_excel_3", "excel_3", "edo_excel3", "excel3"]:
            edo_documents["edo_excel_3"] = document
            excel_documents.append(document)
        elif identity in ["edo_excel_4", "excel_4", "edo_excel4", "excel4"]:
            edo_documents["edo_excel_4"] = document
            excel_documents.append(document)
        # Fallback for dynamic Excel files by name extension or identity keyword
        elif ("excel" in identity or doc_name.endswith((".xlsx", ".xls", ".xlsm", ".csv"))) and len(excel_documents) < 4:
            excel_key = f"edo_excel_{len(excel_documents) + 1}"
            if excel_key not in edo_documents:
                edo_documents[excel_key] = document
                excel_documents.append(document)

    # Retain full list of Excel document collections for downstream iteration
    edo_documents["excel_documents"] = excel_documents

    logging.info("=" * 80)
    logging.info("EDO DOCUMENT CONFIGURATION")
    logging.info("=" * 80)

    for key, doc in edo_documents.items():
        if key == "excel_documents":
            logging.info(f"Loaded Total Excel Documents : {len(doc)}")
            continue
        logging.info(f"{key}")
        logging.info(f"Identity   : {doc.get('document_identity')}")
        logging.info(f"Name       : {doc.get('document_name')}")
        logging.info(f"Collection : {doc.get('collection')}")

    logging.info("=" * 80)

    if "edo_proposed" not in edo_documents:
        raise Exception("Required document EDO_Proposed was not found.")

    if "edo_pdf_new" not in edo_documents:
        logging.warning(
            "Document with identity 'EDO_pdf_new' was not found in the "
            "template's configured documents - New EDO diagram lookup "
            "(Column H) will be skipped for this run."
        )

    if len(excel_documents) < 4:
        logging.warning(
            f"Expected 4 Excel documents, but found {len(excel_documents)}. "
            "Pipeline will proceed with available Chroma DB collections."
        )

    return edo_documents


def initialize_workbook(
    pipeline_config
):
    """
    Opens Excel template.
    """

    workbook = openpyxl.load_workbook(
        pipeline_config["input_file_path"]
    )

    sheet = workbook.active

    return workbook, sheet


# ==========================================================
# IMAGE EXTRACTIONS
# ==========================================================
# DOC image extraction, PDF image extraction, image insertion, and their shared helper functions.

CAPTION_FIGURE_PATTERN = re.compile(r'as\s+below\s*:', re.IGNORECASE)
CAPTION_TABLE_PATTERN = re.compile(r'^\s*table\s+\d+', re.IGNORECASE)
EXCLUDED_LOGO_KEYWORDS = ["hillrom"]

# ---- Last-figure marker ----
# Per requirement, the figure captioned "For 211651, the hose metal ring
# dimension is controlled as below:" is the last real figure that should
# ever be extracted/printed. Duplicates are otherwise allowed as before,
# but once THIS specific figure is reached, nothing after it in the
# document is retrieved - see is_last_figure_caption() /
# extract_docx_figures_only().
LAST_FIGURE_CAPTION_PATTERN = re.compile(
    r'211651.*hose\s+metal\s+ring', re.IGNORECASE | re.DOTALL
)

def _extract_paragraph_texts_and_images(doc_xml):
    """
    Splits word/document.xml into paragraphs (<w:p>...</w:p> blocks),
    extracting each paragraph's plain text and any image relationship
    IDs (r:embed / r:id) referenced within it, in document order.
    Returns a list of {"text": str, "embed_ids": [str, ...]} dicts.
    """
    paragraphs = re.findall(r'<w:p[ >].*?</w:p>', doc_xml, flags=re.DOTALL)
    result = []
    for p in paragraphs:
        texts = re.findall(r'<w:t[^>]*>(.*?)</w:t>', p, flags=re.DOTALL)
        text = "".join(texts).strip()

        # NEW - Word wraps embedded pictures in
        # <mc:AlternateContent><mc:Choice>...</mc:Choice><mc:Fallback>...</mc:Fallback></mc:AlternateContent>.
        # mc:Choice (w:drawing, DrawingML) is what's actually rendered/visible.
        # mc:Fallback is a legacy VML copy (w:pict/v:imagedata) kept only for
        # ancient Word versions - it is NEVER visible in the document, but it
        # still carries its own embed/r:id reference. Without stripping it,
        # that invisible fallback image gets picked up as an extra "figure"
        # that isn't actually in the document as seen. Strip Fallback blocks
        # before scanning for embed ids so only the visible reference remains.
        p_visible = re.sub(r'<mc:Fallback>.*?</mc:Fallback>', '', p, flags=re.DOTALL)

        embed_ids = re.findall(r'embed="([^"]+)"', p_visible) or re.findall(r'r:id="([^"]+)"', p_visible)
        result.append({"text": text, "embed_ids": embed_ids})
    return result


def _resolve_caption_for_image(paragraphs, para_index):
    """
    Looks at the paragraph containing the image, then the paragraph
    immediately before it, then immediately after it, and returns the
    first one with any text - that's treated as the image's caption/
    nearby-text context.
    """
    candidates = []
    if 0 <= para_index < len(paragraphs):
        candidates.append(paragraphs[para_index]["text"])
    if para_index - 1 >= 0:
        candidates.append(paragraphs[para_index - 1]["text"])
    if para_index + 1 < len(paragraphs):
        candidates.append(paragraphs[para_index + 1]["text"])

    for text in candidates:
        if text:
            return text
    return ""


def is_figure_caption(caption_text):
    return bool(CAPTION_FIGURE_PATTERN.search(caption_text or ""))


def is_table_caption(caption_text):
    return bool(CAPTION_TABLE_PATTERN.search(caption_text or ""))


def is_excluded_logo(caption_text):
    lowered = (caption_text or "").lower()
    return any(keyword in lowered for keyword in EXCLUDED_LOGO_KEYWORDS)


def is_last_figure_caption(caption_text):
    """
    True when `caption_text` is the "For 211651, the hose metal ring
    dimension is controlled as below:" caption - the figure that must be
    the LAST one ever extracted. See LAST_FIGURE_CAPTION_PATTERN.
    """
    return bool(LAST_FIGURE_CAPTION_PATTERN.search(caption_text or ""))


def extract_docx_figures_only(docx_file):
    """
    CANONICAL image extractor used by extract_edo_proposed_images().
    See the module-level "FIGURE-ONLY IMAGE EXTRACTION" comment above for
    the full rule set. Every image placement is evaluated independently
    (duplicates allowed/retrieved), and only those with a "Figure N"
    caption nearby are kept - "Table N" captions, uncaptioned images, and
    anything mentioning "Hillrom" are all excluded.

    Per requirement, the figure captioned "For 211651, the hose metal
    ring dimension is controlled as below:" is treated as the LAST real
    figure in the document - once it's been extracted, nothing appearing
    after it is retrieved (see LAST_FIGURE_CAPTION_PATTERN /
    is_last_figure_caption()), even though duplicates before that point
    are still allowed as before.
    """
    logging.info("READING WORD MEDIA IMAGES (FIGURES ONLY)")

    with zipfile.ZipFile(docx_file, "r") as archive:
        try:
            doc_xml = archive.read("word/document.xml").decode("utf-8")
        except KeyError:
            logging.warning("word/document.xml not found - cannot extract figures.")
            return []

        rel_map = {}
        try:
            rel_xml = archive.read("word/_rels/document.xml.rels").decode("utf-8")
            for match in re.finditer(r'Id="([^"]+)"\s+Type="[^"]+/image"\s+Target="([^"]+)"', rel_xml):
                rel_id, target = match.groups()
                rel_map[rel_id] = os.path.basename(target)
        except KeyError:
            logging.warning("word/_rels/document.xml.rels not found - cannot resolve image relationships.")
            return []

        media_bytes = {}
        for file in archive.namelist():
            if file.startswith("word/media/"):
                media_bytes[os.path.basename(file)] = {
                    "extension": os.path.splitext(file)[1],
                    "bytes": archive.read(file)
                }

        paragraphs = _extract_paragraph_texts_and_images(doc_xml)

        figures = []
        skipped_table = 0
        skipped_logo = 0
        skipped_uncaptioned = 0
        reached_last_figure = False

        for para_index, para in enumerate(paragraphs):
            if reached_last_figure:
                break

            for embed_id in para["embed_ids"]:
                filename = rel_map.get(embed_id)
                if not filename or filename not in media_bytes:
                    continue

                caption = _resolve_caption_for_image(paragraphs, para_index)

                if is_excluded_logo(caption):
                    skipped_logo += 1
                    logging.info(f"SKIPPED (Hillrom logo/letterhead) : {filename}")
                    continue

                if is_table_caption(caption):
                    skipped_table += 1
                    logging.info(f"SKIPPED (Table image, not a Figure) : {filename} - caption: {caption!r}")
                    continue

                if not is_figure_caption(caption):
                    skipped_uncaptioned += 1
                    logging.info(f"SKIPPED (no Figure caption found nearby) : {filename} - nearby text: {caption!r}")
                    continue

                media = media_bytes[filename]
                figures.append({
                    "name": filename,
                    "extension": media["extension"],
                    "bytes": media["bytes"],
                    "caption": caption
                })
                logging.info(f"FIGURE FOUND : {filename} - caption: {caption!r}")

                if is_last_figure_caption(caption):
                    reached_last_figure = True
                    logging.info(
                        "LAST FIGURE REACHED - caption "
                        f"{caption!r} matches the 211651 hose metal ring "
                        "marker; no further images later in the document "
                        "will be extracted."
                    )
                    break

    logging.info(
        f"TOTAL FIGURES EXTRACTED : {len(figures)}  "
        f"(skipped {skipped_table} table image(s), {skipped_logo} Hillrom logo/letterhead "
        f"image(s), {skipped_uncaptioned} uncaptioned image(s))"
    )
    return figures


def save_images_to_folder(images, output_folder):
    os.makedirs(output_folder, exist_ok=True)
    saved_paths = []
    for img in images:
        ext = img["extension"] if img["extension"] else ".png"
        out_path = os.path.join(output_folder, img["name"])
        # avoid collisions if names repeat
        base, counter = out_path, 1
        while os.path.exists(out_path):
            root, e = os.path.splitext(base)
            out_path = f"{root}_{counter}{e}"
            counter += 1
        with open(out_path, "wb") as f:
            f.write(img["bytes"])
        saved_paths.append(out_path)
        logging.info(f"Saved : {out_path}")
    return saved_paths


def resolve_edo_source_file(edo_document, document_key, pipeline_config=None):
    """
    SINGLE generic resolver for any EDO source document's local file path.

    Previously this logic existed as two near-identical copies -
    get_edo_proposed_file() (for the "edo_proposed" .docx) and
    get_edo_pdf_new_file() (for the "edo_pdf_new" .pdf) - which was exactly
    the "get proposed file" / "get edo pdf new file" duplication that
    needed to go. Now there is exactly ONE document-file-resolution
    function in the whole pipeline; callers just pass which key they want:

        resolve_edo_source_file(edo_document, "edo_proposed", pipeline_config)
        resolve_edo_source_file(edo_document, "edo_pdf_new", pipeline_config)

    Resolution order (identical to the old behaviour for both callers):
      1. edo_document[document_key] must be present (set once by
         get_edo_document()).
      2. Check common path fields directly on that document's metadata.
      3. Otherwise, search for its document_name/name across
         get_document_search_roots() (pipeline_config["documents_root"],
         EDO_DOCUMENTS_DIR, then the default mount-point fallbacks).

    Every branch that fails to resolve a path logs exactly why, so an
    empty result downstream is traceable back to a specific cause.
    """
    document = edo_document.get(document_key) if edo_document else None
    if not document:
        logging.warning(
            f"{document_key} RESOLUTION FAILED - no document with that "
            "document_identity was configured for this template "
            f"(edo_document has no '{document_key}' key)."
        )
        return None

    for key in ["file_path", "document_path", "path", "local_path", "filepath"]:
        val = document.get(key)
        if not val:
            continue
        if os.path.exists(val):
            logging.info(f"Resolved {document_key} file from document field '{key}' : {val}")
            return val
        logging.warning(
            f"{document_key} document['{key}'] = {val!r} was set but does "
            "not exist on disk in this container - falling back to "
            "filename search."
        )

    doc_name = document.get("document_name") or document.get("name")
    if not doc_name:
        logging.error(
            f"{document_key} RESOLUTION FAILED - the document has no "
            "usable path field (checked file_path/document_path/path/"
            "local_path/filepath) AND no document_name/name field to "
            f"search by. Full document dict: {document}"
        )
        return None

    search_roots = get_document_search_roots(pipeline_config)
    logging.info(f"Searching for {document_key} '{doc_name}' under: {search_roots}")

    for root in search_roots:
        found = find_file_by_name(doc_name, search_dir=root)
        if found:
            logging.info(f"Found {document_key} '{doc_name}' at : {found}")
            return found

    logging.error(
        f"{document_key} RESOLUTION FAILED - could not find '{doc_name}' "
        f"under any of {search_roots}. If this is running in Docker, "
        "that directory needs to actually be mounted into the container "
        "- set EDO_DOCUMENTS_DIR or pipeline_config['documents_root'] to "
        "the correct in-container mount path."
    )
    return None


def extract_edo_proposed_images(edo_document, pipeline_config=None):
    """
    CANONICAL version. Uses the SAME "edo_proposed" document reference
    that extract_edo_tags() already resolves (edo_document["edo_proposed"])
    - no separate document lookup. Resolves its local .docx path via the
    single unified resolve_edo_source_file() call, then extracts its
    embedded FIGURE images (duplicates included, tables/logo excluded)
    via extract_docx_figures_only(). Source documents are always .docx
    here, so no .doc -> .docx conversion step is needed.

    `pipeline_config` is optional and passed straight through to
    resolve_edo_source_file() so pipeline_config["documents_root"] can be
    used to fix Docker deployments where the source .docx files live on
    a mounted volume rather than the container's default CWD - see the
    "DOCUMENT SEARCH ROOTS (Docker fix)" section above.

    """
    try:
        file_path = resolve_edo_source_file(edo_document, "edo_proposed", pipeline_config)
        if not file_path:
            # resolve_edo_source_file() / find_file_by_name() already log
            # exactly which directories were searched and why nothing
            # was found - this just makes the end result unambiguous.
            logging.warning(
                "IMAGE EXTRACTION SKIPPED - the edo_proposed .docx file "
                "itself could not be located (see the search log above)."
            )
            return []

        images = extract_docx_figures_only(file_path)

        if not images:
            # The file WAS found and read - so if this is empty, it's
            # almost certainly the "Figure N" caption filter in
            # extract_docx_figures_only() not matching this document's
            # actual caption style (see its per-image SKIPPED log lines
            # just above this one for exactly why each image was
            # excluded), NOT a missing-file problem.
            logging.warning(
                f"IMAGE EXTRACTION RETURNED 0 FIGURES from a file that WAS "
                f"found and read successfully ({file_path}). This means "
                "every image in the document was excluded by the Figure/"
                "Table/Hillrom caption filter - check the 'SKIPPED (...)' "
                "log lines just above for the reason each one was "
                "excluded, and compare against this document's actual "
                "caption wording (CAPTION_FIGURE_PATTERN currently "
                "requires text starting with 'Figure <number>')."
            )
        else:
            logging.info(
                f"Extracted {len(images)} figure(s) from edo_proposed "
                f"({edo_document.get('edo_proposed', {}).get('document_name')})."
            )

        return images
    except Exception as e:
        logging.error(f"Image extraction failed : {e}")
        return []


def insert_image_below_text(sheet, image, row, column=8, text_offset_px=IMAGE_TEXT_OFFSET_PX):
    """
    CANONICAL - from edo_image.py. Anchors `image` into `sheet` at
    (row, column) - Column H (column=8) by default - positioned
    `text_offset_px` pixels below the top of the cell, so it renders
    underneath whatever text is already in that cell rather than
    overlapping it. Also grows the row height / column width as needed
    so the image isn't clipped.
    """
    if not image:
        return False
    try:
        suffix = image.get("extension") or ".png"
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as temp:
            temp.write(image["bytes"])
            temp_name = temp.name

        excel_image = XLImage(temp_name)
        excel_image.width = IMAGE_WIDTH
        excel_image.height = IMAGE_HEIGHT

        marker = AnchorMarker(
            col=column - 1, colOff=pixels_to_EMU(2),
            row=row - 1, rowOff=pixels_to_EMU(text_offset_px)
        )
        size = XDRPositiveSize2D(cx=pixels_to_EMU(IMAGE_WIDTH), cy=pixels_to_EMU(IMAGE_HEIGHT))
        excel_image.anchor = OneCellAnchor(_from=marker, ext=size)
        sheet.add_image(excel_image)

        required_height = (text_offset_px + IMAGE_HEIGHT) * 0.75
        current_height = sheet.row_dimensions[row].height or 0
        if required_height > current_height:
            sheet.row_dimensions[row].height = required_height

        required_col_width = (IMAGE_WIDTH + 15) * 0.14
        current_width = sheet.column_dimensions['H'].width or 0
        if required_col_width > current_width:
            sheet.column_dimensions['H'].width = required_col_width
        return True
    except Exception as e:
        logging.error(f"Image insert failed at row {row} : {e}")
        return False


def is_valid_rect(rect):
    """True if a PyMuPDF rect is non-empty, finite, and has real size."""
    if rect is None:
        return False
    if rect.is_empty or rect.is_infinite:
        return False
    if rect.width <= 1 or rect.height <= 1:
        return False
    return True


def extract_pdf_page_diagrams(pdf_path, output_dir=IMAGE_OUTPUT_DIR):
    """
    Fresh, content-blind diagram extractor for EDO_pdf_new.

    Does NOT search for RA/FMEA text anywhere. Simply walks every page
    of the PDF, and for any page that contains embedded images or
    vector drawings ("looks like it has a diagram on it"), takes a
    full-page screenshot and appends it to an ORDERED list. No
    identification/matching happens here - that happens later, purely
    by row order, when the Excel is filled.

    Returns (in page order):
        [ {"name": "...", "bytes": b"...", "extension": ".png"}, ... ]
    """
    diagrams = []
    if not pdf_path or not os.path.isfile(pdf_path):
        logging.error(f"PDF PAGE DIAGRAM EXTRACTION FAILED - PDF not found: {pdf_path!r}")
        return diagrams

    os.makedirs(output_dir, exist_ok=True)

    try:
        doc = fitz.open(pdf_path)
    except Exception as e:
        logging.error(f"PDF PAGE DIAGRAM EXTRACTION FAILED - fitz.open() failed: {e!r}")
        return diagrams

    try:
        for page_num in range(len(doc)):
            page = doc[page_num]
            page_no = page_num + 1

            region = _get_drawing_region(page)
            if region is None:
                logging.info(f"PAGE {page_no} SKIPPED - no drawing region detected on this page.")
                continue

            try:
                PAD = 20
                clip_rect = fitz.Rect(
                    max(0, region.x0 - PAD), max(0, region.y0 - PAD),
                    min(page.rect.x1, region.x1 + PAD), min(page.rect.y1, region.y1 + PAD)
                )
                pix = page.get_pixmap(matrix=fitz.Matrix(3, 3), clip=clip_rect, alpha=False)
                filename = f"NEW_EDO_DIAGRAM_p{page_no}.png"
                filepath = os.path.join(output_dir, filename)
                pix.save(filepath)
                with open(filepath, "rb") as f:
                    data = f.read()
                diagrams.append({"name": filename, "bytes": data, "extension": ".png"})
                logging.info(f"PAGE {page_no} CAPTURED (drawing region only, cropped+padded) -> {filepath}")
            except Exception as e:
                logging.warning(f"Failed to rasterize drawing region on page {page_no}: {e}")
    finally:
        doc.close()

    logging.info(f"PDF PAGE DIAGRAM EXTRACTION TOTAL: {len(diagrams)} diagram page(s) captured, in page order.")
    return diagrams


def _get_drawing_region(page, proximity=150, paragraph_char_threshold=80, paragraph_height_threshold=40, band_vertical_tolerance=60):
    """
    Finds the bounding box of just the actual line-art drawing on the
    page - excludes large text paragraphs (DESCRIPTION, RECOMMENDED
    VENDOR, NOTES, title block cells, revision table) and excludes
    page-border/grid lines, but INCLUDES small annotation/callout/
    dimension labels sitting right next to the drawing (e.g.
    "2950±50mm", "NEMA 1-15P BLACK", "SJT 1.00mm x 2") since those are
    part of the drawing, not "other text".

    NOTE ON band_vertical_tolerance:
    A single technical drawing on these sheets is often laid out as
    several DISCONNECTED vector clusters that sit side-by-side in the
    same row (e.g. the main connector/cord diagram on the left, and a
    separate C17-connector + polarity-diagram cluster far to the
    right). Those clusters can be well beyond `proximity` from each
    other horizontally, so the old "only merge what's within
    `proximity` pixels of the main cluster" logic clipped the crop to
    just the left-hand cluster and cut off everything to the right.
    Now, any cluster that overlaps the main cluster's vertical extent
    (within `band_vertical_tolerance`) is merged in regardless of how
    far away it is horizontally, since it's part of the same drawing
    row. `proximity` is still used as a fallback for genuinely nearby
    content that doesn't share the row (e.g. a diagonal callout).

    Returns a fitz.Rect, or None if no drawing content was found.
    """
    page_w, page_h = page.rect.width, page.rect.height
    content_bboxes = []

    # Embedded raster images (if any)
    for img in page.get_images(full=True):
        try:
            xref = img[0]
            for bbox in page.get_image_rects(xref):
                if is_valid_rect(bbox):
                    content_bboxes.append(fitz.Rect(bbox))
        except Exception:
            pass

    # Vector line-art (the CAD drawing itself)
    try:
        raw_rects = []
        for d in page.get_drawings():
            r = fitz.Rect(d["rect"])
            if is_valid_rect(r) and r.width > 3 and r.height > 3:
                raw_rects.append(r)
    except Exception:
        raw_rects = []

    # Drop page-border / zone-grid / table-gridline strokes - these
    # span almost the full page width or height as a thin line, and are
    # not part of the drawing itself.
    raw_rects = [
        r for r in raw_rects
        if not (r.width > 0.85 * page_w and r.height < 0.02 * page_h)
        and not (r.height > 0.85 * page_h and r.width < 0.02 * page_w)
    ]

    used = [False] * len(raw_rects)
    clusters = []
    for i, r in enumerate(raw_rects):
        if used[i]:
            continue
        cluster = fitz.Rect(r)
        changed = True
        while changed:
            changed = False
            for j, r2 in enumerate(raw_rects):
                if used[j]:
                    continue
                expanded = fitz.Rect(cluster.x0 - 40, cluster.y0 - 40, cluster.x1 + 40, cluster.y1 + 40)
                if expanded.intersects(r2):
                    cluster |= r2
                    used[j] = True
                    changed = True
        used[i] = True
        clusters.append(cluster)

    # Keep only clusters that look like a real drawing - not tiny
    # decorations, and not a near-full-page frame.
    clusters = [
        c for c in clusters
        if c.width > 30 and c.height > 30
        and not (c.width > 0.9 * page_w and c.height > 0.9 * page_h)
    ]
    content_bboxes.extend(clusters)

    if not content_bboxes:
        return None

    # Anchor on the single largest piece of drawing content, then pull
    # in every other drawing/image cluster that either:
    #   (a) sits in the same horizontal band/row as the main cluster
    #       (shares vertical extent, within band_vertical_tolerance) -
    #       this is what pulls in far-right content like a separate
    #       C17-connector/polarity-diagram cluster that is part of the
    #       same drawing row but not within `proximity` pixels, or
    #   (b) is simply close to the main cluster (within `proximity`),
    #       same as before, for nearby content that doesn't share a row.
    # This runs iteratively since merging can grow main_bbox's vertical
    # extent, which can then bring a further cluster into the band.
    main_bbox = max(content_bboxes, key=lambda r: r.width * r.height)
    changed = True
    while changed:
        changed = False
        for r in content_bboxes:
            if r is main_bbox or main_bbox.contains(r):
                continue
            vertical_overlap = min(main_bbox.y1, r.y1) - max(main_bbox.y0, r.y0)
            shares_band = vertical_overlap > -band_vertical_tolerance
            expanded = fitz.Rect(main_bbox.x0 - proximity, main_bbox.y0 - proximity, main_bbox.x1 + proximity, main_bbox.y1 + proximity)
            if shares_band or expanded.intersects(r):
                merged = fitz.Rect(main_bbox)
                merged |= r
                if merged != main_bbox:
                    main_bbox = merged
                    changed = True

    # Pull in small nearby labels (dimensions/callouts), but SKIP large
    # paragraph-style text blocks even if they're nearby.
    for block in page.get_text("blocks"):
        bx0, by0, bx1, by1, text = block[0], block[1], block[2], block[3], block[4]
        rect = fitz.Rect(bx0, by0, bx1, by1)
        is_paragraph = (
            len(text.strip()) > paragraph_char_threshold
            or rect.height > paragraph_height_threshold
        )
        if is_paragraph:
            continue
        expanded = fitz.Rect(main_bbox.x0 - proximity, main_bbox.y0 - proximity, main_bbox.x1 + proximity, main_bbox.y1 + proximity)
        if expanded.intersects(rect):
            main_bbox |= rect

    return main_bbox


def extract_new_edo_diagram_queue(edo_document, pipeline_config=None, output_dir=IMAGE_OUTPUT_DIR):
    """
    Resolves the EDO_pdf_new PDF and returns the ordered, content-blind
    diagram list from extract_pdf_page_diagrams(). Replaces every
    previous RA/FMEA-matching diagram function.
    """
    pdf_path = resolve_edo_source_file(edo_document, "edo_pdf_new", pipeline_config)
    if not pdf_path:
        logging.warning("NEW EDO DIAGRAM QUEUE SKIPPED - EDO_pdf_new PDF could not be resolved.")
        return []
    return extract_pdf_page_diagrams(pdf_path, output_dir=output_dir)


# ==========================================================
# PROMPT EXECUTION
# ==========================================================
# LLM invocation, retry logic, response/JSON cleanup, parsing and validation shared by every extraction stage below.

def normalize_text(value):
    if value is None:
        return ""

    value = str(value)

    value = value.replace("\u00A0", " ")
    value = value.replace("\u2007", " ")
    value = value.replace("\u202F", " ")

    value = unicodedata.normalize(
        "NFKC",
        value
    )

    return value.strip()


def clean_llm_response(response):

    response = normalize_text(response)

    return re.sub(
        r"^```json\s*|\s*```$",
        "",
        response,
        flags=re.DOTALL
    ).strip()


def _salvage_truncated_json_array(text):
    """
    Recovers as many complete top-level JSON objects as possible from a
    '[' ... ']' array whose text was cut off partway through (typically
    because the LLM hit its max_tokens limit mid-response on a large
    table). Walks the text tracking brace depth and string/escape state,
    so it isn't fooled by braces or brackets that appear inside quoted
    string values. Every '{ ... }' block that closes cleanly before the
    cutoff is parsed and kept; the dangling partial object at the very
    end (the one that got cut off) is simply dropped instead of causing
    the whole batch to be discarded.
    """
    start = text.find("[")
    if start == -1:
        return []

    objects = []
    depth = 0
    obj_start = None
    in_string = False
    escape = False

    for i in range(start, len(text)):
        ch = text[i]

        if in_string:
            if escape:
                escape = False
            elif ch == "\\":
                escape = True
            elif ch == '"':
                in_string = False
            continue

        if ch == '"':
            in_string = True
            continue

        if ch == "{":
            if depth == 0:
                obj_start = i
            depth += 1
        elif ch == "}":
            depth -= 1
            if depth == 0 and obj_start is not None:
                candidate = text[obj_start:i + 1]
                try:
                    parsed_obj = json.loads(candidate)
                    if isinstance(parsed_obj, dict):
                        objects.append(parsed_obj)
                except Exception:
                    pass
                obj_start = None

    return objects


def parse_json(response):
    try:

        cleaned = clean_llm_response(response)

        logging.info("========== CLEANED JSON ==========")
        logging.info(cleaned)

        return json.loads(cleaned)

    except Exception as e:

        logging.error(f"JSON Parse Error : {e}")

        # FALLBACK 1 - the direct parse failed, most likely because the LLM
        # wrapped the JSON in extra prose/text beyond a plain ```json fence
        # (clean_llm_response only strips a leading/trailing fence, not
        # surrounding text). Try to salvage the first {...} or [...] block
        # in the response before giving up, so a well-formed JSON payload
        # isn't silently discarded just because of text around it.
        try:
            match = re.search(r"(\{.*\}|\[.*\])", normalize_text(response), re.DOTALL)
            if match:
                salvaged = json.loads(match.group(1))
                logging.warning(
                    "JSON Parse Error recovered via fallback extraction - "
                    "the LLM response had extra text around the JSON block."
                )
                return salvaged
        except Exception as fallback_error:
            logging.error(f"JSON fallback extraction also failed: {fallback_error}")

        # FALLBACK 2 - the response is a JSON array that was cut off mid-way
        # (e.g. 'Unterminated string' / 'Expecting value' errors partway
        # through the text) - almost always caused by hitting max_tokens on
        # a large table. Salvage every complete object that DID come through
        # before the cutoff rather than discarding the whole response.
        try:
            salvaged_objects = _salvage_truncated_json_array(normalize_text(response))
            if salvaged_objects:
                logging.warning(
                    f"JSON response appears TRUNCATED (likely hit max_tokens) - "
                    f"recovered {len(salvaged_objects)} complete object(s) before the "
                    f"cutoff point out of a presumably larger table. Consider raising "
                    f"max_tokens for this extraction or chunking the source table into "
                    f"smaller batches to avoid losing the remaining rows."
                )
                return salvaged_objects
        except Exception as salvage_error:
            logging.error(f"Truncated-array salvage also failed: {salvage_error}")

        logging.error(response)

        return {}


def blank(value):
    """
    NOTE: previously substituted the literal text "Blank" for empty
    values. Per requirement, no placeholder text should ever be written
    to the output Excel - a missing value should simply be an empty
    cell. This now just normalizes the text and returns "" for anything
    empty, instead of inserting "Blank".

    BUGFIX: normalize_text() alone does NOT catch the case where the
    LLM itself literally answers with the placeholder word "Blank" (or
    "None") as the VALUE of a design_elements field (location/
    description/reason/sysdd) - that text passed straight through
    unchanged. Two problems resulted:
      1. The literal word "Blank"/"None" got printed into the Excel
         cell instead of a real value or an empty cell.
      2. In format_edo_worksheet(), the split-row forward-fill logic
         (`if description: last_description = description else:
         description = last_description`) only treats a field as
         "missing" when it's falsy/empty - a non-empty string like
         "Blank" is truthy, so forward-fill never kicked in for that
         split row, and the next split row(s) of the same EDO tag kept
         showing the literal "Blank" text instead of inheriting the
         previous real description/reason (exactly the symptom seen in
         Excel: split row 1 shows real text, split row 2+ shows
         "Blank").
    Treating "none"/"blank" (case-insensitive) as empty here - the
    same convention already used by get_llm_value() elsewhere in this
    file - fixes both: the placeholder is never written to Excel, and
    forward-fill correctly carries the last real value down to every
    split row that has no genuine value of its own.
    """
    text = normalize_text(value)
    if text.lower() in ("none", "blank"):
        return ""
    return text


def call_llm(prompt, pipeline_config):
    try:
        llm = pipeline_config["llm"]
        response = llm.generate(
            prompt,
            context="",
            question="risk classification",
            temperature=pipeline_config["temperature"],
            max_tokens=pipeline_config["max_tokens"]
        )
        logging.info(f"LLM raw response: {response!r}")
        return response
    except Exception as e:
        # NOTE: previously returned the literal string "No" here. That
        # sentinel was indistinguishable from a genuine (wrong) LLM
        # answer, so callers such as generate_remarks_and_recommendation()
        # would treat a *failed* call as a valid "No"/short response and
        # write it straight into Column M. Returning "" instead lets
        # every caller's existing emptiness checks correctly detect the
        # failure and fall back cleanly.
        logging.error(f"LLM call failed: {e}")
        return ""


def is_meaningful_llm_text(response):
    """
    True only when `response` is real generated content - i.e. not
    empty, and not one of the placeholder/failure words an LLM (or a
    failed call_llm()) might return instead of an actual answer.
    """
    if not response:
        return False

    text = normalize_text(response).strip().upper()

    if not text:
        return False

    if text in ("NONE", "NO", "N/A", "NA", "NULL", "-"):
        return False

    return True


def clean_response(response):
    """
    Strips whitespace/code-fences from a plain-text LLM response
    (used by classify_risk_status - it only ever expects one bare word
    back, e.g. "High").
    """
    return clean_llm_response(response)


def execute_llm(
    pipeline_config,
    collection,
    prompt_row
):
    """
    Generic LLM wrapper used by every prompt passed via positional arguments.
    """
    return retrieve_content_for_prompt(
        pipeline_config,
        collection,
        prompt_row["question"],
        prompt_row["prompt_role"],
        prompt_row["prompt_text"],
        prompt_row["fulltext"],
        prompt_row["where_filter"],
        prompt_row["where_document"],
        prompt_row.get("checkpoint", ""),
        max_results=prompt_row.get("max_results"),
    )


def execute_llm_retry(
    pipeline_config,
    collection,
    prompt_row
):
    """
    Executes the LLM with retry logic.
    """

    last_exception = None

    for attempt in range(MAX_LLM_RETRIES):

        try:

            docs, metadata, response = execute_llm(
                pipeline_config,
                collection,
                prompt_row
            )

            if response:
                return docs, metadata, response

        except Exception as ex:

            last_exception = ex

            logging.exception(
    f"LLM Retry {attempt+1}/{MAX_LLM_RETRIES} failed: {ex}"
)

            time.sleep(
                INITIAL_RETRY_DELAY * (attempt + 1)
            )

    raise Exception(
        f"LLM failed after retries : {last_exception}"
    )


def get_prompt(
    client,
    product_family,
    product,
    templatename,
    prompt_name,
    db: DatabaseHandler
):
    prompt = db.get_prompt_by_name(
        client,
        product_family,
        product,
        templatename,
        prompt_name
    )

    if not prompt:
        raise Exception(f"Prompt '{prompt_name}' was not found.")

    logging.info(f"Loaded Prompt : {prompt_name}")
    return prompt


def build_prompt_row(
    prompt,
    question,
    fulltext="Yes",
    where_filter="",
    where_document="",
    checkpoint=""
):
    return {
        "prompt_role": prompt["prompt_role"],
        "prompt_text": prompt["prompt_text"],
        "question": question,
        "fulltext": fulltext,
        "where_filter": where_filter,
        "where_document": where_document,
        "checkpoint": checkpoint
    }


def execute_prompt(
    pipeline_config,
    collection,
    prompt
):
    _, _, response = execute_llm_retry(
        pipeline_config,
        collection,
        prompt
    )

    logging.info("=" * 80)
    logging.info("RAW LLM RESPONSE")
    logging.info("=" * 80)
    logging.info(response)

    return parse_json(response)


def deep_extract_records(data):
    """
    Recursive lookup utility to unwrap fluctuating JSON parent containers.
    Merged version: union of the container keys recognised by both source
    files (edo_existing_final.py's set plus generate_EDO_template_copy.py's
    extra "New_EDOs" / "New_Edos" / "new_edos" keys), so this single
    function correctly unwraps both existing-EDO and new-EDO responses.
    """
    if isinstance(data, list):
        return [item for item in data if isinstance(item, dict)]

    if isinstance(data, dict):
        for key in [
            "Records", "records",
            "New_EDOs", "New_Edos", "new_edos",
            "EDO_Table", "EDO_Tag_Values", "Verification_Details"
        ]:
            if key in data and isinstance(data[key], list):
                return [item for item in data[key] if isinstance(item, dict)]

        if any(k in data for k in ["Product_Feature_Function", "RA_Number", "FMEA_Number", "Traceability", "System DFMEA #", "Trace To RAC#"]):
            return [data]

        for val in data.values():
            if isinstance(val, list) and len(val) > 0 and isinstance(val[0], dict):
                return val
            elif isinstance(val, dict):
                res = deep_extract_records(val)
                if res:
                    return res
    return []


# ==========================================================
# EXISTING EDO PIPELINE
# ==========================================================
# CALL 1: Extract Existing EDO Tags -> Extract RA Number -> Extract FMEA Number -> CALL 2: Extract Remaining Existing EDO Details -> CALL 3: Extract Existing EDO Trace Details.

def create_empty_edo():
    """
    CANONICAL version - from edo_existing_final.py.
    Superset of the copy-file version: also carries FMEA_Number so that
    Stage 3B (verification-reference matching by RA/FMEA number) has
    something to match against.

    NOTE: defaults are empty strings, not the literal text "Blank" -
    an unfound value should render as a truly empty cell in the output
    Excel, per requirement.
    """
    return {
        "edo_type": "Existing",
        "edo_tag": "",
        "ra_number": "",
        "FMEA_Number": "",
        "edo_description": "",
        "reason_identified": "",
        "dfmea": "",
        "verification_reference": "",
        "existing_trace": "",
        "location": "",
        "description_2": "",
        "reason_2": "",
        "sysdd": ""
    }


def extract_edo_tags(
    client,
    product_family,
    product,
    templatename,
    pipeline_config,
    edo_document,
    db: DatabaseHandler
):
    """
    CANONICAL - requested function #1, taken from edo_existing_final.py.
    """
    logging.info("=" * 80)
    logging.info("STAGE 3: EXTRACTING EXISTING EDO TAGS")
    logging.info("=" * 80)

    prompt = get_prompt(
        client,
        product_family,
        product,
        templatename,
        "EDO_Existing_Tag",
        db
    )

    prompt_row = build_prompt_row(
        prompt,
        "Extract all Known Active Design Output Tracking Numbers."
    )

    result = execute_prompt(
        pipeline_config,
        edo_document["edo_proposed"]["collection"],
        prompt_row
    )

    items = deep_extract_records(result)
    logging.info(f"Unconditional Record Extraction count: {len(items)}")

    tags = {}
    for item in items:
        if not isinstance(item, dict):
            continue

        tag = (
            item.get("edo_number")
            or item.get("edo_tag")
            or item.get("EDO Number")
            or item.get("EDO_Tag")
            or item.get("EDO")
            or ""
        )
        tag = normalize_text(tag)

        if tag == "" or tag.lower() == "blank" or tag in tags:
            continue

        tags[tag] = create_empty_edo()
        tags[tag]["edo_type"] = "Existing"
        tags[tag]["edo_tag"] = tag

    return tags


def validate_existing_tags(tags):
    validated = {}
    for key, value in tags.items():
        tag = normalize_text(value.get("edo_tag"))
        if tag == "" or tag.lower() == "blank":
            continue
        validated[tag] = value
    return validated


def extract_ra_fmea_from_text(text):
    """
    Pulls RA_Number / FMEA_Number back out of the column D (dfmea) narrative
    that was already extracted for an existing EDO in extract_edo_details().
    This is the "previous edo dictionary" data used for matching in Stage
    3B - reused instead of asking the LLM for tags a second time.

    The RA&C / FMEA documents identify records with bare tokens like
    "RA-180" and "SYS-147" (sometimes written "FMEA SYS-147") - there is
    no literal word "Number" next to them, so we match the token pattern
    directly rather than looking for a "RA Number:" label.
    """
    text = normalize_text(text)

    ra_match = re.search(r"RA[\s-]?(\d+)", text, re.IGNORECASE)
    fmea_match = re.search(r"SYS[\s-]?(\d+)", text, re.IGNORECASE)

    ra_number = f"RA-{ra_match.group(1)}" if ra_match else ""
    fmea_number = f"SYS-{fmea_match.group(1)}" if fmea_match else ""

    return ra_number, fmea_number


def normalize_id(value):
    """
    Canonicalizes an RA/FMEA identifier for comparison purposes:
    uppercase, single spaces, trimmed. Used so that matching between the
    column-D-derived identifiers and whatever format Stage 3B's
    verification prompt returns doesn't fail on case/whitespace noise.
    """
    value = normalize_text(value).upper()
    return re.sub(r"\s+", " ", value).strip()


def extract_edo_details(
    client,
    product_family,
    product,
    templatename,
    pipeline_config,
    edo_document,
    existing_edos,
    db: DatabaseHandler
):
    """
    CANONICAL version - from edo_existing_final.py.
    Comprehensive attribute hydration: extracts description/reason lists,
    dfmea narrative, location, sysdd, and (crucially) RA_Number /
    FMEA_Number - needed downstream by Stage 3B verification matching.

    UPDATED to match the EDO_Existing_Generic prompt's nested JSON shape
    (EDO_Table -> existing_edo_data[] -> design_elements[]):
      - Tag matching now also recognises "edo_no" - the key this prompt
        actually returns. The old code only checked edo_number/edo_tag/
        "EDO Number"/"EDO_Tag", so `result` was NEVER populated for this
        prompt's output and every field silently stayed blank - this is
        why nothing was printing in Excel.
      - "design_elements" (a LIST of per-location dicts) is walked and
        normalized into {"location", "description", "reason", "sysdd",
        "image"} entries, stored as a list on edo["design_elements"] for
        any downstream code (Excel writer / image placement) that wants
        the full per-location breakdown.
      - The single-value fields the rest of the pipeline reads today
        (edo_description / reason_identified / location / sysdd /
        description_2 / reason_2) are populated from the new top-level
        fields, backfilled from the design_elements list where the old
        schema had no top-level equivalent (e.g. location/sysdd only
        ever existed per design element, never at the top level).
    """
    logging.info("=" * 80)
    logging.info("STAGE 3: ATTRIBUTE HYDRATION (EXISTING EDO)")
    logging.info("=" * 80)

    prompt = get_prompt(
        client,
        product_family,
        product,
        templatename,
        "EDO_Existing_Generic",
        db
    )

    def normalize_key(key):
        key = normalize_text(key).lower()
        key = key.replace("&", "and")
        key = re.sub(r"[^a-z0-9]+", "_", key)
        key = re.sub(r"_+", "_", key)
        return key.strip("_")

    def first_value(record, keywords):
        """
        Return first non-empty value whose normalized key contains
        any keyword as a substring.
        """
        normalized_record = {normalize_key(k): v for k, v in record.items()}
        for kw in keywords:
            for nk, v in normalized_record.items():
                if kw in nk or nk in kw:
                    if normalize_text(v):
                        return normalize_text(v)
        return ""

    for edo_tag in existing_edos.keys():

        logging.info(f"Hydrating {edo_tag}")

        question = (
            f"Extract the complete row details for the EDO tag requested by the user. Requested EDO Tag: {edo_tag}  Find the row where the EDO Tag exactly matches the requested EDO Ta. Return all available column values for that row, including EDO number/tag, description, reason identified as EDO, RA&C and/or Sys-DFMEA Trace,EDO Location, EDO Description, reason identified as EDO  If the requested EDO tag is not found, return none"
        )

        prompt_row = build_prompt_row(
            prompt,
            question
        )

        result = {}

        # 1. Fetch data
        for attempt in range(3):
            raw_result = execute_prompt(
                pipeline_config,
                edo_document["edo_proposed"]["collection"],
                prompt_row
            )
            records = deep_extract_records(raw_result)

            for record in records:
                tag = normalize_text(
                    record.get("edo_no")
                    or record.get("edo_number")
                    or record.get("edo_tag")
                    or record.get("EDO Number")
                    or record.get("EDO_Tag")
                ).lower()

                if tag == edo_tag.lower():
                    result = record
                    break

            # 2. Debug log
            logging.info(f"DEBUG: Hydration data for {edo_tag}: {json.dumps(result, indent=2)}")

            # Always accept whatever we found to avoid "Blank"
            if result:
                break
            time.sleep(INITIAL_RETRY_DELAY)

        normalized = {normalize_key(k): v for k, v in result.items()}
        edo = existing_edos[edo_tag]

        # ---- Design elements (new nested structure) ----
        # "design_elements" is a LIST of per-location dicts under the new
        # EDO_Existing_Generic prompt shape - normalize every entry so
        # downstream code has a clean, predictable structure to work
        # from, instead of only ever seeing one location's worth of data.
        raw_design_elements = result.get("design_elements") or result.get("Design_Elements") or []
        if not isinstance(raw_design_elements, list):
            raw_design_elements = []

        design_elements = []
        for element in raw_design_elements:
            if not isinstance(element, dict):
                continue
            design_elements.append({
                "location": blank(
                    element.get("edo_location") or element.get("EDO_Location") or element.get("location")
                ),
                "description": blank(
                    element.get("edo_description") or element.get("EDO_Description") or element.get("description")
                ),
                "reason": blank(
                    element.get("reason identified as edo")
                    or element.get("reason_identified_as_edo")
                    or element.get("reason")
                ),
                "sysdd": blank(
                    element.get("SysDD or HDD Reference") or element.get("sysdd") or element.get("SysDD")
                ),
                "image": blank(element.get("image")),
            })

        edo["design_elements"] = design_elements

        # ---- Top-level description / reason ----
        # The new prompt returns ONE top-level "edo_description" /
        # "reason identified as edo" pair for the EDO itself - fall back
        # to the first design element only if the top-level fields are
        # somehow missing.
        top_description = first_value(result, ["edo_description"])
        top_reason = first_value(result, ["reason_identified_as_edo", "reason"])

        edo["edo_description"] = top_description or (design_elements[0]["description"] if design_elements else "")
        edo["reason_identified"] = top_reason or (design_elements[0]["reason"] if design_elements else "")

        # description_2 / reason_2 - backfilled from the SECOND design
        # element (if any), for backward compatibility with any code
        # still reading these two single-value fields.
        edo["description_2"] = design_elements[1]["description"] if len(design_elements) > 1 else ""
        edo["reason_2"] = design_elements[1]["reason"] if len(design_elements) > 1 else ""

        # ---- RA&C / Sys-DFMEA trace (Column D) ----
        # "RA and FMEA no" holds the real RA&C/System FMEA reference text
        # under the new schema - "Trace" is usually just "Blank"/empty,
        # so prefer the former and only fall back to the latter.
        ra_and_fmea = first_value(result, ["ra_and_fmea_no", "ra_and_fmea", "ra_and"])
        trace_narrative = first_value(result, ["trace"])
        edo["dfmea"] = ra_and_fmea or trace_narrative

        # ---- Location / SysDD ----
        # Neither ever existed at the top level under the new schema -
        # both only ever live per design element - so pull them from the
        # first design element.
        edo["location"] = design_elements[0]["location"] if design_elements else first_value(result, ["location"])
        edo["sysdd"] = design_elements[0]["sysdd"] if design_elements else first_value(result, ["sysdd", "hardware", "design_reference"])

        ra_number = first_value(result, ["ra_number"])
        fmea_number = first_value(result, ["fmea_number"])

        if not ra_number or not fmea_number:
            p_ra, p_fmea = extract_ra_fmea_from_text(edo["dfmea"])
            ra_number = ra_number or p_ra
            fmea_number = fmea_number or p_fmea

        edo["ra_number"] = ra_number or ""
        edo["FMEA_Number"] = fmea_number or ""
        edo["verification_reference"] = ""

    return existing_edos


def extract_existing_edo_trace_details(
    client,
    product_family,
    product,
    templatename,
    pipeline_config,
    edo_document,
    existing_edos,
    db: DatabaseHandler
):
    """
    ADDED - dedicated trace extraction for EXISTING EDOs (CALL 3): its own prompt
    ("EDO_Existing_Trace"), queried against the edo_fmea collection,
    matched by RA/FMEA number that was pulled from the column D (dfmea)
    data in extract_edo_details(). The result is written to Column D,
    below the existing RA/FMEA data, in the same row (see
    apply_existing_edo_trace() and the Column D value construction in
    format_edo_worksheet()).

    CHANGED: now issues ONE LLM call PER EDO instead of a single batched
    call covering all EDOs at once. The batched version could get cut
    short mid-way through the response (partial results - some EDOs'
    tags missing or truncated), because all RA/FMEA targets shared one
    fixed max_tokens budget. Calling per-EDO gives each EDO its own full
    output budget and means one bad/oversized row can't blank out the
    rest of the batch.
    """
    logging.info("=" * 80)
    logging.info("STAGE 3C: EXISTING EDO - TRACE EXTRACTION (per-EDO calls)")
    logging.info("=" * 80)

    if "edo_fmea" not in edo_document:
        raise Exception(
            "EDO_FMEA document/collection not configured - cannot run "
            "trace extraction."
        )

    prompt_data = get_prompt(
        client,
        product_family,
        product,
        templatename,
        "EDO_Existing_Trace",
        db
    )

    results = {}

    for edo_tag, edo in existing_edos.items():
        ra_number = edo.get("ra_number")
        fmea_number = edo.get("FMEA_Number")

        has_ra = ra_number not in (None, "", "Blank")
        has_fmea = fmea_number not in (None, "", "Blank")

        if not has_ra and not has_fmea:
            logging.info(
                f"extract_existing_edo_trace_details: {edo_tag} has no "
                "RA/FMEA number - skipping (empty traces)."
            )
            results[edo_tag] = {"traces": []}
            continue

        target = (
            f"EDO : {edo_tag}\n"
            f"RA_Number : {ra_number}\n"
            f"FMEA_Number : {fmea_number}"
        )

        prompt_row = {
            "prompt_role": prompt_data["prompt_role"],
            "prompt_text": prompt_data["prompt_text"] + "\nTARGET:\n" + target,
            "question": f"Fetch all trace records for {edo_tag} (FMEA_Number: {fmea_number})",
            "fulltext": "Yes",
            "where_filter": "",
            "where_document": "",
            "checkpoint": f"Fetch all trace records for {edo_tag} (FMEA_Number: {fmea_number})"
        }

        try:
            _, _, response = execute_llm_retry(
                pipeline_config,
                edo_document["edo_fmea"]["collection"],
                prompt_row
            )
        except Exception as call_error:
            logging.warning(
                f"extract_existing_edo_trace_details: LLM call failed for "
                f"{edo_tag} after retries - leaving traces empty. "
                f"Reason: {call_error}"
            )
            results[edo_tag] = {"traces": []}
            continue

        edo_result = _parse_single_edo_trace_response(edo_tag, response)
        results[edo_tag] = edo_result

        logging.info(
            f"extract_existing_edo_trace_details: {edo_tag} -> "
            f"{len(edo_result.get('traces', []))} trace(s) extracted."
        )

    logging.info("========== PER-EDO TRACE RESULTS (COMBINED) ==========")
    logging.info(json.dumps(results, indent=2))

    return results


def parse_custom_fmea_format(response):
    """
    Parses custom text format containing:
    Fmea_number: SYS-147  
    [Document_number: NPD37819  
    {Trace: MRS CU FMEA-391, Trace: MRS CU FMEA-463}]

    Returns a list of formatted trace strings, e.g.:
    [
      "NPD37819: MRS CU FMEA-391, MRS CU FMEA-463",
      "NPD36569: MRS Software FMEA-422, MRS Software FMEA-423"
    ]
    """
    text = normalize_text(response)
    traces = []

    # Find each block starting with [Document_number: ... {Trace: ...}]
    # NOTE: doc_num uses a non-greedy `.*?` (bounded by the next "{")
    # instead of a `[^\]\n]+` char class, so it can't accidentally
    # swallow past a brace on multi-line blocks.
    doc_blocks = re.findall(
        r"\[\s*Document_number\s*:\s*(.*?)\s*\{(.*?)\}\s*\]?",
        text,
        re.IGNORECASE | re.DOTALL
    )

    for doc_num, trace_content in doc_blocks:
        doc_num = doc_num.strip()

        # Extract all individual tag values following "Trace:".
        # FIXED: the previous pattern used a character class
        # [^,Trace:\}] which - combined with re.IGNORECASE - excludes
        # any occurrence of the individual letters T/r/a/c/e (any
        # case), NOT the literal substring "Trace:". That truncated
        # every value at its first such letter, e.g. "MRS CU FMEA-391"
        # got cut to just "M" (stopped at the "R"). Instead, capture
        # everything up to the NEXT "Trace:" marker or the end of the
        # block, which correctly preserves the full value.
        raw_tags = re.findall(
            r"Trace\s*:\s*(.*?)(?=,\s*Trace\s*:|$)",
            trace_content,
            re.IGNORECASE | re.DOTALL
        )
        tags = [t.strip().rstrip(",").strip() for t in raw_tags if t.strip()]

        if tags:
            # Format as: NPD37819: MRS CU FMEA-391, MRS CU FMEA-463
            comma_separated_tags = ", ".join(tags)
            traces.append(f"{doc_num}: {comma_separated_tags}")

    return traces

def _parse_single_edo_trace_response(edo_tag, response):
    """
    Parses ONE EDO's raw LLM response into {"traces": [str, ...]}.
    Order of operations:
      1. Direct JSON parse (parse_json / clean_llm_response)
      2. Fenced ```json code blocks (extract_json_code_blocks)
      3. Custom Bracket/Trace parser (parse_custom_fmea_format) -> Handles your prompt format!
      4. Legacy Markdown fallback (parse_markdown_trace_response)
    """
    # 1. Direct JSON parse
    parsed = parse_json(response)

    def normalize_result(obj):
        if isinstance(obj, dict):
            if "traces" in obj and isinstance(obj["traces"], list):
                traces = [normalize_text(t) for t in obj["traces"] if normalize_text(t)]
                return {"traces": traces}
            if edo_tag in obj and isinstance(obj[edo_tag], dict):
                return normalize_result(obj[edo_tag])
            for v in obj.values():
                if isinstance(v, dict) and "traces" in v:
                    return normalize_result(v)
        return None

    result = normalize_result(parsed)
    if result is not None:
        return result

    # 2. Code block parser
    code_block_records = extract_json_code_blocks(response)
    for block in code_block_records:
        result = normalize_result(block)
        if result is not None:
            logging.warning(
                f"_parse_single_edo_trace_response: {edo_tag} recovered via "
                "fenced ```json code block fallback."
            )
            return result

    # 3. Custom Bracket/Trace parser for the custom prompt format
    custom_traces = parse_custom_fmea_format(response)
    if custom_traces:
        logging.info(
            f"_parse_single_edo_trace_response: {edo_tag} successfully extracted "
            f"{len(custom_traces)} document trace group(s) using Custom Bracket parser."
        )
        return {"traces": custom_traces}

    # 4. Markdown fallback parser
    markdown_records = parse_markdown_trace_response(response)
    if markdown_records:
        traces = []
        for rec in markdown_records:
            module_controls = rec.get("Traces to Module DFMEA risk controls")
            if isinstance(module_controls, list):
                traces.extend(normalize_text(t) for t in module_controls if normalize_text(t))
        if traces:
            logging.warning(
                f"_parse_single_edo_trace_response: {edo_tag} recovered via "
                "Markdown fallback parser."
            )
            return {"traces": traces}

    logging.warning(
        f"_parse_single_edo_trace_response: {edo_tag} - could not parse any "
        "traces from the LLM response. Raw response logged below."
    )
    logging.warning(response)
    return {"traces": []}


def extract_json_code_blocks(text):
    """
    Fallback parser for extract_existing_edo_trace_details(). The LLM
    sometimes wraps EACH FMEA number's answer in its own separate
    ```json ... ``` fenced code block (e.g. one block per
    "### **For FMEA_Number: SYS-XXX**" section) instead of returning one
    single JSON document for the whole response. A single json.loads()
    call on the whole response fails in that case, because multiple
    top-level JSON objects side-by-side with Markdown headers/separators
    in between them isn't valid JSON as a whole - even though each
    individual fenced block IS perfectly valid JSON on its own. This
    extracts every ```json fenced block and parses each independently,
    returning the combined list of records.
    """
    text = normalize_text(text)
    blocks = re.findall(r"```(?:json)?\s*(.*?)```", text, re.DOTALL | re.IGNORECASE)

    records = []
    for block in blocks:
        try:
            parsed_block = json.loads(block.strip())
        except Exception:
            continue

        if isinstance(parsed_block, list):
            records.extend(item for item in parsed_block if isinstance(item, dict))
        elif isinstance(parsed_block, dict):
            records.append(parsed_block)

    return records


def parse_markdown_trace_response(response):
    """
    Fallback parser for extract_existing_edo_trace_details(). The
    EDO_Existing_Trace prompt is supposed to return JSON, but the LLM
    sometimes replies with Markdown instead - one "### **<FMEA id>**"
    block per FMEA number, with **bold** labels and "-" bullet lists,
    e.g.:

        ### **FMEA Sys-147**
        **Trace To RAC#:** RA-180
        **Traces to Module DFMEA risk controls:**
        - **Empty Array** (No matching entries found)
        **Additional Risk Control Measures from Module and/or Component DFMEAs:**
        - **Empty Array** (No matching identifiers found)
        **DRS Identifiers:**
        - **DRS-570**

    parse_json() can't recover this (there is no {}/[] JSON anywhere in
    it). This converts each block into the same record shape
    apply_existing_edo_trace() expects from a proper JSON response:
    {"System DFMEA #", "Trace To RAC#", "Traces to Module DFMEA risk
    controls", "Additional Risk Control Measures from Module and/or
    Component DFMEAs", "DRS Identifiers"}.
    """
    text = normalize_text(response)

    def extract_bullets(section_text):
        items = []
        for line in section_text.splitlines():
            line = line.strip()
            if not line.startswith("-"):
                continue
            line = line.lstrip("-").strip()
            line = re.sub(r"^\*+|\*+$", "", line).strip()
            if not line or "empty array" in line.lower() or line.lower() in ("none", "n/a"):
                continue
            items.append(line)
        return items

    def extract_section(block, label, next_labels):
        stop_pattern = "|".join(re.escape(nl) for nl in next_labels) if next_labels else None
        pattern = re.escape(label) + r"[:\*\s]*\n?(.*?)" + (
            f"(?=\\*\\*(?:{stop_pattern})|$)" if stop_pattern else "$"
        )
        match = re.search(pattern, block, re.DOTALL | re.IGNORECASE)
        return match.group(1) if match else ""

    records = []

    # Split on "### **<header>**" markers into (header, body) pairs.
    blocks = re.split(r"^#{1,4}\s*\*\*(.+?)\*\*\s*$", text, flags=re.MULTILINE)

    if len(blocks) > 1:
        for i in range(1, len(blocks), 2):
            header = normalize_text(blocks[i])
            body = blocks[i + 1] if i + 1 < len(blocks) else ""

            ra_match = re.search(r"Trace To RAC#\s*[:\*]*\s*([^\n]+)", body, re.IGNORECASE)
            ra_number = re.sub(r"\*+", "", normalize_text(ra_match.group(1))).strip() if ra_match else ""

            module_section = extract_section(
                body, "Traces to Module DFMEA risk controls",
                ["Additional Risk Control Measures", "DRS Identifiers"]
            )
            additional_section = extract_section(
                body, "Additional Risk Control Measures from Module and/or Component DFMEAs",
                ["DRS Identifiers"]
            )
            drs_section = extract_section(body, "DRS Identifiers", [])

            records.append({
                "System DFMEA #": header,
                "Trace To RAC#": ra_number,
                "Traces to Module DFMEA risk controls": extract_bullets(module_section),
                "Additional Risk Control Measures from Module and/or Component DFMEAs": extract_bullets(additional_section),
                "DRS Identifiers": extract_bullets(drs_section),
            })

    return records


def apply_existing_edo_trace(existing_edos, trace_details):
    """
    ADDED - couples the trace records back onto existing_edos.

    CHANGED: the EDO_Existing_Trace prompt's response (from the per-EDO
    extract_existing_edo_trace_details above) is now keyed DIRECTLY by
    EDO tag - {"EDO-29": {"traces": [...]}, "EDO-31": {"traces": []}, ...}
    - matching existing_edos' own keys exactly (existing_edos is itself
    keyed by edo_tag, e.g. existing_edos["EDO-29"]).

    Previously this went through deep_extract_records() + RA/FMEA-number
    matching, which only worked for a different response shape (a flat
    list of {RA_Number, FMEA_Number, ...} records). Against the current
    dict-of-EDO-tag response, deep_extract_records() returned an empty
    list every time (no recognized container key, and "traces" holds
    plain strings rather than dicts) - so existing_trace was NEVER
    populated and nothing reached the Excel output, even though the LLM
    logs showed correct answers. This version matches by EDO tag
    directly instead.
    """
    matched_count = 0

    if not isinstance(trace_details, dict):
        logging.warning(
            "apply_existing_edo_trace: trace_details was not a dict "
            f"(got {type(trace_details).__name__}) - nothing to apply."
        )
        return existing_edos

    for raw_tag, row in trace_details.items():
        if not isinstance(row, dict):
            continue

        # Format-tolerant lookup: exact key, then case-insensitive match
        # against existing_edos' own keys.
        edo = existing_edos.get(raw_tag)
        if edo is None:
            normalized_tag = normalize_text(raw_tag).lower()
            for key, candidate in existing_edos.items():
                if normalize_text(key).lower() == normalized_tag:
                    edo = candidate
                    break

        if edo is None:
            logging.warning(
                f"apply_existing_edo_trace: no existing EDO found matching "
                f"key {raw_tag!r} - skipping."
            )
            continue

        traces = row.get("traces") or row.get("Traces") or []
        lines = [normalize_text(t) for t in traces if isinstance(t, str) and normalize_text(t)]

        if lines:
            edo["existing_trace"] = "\n".join(lines)
            matched_count += 1
        else:
            edo["existing_trace"] = ""

    logging.info(
        f"apply_existing_edo_trace: populated existing_trace on "
        f"{matched_count} existing-EDO match(es) out of {len(trace_details)} "
        "trace response(s)."
    )

    return existing_edos


# ==========================================================
# NEW EDO PIPELINE
# ==========================================================
# CALL 4: Extract New EDO Tags -> Remove Existing EDO Matches (performed during the merge below) -> CALL 5: Extract New EDO Summary Details (Verification Reference is intentionally excluded here).

def extract_new_edo_tags(
    client,
    product_family,
    product,
    templatename,
    pipeline_config,
    edo_document,
    db
):
    """
    STEP 1 of 3: scans Section 10 of the FMEA against the RA&C and
    identifies every qualifying new-EDO row's {RA_Number, FMEA_Number,
    Status}.

    Per requirement, every RA id found is consolidated into ONE
    dictionary - edo_new_data - keyed by RA_Number (falling back to
    FMEA_Number, then to a positional key, only when RA_Number itself
    is missing). This SAME dictionary is what gets passed into and
    progressively enriched by extract_new_edo_summary_details() and
    extract_new_edo_traceability_details() afterwards, so every
    RA id's data lives in one place from identification all the way
    through to the final merge.
    """
    logging.info("=" * 80)
    logging.info("STAGE 4a: WORKFLOW 2 - NEW EDO IDENTIFICATION (TAGS ONLY)")
    logging.info("=" * 80)

    prompt_data = db.get_prompt_by_name(
        client,
        product_family,
        product,
        templatename,
        "EDO_NEW_tags"
    )

    prompt_row = {
        "prompt_role": prompt_data["prompt_role"],
        "prompt_text": prompt_data["prompt_text"],
        "question": (
            'Get all RA-# entries whose Risk Evaluation status is "See FMEA" '
            'from Appendix A - Risk Assessment and Control Table.'
        ),
        "fulltext": "Yes",
        "where_filter": "",
        "where_document": '{"$contains": "See FMEA"}',
        "checkpoint": ('Extract all Appendix A rows whose Risk Evaluation is "See FMEA".'),
        "max_results": 25,
    }

    # ---- DIAGNOSTIC: log how many chunks/docs are in this collection,
    # so we can tell "collection is empty" apart from "filter/retrieval
    # logic excluded everything".
    try:
        collection_count = edo_document["edo_ra_c"]["collection"].count()
        logging.info(f"edo_ra_c collection count: {collection_count}")
    except Exception as e:
        logging.warning(f"Could not inspect edo_ra_c collection count: {e}")

    docs, metadata, response = execute_llm_retry(
        pipeline_config,
        edo_document["edo_ra_c"]["collection"],
        prompt_row
    )

    # ---- DIAGNOSTIC: confirm content is now actually being retrieved.
    try:
        logging.info(f"RETRIEVED DOCS COUNT: {len(docs) if docs else 0}")
        logging.info(f"RETRIEVED DOCS PREVIEW: {str(docs)[:500]}")
    except Exception as e:
        logging.warning(f"Could not log retrieved docs: {e}")

    tags = parse_json(response)
    tag_records = deep_extract_records(tags)

    logging.info(f"extract_new_edo_tags: {len(tag_records)} raw tag row(s) returned by the LLM.")

    # ---- Consolidate every RA id found into ONE dictionary ----
    edo_new_data = {}

    for index, row in enumerate(tag_records):
        if not isinstance(row, dict):
            continue

        ra_number = normalize_text(row.get("RA_Number") or row.get("RA Number") or "")
        status = normalize_text(row.get("Status") or row.get("status") or "")

        # FIXED: the LLM returns every FMEA number for a given RA as ONE
        # combined, comma-separated string under "FMEA_Numbers" (plural)
        # - e.g. "FMEA Sys-150, FMEA Sys-151, FMEA Sys-167, FMEA Sys-725"
        # - not a single "FMEA_Number". The old code only ever looked
        # for "FMEA_Number"/"FMEA Number" (singular), so that key never
        # matched at all and every record ended up with FMEA_Number =
        # "" downstream - and even if it HAD matched, storing the whole
        # comma-joined string under one dict entry per RA_Number would
        # still collapse every FMEA number for that RA into a single
        # row. Both the plural and singular keys are read here, the
        # value is split into its individual FMEA numbers, and each one
        # becomes its OWN separate edo_new_data entry - one row per
        # RA/FMEA pair, matching the Excel layout where a RA with
        # several FMEA numbers (e.g. RA-141 with FMEA Sys-154 and
        # FMEA Sys-729) prints as separate blocks, not one row with
        # every FMEA number crammed together.
        fmea_numbers_raw = normalize_text(
            row.get("FMEA_Numbers") or row.get("FMEA Numbers")
            or row.get("FMEA_Number") or row.get("FMEA Number") or ""
        )
        fmea_numbers = [f.strip() for f in fmea_numbers_raw.split(",") if f.strip()]
        if not fmea_numbers:
            # Keep the RA even when it has no FMEA number at all, so it
            # isn't silently dropped from edo_new_data.
            fmea_numbers = [""]

        for fmea_number in fmea_numbers:
            key = (
                f"{ra_number}::{fmea_number}" if ra_number and fmea_number
                else (ra_number or fmea_number or f"NEW-EDO-TAG-{index}")
            )
            if key in edo_new_data:
                key = f"{key}_{index}"

            edo_new_data[key] = {
                "RA_Number": ra_number,
                "FMEA_Number": fmea_number,
                "Status": status
            }

    logging.info(
        f"extract_new_edo_tags: consolidated {len(edo_new_data)} RA id(s) "
        f"into edo_new_data."
    )
    print ("NEw edo tags:",edo_new_data)
    return edo_new_data


def filter_new_edo_by_risk_evaluation(
    client,
    product_family,
    product,
    templatename,
    pipeline_config,
    edo_document,
    edo_new_data,
    db
):
    """
    STEP 1.5 of 3 (runs AFTER extract_new_edo_tags / CALL 4 and BEFORE
    extract_new_edo_summary_details / CALL 5).

    For every RA_Number/FMEA_Number pair already collected in
    edo_new_data, this checks TWO things against the EDO_FMEA document:

        1. Does that exact RA_Number/FMEA_Number pair exist as a row
           in EDO_FMEA at all?
        2. If it exists, is that row's Risk_Evaluation value exactly
           "Medium" (case-insensitive)?

    An entry is kept ONLY when BOTH are true:
        - Pair IS found in EDO_FMEA, AND its Risk_Evaluation == "Medium"
          -> entry is kept in edo_new_data, unchanged.
        - Pair is NOT found in EDO_FMEA (or the lookup fails), OR the
          row's Risk_Evaluation is anything other than "Medium"
          (High, Low, blank, missing, non-answer, etc.) -> entry is
          removed from edo_new_data.
    """
    logging.info("=" * 80)
    logging.info("STAGE 4a-2: WORKFLOW 2 - NEW EDO FMEA PRESENCE + RISK EVALUATION FILTER")
    logging.info("=" * 80)

    if not edo_new_data:
        logging.info("No new EDO tags to filter.")
        return edo_new_data

    if "edo_fmea" not in edo_document:
        logging.warning(
            "FMEA PRESENCE FILTER SKIPPED - EDO_FMEA document/collection "
            "not configured. Leaving edo_new_data unfiltered for this run."
        )
        return edo_new_data

    try:
        prompt_data = db.get_prompt_by_name(
            client,
            product_family,
            product,
            templatename,
            "EDO_NEW_risk_evaluation"
        )
    except Exception as e:
        logging.error(
            "FMEA PRESENCE FILTER SKIPPED - could not load the "
            f"'EDO_NEW_risk_evaluation' prompt: {e}. Leaving edo_new_data "
            "unfiltered for this run."
        )
        return edo_new_data

    keys_to_remove = []

    for key, entry in edo_new_data.items():
        ra_number = entry.get("RA_Number", "")
        fmea_number = entry.get("FMEA_Number", "")

        target_text = f"RA_Number : {ra_number}\nFMEA_Number : {fmea_number}"

        prompt_row = {
            "prompt_role": prompt_data["prompt_role"],
            "prompt_text": prompt_data["prompt_text"] + "\nTARGETS:\n" + target_text,
            "question": (
                f"Does a row for RA_Number {ra_number} / FMEA_Number "
                f"{fmea_number} exist in this FMEA document? Return that "
                "row as a single JSON object if it exists, including its "
                "Risk_Evaluation value."
            ),
            "fulltext": "Yes",
            "where_filter": "",
            "where_document": "",
            "checkpoint": ""
        }

        logging.info(
            f"--- FMEA PRESENCE + RISK EVALUATION LOOKUP for RA={ra_number!r} "
            f"FMEA={fmea_number!r} ---"
        )

        found_in_fmea = False
        detail_row = {}
        try:
            docs, metadata, response = execute_llm_retry(
                pipeline_config,
                edo_document["edo_fmea"]["collection"],
                prompt_row
            )

            print(
                f"FMEA FILTER: RAW LLM RESPONSE RA={ra_number!r} "
                f"FMEA={fmea_number!r} -> {response!r}"
            )

            parsed = parse_json(response)
            records = deep_extract_records(parsed)
            if not records and isinstance(parsed, dict):
                records = [parsed]

            detail_row = records[0] if records else {}

            # "Present" means the FMEA document actually returned a row
            # with real content for this RA/FMEA pair - at least one
            # non-empty, non-placeholder value anywhere in it - not
            # just an empty object or a "none"/"blank" non-answer.
            found_in_fmea = any(
                normalize_text(v) and normalize_text(v).strip().lower() not in ("none", "blank")
                for v in _flatten_record_values(detail_row)
            )

        except Exception as e:
            logging.error(
                f"FMEA PRESENCE lookup failed for RA={ra_number!r} "
                f"FMEA={fmea_number!r}: {e}"
            )
            found_in_fmea = False
            detail_row = {}

        # ---- Risk_Evaluation check: only relevant if the row was
        # actually found. Uses the same normalized/case-insensitive
        # key matching as get_llm_value, so "Risk_Evaluation",
        # "Risk Evaluation", "risk_evaluation", etc. all match. ----
        risk_evaluation_value = ""
        is_medium_risk = False
        if found_in_fmea:
            risk_evaluation_value = normalize_text(
                get_llm_value(
                    detail_row,
                    "Risk_Evaluation",
                    "Risk Evaluation",
                    "risk_evaluation",
                    "RiskEvaluation",
                    "Risk_Rating",
                    "Risk Rating"
                )
            )
            is_medium_risk = risk_evaluation_value.strip().lower() == "medium"

            print(
                f"FMEA FILTER: Risk_Evaluation value for RA={ra_number!r} "
                f"FMEA={fmea_number!r} -> {risk_evaluation_value!r} "
                f"(is_medium={is_medium_risk})"
            )

        if found_in_fmea and is_medium_risk:
            logging.info(
                f"FMEA PRESENCE FILTER: keeping RA={ra_number!r} "
                f"FMEA={fmea_number!r} - found in EDO_FMEA with "
                f"Risk_Evaluation={risk_evaluation_value!r}."
            )
        elif found_in_fmea and not is_medium_risk:
            logging.info(
                f"FMEA PRESENCE FILTER: excluding RA={ra_number!r} "
                f"FMEA={fmea_number!r} - found in EDO_FMEA but "
                f"Risk_Evaluation={risk_evaluation_value!r} is not 'Medium'."
            )
            keys_to_remove.append(key)
        else:
            logging.info(
                f"FMEA PRESENCE FILTER: excluding RA={ra_number!r} "
                f"FMEA={fmea_number!r} - not found in EDO_FMEA."
            )
            keys_to_remove.append(key)

    for key in keys_to_remove:
        del edo_new_data[key]

    logging.info(
        f"FMEA PRESENCE FILTER complete: {len(edo_new_data)} entr"
        f"{'y' if len(edo_new_data) == 1 else 'ies'} remaining "
        f"({len(keys_to_remove)} excluded - not found in EDO_FMEA or "
        "Risk_Evaluation was not 'Medium')."
    )
    print("FMEA presence + risk evaluation filtered value:", edo_new_data)
    return edo_new_data

def extract_new_edo_summary_details(
    client,
    product_family,
    product,
    templatename,
    pipeline_config,
    edo_document,
    edo_new_data,
    db
):
    """
    STEP 2 of 3: takes the SAME edo_new_data dictionary built by
    extract_new_edo_tags() (keyed by RA id) and, for EVERY entry in it,
    calls the LLM ONE TIME - inside the loop, one RA_Number/FMEA_Number
    pair per call - to pull that single row's full detail record.
    """
    logging.info("=" * 80)
    logging.info("STAGE 4b: WORKFLOW 2 - NEW EDO FULL DETAIL EXTRACTION (PER RA/FMEA, LOOPED)")
    logging.info("=" * 80)

    if not edo_new_data:
        logging.info("No new EDO tags to extract details for.")
        return edo_new_data

    prompt_data = db.get_prompt_by_name(
        client,
        product_family,
        product,
        templatename,
        "EDO_NEW_details"
    )

    keys_to_remove = []

    for key, entry in edo_new_data.items():
        ra_number = entry.get("RA_Number", "")
        fmea_number = entry.get("FMEA_Number", "")

        target_text = f"RA_Number : {ra_number}\nFMEA_Number : {fmea_number}"

        prompt_row = {
            "prompt_role": prompt_data["prompt_role"],
            "prompt_text": prompt_data["prompt_text"] + "\nTARGETS:\n" + target_text,
            "question": (
                f"For RA_Number {ra_number} / FMEA_Number {fmea_number} "
                "only, extract the full EDO detail record as a single "
                "JSON object - no other RA/FMEA pairs."
            ),
            "fulltext": "Yes",
            "where_filter": "",
           # "where_document": {"$contains": "Safety Hazard DFMEA Table"},
           "where_document":"",
            "checkpoint": ""
        }

        # --- LOGGING: Inspect Prompt Row Filters & Collection state ---
        logging.info(f"--- DEBUG RETRIEVAL ATTEMPT for RA={ra_number!r} FMEA={fmea_number!r} ---")
        logging.info(f"where_document filter being applied: {prompt_row['where_document']}")

        try:
            collection_count = edo_document["edo_fmea"]["collection"].count()
            logging.info(f"edo_fmea collection total count: {collection_count}")
        except Exception as e:
            logging.warning(f"Could not inspect edo_fmea collection count: {e}")

        detail_row = {}
        try:
            docs, metadata, response = execute_llm_retry(
                pipeline_config,
                edo_document["edo_fmea"]["collection"],
                prompt_row
            )

            # --- LOGGING: Detailed Chunk & Metadata Diagnostics ---
            logging.info(f"RA={ra_number!r} FMEA={fmea_number!r} - RETRIEVED DOCS COUNT: {len(docs) if docs else 0}")
            logging.info(f"RA={ra_number!r} FMEA={fmea_number!r} - RETRIEVED DOCS TYPE: {type(docs)}")
            logging.info(f"RA={ra_number!r} FMEA={fmea_number!r} - RETRIEVED DOCS CONTENT PREVIEW: {str(docs)[:500]}")
            logging.info(f"RA={ra_number!r} FMEA={fmea_number!r} - METADATA TYPE: {type(metadata)}")
            logging.info(f"RA={ra_number!r} FMEA={fmea_number!r} - METADATA CONTENT PREVIEW: {str(metadata)[:500]}")
            logging.info(f"RA={ra_number!r} FMEA={fmea_number!r} - RAW LLM RESPONSE PREVIEW: {str(response)[:300]}")

            parsed = parse_json(response)
            records = deep_extract_records(parsed)
            if records:
                detail_row = records[0]
            elif isinstance(parsed, dict):
                detail_row = parsed

        except Exception as e:
            logging.error(
                f"LLM call failed for RA={ra_number!r} FMEA={fmea_number!r}: {e}"
            )
            detail_row = {}

        if not detail_row:
            logging.warning(
                f"No detail record returned for RA={ra_number!r} FMEA={fmea_number!r}. "
                "Excluding this entry from edo_new_data."
            )
            keys_to_remove.append(key)
            continue

        entry["Product_Feature_Function"] = get_llm_value(
            detail_row, "Product_Feature_Function", "Product Feature Function"
        )
        entry["Reason_Identified_as_EDO"] = get_llm_value(
            detail_row, "Reason_Identified_as_EDO", "Reason Identified as EDO"
        )
        entry["Traceability"] = get_llm_value(
            detail_row, "Traceability", "traceability"
        )
        # NOTE: Verification_Reference is intentionally NOT set here.
        # Per the new pipeline order, verification reference extraction
        # for every record (existing + new) happens in exactly ONE place -
        # the unified, per-record CALL 6 (extract_and_apply_verification_details)
        # that runs after Existing + New are merged. See STAGE 6 in
        # generate_edo_template().
        entry["EDO_Location"] = get_llm_value(
            detail_row, "EDO_Location", "location"
        )
        entry["EDO_Description"] = get_llm_value(
            detail_row, "EDO_Description", "description"
        )
        entry["Reason_Identified_as_EDO_ColH"] = get_llm_value(
            detail_row, "Reason_Identified_as_EDO_ColH", "reason_2"
        )

    # Remove entries that had no detail record returned.
    for key in keys_to_remove:
        del edo_new_data[key]

    logging.info(
        f"extract_new_edo_summary_details: enriched {len(edo_new_data)} "
        f"entries in edo_new_data with full detail records "
        f"({len(keys_to_remove)} excluded due to missing details)."
    )

    return edo_new_data


# ==========================================================
# MERGE PIPELINE
# ==========================================================
# Merges Existing EDO records + New EDO records (de-duplicated against Existing) into ONE common list/dict. Everything after this point works on the merged list only.

def get_existing_identifier_set(existing_edos):
    """
    Builds a set of every RA Number / FMEA Number already tracked under
    the Existing EDOs (Stage 3), normalized to uppercase for
    case-insensitive comparison. Used to drop any "New EDO" record whose
    RA id or FMEA id is already represented by an Existing EDO, so the
    same underlying issue is never printed twice.
    """
    identifiers = set()
    for edo in existing_edos.values():
        ra = normalize_text(edo.get("ra_number")).upper()
        fmea = normalize_text(edo.get("FMEA_Number")).upper()
        if ra:
            identifiers.add(ra)
        if fmea:
            identifiers.add(fmea)
    return identifiers


def normalize_edo_record(edo_record):
    default_structure = {
        "edo_type": "Existing",
        "edo_tag": "",
        "RA_Number": "",
        "FMEA_Number": "",
        "edo_description": "",
        "reason_identified": "",
        "dfmea": "",
        "verification_reference": "",
        "existing_trace": "",
        "location": "",
        "description_2": "",
        "reason_2": "",
        "sysdd": ""
    }
    if not isinstance(edo_record, dict):
        return default_structure

    for key, value in default_structure.items():
        if key not in edo_record:
            edo_record[key] = value
    return edo_record


def merge_new_edo_dictionary(
    edo_tags,
    edo_summary_details,
    edo_trace_details
):
    """
    From generate_EDO_template_copy.py - only pipeline that builds NEW EDO
    records, so it is kept as-is and used by the merged pipeline.
    """
    logging.info("=" * 80)
    logging.info("STAGE 5: WORKFLOW MATRIX COUPLING ENGINE")
    logging.info("=" * 80)

    merged = {}

    # Deep Structure Unwrapping
    records = deep_extract_records(edo_summary_details)
    logging.info(f"SUMMARY DETAILS COUNT EXTRICATED : {len(records)}")

    for index, row in enumerate(records):
        ra_number = normalize_text(get_llm_value(row, "RA_Number", "RA Number", "ra_num"))
        fmea_number = normalize_text(get_llm_value(row, "FMEA_Number", "FMEA Number", "fmea_num"))
        status_label = normalize_text(get_llm_value(row, "Status", "status")) or "New EDO"

        key = fmea_number or ra_number
        if not key:
            key = f"NEW-EDO-RECORD-{index}"

        # Safe attribute alignment ensuring literal 'None' strings from the dictionary are preserved
        edo_desc = normalize_text(get_llm_value(row, "Product_Feature_Function", "Product Feature Function", "EDO_Description", "edo_description"))
        reason_id = normalize_text(get_llm_value(row, "Reason_Identified_as_EDO", "Reason Identified as EDO", "reason_identified"))
        dfmea_trace = normalize_text(get_llm_value(row, "Traceability", "dfmea", "traceability"))
        ver_ref = normalize_text(get_llm_value(row, "Verification_Reference", "Verification Reference", "verification_reference"))

        loc_val = normalize_text(get_llm_value(row, "EDO_Location", "location")) or "None"
        desc2_val = normalize_text(get_llm_value(row, "EDO_Description", "Description_2")) or "None"
        reason2_val = normalize_text(get_llm_value(row, "Reason_2")) or "None"
        sys_dd_val = normalize_text(get_llm_value(row, "Reason_Identified_as_EDO_ColH", "sysdd", "SYS_DD")) or "None"

        merged[key] = {
            "edo_type": "New",
            "edo_tag": f"EDO-XX\n{status_label}",
            "edo_description": edo_desc if edo_desc else "Blank",
            "reason_identified": reason_id if reason_id else "Blank",
            "dfmea": dfmea_trace if dfmea_trace else "Blank",
            "verification_reference": ver_ref if ver_ref else "Blank",
            "location": loc_val,
            "description_2": desc2_val,
            "reason_2": reason2_val,
            "sysdd": sys_dd_val,
            "RA_Number": ra_number,
            "FMEA_Number": fmea_number
        }

    # Asymmetric Verification Coupling
    trace_records = deep_extract_records(edo_trace_details)
    for row in trace_records:
        if not isinstance(row, dict):
            continue

        trace_fmea = normalize_text(get_llm_value(row, "FMEA_Number", "FMEA Number"))
        trace_ra = normalize_text(get_llm_value(row, "RA_Number", "RA Number"))
        trace_key = trace_fmea or trace_ra

        verification = normalize_text(get_llm_value(row, "Verification_Reference", "Verification Reference", "verification_reference"))
        if not verification or verification.lower() == "none" or verification == "":
            continue

        if trace_key and trace_key in merged:
            merged[trace_key]["verification_reference"] = verification
        else:
            for main_key, data in merged.items():
                if (trace_ra and data["RA_Number"] == trace_ra) or (trace_fmea and data["FMEA_Number"] == trace_fmea):
                    data["verification_reference"] = verification

    return merged


def merge_new_edo_dictionary_full(
    edo_tags,
    edo_summary_details,
    edo_trace_details,
    edo_ra_details
):
    """
    CANONICAL new-EDO merge - used by generate_edo_template().

    merge_new_edo_dictionary() (above) is left completely UNTOUCHED, but
    it silently drops any RA record whose Status is "Medium", because it
    only ever looks at edo_summary_details / edo_trace_details, both of
    which come exclusively from the EDO_FMEA collection.

    This function fixes that: it builds the output starting from
    edo_tags itself (the full, unfiltered list returned by
    extract_new_edo_tags), so EVERY RA Number is guaranteed a row in the
    final Excel output - none are excluded.

    For each tag:
      - Status contains "Medium"  -> description/reason are pulled ONLY
        from edo_ra_details (EDO_RA_C document) - the FMEA-sourced
        edo_summary_details is deliberately NOT consulted for these,
        exactly as requested.
      - Any other Status (e.g. "See FMEA") -> description/reason/trace/
        verification are pulled from edo_summary_details / edo_trace_details,
        exactly like the original merge_new_edo_dictionary() behaviour.
      - If no matching record is found in the relevant source at all, the
        tag is still written to the output - with empty ("") fields
        rather than a "Blank"/"None" placeholder string, and rather than
        being skipped/excluded.
      - Column A (edo_tag) is always the fixed literal "EDO-XX\\nNew EDO"
        for every New EDO record - the risk Status (Medium / See FMEA) is
        used internally to choose the data source but is never printed.
      - Column D (dfmea) for "Medium" records is always the RA Number
        itself (there is no FMEA trace to show for these), so every RA
        Number is visibly represented in Column D regardless of Status.
    """
    logging.info("=" * 80)
    logging.info("STAGE 5: WORKFLOW MATRIX COUPLING ENGINE (FULL - INCLUDES MEDIUM RA)")
    logging.info("=" * 80)

    def index_records(records):
        index = {}
        for row in records:
            if not isinstance(row, dict):
                continue
            ra = normalize_text(get_llm_value(row, "RA_Number", "RA Number", "ra_num"))
            fmea = normalize_text(get_llm_value(row, "FMEA_Number", "FMEA Number", "fmea_num"))
            if ra and ra not in index:
                index[ra] = row
            if fmea and fmea not in index:
                index[fmea] = row
        return index

    summary_index = index_records(deep_extract_records(edo_summary_details))
    ra_index = index_records(deep_extract_records(edo_ra_details))

    # Verification lookups from the FMEA-sourced traceability extraction -
    # only ever applied to non-Medium ("See FMEA") records.
    trace_index = {}
    for row in deep_extract_records(edo_trace_details):
        if not isinstance(row, dict):
            continue
        trace_ra = normalize_text(get_llm_value(row, "RA_Number", "RA Number"))
        trace_fmea = normalize_text(get_llm_value(row, "FMEA_Number", "FMEA Number"))
        verification = normalize_text(get_llm_value(row, "Verification_Reference", "Verification Reference", "verification_reference"))
        if not verification or verification.lower() == "none":
            continue
        if trace_ra:
            trace_index.setdefault(trace_ra, verification)
        if trace_fmea:
            trace_index.setdefault(trace_fmea, verification)

    merged = {}

    for index, item in enumerate(edo_tags):
        ra_number = normalize_text(item.get("RA_Number"))
        fmea_number = normalize_text(item.get("FMEA_Number"))
        status_label = normalize_text(item.get("Status")) or "New EDO"
        is_medium = "medium" in status_label.lower()

        key = fmea_number or ra_number
        if not key:
            key = f"NEW-EDO-RECORD-{index}"
        # avoid clobbering an existing key on duplicate RA/FMEA numbers
        if key in merged:
            key = f"{key}_{index}"

        if is_medium:
            # Medium risk value -> EDO_RA_C ONLY, never the FMEA document.
            # Keyed by RA Number ONLY - the tag's FMEA_Number is a
            # placeholder ("FMEA-UNKNOWN") for Medium records and is not
            # a reliable lookup key.
            source_row = ra_index.get(ra_number)
        else:
            # See FMEA / anything else -> FMEA-sourced summary, as before
            source_row = summary_index.get(fmea_number) or summary_index.get(ra_number)

        if source_row:
            edo_desc = normalize_text(get_llm_value(source_row, "Product_Feature_Function", "Product Feature Function", "EDO_Description", "edo_description"))
            reason_id = normalize_text(get_llm_value(source_row, "Reason_Identified_as_EDO", "Reason Identified as EDO", "reason_identified"))
            dfmea_trace = normalize_text(get_llm_value(source_row, "Traceability", "dfmea", "traceability"))
            ver_ref = normalize_text(get_llm_value(source_row, "Verification_Reference", "Verification Reference", "verification_reference"))
            loc_val = normalize_text(get_llm_value(source_row, "EDO_Location", "location"))
            desc2_val = normalize_text(get_llm_value(source_row, "EDO_Description", "Description_2"))
            reason2_val = normalize_text(get_llm_value(source_row, "Reason_2"))
        else:
            edo_desc = reason_id = dfmea_trace = ver_ref = ""
            loc_val = desc2_val = reason2_val = ""

        # Per requirement: for "Medium" records, Column D (dfmea) always
        # shows the RA Number itself - there is no FMEA trace document to
        # pull a narrative from, so the RA id is the traceability value.
        if is_medium:
            dfmea_trace = ra_number

        # Verification backfill only applies to non-Medium records, since
        # the trace/verification extraction is itself FMEA-sourced.
        if not is_medium and not ver_ref:
            ver_ref = trace_index.get(fmea_number) or trace_index.get(ra_number) or ver_ref

        # Column J (SYSDD / HDD Reference) is a fixed document reference
        # for this template, not an LLM-extracted value - see
        # get_fixed_sysdd_reference().
        sys_dd_val = get_fixed_sysdd_reference()

        merged[key] = {
            "edo_type": "New",
            # Per requirement: Column A never prints the risk Status
            # (Medium / See FMEA) - always the fixed literal tag text.
            "edo_tag": "EDO-XX\nNew EDO",
            "edo_description": edo_desc,
            "reason_identified": reason_id,
            "dfmea": dfmea_trace,
            "verification_reference": ver_ref,
            "location": loc_val,
            "description_2": desc2_val,
            "reason_2": reason2_val,
            "sysdd": sys_dd_val,
            "RA_Number": ra_number,
            "FMEA_Number": fmea_number
        }

    return merged


def merge_new_edo_records(new_records, existing_edos):
    """
    CANONICAL new-EDO merge - used by generate_edo_template(), replacing
    the old 4-source merge_new_edo_dictionary_full() now that
    extract_new_edo_tags() has been removed and extract_new_edo_summary_details()
    returns complete records directly from the single merged "EDO_NEW_details"
    prompt.

    Fixes the "column mapped wrongly" bug: the previous merge logic looked
    for keys like "Description_2" / "Reason_2" that never actually existed
    in the LLM's JSON output - the real keys are "EDO_Description" and
    "Reason_Identified_as_EDO_ColH". This function maps directly off the
    ACTUAL field names the prompt returns:

        edo_description   <- Product_Feature_Function
        reason_identified <- Reason_Identified_as_EDO
        dfmea             <- Traceability
        verification_reference <- Verification_Reference
        location          <- EDO_Location
        description_2     <- EDO_Description
        reason_2          <- Reason_Identified_as_EDO_ColH
        sysdd             <- fixed constant (get_fixed_sysdd_reference())

    Also applies requirement #2: any new record whose RA Number or FMEA
    Number already belongs to an Existing EDO is dropped from the output,
    so the same issue is never printed twice under both types.

    Column A (edo_tag) is always the fixed literal "EDO-XX\\nNew" - the
    risk Status is never printed there.
    """
    logging.info("=" * 80)
    logging.info("STAGE 5: NEW EDO MERGE (SINGLE-PASS RECORDS, DE-DUPED AGAINST EXISTING)")
    logging.info("=" * 80)

    existing_identifiers = get_existing_identifier_set(existing_edos)

    records = deep_extract_records(new_records)
    merged = {}

    for index, row in enumerate(records):
        if not isinstance(row, dict):
            continue

        ra_number = normalize_text(get_llm_value(row, "RA_Number", "RA Number"))
        fmea_number = normalize_text(get_llm_value(row, "FMEA_Number", "FMEA Number"))

        # Requirement #2: skip any new record already tracked as an
        # Existing EDO (same RA id or FMEA id), so it isn't duplicated.
        ra_key = ra_number.upper()
        fmea_key = fmea_number.upper()
        if (ra_key and ra_key in existing_identifiers) and (fmea_key and fmea_key in existing_identifiers):
            logging.info(
                f"Skipping New EDO record RA={ra_number!r} FMEA={fmea_number!r} - "
                f"already tracked as an Existing EDO."
            )
            continue

        key = fmea_number or ra_number or f"NEW-EDO-RECORD-{index}"
        if key in merged:
            key = f"{key}_{index}"

        merged[key] = {
            "edo_type": "New",
            # Per requirement: Column A never prints the risk Status -
            # always the fixed literal tag text, correctly cased.
            "edo_tag": "EDO-XX\nNew",
            "edo_description": normalize_text(get_llm_value(row, "Product_Feature_Function", "Product Feature Function")),
            "reason_identified": normalize_text(get_llm_value(row, "Reason_Identified_as_EDO", "Reason Identified as EDO")),
            "dfmea": normalize_text(get_llm_value(row, "Traceability", "traceability")),
            "verification_reference": normalize_text(get_llm_value(row, "Verification_Reference", "Verification Reference")),
            "location": normalize_text(get_llm_value(row, "EDO_Location", "location")),
            "description_2": normalize_text(get_llm_value(row, "EDO_Description", "description")),
            "reason_2": normalize_text(get_llm_value(row, "Reason_Identified_as_EDO_ColH", "reason_2")),
            "sysdd": get_fixed_sysdd_reference(),
            "RA_Number": ra_number,
            "FMEA_Number": fmea_number
        }

    return merged


def merge_existing_edo_dictionary(existing_details):
    """
    CANONICAL version - from edo_existing_final.py.
    Richer than the copy-file version: carries edo_type, and correctly
    threads RA_Number/FMEA_Number through (needed since Stage 3B already
    populated ra_number/FMEA_Number on each existing EDO).

    NOTE: missing values default to "" (a truly empty cell), not the
    literal text "Blank". Column J (sysdd) is always the fixed reference
    from get_fixed_sysdd_reference(), per requirement.

    Also carries "design_elements" (the full per-location list built by
    extract_edo_details() from the EDO_Existing_Generic prompt's nested
    design_elements[] array) straight through - previously this dict was
    rebuilt with only a fixed set of keys, so design_elements was
    silently dropped here and format_edo_worksheet() only ever saw a
    single backfilled location instead of every location.
    """
    final = {}
    for tag, data in existing_details.items():
        final[tag] = {
            "edo_type": "Existing",
            "edo_tag": data.get("edo_tag", tag),
            "edo_description": data.get("edo_description", ""),
            "reason_identified": data.get("reason_identified", ""),
            "dfmea": data.get("dfmea", ""),
            "location": data.get("location", ""),
            "description_2": data.get("description_2", ""),
            "reason_2": data.get("reason_2", ""),
            "sysdd": get_fixed_sysdd_reference(),
            "verification_reference": data.get("verification_reference", ""),
            "existing_trace": data.get("existing_trace", ""),
            "RA_Number": data.get("ra_number", ""),
            "FMEA_Number": data.get("FMEA_Number", ""),
            "design_elements": data.get("design_elements", [])
        }
    return final


def merge_all_edos(existing_edos, new_edos):
    """
    From generate_EDO_template_copy.py - only pipeline that combines
    Existing + New EDO dictionaries into one, so it is kept as-is.
    """
    final_edos = {}

    for key, value in existing_edos.items():
        normalized = normalize_edo_record(value)
        normalized["edo_type"] = "Existing"
        final_edos[key] = normalized

    for key, value in new_edos.items():
        normalized = normalize_edo_record(value)
        normalized["edo_type"] = "New"

        final_key = key
        if final_key in final_edos:
            counter = 1
            while f"{key}_{counter}" in final_edos:
                counter += 1
            final_key = f"{key}_{counter}"

        final_edos[final_key] = normalized
    print("All nerged values:",final_edos)
    return final_edos


def validate_final_edos(final_edos):
    """
    CANONICAL version - from generate_EDO_template_copy.py.
    General-purpose: handles both "Existing" and "New" typed records
    (assigns a placeholder tag for blank New EDOs). Needed because the
    merged pipeline's final_edos dictionary contains both types.
    """
    validated = {}
    for key, value in final_edos.items():
        if not isinstance(value, dict):
            continue

        edo_tag = normalize_text(value.get("edo_tag"))
        if edo_tag == "" and value.get("edo_type") == "New":
            value["edo_tag"] = "EDO-XX\nNew EDO"

        normalized = normalize_edo_record(value)
        validated[key] = normalized
        print(f"final values:", final_edos)
    return validated


# ==========================================================
# VERIFICATION REFERENCE
# ==========================================================
# CALL 6 - SINGLE unified, post-merge, per-record verification reference
# extraction, shared between Existing and New EDOs. Runs on final_edos
# (the already-merged dictionary) AFTER Existing (CALL 1-3) and New
# (CALL 4-5) have been combined - see merge/validate above and STAGE 6 in
# generate_edo_template(). Looped one LLM call per RA/FMEA pair, the
# exact same call pattern as extract_new_edo_summary_details() (CALL 5),
# instead of the old two separate bulk calls (one for Existing, one for
# New).

VERIFICATION_LABELED_LINE_PATTERN = re.compile(
    r'^\s*(?P<prefix>\([^)]*\))?\s*'
    r'Source\s*:\s*(?P<source>.*?)\s*'
    r'(?:\|\s*Location\s*:\s*(?P<location>.*?)\s*)?'
    r'(?:\|\s*File\s*:\s*(?P<file>.*?)\s*)?'
    r'(?:\|\s*Result\s*:\s*(?P<result>.*?)\s*)?$',
    re.IGNORECASE
)


def clean_verification_reference_line(line):
    """
    Per requirement: a raw Verification_Reference line shaped like
        "(DRS-570) Source: Vest APX ... .xlsx | Location: NPD45678 ... | File: ... | Result: PASS"
    is stripped down to just the trace code (if present) plus the
    Location and File values - the "Source:" and "Result:" labels/
    values are dropped entirely, and no label text is printed at all:
        "(DRS-570) NPD45678 Vest APX System Verification Traceability Report ..."
    A line that doesn't match this labeled "Source: ... | Location: ...
    | File: ... | Result: ..." shape is returned unchanged, so genuine
    free-form verification text still prints as-is.
    """
    stripped = line.strip()
    if not stripped:
        return stripped

    match = VERIFICATION_LABELED_LINE_PATTERN.match(stripped)
    if not match:
        return stripped

    prefix = normalize_text(match.group("prefix"))
    location = normalize_text(match.group("location"))
    file_name = normalize_text(match.group("file"))

    body = " ".join(part for part in [location, file_name] if part)
    if not body:
        # Nothing usable besides the label text itself - fall back to
        # the original line rather than printing an empty result.
        return stripped

    return f"{prefix} {body}".strip() if prefix else body


def clean_verification_reference_text(raw_text):
    """
    Applies clean_verification_reference_line() to every line of a
    (possibly multi-line, multi-code) raw Verification_Reference value.
    """
    if not raw_text:
        return raw_text
    lines = [ln for ln in raw_text.split("\n") if ln.strip()]
    return "\n".join(clean_verification_reference_line(ln) for ln in lines)


# BUGFIX: the LLM sometimes answers with its own search narration or a
# plain "I couldn't find anything" statement - e.g. "Searching in RA
# document for RA_Number 12345, no verification reference found." -
# instead of either a real reference or the literal word "None". The
# only guard before this was `verification.lower() != "none"`, so that
# narration text passed straight through and got written into Column E
# as if it were real data. NON_ANSWER_VERIFICATION_PATTERN recognizes
# this shape of response so it can be discarded like "None" already is.
NON_ANSWER_VERIFICATION_PATTERN = re.compile(
    r'^\s*(searching|search(?:ed|ing)?\s+(?:in|the|for)|no\s+(?:verification|reference|'
    r'matching|relevant|explicit)|not\s+found|unable\s+to\s+(?:find|locate)|'
    r'could\s+not\s+(?:find|locate)|no\s+(?:information|data|match|result)|'
    r'not\s+(?:available|mentioned|provided|present)|n/?a)\b',
    re.IGNORECASE
)


def is_non_answer_verification_text(text):
    """
    True when `text` reads like the LLM describing its search process or
    reporting that it found nothing (see NON_ANSWER_VERIFICATION_PATTERN),
    rather than an actual Verification Reference value that should be
    written to Column E.
    """
    return bool(NON_ANSWER_VERIFICATION_PATTERN.match((text or "").strip()))


# BUGFIX: EDO_RA_C is the Risk Assessment and Control document itself,
# not a verification/test report - it has no genuine "Verification
# Reference" field. When a record is searched there via the RA_Number-
# only fallback and no real reference exists, the search still returns
# a labeled line pointing back at the RA&C document's own identity, e.g.
# "NPD36702 Vest APX Risk Assessment and Control RA-124" - the RA&C
# document's own Location/File name, not a distinct verification code.
# This is self-referential noise ("this is the document I searched"),
# not data, and must not be written to Column E.
#
# NARROWED: only treat this as self-referential noise when the "Risk
# Assessment and Control" phrase is immediately followed by THIS
# record's own RA_Number (i.e. it's just echoing back what was
# searched for). Previously ANY answer containing that phrase was
# discarded, which also threw away genuinely distinct verification
# text that happened to mention it (e.g. a real reference quoting the
# RA&C document's title alongside its own separate report number) -
# that over-filtering was blanking Column E for records that actually
# had a real answer.
SELF_REFERENTIAL_RA_DOCUMENT_PATTERN = re.compile(
    r'risk\s+assessment\s+(?:and|&)\s+control\s*[-:]?\s*(RA[\s-]?\d+|\d+)',
    re.IGNORECASE
)


def is_self_referential_ra_document_text(text, ra_number=""):
    """
    True when `text` is just the RA&C document naming itself back using
    THIS record's own RA_Number (e.g. searched for "RA-124" and got back
    "... Risk Assessment and Control RA-124") rather than a genuine,
    distinct Verification Reference. Only fires when the number right
    after "Risk Assessment and Control" matches `ra_number` - if it's a
    different number, or there's no `ra_number` to compare against, this
    returns False so real data is never discarded.
    """
    if not text:
        return False

    match = SELF_REFERENTIAL_RA_DOCUMENT_PATTERN.search(text)
    if not match:
        return False

    if not ra_number:
        return False

    found_digits = re.sub(r"\D", "", match.group(1))
    target_digits = re.sub(r"\D", "", ra_number)
    return bool(found_digits) and found_digits == target_digits


def extract_and_apply_verification_details(
    client,
    product_family,
    product,
    templatename,
    pipeline_config,
    edo_document,
    final_edos,
    db: DatabaseHandler
):
    """
    CALL 6 - for every record in the already-merged `final_edos` dict
    (Existing + New together), use its RA_Number/FMEA_Number to search
    for the Verification Reference and write it onto each record's
    "verification_reference" field - ONE (or, on fallback, two) LLM
    call(s) per record, same loop-per-entry pattern as
    extract_new_edo_summary_details(). Returns final_edos.

    OR fallback (BUGFIX): previously a record only ever searched
    EDO_RA_C instead of EDO_FMEA when FMEA_Number was textually blank -
    if FMEA_Number WAS present but the EDO_FMEA document simply had no
    entry for it (a very common case even with FMEA configured), the
    record was left with nothing and Column E came back blank even
    though EDO_RA_C was fully available and configured. Now every
    record with an FMEA_Number tries EDO_FMEA first, and EDO_RA_C (via
    RA_Number alone) is tried right after as a genuine OR fallback
    whenever EDO_FMEA comes back with no usable answer - not only when
    FMEA_Number was missing to begin with.

    A non-answer is filtered out before accepting either source's
    response: the LLM's own search narration ("Searching in RA
    document...", "No verification reference found.", etc. - see
    is_non_answer_verification_text()). Per requirement, EDO_RA_C's
    Location/File identification of itself (e.g. "NPD36702 Vest APX Risk
    Assessment and Control RA-124") is printed to Column E as-is when
    that's all EDO_RA_C has for this record - it is no longer treated as
    a non-answer.

    DEBUG (added): the raw LLM response text for each record is now
    printed at every stage of parsing (raw -> parsed JSON -> extracted
    row -> cleaned value) and also stored on the record itself under
    "verification_raw_llm_response", so it's visible in final_edos even
    when the final verification_reference ends up empty. This makes it
    possible to tell apart "LLM genuinely found nothing" from "LLM found
    something but parsing/key-matching dropped it".
    """
    logging.info("=" * 80)
    logging.info("CALL 6: UNIFIED VERIFICATION REFERENCE EXTRACTION (POST-MERGE, PER-RECORD)")
    logging.info("=" * 80)

    if not final_edos:
        logging.info("No merged EDO records to extract verification details for.")
        return final_edos

    if "edo_fmea" not in edo_document and "edo_ra_c" not in edo_document:
        raise Exception(
            "Neither EDO_FMEA nor EDO_RA_C document/collection is "
            "configured - cannot run verification reference extraction."
        )

    prompt_data = get_prompt(
        client,
        product_family,
        product,
        templatename,
        "EDO_NEW_Verification_details",
        db
    )

    def _query_verification_reference(source_key, ra_number, fmea_number, ra_only):
        """
        Runs the CALL 6 verification-reference prompt against a single
        source document/collection (edo_fmea or edo_ra_c) for one
        RA_Number/FMEA_Number pair. `ra_only` builds a clean RA_Number-
        only target/question (used for EDO_RA_C, which has no
        FMEA_Number field at all - a target that still mentions a blank
        "FMEA_Number : " line confuses the match). Returns a tuple:
        (cleaned_verification_text_or_empty, raw_llm_response_text).
        """
        # BUGFIX: these were previously plain strings, not f-strings, so
        # the literal text "{ra_number}"/"{fmea_number}" was sent to the
        # LLM instead of the actual RA/FMEA values - the search query
        # never told the model which record to look for. This was the
        # root cause of the RA&C (EDO_RA_C) fallback - "if FMEA number
        # not available, use RA number with the RA document" - not
        # reliably returning the right answer: the RA-only fallback
        # query for EDO_RA_C carried no real RA number either. `target_text`
        # was also built but never actually included in the prompt (every
        # other CALL in this file appends it to prompt_text via
        # "\nTARGETS:\n" + target_text - this one alone omitted it).
        if ra_only:
            target_text = f"RA_Number : {ra_number}"
            verification_question = f"Fetch verification records for {ra_number}"
        else:
            target_text = f"RA_Number : {ra_number}\nFMEA_Number : {fmea_number}"
            verification_question = f"Fetch verification records for {fmea_number} and {ra_number}"

        prompt_row = {
            "prompt_role": prompt_data["prompt_role"],
            "prompt_text": prompt_data["prompt_text"] + "\nTARGETS:\n" + target_text,
            "question": verification_question,
            "fulltext": "No",
            "where_filter": "",
            "where_document": "",
            "checkpoint": f"Fetch verification records for {fmea_number} and {ra_number}"
        }

        try:
            _, _, response = execute_llm_retry(
                pipeline_config,
                edo_document[source_key]["collection"],
                prompt_row
            )
            print(
                f"CALL 6: RAW LLM RESPONSE [{source_key}] "
                f"RA={ra_number!r} FMEA={fmea_number!r} -> {response!r}"
            )
            parsed = parse_json(response)
            print(
                f"CALL 6: PARSED JSON [{source_key}] "
                f"RA={ra_number!r} FMEA={fmea_number!r} -> {parsed!r}"
            )
        except Exception as e:
            logging.error(
                f"CALL 6: verification LLM call against {source_key} failed for "
                f"RA={ra_number!r} FMEA={fmea_number!r}: {e}"
            )
            return "", ""

        records = deep_extract_records(parsed)
        row = records[0] if records else (parsed if isinstance(parsed, dict) else {})
        print(f"CALL 6: EXTRACTED ROW [{source_key}] RA={ra_number!r} FMEA={fmea_number!r} -> {row!r}")

        # FIXED (see get_llm_value below): this used to only find the
        # answer when the LLM's JSON key was spelled EXACTLY
        # "Verification_Reference" or "Verification Reference". A real
        # response spelled e.g. "Verification_reference" (lowercase
        # "r") was silently missed, leaving Column E blank even though
        # the raw LLM response had a valid answer.
        verification = normalize_text(
            get_llm_value(row, "Verification_Reference", "Verification Reference", "verification_reference")
        )
        verification = clean_verification_reference_text(verification)
        print(
            f"CALL 6: CLEANED VERIFICATION VALUE [{source_key}] "
            f"RA={ra_number!r} FMEA={fmea_number!r} -> {verification!r}"
        )

        # NOTE: the self-referential-RA&C-echo check (see
        # is_self_referential_ra_document_text() above) used to also be
        # applied here and discard an EDO_RA_C answer that was just the
        # RA&C document naming itself back (e.g. "NPD36702 Vest APX Risk
        # Assessment and Control RA-124"). Per requirement, EDO_RA_C's
        # answer should now be printed to Column E even when it's only
        # that Location/File identification and nothing more specific -
        # a real answer from the RA&C document is preferred over leaving
        # Column E blank, so that check is no longer applied here.
        is_bad_answer = (
            not verification
            or verification.lower() == "none"
            or is_non_answer_verification_text(verification)
        )
        result = "" if is_bad_answer else verification
        return result, response

    for key, edo in final_edos.items():
        ra_number = edo.get("RA_Number", "")
        fmea_number = edo.get("FMEA_Number", "")

        if ra_number in (None, "", "Blank") and fmea_number in (None, "", "Blank"):
            logging.info(
                f"CALL 6: skipping {edo.get('edo_tag', key)!r} - no "
                "RA_Number or FMEA_Number available to search with, so "
                "Column E is left blank for this record."
            )
            continue

        verification = ""
        raw_response = ""

        # ---- 1st attempt: EDO_FMEA, whenever this record has an
        # FMEA_Number and the document is configured. ----
        if fmea_number not in (None, "", "Blank") and "edo_fmea" in edo_document:
            verification, raw_response = _query_verification_reference(
                "edo_fmea", ra_number, fmea_number, ra_only=False
            )

        # ---- OR fallback: EDO_RA_C via RA_Number alone, tried whenever
        # EDO_FMEA didn't produce a usable answer - whether because
        # FMEA_Number was blank to begin with, EDO_FMEA isn't
        # configured, or EDO_FMEA simply had no entry for this
        # RA/FMEA pair. ----
        if not verification and ra_number not in (None, "", "Blank") and "edo_ra_c" in edo_document:
            logging.info(
                f"CALL 6: no usable answer from EDO_FMEA for RA={ra_number!r} "
                f"FMEA={fmea_number!r} - trying EDO_RA_C via RA_Number instead."
            )
            verification, raw_response_rac = _query_verification_reference(
                "edo_ra_c", ra_number, fmea_number, ra_only=True
            )
            # Keep whichever attempt actually produced a response, so
            # the debug field isn't overwritten with an empty string
            # when the RA_C fallback itself errored out but EDO_FMEA
            # had returned something (even if unusable).
            raw_response = raw_response_rac or raw_response

        # Store the raw LLM response text on the record itself so it's
        # visible in final_edos for debugging, regardless of whether
        # verification parsing succeeded.
        edo["verification_raw_llm_response"] = raw_response

        if verification:
            edo["verification_reference"] = verification
            logging.info(
                f"CALL 6: verification reference for RA={ra_number!r} "
                f"FMEA={fmea_number!r} -> {verification!r}"
            )
        else:
            logging.info(
                f"CALL 6: no usable verification reference found for "
                f"RA={ra_number!r} FMEA={fmea_number!r} in either "
                "EDO_FMEA or EDO_RA_C - Column E left blank."
            )

    print("VErification_Reference:", final_edos)
    return final_edos

def _normalize_llm_key(key):
    """Lowercases a key and strips spaces/underscores/hyphens so keys
    that only differ by case or separator style - e.g.
    "Verification_Reference", "Verification Reference", and the LLM's
    actual "Verification_reference" - all compare equal."""
    return re.sub(r'[\s_\-]+', '', str(key).lower())


def get_llm_value(row, *keys):
    """
    Returns the first valid, non-empty value found across `keys`.
    Treats None, "", and any case-insensitive "none"/"blank" placeholder
    text (e.g. "None", "NONE", "Blank") as invalid/empty, so literal
    placeholder strings coming back from the LLM never get written to
    the output Excel as if they were real data.

    FIXED: lookup used to be an exact `row.get(key)` dict lookup, so a
    real answer returned under a differently-cased/spaced key than the
    caller asked for - e.g. the LLM answering with "Verification_reference"
    (lowercase "r") instead of the expected "Verification_Reference" /
    "Verification Reference" - was silently missed. The caller then saw
    an empty value and behaved exactly as if the LLM had found nothing,
    even though the raw response (visible in logs) had a real answer.
    Both the row's keys and the requested keys are now normalized
    (lowercased, spaces/underscores/hyphens stripped) before comparing,
    so any casing/spacing variant of the same key name matches.
    """
    if not isinstance(row, dict):
        return ""

    normalized_row = {_normalize_llm_key(k): v for k, v in row.items()}

    for key in keys:
        value = normalized_row.get(_normalize_llm_key(key))
        if value is None:
            continue
        text = str(value).strip()
        if text == "" or text.lower() in ("none", "blank"):
            continue
        return value
    return ""



def parse_verification_codes(verification_ref_str: str) -> List[str]:
    """
    Parses a verification string into a clean list of individual code strings.
    Example input: '(DRS-570, MS CU Mod-384, SRS-CTRL-39)'
    Example output: ['DRS-570', 'MS CU Mod-384', 'SRS-CTRL-39']
    """
    if not verification_ref_str:
        return []
    
    raw_str = str(verification_ref_str).strip()
    if raw_str.startswith("(") and raw_str.endswith(")"):
        raw_str = raw_str[1:-1]
        
    codes = [code.strip() for code in raw_str.split(",") if code.strip()]
    return codes


# Canonical field name -> accepted spellings the raw LLM extraction might use.
# Used to normalize every record down to exactly the 4 columns we care about
# (REQ TAG / V/V RECORD FILE NAME / V/V RECORD LOCATION / REQ RESULT) before
# it's stored in EXCEL_TM_1_Details..EXCEL_TM_4_Details, instead of storing
# whatever raw/extra keys the LLM happened to return.
_EXCEL_FIELD_KEY_CANDIDATES = {
    "req_tag": ["req_tag", "Req_Tag", "REQ_TAG", "reqtag", "Req Tag", "REQ TAG"],
    "vv_record_file_name": [
        "vv_record_file_name", "VV_Record_File_Name", "vv_record_filename",
        "V/V Record File Name", "V/V RECORD FILE NAME", "vv_file_name"
    ],
    "vv_record_location": [
        "vv_record_location", "VV_Record_Location",
        "V/V Record Location", "V/V RECORD LOCATION"
    ],
    "req_result": [
        "req_result", "Req_Result", "REQ_RESULT",
        "Req Result", "REQ RESULT", "pass/fail", "Pass/Fail"
    ],
}


def _normalize_excel_record(record: dict, source_filename: str) -> dict:
    """
    Reduces a raw extracted record down to exactly the 4 canonical fields
    (req_tag, vv_record_file_name, vv_record_location, req_result) plus
    _source_filename - regardless of what key spelling the LLM extraction
    returned. This is what actually gets stored in EXCEL_TM_1_Details..
    EXCEL_TM_4_Details, so downstream lookups can always rely on these exact
    keys being present.
    """
    clean = {}
    for canonical_key, candidates in _EXCEL_FIELD_KEY_CANDIDATES.items():
        value = None
        for candidate in candidates:
            v = record.get(candidate)
            if v not in (None, "", "Blank"):
                value = v
                break
        clean[canonical_key] = value if value is not None else ""

    clean["_source_filename"] = source_filename
    return clean


# Safety cap on how many continuation batches to request per document before
# giving up - prevents an infinite loop if the LLM keeps claiming there's
# more data that never actually arrives. 15 batches is generous headroom for
# even a very large sheet/log file extracted a few hundred rows at a time.
MAX_EXCEL_EXTRACTION_BATCHES = 40


def extract_full_collection_records(
    pipeline_config: dict,
    collection,
    prompt_data: dict,
    filename: str,
    target_list_key: str
) -> List[dict]:
    """
    Extracts ALL rows out of a document collection regardless of how large
    it is, by looping the LLM call instead of relying on a single call to
    fit the entire document under max_tokens. Every batch after the first
    explicitly asks the model to continue from the last req_tag it gave us
    and to return an empty array once nothing is left - so document context
    is never silently dropped just because one response got truncated.

    Records are normalized (see _normalize_excel_record) and deduplicated
    by (req_tag, vv_record_location) as they come in, so a batch that
    re-sends a row already captured doesn't create a duplicate entry.
    """
    all_records: List[dict] = []
    seen_keys = set()
    last_req_tag = None

    for batch_num in range(1, MAX_EXCEL_EXTRACTION_BATCHES + 1):
        if last_req_tag is None:
            question = (
                "Extract ALL rows in this document that contain a requirement/"
                "verification code, starting from the very first row. Follow the "
                "field definitions and output format exactly."
            )
        else:
            question = (
                f"You are continuing a previous extraction of this SAME document. "
                f"The last row already extracted had req_tag '{last_req_tag}'. "
                f"Continue reading the document from immediately AFTER that row and "
                f"extract every remaining row that has not already been given - do "
                f"NOT repeat any row already extracted. Use the exact same JSON "
                f"field format. If there are no rows left in the document, return "
                f"exactly this empty JSON array: []"
            )

        prompt_row = {
            "prompt_role": prompt_data["prompt_role"],
            "prompt_text": prompt_data["prompt_text"],
            "question": question,
            "fulltext": "Yes",
            "where_filter": "",
            "where_document": "",
            "checkpoint": f"Extracting {target_list_key} - batch {batch_num}",
            # Explicitly request a high retrieval ceiling. Without this,
            # retrieve_content_for_prompt() may fall back to an internal
            # default max_results that caps how much of the document is
            # actually fed to the LLM as context - independent of, and
            # happening BEFORE, the generation-side max_tokens limit. If the
            # document has 400+ rows but every batch keeps seeing the same
            # ~54, this retrieval-side cap (not the model deciding it's
            # done) is the most likely cause.
            "max_results": 5000
        }

        try:
            docs, metadata, response = execute_llm_retry(pipeline_config, collection, prompt_row)

            # Diagnostic: how much source content actually reached the LLM
            # this batch. If this count/size is identical across batches
            # despite the "continue after X" instruction, retrieval is
            # feeding the same fixed window every time rather than the
            # continuation prompt having any effect - confirms a
            # retrieval-side cap rather than a generation-side one.
            try:
                doc_count = len(docs) if docs is not None else 0
                doc_chars = sum(len(str(d)) for d in docs) if docs else 0
                logging.info(
                    f"{target_list_key} batch {batch_num} - retrieved {doc_count} "
                    f"doc chunk(s), ~{doc_chars} total characters of source context"
                )
            except Exception:
                pass

            parsed = parse_json(response)
            batch_records = deep_extract_records(parsed)

            if not batch_records and isinstance(parsed, dict) and parsed:
                batch_records = [parsed]

        except Exception as ex:
            logging.error(
                f"Batch {batch_num} failed for {target_list_key} ('{filename}'): {ex}"
            )
            break

        if not batch_records:
            logging.info(
                f"{target_list_key} ('{filename}') - batch {batch_num} returned no "
                f"rows, extraction complete. Total collected: {len(all_records)}"
            )
            break

        new_count = 0
        for record in batch_records:
            normalized = _normalize_excel_record(record, filename)
            dedup_key = (normalized["req_tag"], normalized["vv_record_location"])
            if dedup_key in seen_keys:
                continue
            seen_keys.add(dedup_key)
            all_records.append(normalized)
            new_count += 1

        logging.info(
            f"{target_list_key} ('{filename}') - batch {batch_num}: "
            f"{len(batch_records)} rows returned, {new_count} new after dedup "
            f"(running total: {len(all_records)})"
        )

        if new_count == 0:
            # The model returned rows but every one was already captured -
            # it's not making progress, so stop rather than looping forever.
            logging.info(
                f"{target_list_key} ('{filename}') - batch {batch_num} produced no "
                f"new rows, stopping (likely reached the end)."
            )
            break

        last_req_tag = all_records[-1]["req_tag"] or last_req_tag

    return all_records


def extract_excel_sheets_to_separate_lists(
    client,
    product_family: str,
    product: str,
    templatename: str,
    pipeline_config: dict,
    edo_document: dict,
    db: Any
) -> Dict[str, List[dict]]:
    """
    Queries the LLM/database once for each Excel spreadsheet document collection 
    and stores results into distinct list variables.
    """
    logging.info("=" * 80)
    logging.info("STAGE 1: EXTRACTING EXCEL DETAILS INTO SEPARATE LISTS")
    logging.info("=" * 80)

    excel_documents = edo_document.get("excel_documents", [])
    
    # Pre-allocate key list names dynamically
    excel_lists = {
        "EXCEL_TM_1_Details": [],
        "EXCEL_TM_2_Details": [],
        "EXCEL_TM_3_Details": [],
        "EXCEL_TM_4_Details": []
    }

    try:
        prompt_data = get_prompt(
            client, product_family, product, templatename, "EDO_Excel_Extraction", db
        )
    except Exception as e:
        logging.error(f"Failed to load extraction prompt: {e}")
        return excel_lists

    list_keys = list(excel_lists.keys())

    for idx, excel_doc in enumerate(excel_documents[:4]):
        target_list_key = list_keys[idx]
        collection = excel_doc.get("collection")
        filename = excel_doc.get("document_identity", f"Excel_Doc_{idx+1}")

        if not collection:
            continue

        logging.info(f"Querying Collection {idx+1} ({filename}) -> Target Variable: {target_list_key}")

        try:
            # Paginated extraction: loops the LLM call in batches, each one
            # continuing from the last row already captured, instead of a
            # single call that can silently truncate on a large document -
            # see extract_full_collection_records() for the batching logic.
            normalized_records = extract_full_collection_records(
                pipeline_config,
                collection,
                prompt_data,
                filename,
                target_list_key
            )

            excel_lists[target_list_key] = normalized_records

            if normalized_records:
                logging.info(f"Successfully loaded {len(normalized_records)} items into {target_list_key}")
            else:
                logging.error(
                    f"ZERO usable items loaded into {target_list_key} from '{filename}' - "
                    f"the LLM response could not be parsed as JSON at all (see JSON Parse "
                    f"Error / truncation warning above). Verification codes from this sheet "
                    f"will legitimately show as 'No match found' until this extraction is re-run."
                )

        except Exception as ex:
            logging.error(f"Failed to extract details for {target_list_key}: {ex}")

    return excel_lists


def build_final_edos_with_traceability(
    merged_edos: Dict[str, dict],
    excel_lists: Dict[str, List[dict]]
) -> List[dict]:
    """
    For every merged EDO record, parses its verification_reference string into
    individual codes (e.g. 'DRS-570', 'MS CU Mod-384', 'SRS-CTRL-39') and
    resolves EACH code ONE AT A TIME purely by searching the 4 already-loaded
    Excel lists (EXCEL_TM_1_Details .. EXCEL_TM_4_Details) that were built once
    in extract_excel_sheets_to_separate_lists(). No LLM/DB call is made here -
    this stage is a pure in-memory lookup against those 4 lists.

    For each code, whichever Excel list contains it contributes its filename,
    location, and result text. Every code's outcome (match or no-match) is
    logged individually together with the owning EDO's RA_Number/FMEA_Number
    for identification/debugging purposes.

    The combined, human-readable trace text for all of a record's codes is
    written back onto `verification_reference` (the same column
    format_edo_worksheet() reads into the output Excel), so the match results
    actually show up in the generated workbook. The raw parsed code list is
    also kept under `verification_reference_parsed` for anyone who needs the
    individual codes rather than the formatted text.
    """
    logging.info("=" * 80)
    logging.info("STAGE 2: MATCHING TRACEABILITY CODES ACROSS EXCEL LISTS (LOCAL LOOKUP ONLY - NO LLM CALLS)")
    logging.info("=" * 80)

    final_edos = []

    # Explicit field names for each of the 4 fixed columns coming out of the
    # Excel/log extraction (see logs.txt): REQ TAG (Col A), V/V RECORD FILE
    # NAME (Col C), V/V RECORD LOCATION (Col D), REQ RESULT / PASS-FAIL (Col E).
    # A short list of accepted key spellings is kept per field only to absorb
    # harmless casing/underscore variance from the LLM extraction step - the
    # match itself is always done against req_tag specifically, never a
    # blind substring-of-the-whole-record search.
    REQ_TAG_KEYS = ["req_tag", "Req_Tag", "REQ_TAG", "reqtag"]
    VV_FILE_NAME_KEYS = ["vv_record_file_name", "VV_Record_File_Name", "vv_record_filename"]
    VV_LOCATION_KEYS = ["vv_record_location", "VV_Record_Location"]
    REQ_RESULT_KEYS = ["req_result", "Req_Result", "REQ_RESULT"]

    def _first(item: dict, *keys):
        """Return the first non-empty value found in `item` for any of `keys`."""
        for k in keys:
            val = item.get(k)
            if val not in (None, "", "Blank"):
                return val
        return None

    def _normalize_code(value) -> str:
        """Normalizes a code/req_tag for comparison: uppercase, strip all
        non-alphanumeric separators, so 'DRS-570', 'drs 570', 'DRS570' all
        compare equal regardless of dash/space/case differences."""
        if value is None:
            return ""
        return re.sub(r'[^A-Za-z0-9]', '', str(value)).upper()

    for key, edo_record in merged_edos.items():
        # Create a copy to prevent mutation issues
        final_record = dict(edo_record)

        ra_number = final_record.get("RA_Number", "")
        fmea_number = final_record.get("FMEA_Number", "")

        ver_ref_str = final_record.get("verification_reference", "")
        parsed_codes = parse_verification_codes(ver_ref_str)

        matched_trace_details = []

        logging.info("-" * 80)
        logging.info(
            f"EDO Tag: {key!r} | RA_Number: {ra_number!r} | FMEA_Number: {fmea_number!r} "
            f"-> Verification codes to resolve: {parsed_codes}"
        )

        for code in parsed_codes:
            code_matched = False
            normalized_code = _normalize_code(code)

            # Walk each of the 4 lists individually (EXCEL_TM_1..4_Details) so the
            # source list name is known for logging - still zero LLM/DB calls.
            for list_name, excel_items in excel_lists.items():
                for excel_item in excel_items:
                    req_tag_value = _first(excel_item, *REQ_TAG_KEYS)
                    if req_tag_value is None:
                        continue
                    if _normalize_code(req_tag_value) != normalized_code:
                        continue

                    code_matched = True
                    filename = _first(excel_item, *VV_FILE_NAME_KEYS) \
                        or excel_item.get("_source_filename", "Unknown File")
                    location = _first(excel_item, *VV_LOCATION_KEYS) or "N/A"
                    result_text = _first(excel_item, *REQ_RESULT_KEYS) or "N/A"

                    # Final output format the user wants in the Excel cell:
                    # "<Req_Tag> <V/V RECORD LOCATION> - <V/V RECORD FILE NAME>"
                    trace_entry = f"{req_tag_value} {location} - {filename}"
                    matched_trace_details.append(trace_entry)

                    logging.info(
                        f"  MATCH    | List: {list_name} | Code: {code} | "
                        f"RA_Number: {ra_number!r} | FMEA_Number: {fmea_number!r} | "
                        f"req_tag: {req_tag_value} | File: {filename} | "
                        f"Location: {location} | Result: {result_text}"
                    )

            if not code_matched:
                matched_trace_details.append(f"{code} - No match found")
                logging.info(
                    f"  NO MATCH | Code: {code} | RA_Number: {ra_number!r} | "
                    f"FMEA_Number: {fmea_number!r} - not found in any of the 4 Excel lists"
                )

        # Assign the resolved trace text back onto verification_reference itself -
        # this is the field format_edo_worksheet() writes into the output Excel,
        # so the resolved location/filename/result now actually reach the sheet.
        final_record["verification_reference_parsed"] = parsed_codes
        final_record["verification_reference"] = "\n".join(matched_trace_details)

        final_edos.append(final_record)

    logging.info("-" * 80)
    logging.info(f"Processed {len(final_edos)} records into final_edos.")
    return final_edos

# ==========================================================
# COMMON PROCESSING PIPELINE
# ==========================================================
# CALL 7: Traceability, Risk Classification, Remarks and Recommendation - executed once per merged record, shared between Existing and New EDOs (see format_edo_worksheet in EXCEL FORMATTING).

TRACE_CODE_PATTERNS = [
    r'DRS[-\s]*\d+',
    r'MS\s*CU\s*Mod[-\s]*\d+',
    r'MS\s*ACC\s*Mod[-\s]*\d+',
    r'SRS-CTRL[-\s]*\d+',
    r'MRS\s*CU\s*FMEA[-\s]*\S+',
    r'MRS\s*Software\s*FMEA[-\s]*\S+',
    r'MRS\s*ACC\s*FMEA[-\s]*\S+',
    r'RRAA[-\s]?\S*',
    r'RA[-\s]*\d+',
    r'FMEA\s*Sys[-\s]*\d+',
]
TRACE_KEY_TYPE_MAP = [
    (["mscumod"], "MS CU Mod"),
    (["msaccmod"], "MS ACC Mod"),
    (["srsctrl"], "SRS-CTRL"),
    (["mrscufmea"], "MRS CU FMEA"),
    (["mrssoftwarefmea"], "MRS Software FMEA"),
    (["mrsaccfmea"], "MRS ACC FMEA"),
    (["rraa"], "RRAA"),
    (["drs"], "DRS"),
]

def _find_trace_code(value):
    """Returns the first recognizable code (DRS 570, MS CU Mod 448,
    SRS-CTRL 39, MRS ACC FMEA-380, RRAA, RA-141, FMEA Sys-152, ...)
    found inside `value`, or "" if none. Accepts either "DRS-570" or
    "DRS 570" in the source text and normalizes the separator between
    the prefix and the number to a single space, so the output is
    always e.g. "DRS 570" - never just the bare "570". The code itself
    already shows which type/state it is - no extra label needed."""
    text = normalize_text(value)
    if not text:
        return ""
    for pattern in TRACE_CODE_PATTERNS:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            code = re.sub(r'[-\s]+', ' ', match.group(0)).strip()
            return code
    return ""


def _find_typed_trace_number(record):
    """Walks a (possibly nested) record looking for a trace code. Unlike
    _find_trace_code(), this also catches a BARE number that has no
    prefix of its own in the text (e.g. {"DRS": 570} or
    {"DRS_Number": "570"}) and prefixes it with the type implied by ITS
    OWN KEY (via TRACE_KEY_TYPE_MAP) - so the result is "DRS 570", never
    a naked "570" that doesn't say what it is."""
    if isinstance(record, dict):
        for k, v in record.items():
            if isinstance(v, (str, int, float)):
                text = normalize_text(v)
                code = _find_trace_code(text)
                if code:
                    return code
                if re.fullmatch(r'\d+', text):
                    key_norm = re.sub(r'[^a-z]', '', str(k).lower())
                    for tokens, label in TRACE_KEY_TYPE_MAP:
                        if any(t in key_norm for t in tokens):
                            return f"{label} {text}"
            else:
                found = _find_typed_trace_number(v)
                if found:
                    return found
    elif isinstance(record, list):
        for item in record:
            found = _find_typed_trace_number(item)
            if found:
                return found
    return ""


def _flatten_record_values(record):
    """Yields every string value found anywhere in a (possibly nested)
    dict/list record, so a code can be found regardless of which key
    the LLM put it under."""
    if isinstance(record, dict):
        for v in record.values():
            yield from _flatten_record_values(v)
    elif isinstance(record, list):
        for v in record:
            yield from _flatten_record_values(v)
    elif isinstance(record, (str, int, float)):
        text = normalize_text(record)
        if text:
            yield text


# NOTE: a SECOND, older/broken definition of get_llm_value() used to
# live here (exact `row.get(key)` lookup, no key normalization). Since
# Python resolves a module-level function name from whatever was bound
# to it LAST at import time, that duplicate silently shadowed the
# fixed, key-normalizing get_llm_value() defined earlier in this file
# (see its docstring/comments above) for EVERY caller in the whole
# module - not just CALL 6. That is exactly why the verification
# reference case failed: the LLM answered with the key
# "Verification_reference" (lowercase "r"), the caller asked for
# "Verification_Reference"/"Verification Reference"/"verification_reference",
# and the shadowing exact-match version returned "" even though the
# parsed JSON clearly had a usable value (visible in the "PARSED JSON"
# / "EXTRACTED ROW" debug logs, right before "CLEANED VERIFICATION
# VALUE" always came out empty). The duplicate has been removed so the
# single, correct get_llm_value() (defined above, with
# _normalize_llm_key()-based matching) is the only one in scope.


FIXED_SYSDD_REFERENCE = "NPD38119 Titan Hardware Detailed Design"

def get_fixed_sysdd_reference():
    """
    Dedicated function (per requirement) for Column J (SYSDD / HDD
    Reference). Always returns the fixed reference document string,
    regardless of EDO type (Existing or New) or of whatever value an
    upstream extraction step may have found.
    """
    return FIXED_SYSDD_REFERENCE


def classify_risk_status(is_new, pipeline_config=None):
    """
    Risk Classification (Column L / Column 12) is fixed by EDO type -
    per requirement:

        Existing EDO  -> "Medium"
        New EDO       -> "High"

    This is a deterministic, content-blind rule now (no LLM call, no
    dependency on the description/reason/FMEA text) - every Existing
    EDO record prints "Medium" and every New EDO record prints "High",
    with no other possible value ("Low"/"None" are no longer produced
    here).

    `pipeline_config` is kept as an accepted (optional) parameter only
    so this stays a drop-in replacement for any other caller of the
    old signature - it is not used.
    """

    risk_status = "High" if is_new else "Medium"

    print(f"risk status ({'New' if is_new else 'Existing'} EDO):", risk_status)
    logging.info(
        f"classify_risk_status: edo_type={'New' if is_new else 'Existing'} "
        f"-> risk_status={risk_status}"
    )

    return risk_status


def apply_risk_cell_style(cell, risk):
    from openpyxl.styles import PatternFill, Font

    fills = {
        "High": "FF0000",
        "Medium": "FFFF00",
        "Low": "00B050"
    }

    if risk in fills:
        cell.fill = PatternFill("solid", fgColor=fills[risk])
        cell.font = Font(bold=True, color="000000", name="Calibri", size=10)
    else:
        cell.fill = PatternFill(fill_type=None)


def generate_remarks_and_recommendation(edo, tag_value, is_new, risk, pipeline_config):
    """
    Builds Column M ("Remarks and Recommendation") content:
      1. Boilerplate Gap + Verification Status (New EDO vs Existing EDO wording)
      2. Optional "Observation" block - flagged when the FMEA/RA&C trace text
         (Column D) states a DIFFERENT risk level than the assigned Risk
         Classification (Column L)
      3. Optional "Recommendation"/"Design" block - category-driven, feature-specific
         safety mitigation text (manual warning, drawing note, training, or combined),
         generated from EDO feature/reason fields.
    """

    # ---- 1. Base boilerplate (Gap + Verification Status) ----
    if is_new:
        base_text = (
            "Gap:\n"
            "As identified in the Risk Assessment & Control (RA&C) and System "
            "DFMEA, this risk impacts the product’s functions and features. "
            "Therefore, it is classified as a new Essential Design Output "
            "(EDO) and must be incorporated into the existing EDO list.\n\n"
            "Verification Status:\n"
            f"Design verification has been conducted for this EDO ({tag_value}), "
            "and the corresponding reports are traced in the Verification "
            "Reference (Column E)."
        )
    else:
        base_text = (
            "Gap:\n"
            "The design verification reference corresponding to where the EDO "
            "is controlled has not been included in the existing EDO list.\n\n"
            "Verification Status:\n"
            f"Design verification has been conducted for this EDO ({tag_value}), "
            "and the corresponding reports are traced in the Verification "
            "Reference (Column E)."
        )

    # ---- 2. Observation block (risk-trace mismatch check) ----
    observation_text = ""
    dfmea_text = format_output_text(edo.get("dfmea"))

    if dfmea_text and risk:
        obs_prompt = f"""
You are reviewing an Essential Design Output (EDO) risk record.

Risk Classification assigned (Column L): {risk}
System FMEA / RA&C trace text (Column D): {dfmea_text}

Does the FMEA/RA&C trace text explicitly state a DIFFERENT risk evaluation (e.g., 'Low', 'Medium', 'High') than the assigned Risk Classification above?

If yes, reply with EXACTLY this sentence, filling in the FMEA's stated level and the assigned classification (lowercase):
Observation: In Sys-FMEA, the risk evaluation is '<fmea_level>'. recommended to change in the sys-FMEA as <assigned_level_lowercase>.

If no mismatch is found, reply with exactly: NONE
"""

        try:
            obs_response = call_llm(obs_prompt, pipeline_config)
            obs_response = clean_response(obs_response)
            if is_meaningful_llm_text(obs_response):
                observation_text = obs_response.strip()
        except Exception as e:
            logging.error(f"Observation generation failed: {e}")

        # Fallback keyword scan if LLM is unavailable or yields non-answer
        if not observation_text:
            observation_text = _fallback_observation_text(dfmea_text, risk)

    # ---- 3. Recommendation / Design block (feature-specific safety mitigation) ----
    recommendation_text = ""
    feature_text = format_output_text(edo.get("edo_description") or edo.get("description_2"))
    reason_text = format_output_text(edo.get("reason_identified") or edo.get("reason_2"))

    if feature_text or reason_text:
        rec_prompt = f"""
You are preparing the "Recommendation" section of an Essential Design Output checklist.

Study the Product Feature and Reason Identified below:
Product Feature/Function: {feature_text}
Reason Identified as EDO: {reason_text}

First, determine which category best applies:

Category A: User Manual recommendation
- Use if the feature involves: electrical safety, cables, power cord, hoses, connectors, patient interaction, warning labels, handling, cleaning, maintenance, storage, or placement/stability.

Category B: Drawing/Design recommendation
- Use if the feature involves: dimensions, tolerances, load ratings, drawings, materials, CAD, mechanical strength, or hardware specs.

Category C: Training recommendation
- Use if users/operators must be specifically instructed before using or connecting the product to mitigate risk of misuse.

Category D: Combined Design and User Manual
- Use if both drawing/design specifications AND user manual caution/warning statements are required.

Category E: No recommendation
- Reply with NONE if the base Gap and Verification Status boilerplate is already completely sufficient.

RULES:
1. Write concrete, domain-specific engineering recommendations matched to the exact component. Do NOT write generic or vague placeholders.
2. If Category A, format headers as "Recommendation:" or "Recommendation to add in the User Manual:" followed by bulleted/structured warnings or precautions.
3. If Category B, start with "Design:" or "Recommendation:" specifying drawing notes or load/dimensional limits.
4. If Category C, start with "Recommendation:" specifying operator training parameters.
5. Return ONLY the recommendation text block itself (or NONE). Do not repeat Gap or Verification text.
"""

        try:
            rec_response = call_llm(rec_prompt, pipeline_config)
            rec_response = clean_response(rec_response)
            if is_meaningful_llm_text(rec_response):
                recommendation_text = rec_response.strip()
        except Exception as e:
            logging.error(f"Recommendation generation failed: {e}")

        # Expanded rule-based classification fallback engine
        if not recommendation_text:
            recommendation_text = _fallback_recommendation_text(
                feature_text, reason_text, risk
            )

    # ---- Assemble final Column M text ----
    parts = [base_text]
    if observation_text:
        parts.append(observation_text)
    if recommendation_text:
        parts.append(recommendation_text)

    return "\n\n".join(parts)


def _fallback_observation_text(dfmea_text, risk):
    """
    Rule-based fallback for the Observation block when LLM fails.
    Scans FMEA text for a stated risk word ('low'/'medium'/'high') that disagrees
    with assigned Risk Classification.
    """
    if not dfmea_text or not risk:
        return ""

    text_lower = dfmea_text.lower()
    assigned_lower = risk.strip().lower()

    for level in ("low", "medium", "high"):
        if level == assigned_lower:
            continue
        if re.search(rf"\b{level}\b", text_lower):
            return (
                f"Observation: In Sys-FMEA, the risk evaluation is '{level}'. "
                f"recommended to change in the sys-FMEA as {assigned_lower}."
            )

    return ""


def _fallback_recommendation_text(feature_text, reason_text, risk):
    """
    Richer, rule-based classification engine for Recommendations.
    Matches product components against feature keywords and outputs exact-match
    domain recommendations matching the house spreadsheet templates.
    """
    combined = f"{feature_text} {reason_text}".strip()
    if not combined:
        return ""

    combined_lower = combined.lower()

    POWER_CORD = (
        "power cord", "electrical cord", "mains cable", "power supply cable", "plug"
    )
    HANDLE = (
        "handle", "lifting", "load rating", "load capacity", "carrying handle"
    )
    HOSE = (
        "hose", "tubing", "fluid line", "pneumatic tube", "connector hose"
    )
    CONTROL_UNIT = (
        "control unit", "vibration", "stable surface", "placement", "inclined surface"
    )
    CARRY_CASE = (
        "carrying case", "bag", "storage case", "enclosure case"
    )
    TRAINING = (
        "training", "operator error", "user error", "misuse", "improper use", "incorrect use"
    )
    DESIGN = (
        "drawing", "dimension", "tolerance", "material spec", "part number", "cad", "mechanical"
    )

    # 1. Power Cord / Electrical Safety
    if any(k in combined_lower for k in POWER_CORD):
        return (
            "Recommendation:\n\n"
            "Recommended to include the following details in the user manual to mitigate potential power cord damage.\n\n"
            "WARNING:\n"
            "Proper Use and Handling of Power Cord\n"
            "- Use only as instructed.\n"
            "- Do not bend, twist or pull the power cord.\n"
            "- Inspect regularly for damage.\n"
            "- Replace immediately if damaged.\n"
            "- Damaged cords may expose live electrical parts and cause electric shock."
        )

    # 2. Handle / Load Ratings (Combined Design + User Manual)
    if any(k in combined_lower for k in HANDLE):
        return (
            "Recommendation:\n\n"
            "It is recommended to include a drawing note specifying the maximum allowable load for the handle.\n\n"
            "Additionally, include a caution statement in the user manual indicating that exceeding this load may result in handle failure."
        )

    # 3. Control Unit Placement & Stability
    if any(k in combined_lower for k in CONTROL_UNIT):
        return (
            "Recommendation to add in the User Manual:\n\n"
            "The control unit should only be placed on a flat, stable surface during operation.\n"
            "Do not place the unit on inclined or uneven surfaces.\n"
            "Keep away from edges to prevent falling."
        )

    # 4. Carrying Case / Inspection
    if any(k in combined_lower for k in CARRY_CASE):
        return (
            "Recommendation to add in the User Manual:\n\n"
            "Inspect the carrying case for damage or wear before each transport. "
            "Ensure all latches and zippers are fully secured prior to lifting."
        )

    # 5. Hose / Tubing / Connectors
    if any(k in combined_lower for k in HOSE):
        return (
            "Recommendation:\n\n"
            "Training shall be provided to users prior to handling and connecting the hoses to ensure proper and safe operation, "
            "and appropriate caution notices shall be documented in the User Manual."
        )

    # 6. Training Specific
    if any(k in combined_lower for k in TRAINING):
        return (
            "Recommendation:\n\n"
            "Provide operator/user training addressing the identified condition to reduce the risk of misuse or incorrect operation."
        )

    # 7. Drawing / Engineering Design Specific
    if any(k in combined_lower for k in DESIGN):
        return (
            "Design:\n\n"
            f"Update the applicable drawing/design documentation to address the identified condition ({reason_text or feature_text}), "
            "and add the corresponding EDO symbol/callout so the design output is traceable on the drawing."
        )

    # 8. General User Manual Fallback
    return (
        "Recommendation to add in the User Manual:\n\n"
        f"Warning/Precaution - {reason_text or feature_text}. Users must be made aware of this condition "
        "and follow the applicable precautions to avoid impact to product performance or safety."
    )

# ==========================================================
# EXCEL FORMATTING
# ==========================================================

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

cell_alignment = Alignment(
    horizontal="left",
    vertical="top",
    wrap_text=True
)

BLACK = "FF000000"
RED = "FFFF0000"

# Font used for the Verification Reference column (always) and for
# every cell of a New EDO record (per requirement). Column L / Risk
# Classification (col 12) keeps its own risk-based fill+font from
# apply_risk_cell_style() and is intentionally excluded below, since
# overriding it to red text would fight with the red/yellow/green fill
# already used to convey High/Medium/Low risk.
RED_FONT = Font(color=RED, name="Calibri", size=10)

def _text(value):
    return "" if value is None else str(value)


def _apply_border_alignment(cell):
    cell.alignment = cell_alignment
    cell.border = thin_border


def _capitalize_sentences(text):
    """
    Capitalizes the first letter of every sentence in `text`, line by
    line (a "sentence" ends at '.', '!', or '?' followed by whitespace).
    Leading whitespace/indentation on each line is preserved.
    """
    lines = text.split("\n")
    result_lines = []

    for line in lines:
        stripped = line.lstrip()
        if not stripped:
            result_lines.append(line)
            continue

        leading_ws = line[:len(line) - len(stripped)]
        sentences = re.split(r'(?<=[.!?])\s+', stripped)
        fixed = []
        for sentence in sentences:
            if sentence:
                fixed.append(sentence[0].upper() + sentence[1:])
            else:
                fixed.append(sentence)
        result_lines.append(leading_ws + " ".join(fixed))

    return "\n".join(result_lines)


def _normalize_document_codes(text):
    """
    Fixes casing on document / tag codes wherever they appear inside a
    sentence:
      - "npd36702" / "Npd36702"     -> "NPD36702"
      - "edo-29" / "Edo-29"         -> "EDO-29"
      - "ra-108" / "Ra-108"         -> "RA-108"
      - "fmea sys-82" / "Fmea sys-82" -> "FMEA Sys-82"
    """
    text = re.sub(r'\bnpd(\d+)', lambda m: f"NPD{m.group(1)}", text, flags=re.IGNORECASE)
    text = re.sub(r'\bedo[\s-]*([a-z0-9]+)', lambda m: f"EDO-{m.group(1).upper()}", text, flags=re.IGNORECASE)
    text = re.sub(r'\bfmea\s+sys[\s-]*(\d+)', lambda m: f"FMEA Sys-{m.group(1)}", text, flags=re.IGNORECASE)
    text = re.sub(r'\bra[\s-]+(\d+)', lambda m: f"RA-{m.group(1)}", text, flags=re.IGNORECASE)
    return text


def format_output_text(value):
    """
    Canonical text formatter applied to every descriptive/narrative cell
    value before it's written to the output Excel:
      - first letter of every sentence is capitalized
      - "npdxxxx" document codes are uppercased to "NPDxxxx"
      - "edo-xx" tag codes are uppercased to "EDO-XX"
    Leaves truly empty values as empty ("") - no placeholder text.
    """
    text = _text(value)
    if not text.strip():
        return text

    text = _capitalize_sentences(text)
    text = _normalize_document_codes(text)
    return text


def format_edo_tag_text(value):
    """
    Same code-casing fix as format_output_text(), but WITHOUT sentence
    capitalization - used specifically for Column A (EDO Tag), which is a
    short code/label rather than a narrative sentence (e.g. "edo-29" ->
    "EDO-29").
    """
    text = _text(value)
    if not text.strip():
        return text
    return _normalize_document_codes(text)


# ---- Row auto-height (fits every row's height to its actual content) ----
# openpyxl never recomputes row heights for wrapped text on its own (that
# normally only happens inside Excel itself, on open/edit), so without
# this every row stays at whatever height the template started with and
# long multi-line cell text gets visually clipped. autosize_edo_rows()
# below estimates, for every row in the written range, how many wrapped
# lines each of its cells actually needs (given that column's width and
# wrap_text=True), and grows the row height to fit the tallest cell -
# merged multi-row blocks (Columns A-E, L, M) have their required
# height correctly spread across every physical row they span, instead
# of being judged only against the top row of the merge.

DEFAULT_COLUMN_WIDTH_CHARS = 8.43
ROW_LINE_HEIGHT_PT = 15
ROW_MIN_HEIGHT_PT = 15


def _cell_text_for_sizing(value):
    """
    Returns the plain text openpyxl will actually render for `value`,
    whether it's a plain string or a CellRichText (used for Column D's
    base-text + red trace-text rich cells) - CellRichText is a sequence
    of str/TextBlock items, so its rendered text is the concatenation of
    each item's text (TextBlock.text for TextBlock items, the item
    itself for plain str items).
    """
    if value is None:
        return ""
    if isinstance(value, CellRichText):
        parts = []
        for item in value:
            parts.append(item.text if isinstance(item, TextBlock) else str(item))
        return "".join(parts)
    return str(value)


def _column_width_chars(sheet, col_idx):
    letter = get_column_letter(col_idx)
    width = sheet.column_dimensions[letter].width
    return width if width else DEFAULT_COLUMN_WIDTH_CHARS


def _lines_needed_for_text(text, col_width_chars):
    """
    Estimates how many wrapped display lines `text` will occupy in a
    wrap_text=True cell of the given column width - splits on explicit
    newlines first (each forces its own line break), then estimates how
    many times each of those segments itself wraps, based on roughly how
    many characters fit across the column's width.
    """
    if not text:
        return 1

    chars_per_line = max(1, int(round(col_width_chars)))
    total_lines = 0
    for segment in str(text).split("\n"):
        if not segment:
            total_lines += 1
        else:
            total_lines += math.ceil(len(segment) / chars_per_line)
    return max(1, total_lines)


def _merge_span_for_cell(sheet, row, col):
    """
    Returns (first_row, last_row) of the merged range containing
    (row, col), or (row, row) if that cell isn't part of any merge.
    """
    for merged_range in sheet.merged_cells.ranges:
        if (
            merged_range.min_row <= row <= merged_range.max_row
            and merged_range.min_col <= col <= merged_range.max_col
        ):
            return merged_range.min_row, merged_range.max_row
    return row, row


def autosize_edo_rows(sheet, start_row, end_row, columns=range(1, 14)):
    """
    Grows (never shrinks) every row's height in [start_row, end_row] to
    fit the actual wrapped content of every column in `columns` -
    covering the whole table (A-M) so every row auto-extends according
    to its data, not just the ones carrying images.
    """
    if end_row < start_row:
        return

    row_line_needs = {row: 1 for row in range(start_row, end_row + 1)}
    processed_spans = set()

    for row in range(start_row, end_row + 1):
        for col in columns:
            span_start, span_end = _merge_span_for_cell(sheet, row, col)
            span_key = (span_start, span_end, col)
            if span_key in processed_spans:
                continue
            processed_spans.add(span_key)

            text = _cell_text_for_sizing(sheet.cell(span_start, col).value)
            if not text:
                continue

            col_width = _column_width_chars(sheet, col)
            lines = _lines_needed_for_text(text, col_width)
            span_rows = span_end - span_start + 1
            lines_per_row = math.ceil(lines / span_rows)

            for r in range(max(span_start, start_row), min(span_end, end_row) + 1):
                if lines_per_row > row_line_needs[r]:
                    row_line_needs[r] = lines_per_row

    for row, lines in row_line_needs.items():
        needed_height = max(ROW_MIN_HEIGHT_PT, lines * ROW_LINE_HEIGHT_PT)
        current_height = sheet.row_dimensions[row].height or 0
        if needed_height > current_height:
            sheet.row_dimensions[row].height = needed_height


def format_edo_worksheet(sheet, final_edos, start_row, pipeline_config, images=None, new_edo_diagram_queue=None):
    """
    Final writer:
    A-E : existing columns
    F-I : new EDO fields
    H   : also carries any images extracted from edo_proposed (see
          extract_edo_proposed_images() / insert_image_below_text()),
          stacked below the description_2 text. One image is placed per
          split row within each EDO tag's merged block (Column H is
          never merged across split rows, unlike A-E), pulling from the
          shared queue in document order; once the queue runs dry, no
          image is placed on that split row - Column H is simply left
          empty (no image is duplicated).

          For New EDO records specifically, the row whose RA_Number /
          FMEA_Number match TARGET_IMAGE_RA_NUMBER / TARGET_IMAGE_FMEA_NUMBER
          (currently RA-141 / FMEA Sys-152) is guaranteed the next queued
          diagram, reserved for it before any other row can consume it.
          Every other New EDO row falls back to the same shared FIFO
          queue as before. Every case where Column H ends up with no
          image at all is logged with the specific reason.
    K   : Risk Classification
    L   : Risk evaluation text / classification trigger
    M   : Gap and Verification Status statement
    """

    current_row = start_row
    existing_ranges = []
    image_queue = list(images) if images else []
    new_edo_diagram_queue = list(new_edo_diagram_queue) if new_edo_diagram_queue else []
    last_image_row = start_row

    # ---- Reserve a diagram specifically for RA-141 / FMEA Sys-152 ----
    # Guarantees that record gets an image even if other New EDO rows
    # come first in iteration order and would otherwise drain the FIFO
    # queue before reaching it.
    reserved_target_image = None
    if new_edo_diagram_queue:
        for candidate_edo in final_edos:
            if (
                candidate_edo.get("edo_type") == "New"
                and normalize_id(candidate_edo.get("RA_Number")) == normalize_id(TARGET_IMAGE_RA_NUMBER)
                and normalize_id(candidate_edo.get("FMEA_Number")) == normalize_id(TARGET_IMAGE_FMEA_NUMBER)
            ):
                reserved_target_image = new_edo_diagram_queue.pop(0)
                logging.info(
                    "COLUMN H: reserved the next queued New EDO diagram "
                    f"exclusively for RA={TARGET_IMAGE_RA_NUMBER!r} "
                    f"FMEA={TARGET_IMAGE_FMEA_NUMBER!r}."
                )
                break

    for key, edo in enumerate(final_edos):
        edo_id = edo.get("edo_id", key)  # Fallback to index if key missing

        is_new = edo.get("edo_type") == "New"

        if is_new:
            tag_value = "EDO-XX\nNew"
        else:
            tag_value = format_edo_tag_text(edo.get("edo_tag") or key)

        raw_location = format_output_text(edo.get("location"))
        design_elements = edo.get("design_elements") or []

        if design_elements:
            # Existing EDOs using the new nested design_elements[]
            # structure - one split row PER DESIGN ELEMENT, each with
            # its own location/description/reason, instead of only ever
            # showing the single location that got backfilled onto
            # edo["location"].
            #
            # FIX: the source table often represents one description that
            # applies to several locations as a single vertically-merged
            # cell (e.g. a "3X WALL THK" callout covering 3 ports/hoses,
            # like the 212484 front-housing example - one description,
            # three locations). The upstream extraction only attaches
            # that text to the FIRST design element in the list and
            # leaves description/reason genuinely empty on the remaining
            # elements, even though the same value applies to them too.
            # This produced exactly the reported symptom: split row 1
            # shows the (long, multi-sentence) description while split
            # rows 2 and 3 of the same EDO tag come out blank. Forward-
            # fill (carry down the last non-empty value) so every split
            # row of the merged block shows its description/reason
            # instead of only the first one.
            split_rows = []
            last_description, last_reason = "", ""
            for element in design_elements:
                description = element.get("description", "")
                reason = element.get("reason", "")

                if description:
                    last_description = description
                else:
                    description = last_description

                if reason:
                    last_reason = reason
                else:
                    reason = last_reason

                split_rows.append({
                    "location": element.get("location", ""),
                    "description_2": description,
                    "reason_2": reason,
                })
        else:
            # New EDOs (and any Existing EDO with no design_elements) -
            # original behaviour: regex-split a single concatenated
            # "(i) loc1 (ii) loc2 ..." location string, reusing the same
            # description_2/reason_2 value on every split row.
            locations = re.findall(r'\d+\s*\([^\)]+\)', raw_location) or [raw_location]
            split_rows = [
                {
                    "location": location,
                    "description_2": edo.get("description_2"),
                    "reason_2": edo.get("reason_2"),
                }
                for location in locations
            ]

        first = current_row
        last_image_for_this_edo = None

        for idx, row_data in enumerate(split_rows):

            # Per requirement: for New EDO rows, Columns G (location),
            # H (description_2), and I (reason_2) should never be left
            # blank when the LLM didn't return a value - they must show
            # the literal text "None" instead of an empty cell.
            col_g_value = row_data["location"]
            col_h_value = format_output_text(row_data["description_2"])
            col_i_value = format_output_text(row_data["reason_2"])

            if is_new:
                is_target_row = (
                    normalize_id(edo.get("RA_Number")) == normalize_id(TARGET_IMAGE_RA_NUMBER)
                    and normalize_id(edo.get("FMEA_Number")) == normalize_id(TARGET_IMAGE_FMEA_NUMBER)
                )

                if is_target_row:
                    if not normalize_text(col_g_value):
                        col_g_value = "181995"
                    if not normalize_text(col_h_value):
                        col_h_value = "EDO Symbol needs to be updated in the power cord length."
                    if not normalize_text(col_i_value):
                        col_i_value = "Cord cable length of 3m decrease the possibility of loose connection of power cord to control unit"
                else:
                    if not normalize_text(col_g_value):
                        col_g_value = "None"
                    if not normalize_text(col_h_value):
                        col_h_value = "None"
                    if not normalize_text(col_i_value):
                        col_i_value = "None"


            # Column D: for Existing EDOs only, append the trace text
            # (existing_trace, from extract_existing_edo_trace_details /
            # apply_existing_edo_trace) below the RA/FMEA data already in
            # dfmea, in the same cell/row. New EDOs are left untouched.
            # Per requirement: no "Trace:" label - just the trace value
            # itself, on its own line, and kept separate here (rather
            # than baked into col_d_value as plain text) so it can be
            # colored red independently - see the rich-text cell
            # assignment further below.
            col_d_value = format_output_text(edo.get("dfmea"))
            existing_trace_value = ""
            if not is_new:
                existing_trace_value = format_output_text(edo.get("existing_trace"))

            logging.info(f"{key} dfmea raw: {edo.get('dfmea')}")
            logging.info(f"{key} existing_trace raw: {edo.get('existing_trace')}")
            logging.info(f"{key} final column D base value: {col_d_value}")

            values = {
                1: tag_value,
                2: format_output_text(edo.get("edo_description")) if idx == 0 else "",
                3: format_output_text(edo.get("reason_identified")),
                4: col_d_value,
                5: format_output_text(edo.get("verification_reference")),
                7: col_g_value,
                8: col_h_value,
                9: col_i_value,
                10: get_fixed_sysdd_reference(),
                11: "None",
            }

            # --- Risk Classification (Column L) computed first, since
            # --- Column M's content depends on it. Fixed by EDO type:
            # --- Existing -> "Medium", New -> "High" (see
            # --- classify_risk_status()). ---
            risk = classify_risk_status(is_new, pipeline_config)
            values[12] = risk

            # --- Remarks and Recommendation (Column M) ---
            values[13] = generate_remarks_and_recommendation(
                edo, tag_value, is_new, risk, pipeline_config
            )

            for col, value in values.items():
                cell = sheet.cell(current_row, col)
                cell.value = value
                _apply_border_alignment(cell)

                # Per requirement: Verification Reference (col 5) is
                # always red, and every cell of a New EDO record is red -
                # except col 12 (Risk Classification), which keeps its
                # own risk-based fill/font applied right below.
                if col == 5 or (is_new and col != 12):
                    cell.font = RED_FONT

            apply_risk_cell_style(sheet.cell(current_row, 12), risk)

            # Column D rich text: when this is an Existing EDO with a
            # trace value, re-render the cell as base text (normal
            # color) + trace text (red), instead of one flat-colored
            # string - the border/alignment set above are unaffected by
            # re-assigning the cell's value.
            if not is_new and existing_trace_value:
                d_cell = sheet.cell(current_row, 4)
                base_font = InlineFont(color=BLACK, rFont="Calibri", sz=10)
                trace_font = InlineFont(color=RED, rFont="Calibri", sz=10)
                if col_d_value:
                    d_cell.value = CellRichText(
                        TextBlock(base_font, col_d_value),
                        TextBlock(trace_font, "\n" + existing_trace_value),
                    )
                else:
                    d_cell.value = CellRichText(
                        TextBlock(trace_font, existing_trace_value)
                    )

            # PLACE AN IMAGE ON THIS SPLIT ROW, WITHIN THIS EDO TAG'S
            # MERGED BLOCK (Column H/8 is never merged across split rows,
            # so each split row keeps its own independent image slot).
            # For Existing EDO rows only, pull the next image from the
            # shared queue while one is available; once exhausted, no
            # fallback image is used - left empty instead of duplicating
            # a previous one. New EDO rows NEVER fall back to this
            # shared/generic queue - they only ever get an image from
            # new_edo_diagram_queue (their own matched diagram), so a
            # New EDO row never ends up showing an unrelated Existing
            # EDO image.
            row_image = None

            # Column D carries RA_Number/FMEA_Number for this row. Only
            # place a New EDO diagram on rows where that identifier is
            # actually present (i.e. real New EDO rows).
            has_ra_fmea_in_col_d = bool(
                normalize_text(edo.get("RA_Number")) or normalize_text(edo.get("FMEA_Number"))
            )

            is_target_image_row = (
                is_new
                and normalize_id(edo.get("RA_Number")) == normalize_id(TARGET_IMAGE_RA_NUMBER)
                and normalize_id(edo.get("FMEA_Number")) == normalize_id(TARGET_IMAGE_FMEA_NUMBER)
            )

            if is_target_image_row and reserved_target_image:
                row_image = reserved_target_image
                reserved_target_image = None
                logging.info(
                    f"COLUMN H (row {current_row}, key {key!r}): placed the "
                    f"RESERVED New EDO diagram for RA={TARGET_IMAGE_RA_NUMBER!r} "
                    f"FMEA={TARGET_IMAGE_FMEA_NUMBER!r}."
                )
            elif is_new and has_ra_fmea_in_col_d and new_edo_diagram_queue:
                row_image = new_edo_diagram_queue.pop(0)
                logging.info(
                    f"COLUMN H (row {current_row}, key {key!r}): placed next "
                    f"queued New EDO diagram (RA={edo.get('RA_Number')!r} "
                    f"FMEA={edo.get('FMEA_Number')!r})."
                )
            elif is_new and has_ra_fmea_in_col_d:
                logging.warning(
                    f"COLUMN H (row {current_row}, key {key!r}): New EDO "
                    "diagram queue is empty - no more diagrams to place."
                )

            if not row_image and not is_new:
                if image_queue:
                    row_image = image_queue.pop(0)
                    last_image_for_this_edo = row_image
                # else: queue is empty - no fallback/duplicate image is
                # used, row_image stays None and Column H is left empty.

            if row_image:
                insert_image_below_text(sheet, row_image, row=current_row, column=8, text_offset_px=IMAGE_TEXT_OFFSET_PX)
                last_image_row = current_row
                last_image_for_this_edo = row_image
            else:
                logging.error(
                    f"COLUMN H WILL BE EMPTY at row {current_row} for "
                    f"key {key!r} (RA={edo.get('RA_Number')!r} FMEA="
                    f"{edo.get('FMEA_Number')!r}, is_new={is_new}) - no "
                    "New EDO diagram available AND the generic image "
                    "queue is exhausted."
                )

            current_row += 1

        if not is_new:
            existing_ranges.append((first, current_row - 1))

    # Any images still left in image_queue at this point have no split
    # row left to go into (every split row of every EDO tag already
    # either got one or was intentionally left empty per the no-
    # duplicate rule above). Per requirement, these leftovers are no
    # longer stacked below the last image row - they are simply left
    # unplaced.

    for first, last in existing_ranges:
        if last > first:
            for col in range(1, 6):
                sheet.merge_cells(
                    start_row=first,
                    start_column=col,
                    end_row=last,
                    end_column=col
                )
                sheet.cell(first, col).alignment = cell_alignment
            for col in (12, 13):
                sheet.merge_cells(
                    start_row=first,
                    start_column=col,
                    end_row=last,
                    end_column=col
                )
                sheet.cell(first, col).alignment = cell_alignment

    # ---- Auto-extend every row's height to fit its actual data ----
    # Runs last, after every value AND every merge is already in place,
    # so merged multi-row blocks (Columns A-E, L, M) are measured against
    # the full range they span rather than just their first physical row.
    autosize_edo_rows(sheet, start_row, current_row - 1)

    print(f"current _value:")
    return current_row


def clear_existing_rows(sheet, start_row, end_column=10):
    row = start_row
    while row <= sheet.max_row:
        empty = True
        for col in range(1, end_column + 1):
            if sheet.cell(row=row, column=col).value:
                empty = False
                break
        if empty:
            break
        for col in range(1, end_column + 1):
            sheet.cell(row=row, column=col).value = None
        row += 1


# ==========================================================
# EXCEL WRITER
# ==========================================================

def save_edo_workbook(workbook, pipeline_config):
    output_path = pipeline_config["output_file_path"]
    workbook.save(output_path)
    return output_path


# ==========================================================
# MAIN PIPELINE
# ==========================================================
# START -> load all documents from the database -> load template -> image extractions -> prompt execution -> Existing EDO pipeline (CALL 1-3) -> New EDO pipeline (CALL 4-5) -> merge -> verification reference (CALL 6) -> common processing (CALL 7) -> Excel formatting -> write Excel -> END.

def generate_edo_template(
    client,
    product_family,
    product,
    templatename,
    pipeline_config,
    db: DatabaseHandler
):
    logging.info("=" * 80)
    logging.info("STARTING COMPLETE EDO TEMPLATE GENERATION PIPELINE (MERGED)")
    logging.info("=" * 80)

    try:
        # ---------------------------------------------------
        # Load all documents from the database first
        # ---------------------------------------------------
        edo_document = get_edo_document(
            client,
            product_family,
            product,
            templatename,
            db
        )

        workbook, sheet = initialize_workbook(pipeline_config)
        start_row = pipeline_config.get("templatestartrow", 4)

        # Clear active table grid space exclusively up to Column J
        clear_existing_rows(sheet, start_row, end_column=10)

        # ---------------------------------------------------
        # Image extraction
        # ---------------------------------------------------
        try:
            edo_proposed_images = extract_edo_proposed_images(edo_document, pipeline_config)

            images_output_dir = os.path.join(
                os.path.dirname(pipeline_config.get("output_file_path", ".")) or ".",
                "edo_proposed_images"
            )
            if edo_proposed_images:
                save_images_to_folder(edo_proposed_images, images_output_dir)

        except Exception as image_error:
            edo_proposed_images = []
            logging.warning(
                "IMAGE EXTRACTION SKIPPED - could not extract images from "
                f"edo_proposed for this run. Reason: {image_error}"
            )

        # ---------------------------------------------------
        # CALL 1: Extract Existing EDO tags
        # ---------------------------------------------------
        existing_edos = extract_edo_tags(
            client,
            product_family,
            product,
            templatename,
            pipeline_config,
            edo_document,
            db
        )
        existing_edos = validate_existing_tags(existing_edos)

        if existing_edos:
            # -----------------------------------------------
            # CALL 2: Extract Existing EDO details
            # -----------------------------------------------
            existing_edos = extract_edo_details(
                client,
                product_family,
                product,
                templatename,
                pipeline_config,
                edo_document,
                existing_edos,
                db
            )

            # -----------------------------------------------
            # CALL 3: Extract Existing EDO trace details
            # -----------------------------------------------
            try:
                existing_trace_details = extract_existing_edo_trace_details(
                    client,
                    product_family,
                    product,
                    templatename,
                    pipeline_config,
                    edo_document,
                    existing_edos,
                    db
                )

                existing_edos = apply_existing_edo_trace(
                    existing_edos,
                    existing_trace_details
                )
            
            except Exception as trace_error:
                logging.warning(
                    "CALL 3 SKIPPED - trace extraction failed, leaving "
                    "the trace part of column D blank for this run. "
                    f"Reason: {trace_error}"
                )

            print(f"extracted existing edo trace: ", existing_edos)

        # ---------------------------------------------------
        # CALL 4: Extract New EDO tags
        # ---------------------------------------------------
        edo_new_data = extract_new_edo_tags(
            client,
            product_family,
            product,
            templatename,
            pipeline_config,
            edo_document,
            db
        )
        print(f"extract_new_edo_tags: ", edo_new_data)

        # ---------------------------------------------------
        # CALL 4.5: Filter New EDO tags by Risk Evaluation
        # ---------------------------------------------------
        edo_new_data = filter_new_edo_by_risk_evaluation(
            client,
            product_family,
            product,
            templatename,
            pipeline_config,
            edo_document,
            edo_new_data,
            db
        )
        print(f"filter_new_edo_by_risk_evaluation: ", edo_new_data)

        # ---------------------------------------------------
        # CALL 5: Extract New EDO summary details
        # ---------------------------------------------------
        edo_new_data = extract_new_edo_summary_details(
            client,
            product_family,
            product,
            templatename,
            pipeline_config,
            edo_document,
            edo_new_data,
            db
        )
        print(f"extract_new_edo_summary_details: ", edo_new_data)

        new_records = list(edo_new_data.values())

        # ---------------------------------------------------
        # New EDO Diagram Extraction
        # ---------------------------------------------------
        try:
            new_edo_diagram_queue = extract_new_edo_diagram_queue(
                edo_document,
                pipeline_config
            )
        except Exception as new_diagram_error:
            new_edo_diagram_queue = []
            logging.warning(
                "NEW EDO DIAGRAM EXTRACTION SKIPPED - could not extract "
                f"diagrams from EDO_pdf_new for this run. Reason: {new_diagram_error}"
            )

        # ---------------------------------------------------
        # MERGE: Combine Existing and New EDOs
        # ---------------------------------------------------
        new_final = merge_new_edo_records(
            new_records,
            existing_edos
        )

        existing_edo_final = merge_existing_edo_dictionary(existing_edos)
        final_edos = merge_all_edos(existing_edo_final, new_final)
        final_edos = validate_final_edos(final_edos)

        if not final_edos:
            raise Exception("No EDO records generated.")

        # ---------------------------------------------------
        # CALL 6: Verification Reference extraction
        # ---------------------------------------------------
        try:
            final_edos = extract_and_apply_verification_details(
                client,
                product_family,
                product,
                templatename,
                pipeline_config,
                edo_document,
                final_edos,
                db
            )
        except Exception as verification_error:
            logging.warning(
                "CALL 6 SKIPPED - verification reference extraction "
                "failed, leaving column E as 'Blank' for this run. "
                f"Reason: {verification_error}"
            )
        print(f"verification_details: ", final_edos)

        # ---------------------------------------------------
        # CALL 6B: Stage 1 - Pre-Extract 4 Excel Collections into Separate Lists
        # ---------------------------------------------------
        excel_details = {
            "EXCEL_TM_1_Details": [],
            "EXCEL_TM_2_Details": [],
            "EXCEL_TM_3_Details": [],
            "EXCEL_TM_4_Details": []
        }
        
        try:
            excel_details = extract_excel_sheets_to_separate_lists(
                client,
                product_family,
                product,
                templatename,
                pipeline_config,
                edo_document,
                db
            )
        except Exception as excel_extract_error:
            logging.warning(
                "EXCEL EXTRACTION SKIPPED - failed to extract details into 4 lists. "
                f"Reason: {excel_extract_error}"
            )

        # ---------------------------------------------------
        # CALL 6B: Stage 2 - In-Memory Traceability Matching
        # ---------------------------------------------------
        try:
            final_edos = build_final_edos_with_traceability(
                final_edos,
                excel_details
            )
        except Exception as traceability_reference_error:
            logging.warning(
                "CALL 6B SKIPPED - traceability reference lookup failed. "
                f"Reason: {traceability_reference_error}"
            )
        print(f"Traceability details: ", final_edos)

        # ---------------------------------------------------
        # CALL 7: Output Mapping, Formatting, and Storage
        # ---------------------------------------------------
        format_edo_worksheet(
            sheet,
            final_edos,
            start_row,
            pipeline_config,
            images=edo_proposed_images,
            new_edo_diagram_queue=new_edo_diagram_queue
        )

        output_file = save_edo_workbook(
            workbook,
            pipeline_config
        )

        logging.info("=" * 80)
        logging.info("EDO PIPELINE COMPLETED SUCCESSFULLY")
        logging.info(f"OUTPUT FILE PERSISTED AT : {output_file}")
        logging.info("=" * 80)

        return output_file

    except Exception as e:
        logging.error("=" * 80)
        logging.error(f"EDO PIPELINE CRITICAL RUNTIME FAILURE : {str(e)}")
        logging.error("=" * 80)
        raise e