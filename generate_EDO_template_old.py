import os
import re
import json
import time
import logging
import sys
import unicodedata
import openpyxl

from datetime import datetime

from openpyxl.styles import (
    Alignment,
    Border,
    Side
)

from Files.database import DatabaseHandler
from retrieval.retrieve_content_prompt import retrieve_content_for_prompt


# ==========================================================
# GLOBALS
# ==========================================================

CURRENT_FOLDER = os.path.dirname(os.path.abspath(__file__))

MAX_LLM_RETRIES = 3
INITIAL_RETRY_DELAY = 1


logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    stream=sys.stdout
)


# ==========================================================
# EXCEL FORMATTING
# ==========================================================

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

cell_alignment = Alignment(
    horizontal="left",
    vertical="top",
    wrap_text=True
)


# ==========================================================
# TEXT HELPERS
# ==========================================================

def normalize_text(value):
    if value is None:
        return ""

    value = str(value)

    value = value.replace("\u00A0", " ")
    value = value.replace("\u2007", " ")
    value = value.replace("\u202F", " ")

    value = unicodedata.normalize(
        "NFKC",
        value
    )

    return value.strip()


def clean_llm_response(response):

    response = normalize_text(response)

    return re.sub(
        r"^```json\s*|\s*```$",
        "",
        response,
        flags=re.DOTALL
    ).strip()


def parse_json(response):
    try:

        cleaned = clean_llm_response(response)

        logging.info("========== CLEANED JSON ==========")
        logging.info(cleaned)

        return json.loads(cleaned)

    except Exception as e:

        logging.error(f"JSON Parse Error : {e}")
        logging.error(response)

        return {}


def blank(value):

    value = normalize_text(value)

    if value == "":
        return "Blank"

    return value


# ==========================================================
# LLM EXECUTION
# ==========================================================

def execute_llm(
    pipeline_config,
    collection,
    prompt_row
):
    """
    Generic LLM wrapper used by every prompt.
    """

    return retrieve_content_for_prompt(

        pipeline_config,

        collection,

        prompt_row["question"],

        prompt_row["prompt_role"],

        prompt_row["prompt_text"],

        prompt_row["fulltext"],

        prompt_row["where_filter"],

        prompt_row["where_document"],

        prompt_row.get(
            "checkpoint",
            ""
        )
    )


# ==========================================================
# RETRY WRAPPER
# ==========================================================

def execute_llm_retry(
    pipeline_config,
    collection,
    prompt_row
):
    """
    Executes the LLM with retry logic.
    """

    last_exception = None

    for attempt in range(MAX_LLM_RETRIES):

        try:

            docs, metadata, response = execute_llm(
                pipeline_config,
                collection,
                prompt_row
            )

            if response:

                return docs, metadata, response

        except Exception as ex:

            last_exception = ex

            logging.warning(
                f"LLM Retry {attempt+1}/{MAX_LLM_RETRIES}"
            )

            time.sleep(
                INITIAL_RETRY_DELAY * (attempt + 1)
            )

    raise Exception(
        f"LLM failed after retries : {last_exception}"
    )


# ==========================================================
# WORKBOOK
# ==========================================================

def initialize_workbook(
    pipeline_config
):
    """
    Opens Excel template.
    """

    workbook = openpyxl.load_workbook(
        pipeline_config["input_file_path"]
    )

    sheet = workbook.active

    return workbook, sheet


# ==========================================================
# FORMAT SINGLE ROW
# ==========================================================

def format_excel_row(
    sheet,
    row
):
    """
    Apply formatting to one Excel row.
    """

    for col in range(1, 15):

        cell = sheet.cell(
            row=row,
            column=col
        )

        cell.alignment = cell_alignment
        cell.border = thin_border

    sheet.row_dimensions[row].height = 32


# ==========================================================
# JSON HELPERS
# ==========================================================

def unwrap_existing_tags(result):
    """
    Existing EDO Tag prompt returns

    {
        "EDO_Table":
        {
            "EDO_Tag_Values":[]
        }
    }

    This function extracts the list.
    """

    if not isinstance(result, dict):
        return []

    if "EDO_Table" in result:
        result = result["EDO_Table"]

    if isinstance(result, dict):
        return result.get(
            "EDO_Tag_Values",
            []
        )

    return []


def unwrap_new_edos(result):
    """
    New EDO prompt should return

    {
        "New_EDOs":[]
    }
    """

    if not isinstance(result, dict):
        return []

    return result.get(
        "New_EDOs",
        []
    )


# ==========================================================
# COMMON EMPTY EDO STRUCTURE
# ==========================================================

def create_empty_edo():
    """
    Standard dictionary used for every EDO
    (Existing and New).
    """

    return {

        "edo_type": "Existing",

        "edo_tag": "Blank",

        "ra_number": "Blank",

        "edo_description": "Blank",

        "reason_identified": "Blank",

        "dfmea": "Blank",

        "location": "Blank",

        "description_2": "Blank",

        "reason_2": "Blank",

        "sysdd": "Blank"
    }


# ==========================================================
# LOGGING
# ==========================================================

def log_dictionary(title, dictionary):
    """
    Pretty logging for dictionaries.
    """

    logging.info("=" * 80)
    logging.info(title)

    for key, value in dictionary.items():

        logging.info(f"{key}")

        logging.info(
            json.dumps(
                value,
                indent=4
            )
        )

    logging.info("=" * 80)

# ==========================================================
# DOCUMENT HELPERS
# ==========================================================

def get_template_documents(
    client,
    product_family,
    product,
    templatename,
    db: DatabaseHandler
):
    """
    Retrieve all documents associated with the template.
    """

    docs = db.get_template_documents(
        client,
        product_family,
        product,
        templatename
    )

    if not docs:
        raise Exception(
            "No template documents were found."
        )

    logging.info("=" * 80)
    logging.info("AVAILABLE TEMPLATE DOCUMENTS")
    logging.info("=" * 80)

    for doc in docs:

        logging.info(
            f"{doc.get('document_identity')}  -->  "
            f"{doc.get('document_name')}"
        )

    logging.info("=" * 80)

    return docs


# ==========================================================
# FIND DOCUMENT BY IDENTITY
# ==========================================================

def find_document(
    documents,
    identity
):
    """
    Find a document by document_identity.
    """

    for document in documents:

        if normalize_text(
            document.get("document_identity")
        ) == normalize_text(identity):

            return document

    return None


# ==========================================================
# GET EDO DOCUMENT
# ==========================================================
def get_edo_document(
    client,
    product_family,
    product,
    templatename,
    db: DatabaseHandler
):
    """
    Returns all EDO related documents:

    {
        "edo_proposed": {...},
        "edo_ra_c": {...},
        "edo_fmea": {...}
    }

    Used by:
        - Existing EDO extraction
        - New EDO extraction
    """

    documents = get_template_documents(
        client,
        product_family,
        product,
        templatename,
        db
    )


    if not documents:

        raise Exception(
            "No EDO template documents found."
        )


    edo_documents = {}


    for document in documents:


        identity = normalize_text(
            document.get("document_identity")
        ).lower()


        logging.info(
            f"AVAILABLE DOCUMENT : {identity}"
        )


        # Existing EDO source
        if identity == "edo_proposed":


            edo_documents["edo_proposed"] = document



        # New EDO RA&C source
        elif identity in [
            "edo_ra_c",
            "edo_ra&c",
            "edo_rac",
            "edo_ra"
        ]:


            edo_documents["edo_ra_c"] = document



        # New EDO FMEA source
        elif identity in [
            "edo_fmea",
            "fmea",
            "system_fmea"
        ]:


            edo_documents["edo_fmea"] = document



    logging.info("=" * 80)
    logging.info("EDO DOCUMENT CONFIGURATION")
    logging.info("=" * 80)


    for key, doc in edo_documents.items():

        logging.info(
            f"{key}"
        )

        logging.info(
            f"Identity   : {doc.get('document_identity')}"
        )

        logging.info(
            f"Name       : {doc.get('document_name')}"
        )

        logging.info(
            f"Collection : {doc.get('collection')}"
        )


    logging.info("=" * 80)



    # Validation

    if "edo_proposed" not in edo_documents:

        logging.warning(
            "EDO_Proposed document not found. Existing EDO extraction may fail."
        )


    if "edo_ra_c" not in edo_documents:

        logging.warning(
            "EDO_RA_C document not found. New EDO extraction will be skipped."
        )


    if "edo_fmea" not in edo_documents:

        logging.warning(
            "EDO_FMEA document not found. New EDO trace extraction may fail."
        )


    return edo_documents
# ==========================================================
# GET PROMPT
# ==========================================================

def get_prompt(
    client,
    product_family,
    product,
    templatename,
    prompt_name,
    db: DatabaseHandler
):
    """
    Loads a prompt from the database and validates it.
    """

    prompt = db.get_prompt_by_name(
        client,
        product_family,
        product,
        templatename,
        prompt_name
    )

    if not prompt:

        raise Exception(
            f"Prompt '{prompt_name}' was not found."
        )

    logging.info(
        f"Loaded Prompt : {prompt_name}"
    )

    return prompt


# ==========================================================
# BUILD PROMPT ROW
# ==========================================================

def build_prompt_row(
    prompt,
    question,
    fulltext="Yes",
    where_filter="",
    where_document="",
    checkpoint=""
):
    """
    Converts a prompt from the database into the
    structure expected by execute_llm().
    """

    return {

        "prompt_role": prompt["prompt_role"],

        "prompt_text": prompt["prompt_text"],

        "question": question,

        "fulltext": fulltext,

        "where_filter": where_filter,

        "where_document": where_document,

        "checkpoint": checkpoint
    }


# ==========================================================
# EXECUTE PROMPT
# ==========================================================

def execute_prompt(
    pipeline_config,
    collection,
    prompt
):
    """
    Executes a prompt using the retry wrapper.

    Returns parsed JSON.
    """

    _, _, response = execute_llm_retry(
        pipeline_config,
        collection,
        prompt
    )

    logging.info("=" * 80)
    logging.info("RAW LLM RESPONSE")
    logging.info("=" * 80)

    logging.info(response)

    return parse_json(response)

# ==========================================================
# EXISTING EDO TAG EXTRACTION
# ==========================================================

def extract_edo_tags(
    client,
    product_family,
    product,
    templatename,
    pipeline_config,
    edo_document,
    db: DatabaseHandler
):
    """
    Extract all Existing EDO tags.

    Returns

    {
        "EDO-29":
        {
            "edo_type":"Existing",
            "edo_tag":"EDO-29"
        },

        "EDO-31":
        {
            "edo_type":"Existing",
            "edo_tag":"EDO-31"
        }
    }
    """

    logging.info("=" * 80)
    logging.info("EXTRACTING EXISTING EDO TAGS")
    logging.info("=" * 80)

    prompt = get_prompt(
        client,
        product_family,
        product,
        templatename,
        "EDO_Existing_Tag",
        db
    )

    prompt_row = build_prompt_row(
        prompt,
        "Extract all Existing EDO tags from the document."
    )

    result = execute_prompt(
        pipeline_config,
        edo_document["edo_proposed"]["collection"],
        prompt_row
    )

    items = unwrap_existing_tags(result)

    logging.info(f"Rows returned : {len(items)}")

    tags = {}

    for item in items:

        if not isinstance(item, dict):
            continue

        tag = (
            item.get("edo_number")
            or item.get("edo_tag")
            or item.get("EDO Number")
            or item.get("EDO_Tag")
            or item.get("EDO")
            or ""
        )

        tag = normalize_text(tag)

        if tag == "":
            continue

        if tag.lower() == "blank":
            continue

        if tag in tags:
            continue

        tags[tag] = create_empty_edo()

        tags[tag]["edo_type"] = "Existing"
        tags[tag]["edo_tag"] = tag

        logging.info(f"Found Existing EDO : {tag}")

    logging.info("=" * 80)
    logging.info(f"TOTAL EXISTING EDO TAGS : {len(tags)}")
    logging.info("=" * 80)

    return tags


# ==========================================================
# VALIDATE EXISTING TAGS
# ==========================================================

def validate_existing_tags(tags):
    """
    Removes invalid tags and duplicates.

    Mostly used for safety before detail extraction.
    """

    validated = {}

    for key, value in tags.items():

        tag = normalize_text(value.get("edo_tag"))

        if tag == "":
            continue

        if tag.lower() == "blank":
            continue

        validated[tag] = value

    logging.info(
        f"Validated Existing Tags : {len(validated)}"
    )

    return validated


# ==========================================================
# PRINT EXISTING TAGS
# ==========================================================

def log_existing_tags(tags):
    """
    Pretty logging.
    """

    logging.info("=" * 80)
    logging.info("EXISTING EDO TAGS")
    logging.info("=" * 80)

    for tag in tags:

        logging.info(tag)

    logging.info("=" * 80)


# ==========================================================
# EXISTING EDO DETAIL EXTRACTION
# ==========================================================

def extract_edo_details(
    client,
    product_family,
    product,
    templatename,
    pipeline_config,
    edo_document,
    existing_edos,
    db: DatabaseHandler
):
    """
    Populates every Existing EDO with complete information.

    Input:

    {
        "EDO-29": {...},
        "EDO-31": {...}
    }

    Output:

    {
        "EDO-29":
        {
            ...
            "edo_description": "...",
            "reason_identified":"...",
            "dfmea":"...",
            "location":"...",
            "description_2":"...",
            "reason_2":"...",
            "sysdd":"..."
        }
    }
    """

    logging.info("=" * 80)
    logging.info("EXTRACTING EXISTING EDO DETAILS")
    logging.info("=" * 80)

    prompt = get_prompt(
        client,
        product_family,
        product,
        templatename,
        "EDO_Existing_Generic",
        db
    )

    for edo_tag in existing_edos.keys():

        logging.info(f"Processing {edo_tag}")

        question = f"""
Extract ALL information for Existing EDO Tag: {edo_tag}

Return ONLY valid JSON.

Required Keys:

edo_description
reason_identified
dfmea
location
description_2
reason_2
sysdd
"""

        prompt_row = build_prompt_row(
            prompt,
            question
        )

        result = execute_prompt(
            pipeline_config,
            edo_document["edo_proposed"]["collection"],
            prompt_row
        )

        if isinstance(result, list):

            if len(result) == 0:
                result = {}
            else:
                result = result[0]

        edo = existing_edos[edo_tag]

        edo["edo_description"] = blank(
            result.get("edo_description")
            or result.get("EDO Description")
            or result.get("description")
        )

        edo["reason_identified"] = blank(
            result.get("reason_identified")
            or result.get("Reason Identified")
            or result.get("reason")
        )

        edo["dfmea"] = blank(
            result.get("dfmea")
            or result.get("DFMEA")
            or result.get("RA&C and/or Sys-DFMEA Trace")
        )

        edo["location"] = blank(
            result.get("location")
            or result.get("EDO Location")
            or result.get("Location")
        )

        edo["description_2"] = blank(
            result.get("description_2")
            or result.get("EDO Description 2")
            or result.get("Description 2")
        )

        edo["reason_2"] = blank(
            result.get("reason_2")
            or result.get("Reason 2")
        )

        edo["sysdd"] = blank(
            result.get("sysdd")
            or result.get("SYSDD")
            or result.get("SYS DD")
            or result.get("SYSDD or HDD Reference")
            or result.get("sysdd_reference")
        )

        logging.info(
            f"Completed {edo_tag}"
        )

    logging.info("=" * 80)
    logging.info("EXISTING EDO DETAIL EXTRACTION COMPLETE")
    logging.info("=" * 80)

    return existing_edos


# ==========================================================
# OPTIONAL DEBUG FUNCTION
# ==========================================================

def log_existing_edo_details(existing_edos):
    """
    Pretty-print Existing EDO dictionary.
    """

    logging.info("=" * 80)
    logging.info("EXISTING EDO DETAILS")
    logging.info("=" * 80)

    for tag, data in existing_edos.items():

        logging.info(tag)

        logging.info(
            json.dumps(
                data,
                indent=4
            )
        )

    logging.info("=" * 80)

# ==========================================================
# NEW EDO EXTRACTION
# ==========================================================

def extract_new_edo_tags(
    client,
    product_family,
    product,
    templatename,
    pipeline_config,
    edo_document,
    db
):

    logging.info(
        "NEW EDO CALL 1 - Extract RA/FMEA targets"
    )


    prompt_data = db.get_prompt_by_name(
        client,
        product_family,
        product,
        templatename,
        "EDO_NEW_RA"
    )


    prompt_row = {

        "prompt_role":
            prompt_data["prompt_role"],

        "prompt_text":
            prompt_data["prompt_text"],

        "question":
            "Extract RA records with Medium or See FMEA",

        "fulltext":
            "Yes",

        "where_filter":
            "",

        "where_document":
            "",

        "checkpoint":
            ""
    }


    _,_,response = execute_llm(
        pipeline_config,
        edo_document["edo_ra_c"]["collection"],
        prompt_row
    )


    result=parse_json(response)


    records=result.get(
        "Records",
        []
    )


    output=[]


    for row in records:

        fmea=normalize_text(
            row.get("FMEA_Number")
        )

        if not fmea:
            continue


        output.append({

            "RA_Number":
                normalize_text(
                    row.get("RA_Number")
                ),

            "Status":
                normalize_text(
                    row.get("Status")
                ),

            "FMEA_Number":
                fmea

        })


    return output

def extract_new_edo_summary_details(
    client,
    product_family,
    product,
    templatename,
    pipeline_config,
    edo_document,
    edo_tags,
    db
):

    logging.info("=" * 80)
    logging.info("NEW EDO CALL 2 - EXTRACT EDO DETAILS")
    logging.info("=" * 80)


    prompt_data = db.get_prompt_by_name(
        client,
        product_family,
        product,
        templatename,
        "EDO_NEW_details"
    )


    if not prompt_data:
        raise Exception(
            "Missing prompt : EDO_NEW_details"
        )


    targets = "\n".join(
        [
            f"""
RA_Number : {item.get('RA_Number')}
FMEA_Number : {item.get('FMEA_Number')}
"""
            for item in edo_tags
        ]
    )


    prompt_row = {

        "prompt_role":
            prompt_data["prompt_role"],


        "prompt_text":
            prompt_data["prompt_text"]
            +
            """

IMPORTANT OUTPUT RULE:

Return ONLY JSON.

Required keys:

RA_Number
FMEA_Number
EDO_Number
Status
Product_Feature_Function
Reason_Identified_as_EDO
Traceability
EDO_Location
EDO_Description
Reason_Identified_as_EDO_ColH

Do not change key names.

"""
            +
            "\nTARGETS:\n"
            +
            targets,


        "question":
            "Extract complete New EDO details",


        "fulltext":
            "Yes",

        "where_filter":
            "",

        "where_document":
            "",

        "checkpoint":
            ""
    }


    _,_,response = execute_llm(
        pipeline_config,
        edo_document["edo_fmea"]["collection"],
        prompt_row
    )


    logging.info("=" * 80)
    logging.info("RAW NEW EDO SUMMARY RESPONSE")
    logging.info(response)
    logging.info("=" * 80)


    result = parse_json(response)


    logging.info(
        json.dumps(
            result,
            indent=4
        )
    )


    if isinstance(result,dict):

        if "Records" in result:
            return result["Records"]


        return [result]


    if isinstance(result,list):
        return result


    return []

def extract_new_edo_traceability_details(
    client,
    product_family,
    product,
    templatename,
    pipeline_config,
    edo_document,
    edo_tags,
    db
):

    logging.info("=" * 80)
    logging.info("NEW EDO CALL 3 - TRACEABILITY EXTRACTION")
    logging.info("=" * 80)


    prompt_data = db.get_prompt_by_name(
        client,
        product_family,
        product,
        templatename,
        "EDO_NEW_Verification_details"
    )


    if not prompt_data:
        raise Exception(
            "Missing prompt : EDO_NEW_Verification_details"
        )


    targets = "\n".join(
        [
            f"""
RA_Number : {item.get('RA_Number')}
FMEA_Number : {item.get('FMEA_Number')}
"""
            for item in edo_tags
        ]
    )


    prompt_row = {

        "prompt_role":
            prompt_data["prompt_role"],


        "prompt_text":
            prompt_data["prompt_text"]
            +
            """

IMPORTANT OUTPUT RULE:

Return ONLY JSON.

Required keys:

RA_Number
FMEA_Number
Verification_Reference

Do not rename keys.

"""
            +
            "\nTARGETS:\n"
            +
            targets,


        "question":
            "Extract verification references",


        "fulltext":
            "Yes",

        "where_filter":
            "",

        "where_document":
            "",

        "checkpoint":
            ""
    }


    _,_,response = execute_llm(
        pipeline_config,
        edo_document["edo_fmea"]["collection"],
        prompt_row
    )


    logging.info(
        "RAW TRACEABILITY RESPONSE"
    )

    logging.info(response)


    result = parse_json(response)


    logging.info(
        json.dumps(
            result,
            indent=4
        )
    )


    if isinstance(result,dict):

        if "Records" in result:
            return result["Records"]

        return [result]


    if isinstance(result,list):
        return result


    return []
# ==========================================================
# UNIFIED EDO DICTIONARY MODEL
# ==========================================================
def normalize_edo_record(
    edo_record
):

    default_structure = {

        "edo_type": "Existing",

        "edo_tag": "",

        "RA_Number": "",

        "FMEA_Number": "",

        "edo_description": "",

        "reason_identified": "",

        "dfmea": "",

        "location": "",

        "description_2": "",

        "reason_2": "",

        "sysdd": ""
    }


    if not isinstance(
        edo_record,
        dict
    ):
        return default_structure


    for key,value in default_structure.items():

        if key not in edo_record:

            edo_record[key] = value


    return edo_record
# ==========================================================
# MERGE EXISTING + NEW EDOs
# ==========================================================
def get_llm_value(
    row,
    *keys
):

    for key in keys:

        value = row.get(key)

        if value not in [
            None,
            "",
            "Blank"
        ]:

            return value


    return ""

def merge_all_edos(
    existing_edos,
    new_edos
):
    """
    Combine Existing and New EDO dictionaries.

    Existing:

    {
        "EDO-29": {...}
    }


    New:

    {
        "RA-104": {...}
    }


    Returns:

    {
        "EDO-29": {...},
        "RA-104": {...}
    }
    """

    logging.info("=" * 80)
    logging.info("MERGING EXISTING AND NEW EDOS")
    logging.info("=" * 80)


    final_edos = {}



    # ------------------------------------------------------
    # Add Existing EDOs
    # ------------------------------------------------------

    for key, value in existing_edos.items():


        normalized = normalize_edo_record(
            value
        )


        normalized["edo_type"] = "Existing"


        final_edos[key] = normalized


        logging.info(
            f"Added Existing EDO : {key}"
        )



    # ------------------------------------------------------
    # Add New EDOs
    # ------------------------------------------------------

    for key, value in new_edos.items():


        normalized = normalize_edo_record(
            value
        )


        normalized["edo_type"] = "New"



        # Avoid accidental overwrite
        final_key = key


        if final_key in final_edos:

            counter = 1

            while f"{key}_{counter}" in final_edos:

                counter += 1


            final_key = (
                f"{key}_{counter}"
            )


        final_edos[final_key] = normalized


        logging.info(
            f"Added New EDO : {final_key}"
        )



    logging.info("=" * 80)

    logging.info(
        f"TOTAL FINAL EDO COUNT : {len(final_edos)}"
    )

    logging.info("=" * 80)


    return final_edos



# ==========================================================
# FINAL EDO VALIDATION
# ==========================================================
def validate_final_edos(
    final_edos
):
    """
    Final validation before Excel writing.

    Removes invalid empty records and
    ensures required EDO fields exist.
    """

    validated = {}


    for key, value in final_edos.items():


        if not isinstance(value, dict):

            continue



        # Remove completely empty records
        ra_number = normalize_text(
            value.get("RA_Number")
        )

        fmea_number = normalize_text(
            value.get("FMEA_Number")
        )


        edo_tag = normalize_text(
            value.get("edo_tag")
        )



        # Skip invalid records
        # Example:
        # key = ""
        # RA_Number = ""
        # FMEA_Number = ""

        if (
            not key
            and not ra_number
            and not fmea_number
        ):
            logging.warning(
                "Skipping empty EDO record"
            )
            continue



        # New EDO default tag
        if edo_tag == "":

            value["edo_tag"] = (
                "EDO-XX\nNew EDO"
            )



        # Normalize all fields
        normalized = normalize_edo_record(
            value
        )


        validated[key] = normalized



    logging.info(
        f"Validated Final EDO Count : {len(validated)}"
    )


    return validated
# ==========================================================
# FINAL DICTIONARY LOGGER
# ==========================================================

def log_final_edo_dictionary(
    final_edos
):
    """
    Logs the final structure.
    """

    logging.info("=" * 80)
    logging.info("FINAL EDO DICTIONARY")
    logging.info("=" * 80)


    for key, value in final_edos.items():


        logging.info(
            f"\n{key}"
        )


        logging.info(
            json.dumps(
                value,
                indent=4
            )
        )


    logging.info("=" * 80)


# ==========================================================
# SINGLE EXCEL WRITER
# ==========================================================

def write_edo_excel(
    sheet,
    final_edos,
    start_row
):
    """
    Writes Existing + New EDOs into Excel.

    Expected Excel Mapping:

    Column A:
        EDO Number

    Column B:
        EDO Description

    Column C:
        Reason Identified

    Column D:
        DFMEA / RA Trace

    Column E:
        Existing workbook column (unused)

    Column F:
        Existing workbook column (unused)

    Column G:
        EDO Location

    Column H:
        Description 2

    Column I:
        Reason 2

    Column J:
        SYS DD Reference


    Returns:
        next available row
    """

    logging.info("=" * 80)
    logging.info("WRITING EDO DATA TO EXCEL")
    logging.info("=" * 80)


    current_row = start_row



    for key, edo in final_edos.items():


        logging.info(
            f"Writing row {current_row}: {key}"
        )


        # --------------------------------------------------
        # Column A
        # --------------------------------------------------

        if edo.get("edo_type") == "New":

            value = edo.get(
                "edo_tag",
                "EDO-XX New EDO"
            )


        else:

            value = edo.get(
                "edo_tag",
                key
            )


        sheet.cell(
            row=current_row,
            column=1
        ).value = value



        # --------------------------------------------------
        # Column B
        # --------------------------------------------------

        sheet.cell(
            row=current_row,
            column=2
        ).value = edo.get(
            "edo_description",
            "Blank"
        )



        # --------------------------------------------------
        # Column C
        # --------------------------------------------------

        sheet.cell(
            row=current_row,
            column=3
        ).value = edo.get(
            "reason_identified",
            "Blank"
        )



        # --------------------------------------------------
        # Column D
        # --------------------------------------------------

        sheet.cell(
            row=current_row,
            column=4
        ).value = edo.get(
            "dfmea",
            "Blank"
        )



        # --------------------------------------------------
        # Column G
        # --------------------------------------------------

        sheet.cell(
            row=current_row,
            column=7
        ).value = edo.get(
            "location",
            "Blank"
        )



        # --------------------------------------------------
        # Column H
        # --------------------------------------------------

        sheet.cell(
            row=current_row,
            column=8
        ).value = edo.get(
            "description_2",
            "Blank"
        )



        # --------------------------------------------------
        # Column I
        # --------------------------------------------------

        sheet.cell(
            row=current_row,
            column=9
        ).value = edo.get(
            "reason_2",
            "Blank"
        )



        # --------------------------------------------------
        # Column J
        # --------------------------------------------------

        sheet.cell(
            row=current_row,
            column=10
        ).value = edo.get(
            "sysdd",
            "Blank"
        )



        # --------------------------------------------------
        # Formatting
        # --------------------------------------------------

        format_excel_row(
            sheet,
            current_row
        )


        current_row += 1



    logging.info("=" * 80)

    logging.info(
        f"TOTAL EXCEL ROWS WRITTEN : "
        f"{current_row-start_row}"
    )

    logging.info("=" * 80)


    return current_row



# ==========================================================
# EXCEL CLEANUP HELPER
# ==========================================================

def clear_existing_rows(
    sheet,
    start_row,
    end_column=10
):
    """
    Clears old EDO data before writing.

    Useful when regenerating templates.
    """

    row = start_row


    while row <= sheet.max_row:


        empty = True


        for col in range(
            1,
            end_column + 1
        ):

            if sheet.cell(
                row=row,
                column=col
            ).value:

                empty = False
                break


        if empty:

            break


        for col in range(
            1,
            end_column + 1
        ):

            sheet.cell(
                row=row,
                column=col
            ).value = None


        row += 1



# ==========================================================
# SAVE WORKBOOK
# ==========================================================

def save_edo_workbook(
    workbook,
    pipeline_config
):
    """
    Saves final workbook.
    """

    output_path = pipeline_config[
        "output_file_path"
    ]


    workbook.save(
        output_path
    )


    logging.info("=" * 80)

    logging.info(
        f"EXCEL SAVED : {output_path}"
    )

    logging.info("=" * 80)


    return output_path

# ==========================================================
# MAIN EDO PIPELINE
# ==========================================================
def merge_existing_edo_dictionary(existing_details):

    final = {}

    for tag, data in existing_details.items():

        final[tag] = {

            "edo_tag": data.get(
                "edo_tag",
                tag
            ),

            "edo_description":
                data.get(
                    "edo_description",
                    "Blank"
                ),

            "reason_identified":
                data.get(
                    "reason_identified",
                    "Blank"
                ),

            "dfmea":
                data.get(
                    "dfmea",
                    "Blank"
                ),

            "location":
                data.get(
                    "location",
                    "Blank"
                ),

            "description_2":
                data.get(
                    "description_2",
                    "Blank"
                ),

            "reason_2":
                data.get(
                    "reason_2",
                    "Blank"
                ),

            "sysdd":
                data.get(
                    "sysdd",
                    "Blank"
                )
        }

    return final
def merge_new_edo_dictionary(
    edo_tags,
    edo_summary_details,
    edo_trace_details
):


    logging.info(
        "========== MERGING NEW EDO DICTIONARY =========="
    )


    merged = {}


    if isinstance(
        edo_summary_details,
        dict
    ):

        edo_summary_details = (
            edo_summary_details.get(
                "Records",
                []
            )
        )


    if not isinstance(
        edo_summary_details,
        list
    ):

        edo_summary_details = []


    logging.info(
        f"SUMMARY COUNT : {len(edo_summary_details)}"
    )


    for row in edo_summary_details:


        if not isinstance(row,dict):
            continue


        logging.info(
            "PROCESSING SUMMARY ROW"
        )

        logging.info(
            json.dumps(
                row,
                indent=4
            )
        )


        ra_number = get_llm_value(
            row,
            "RA_Number",
            "RA Number"
        )


        fmea_number = get_llm_value(
            row,
            "FMEA_Number",
            "FMEA Number"
        )


        key = (
            fmea_number
            or
            ra_number
        )


        if not key:

            logging.warning(
                "Skipping record without RA/FMEA"
            )

            continue



        merged[key] = {

            "edo_type":
                "New",


            "edo_tag":
                "EDO-XX\nNew EDO",


            "edo_description":
                get_llm_value(
                    row,
                    "Product_Feature_Function",
                    "Product Feature Function"
                ),


            "reason_identified":
                get_llm_value(
                    row,
                    "Reason_Identified_as_EDO",
                    "Reason Identified as EDO"
                ),


            "dfmea":
                get_llm_value(
                    row,
                    "Traceability"
                ),


            "location":
                "",


            "description_2":
                get_llm_value(
                    row,
                    "EDO_Location",
                    "EDO Location"
                ),


            "reason_2":
                get_llm_value(
                    row,
                    "EDO_Description",
                    "EDO Description"
                ),


            "sysdd":
                get_llm_value(
                    row,
                    "Reason_Identified_as_EDO_ColH"
                ),


            "RA_Number":
                ra_number,


            "FMEA_Number":
                fmea_number
        }



    # Add traceability

    if isinstance(
        edo_trace_details,
        dict
    ):

        edo_trace_details = (
            edo_trace_details.get(
                "Records",
                []
            )
        )


    for row in edo_trace_details:


        fmea_number = get_llm_value(
            row,
            "FMEA_Number"
        )


        ra_number = get_llm_value(
            row,
            "RA_Number"
        )


        key = (
            fmea_number
            or
            ra_number
        )


        if key in merged:


            verification = get_llm_value(
                row,
                "Verification_Reference"
            )


            if verification:

                merged[key]["location"] = verification



    logging.info(
        "FINAL NEW EDO DATA"
    )


    logging.info(
        json.dumps(
            merged,
            indent=4
        )
    )


    return merged
def generate_edo_template(
    client,
    product_family,
    product,
    templatename,
    pipeline_config,
    db: DatabaseHandler
):
    """
    Complete EDO generation pipeline.

    Generates Excel containing:

    - Existing EDOs
    - New EDOs

    using one unified dictionary.
    """


    logging.info("=" * 80)
    logging.info("STARTING COMPLETE EDO PIPELINE")
    logging.info("=" * 80)



    try:


        # --------------------------------------------------
        # Initialize Workbook
        # --------------------------------------------------

        workbook, sheet = initialize_workbook(
            pipeline_config
        )


        start_row = pipeline_config.get(
            "templatestartrow",
            4
        )



        # --------------------------------------------------
        # Get EDO Document
        # --------------------------------------------------

        edo_document = get_edo_document(
            client,
            product_family,
            product,
            templatename,
            db
        )



        # --------------------------------------------------
        # EXISTING EDO EXTRACTION
        # --------------------------------------------------

        logging.info(
            "STARTING EXISTING EDO EXTRACTION"
        )


        existing_edos = extract_edo_tags(
            client,
            product_family,
            product,
            templatename,
            pipeline_config,
            edo_document,
            db
        )


        existing_edos = validate_existing_tags(
            existing_edos
        )


        if existing_edos:


            existing_edos = extract_edo_details(
                client,
                product_family,
                product,
                templatename,
                pipeline_config,
                edo_document,
                existing_edos,
                db
            )


        else:

            logging.warning(
                "No Existing EDOs found"
            )



        # --------------------------------------------------
        # NEW EDO EXTRACTION
        # --------------------------------------------------

        logging.info(
            "STARTING NEW EDO EXTRACTION"
        )


        new_tags = extract_new_edo_tags(
            client,
            product_family,
            product,
            templatename,
            pipeline_config,
            edo_document,
            db
        )

        new_summary = extract_new_edo_summary_details(
            client,
            product_family,
            product,
            templatename,
            pipeline_config,
            edo_document,
            new_tags,
            db
        )

        new_trace = extract_new_edo_traceability_details(
            client,
            product_family,
            product,
            templatename,
            pipeline_config,
            edo_document,
            new_tags,
            db
        )

        new_final = merge_new_edo_dictionary(
            new_tags,
            new_summary,
            new_trace
        )

        # --------------------------------------------------
        # MERGE BOTH TYPES
        # --------------------------------------------------

        existing_edo_final = merge_existing_edo_dictionary(
            existing_edos
        )


        final_edos = {}


        final_edos.update(
            existing_edo_final
        )


        final_edos.update(
            new_final
        )


        logging.info(
            f"TOTAL FINAL EDO RECORDS : {len(final_edos)}"
        )


        if not final_edos:

            raise Exception(
                "No EDO records generated."
            )



        # --------------------------------------------------
        # WRITE EXCEL
        # --------------------------------------------------

        write_edo_excel(
            sheet,
            final_edos,
            start_row
        )



        # --------------------------------------------------
        # SAVE FILE
        # --------------------------------------------------

        output_file = save_edo_workbook(
            workbook,
            pipeline_config
        )



        logging.info("=" * 80)

        logging.info(
            "EDO PIPELINE COMPLETED SUCCESSFULLY"
        )

        logging.info(
            f"OUTPUT FILE : {output_file}"
        )

        logging.info("=" * 80)



        return output_file



    except Exception as e:


        logging.error("=" * 80)

        logging.error(
            "EDO PIPELINE FAILED"
        )

        logging.error(
            str(e)
        )

        logging.error("=" * 80)


        raise e