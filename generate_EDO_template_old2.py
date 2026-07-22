"""
================================================================================
MERGED EDO TEMPLATE GENERATION PIPELINE
================================================================================
This file merges:
  - edo_existing_final.py            (EXISTING-EDO-ONLY pipeline)
  - generate_EDO_template_copy.py    (EXISTING + NEW EDO combined pipeline)

Per requirement, NO function from either source file has been deleted.
Where both files defined a function with the SAME NAME but a DIFFERENT BODY,
both versions are kept:
    - the version chosen to run in the merged pipeline keeps its original
      name (the "canonical" version), and
    - the other version is kept under a suffixed name (e.g. "_legacy",
      "_basic", "_lenient", "_minimal", "_existing_style", "_new_style",
      "_existing_only") purely for reference / backward compatibility.

Canonical selection for the 6 functions originally requested by the user:
    1. extract_edo_tags                          -> from edo_existing_final.py
    2. extract_existing_edo_verification_details  -> from edo_existing_final.py
    3. apply_existing_edo_verification            -> from edo_existing_final.py
    4. extract_new_edo_tags                       -> REMOVED (see "NEW-EDO
       WORKFLOW RESTRUCTURE" below) - merged into extract_new_edo_summary_details
    5. extract_new_edo_summary_details            -> from generate_EDO_template_copy.py,
       now self-contained (see restructure notes below)
    6. extract_new_edo_traceability_details        -> from generate_EDO_template_copy.py,
       kept, unused (see restructure notes below)

Everything else (normalize_text, blank, deep_extract_records, merge
functions, write_edo_excel, etc.) has been merged / reconciled so that the
combined pipeline (generate_edo_template) can run BOTH the existing-EDO
workflow and the new-EDO workflow end to end, exactly the way
generate_EDO_template_copy.py's main pipeline did - but now using the
richer Stage 3 / Stage 3B logic that only existed in edo_existing_final.py.

The original standalone "existing-EDO-only" pipeline
(generate_and_download_edo1) is also preserved and still works on its own.

ADDITIONALLY MERGED: the dedicated Excel rich-text formatting module
(edo_excel_formatting.py) is folded in as format_edo_worksheet() - the
Stage 6 writer now used by generate_edo_template(). It reproduces the
reference-screenshot styling (black-bold tag for Existing EDOs, red-bold
tag + red-bold "None" fields for New EDOs, per-line red/black rich text
for the RA&C Trace and Verification Reference columns). The older plain
writers (write_edo_excel, write_edo_excel_existing_style,
write_edo_excel_new_style) are kept, unused, for backward compatibility -
nothing was deleted.

================================================================================
NEW-EDO WORKFLOW RESTRUCTURE (latest revision)
================================================================================
Per explicit requirement, extract_new_edo_tags() has been REMOVED. It used
to pre-list {RA_Number, Status, FMEA_Number} from the RA&C document before
extract_new_edo_summary_details() could run its own separate FMEA query.
Those two prompts have been merged into ONE prompt - "EDO_NEW_details" -
that scans the FMEA document's "10.3 Safety Hazard DFMEA Table" directly
and returns complete, ready-to-use EDO records in a single pass.
extract_new_edo_summary_details() is now self-contained (no more edo_tags
parameter / no more "TARGETS" list).

extract_new_edo_traceability_details() and extract_new_edo_ra_details()
are kept, defined, but are no longer called from generate_edo_template()
- their old edo_tags-based calling contract no longer applies now that
extract_new_edo_tags() is gone, and their functionality (verification
references / Medium-risk backfill) is already covered by the single
merged extraction. They remain available for backward compatibility.

merge_new_edo_records() is the new canonical Stage 5 merge for New EDOs -
it maps directly off the ACTUAL JSON keys the merged prompt returns
(Product_Feature_Function, Reason_Identified_as_EDO, Traceability,
Verification_Reference, EDO_Location, EDO_Description,
Reason_Identified_as_EDO_ColH), fixing a prior "column mapped wrongly" bug
where the merge logic looked for keys that didn't actually exist in the
LLM output. It also de-duplicates: any New EDO record whose RA Number or
FMEA Number already belongs to an Existing EDO is dropped, so the same
issue is never printed twice. merge_new_edo_dictionary_full() is kept,
unused, for backward compatibility.

format_edo_worksheet() now also runs every descriptive/narrative cell
through format_output_text() (capitalizes the first letter of every
sentence; uppercases "npdxxxx"/"edo-xx" document codes to
"NPDxxxx"/"EDO-XX"), and the New EDO tag is now the fixed literal
"EDO-XX\nNew" (previously "EDO-XX\nNew EDO").
================================================================================
"""

import os
import re
import json
import time
import logging
import sys
import unicodedata
import openpyxl

from datetime import datetime

from openpyxl.styles import (
    Alignment,
    Border,
    Side,
    Font
)
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText

from Files.database import DatabaseHandler
from retrieval.retrieve_content_prompt import retrieve_content_for_prompt


# ==========================================================
# GLOBALS
# ==========================================================

CURRENT_FOLDER = os.path.dirname(os.path.abspath(__file__))

MAX_LLM_RETRIES = 3
INITIAL_RETRY_DELAY = 1


logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    stream=sys.stdout
)


# ==========================================================
# EXCEL FORMATTING
# ==========================================================

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

cell_alignment = Alignment(
    horizontal="left",
    vertical="top",
    wrap_text=True
)

aptos_font = Font(name="Aptos Narrow", size=8)


# ==========================================================
# RICH-TEXT / COLOR FORMATTING (drives format_edo_worksheet)
# ==========================================================
# Tweak these to match your exact template look-and-feel.

FONT_NAME = "Calibri"
FONT_SIZE = 10

BLACK = "FF000000"
RED = "FFFF0000"

# Lines matching these patterns (inside Column D / Column E multi-line
# text) are treated as "reference / traceability ID" lines -> RED + BOLD.
RED_BOLD_PATTERNS = [
    r'^\(DRS-\d+\)',        # (DRS-95) NPD44078 Rev 1 ...
    r'^\(MS\s*CU',          # (MS CU Mod-448) NPD45862 Rev 3 ...
    r'^RA-\d+',             # RA-133
    r'^FMEA\s*Sys-\d+',     # FMEA Sys-725
    r'^RRAA',               # RRAA - NPD35987
]

# Lines matching these patterns are treated as "organisation / document
# name" lines -> BLACK + BOLD.
BLACK_BOLD_PATTERNS = [
    r'^\(MS\s*ACC',         # (MS ACC Mod-413, ...) - NPD44001
    r'^NPD\d+',             # NPD36702 Vest APX Risk Assessment and Control
]


def _font(color=BLACK, bold=False, size=FONT_SIZE, name=FONT_NAME):
    return Font(name=name, size=size, color=color, bold=bold)


def _inline_font(color=BLACK, bold=False, size=FONT_SIZE, name=FONT_NAME):
    return InlineFont(rFont=name, sz=size, color=color, b=bold)


def _text(value):
    return "" if value is None else str(value)


def _apply_border_alignment(cell):
    cell.alignment = cell_alignment
    cell.border = thin_border


def classify_line(line):
    """
    Returns (color, bold) for a single line of a multi-line Column D /
    Column E cell value, based on simple regex rules. Anything that
    doesn't match a rule falls back to plain BLACK, regular text
    (continuation / description lines, e.g. "Rev 1 - Vest APX ... TDR").
    """
    stripped = line.strip()

    if not stripped:
        return BLACK, False

    for pattern in RED_BOLD_PATTERNS:
        if re.search(pattern, stripped, re.IGNORECASE):
            return RED, True

    for pattern in BLACK_BOLD_PATTERNS:
        if re.search(pattern, stripped, re.IGNORECASE):
            return BLACK, True

    return BLACK, False


def build_rich_text(value):
    """
    Splits a multi-line cell value on "\\n" and builds a CellRichText
    object, giving each line its own color/bold via classify_line().
    Falls back to a plain string when there's nothing meaningful to
    style, so single-line / blank cells don't get needless rich-text
    wrapping.
    """
    text = _text(value)

    if not text.strip() or "\n" not in text:
        color, bold = classify_line(text)
        if bold or color == RED:
            return CellRichText(TextBlock(_inline_font(color=color, bold=bold), text))
        return text

    lines = text.split("\n")
    blocks = []

    for idx, line in enumerate(lines):
        color, bold = classify_line(line)
        blocks.append(TextBlock(_inline_font(color=color, bold=bold), line))
        if idx != len(lines) - 1:
            blocks.append("\n")

    return CellRichText(*blocks)


# ==========================================================
# TEXT HELPERS (identical in both source files - kept once)
# ==========================================================

def normalize_text(value):
    if value is None:
        return ""

    value = str(value)

    value = value.replace("\u00A0", " ")
    value = value.replace("\u2007", " ")
    value = value.replace("\u202F", " ")

    value = unicodedata.normalize(
        "NFKC",
        value
    )

    return value.strip()


def clean_llm_response(response):

    response = normalize_text(response)

    return re.sub(
        r"^```json\s*|\s*```$",
        "",
        response,
        flags=re.DOTALL
    ).strip()


def parse_json(response):
    try:

        cleaned = clean_llm_response(response)

        logging.info("========== CLEANED JSON ==========")
        logging.info(cleaned)

        return json.loads(cleaned)

    except Exception as e:

        logging.error(f"JSON Parse Error : {e}")
        logging.error(response)

        return {}


def blank(value):
    """
    NOTE: previously substituted the literal text "Blank" for empty
    values. Per requirement, no placeholder text should ever be written
    to the output Excel - a missing value should simply be an empty
    cell. This now just normalizes the text and returns "" for anything
    empty, instead of inserting "Blank".
    """
    return normalize_text(value)


# ==========================================================
# LLM EXECUTION (identical in both source files - kept once)
# ==========================================================

# NOTE: classify_risk_status() needs a plain "send this single prompt
# string, get a text answer back" call - it isn't tied to any document
# collection, so it can't go through execute_llm/execute_llm_retry
# (those require a RAG "collection" argument). call_llm() below matches
# the same pipeline_config["llm"].generate(...) pattern already used by
# evaluate_comparison() elsewhere in this app.
def call_llm(prompt, pipeline_config):
    try:
        llm = pipeline_config["llm"]
        return llm.generate(
            prompt,
            context="",
            question="risk classification",
            temperature=pipeline_config["temperature"],
            max_tokens=pipeline_config["max_tokens"]
        )
    except Exception as e:
        logging.error(f"LLM risk classification call failed: {e}")
        return "No"


def clean_response(response):
    """
    Strips whitespace/code-fences from a plain-text LLM response
    (used by classify_risk_status - it only ever expects one bare word
    back, e.g. "High").
    """
    return clean_llm_response(response)


def execute_llm(
    pipeline_config,
    collection,
    prompt_row
):
    """
    Generic LLM wrapper used by every prompt passed via positional arguments.
    """
    return retrieve_content_for_prompt(
        pipeline_config,
        collection,
        prompt_row["question"],
        prompt_row["prompt_role"],
        prompt_row["prompt_text"],
        prompt_row["fulltext"],
        prompt_row["where_filter"],
        prompt_row["where_document"],
        prompt_row.get("checkpoint", "")
    )


def execute_llm_retry(
    pipeline_config,
    collection,
    prompt_row
):
    """
    Executes the LLM with retry logic.
    """

    last_exception = None

    for attempt in range(MAX_LLM_RETRIES):

        try:

            docs, metadata, response = execute_llm(
                pipeline_config,
                collection,
                prompt_row
            )

            if response:
                return docs, metadata, response

        except Exception as ex:

            last_exception = ex

            logging.warning(
                f"LLM Retry {attempt+1}/{MAX_LLM_RETRIES}"
            )

            time.sleep(
                INITIAL_RETRY_DELAY * (attempt + 1)
            )

    raise Exception(
        f"LLM failed after retries : {last_exception}"
    )


# ==========================================================
# WORKBOOK (identical in both source files - kept once)
# ==========================================================

def initialize_workbook(
    pipeline_config
):
    """
    Opens Excel template.
    """

    workbook = openpyxl.load_workbook(
        pipeline_config["input_file_path"]
    )

    sheet = workbook.active

    return workbook, sheet


# ==========================================================
# COMMON EMPTY EDO STRUCTURE
# ==========================================================

def create_empty_edo():
    """
    CANONICAL version - from edo_existing_final.py.
    Superset of the copy-file version: also carries FMEA_Number so that
    Stage 3B (verification-reference matching by RA/FMEA number) has
    something to match against.

    NOTE: defaults are empty strings, not the literal text "Blank" -
    an unfound value should render as a truly empty cell in the output
    Excel, per requirement.
    """
    return {
        "edo_type": "Existing",
        "edo_tag": "",
        "ra_number": "",
        "FMEA_Number": "",
        "edo_description": "",
        "reason_identified": "",
        "dfmea": "",
        "verification_reference": "",
        "location": "",
        "description_2": "",
        "reason_2": "",
        "sysdd": ""
    }


def create_empty_edo_minimal():
    """
    Original version from generate_EDO_template_copy.py - kept for
    backward compatibility. Does NOT carry FMEA_Number. Not used by the
    merged pipeline (create_empty_edo is used instead).
    """
    return {
        "edo_type": "Existing",
        "edo_tag": "Blank",
        "ra_number": "Blank",
        "edo_description": "Blank",
        "reason_identified": "Blank",
        "dfmea": "Blank",
        "verification_reference": "Blank",
        "location": "Blank",
        "description_2": "Blank",
        "reason_2": "Blank",
        "sysdd": "Blank"
    }


# ==========================================================
# LOGGING (identical in both source files - kept once)
# ==========================================================

def log_dictionary(title, dictionary):
    logging.info("=" * 80)
    logging.info(title)

    for key, value in dictionary.items():
        logging.info(f"{key}")
        logging.info(json.dumps(value, indent=4))

    logging.info("=" * 80)


# ==========================================================
# DOCUMENT HELPERS
# ==========================================================

def get_template_documents(
    client,
    product_family,
    product,
    templatename,
    db: DatabaseHandler
):
    docs = db.get_template_documents(
        client,
        product_family,
        product,
        templatename
    )

    if not docs:
        raise Exception("No template documents were found.")

    logging.info("=" * 80)
    logging.info("AVAILABLE TEMPLATE DOCUMENTS")
    logging.info("=" * 80)

    for doc in docs:
        logging.info(
            f"{doc.get('document_identity')}  -->  "
            f"{doc.get('document_name')}"
        )

    logging.info("=" * 80)

    return docs


def find_document(documents, identity):
    for document in documents:
        if normalize_text(document.get("document_identity")) == normalize_text(identity):
            return document
    return None


def get_edo_document(
    client,
    product_family,
    product,
    templatename,
    db: DatabaseHandler
):
    """
    CANONICAL version - from edo_existing_final.py.
    Still resolves EDO_Proposed, EDO_RA_C and EDO_FMEA. Raises if
    EDO_Proposed (required by every stage) is missing.
    """
    documents = get_template_documents(
        client,
        product_family,
        product,
        templatename,
        db
    )

    if not documents:
        raise Exception("No EDO template documents found.")

    edo_documents = {}

    for document in documents:
        identity = normalize_text(document.get("document_identity")).lower()
        logging.info(f"AVAILABLE DOCUMENT : {identity}")

        if identity == "edo_proposed":
            edo_documents["edo_proposed"] = document
        elif identity in ["edo_ra_c", "edo_ra&c", "edo_rac", "edo_ra"]:
            edo_documents["edo_ra_c"] = document
        elif identity in ["edo_fmea", "fmea", "system_fmea"]:
            edo_documents["edo_fmea"] = document

    logging.info("=" * 80)
    logging.info("EDO DOCUMENT CONFIGURATION")
    logging.info("=" * 80)

    for key, doc in edo_documents.items():
        logging.info(f"{key}")
        logging.info(f"Identity   : {doc.get('document_identity')}")
        logging.info(f"Name       : {doc.get('document_name')}")
        logging.info(f"Collection : {doc.get('collection')}")

    logging.info("=" * 80)

    if "edo_proposed" not in edo_documents:
        raise Exception("Required document EDO_Proposed was not found.")

    return edo_documents


def get_edo_document_lenient(
    client,
    product_family,
    product,
    templatename,
    db: DatabaseHandler
):
    """
    Original version from generate_EDO_template_copy.py - kept for
    backward compatibility. Does NOT raise if EDO_Proposed is missing.
    Not used by the merged pipeline (get_edo_document is used instead).
    """
    documents = get_template_documents(
        client,
        product_family,
        product,
        templatename,
        db
    )

    if not documents:
        raise Exception("No EDO template documents found.")

    edo_documents = {}

    for document in documents:
        identity = normalize_text(document.get("document_identity")).lower()
        logging.info(f"AVAILABLE DOCUMENT : {identity}")

        if identity == "edo_proposed":
            edo_documents["edo_proposed"] = document
        elif identity in ["edo_ra_c", "edo_ra&c", "edo_rac", "edo_ra"]:
            edo_documents["edo_ra_c"] = document
        elif identity in ["edo_fmea", "fmea", "system_fmea"]:
            edo_documents["edo_fmea"] = document

    logging.info("=" * 80)
    logging.info("EDO DOCUMENT CONFIGURATION")
    logging.info("=" * 80)

    for key, doc in edo_documents.items():
        logging.info(f"{key}")
        logging.info(f"Identity   : {doc.get('document_identity')}")
        logging.info(f"Name       : {doc.get('document_name')}")
        logging.info(f"Collection : {doc.get('collection')}")

    logging.info("=" * 80)

    return edo_documents


def get_prompt(
    client,
    product_family,
    product,
    templatename,
    prompt_name,
    db: DatabaseHandler
):
    prompt = db.get_prompt_by_name(
        client,
        product_family,
        product,
        templatename,
        prompt_name
    )

    if not prompt:
        raise Exception(f"Prompt '{prompt_name}' was not found.")

    logging.info(f"Loaded Prompt : {prompt_name}")
    return prompt


def build_prompt_row(
    prompt,
    question,
    fulltext="Yes",
    where_filter="",
    where_document="",
    checkpoint=""
):
    return {
        "prompt_role": prompt["prompt_role"],
        "prompt_text": prompt["prompt_text"],
        "question": question,
        "fulltext": fulltext,
        "where_filter": where_filter,
        "where_document": where_document,
        "checkpoint": checkpoint
    }


def execute_prompt(
    pipeline_config,
    collection,
    prompt
):
    _, _, response = execute_llm_retry(
        pipeline_config,
        collection,
        prompt
    )

    logging.info("=" * 80)
    logging.info("RAW LLM RESPONSE")
    logging.info("=" * 80)
    logging.info(response)

    return parse_json(response)


# ==========================================================
# UNCONDITIONAL RECORD DEEP EXTRACTION
# ==========================================================

def deep_extract_records(data):
    """
    Recursive lookup utility to unwrap fluctuating JSON parent containers.
    Merged version: union of the container keys recognised by both source
    files (edo_existing_final.py's set plus generate_EDO_template_copy.py's
    extra "New_EDOs" / "New_Edos" / "new_edos" keys), so this single
    function correctly unwraps both existing-EDO and new-EDO responses.
    """
    if isinstance(data, list):
        return [item for item in data if isinstance(item, dict)]

    if isinstance(data, dict):
        for key in [
            "Records", "records",
            "New_EDOs", "New_Edos", "new_edos",
            "EDO_Table", "EDO_Tag_Values", "Verification_Details"
        ]:
            if key in data and isinstance(data[key], list):
                return [item for item in data[key] if isinstance(item, dict)]

        if any(k in data for k in ["Product_Feature_Function", "RA_Number", "FMEA_Number", "Traceability"]):
            return [data]

        for val in data.values():
            if isinstance(val, list) and len(val) > 0 and isinstance(val[0], dict):
                return val
            elif isinstance(val, dict):
                res = deep_extract_records(val)
                if res:
                    return res
    return []


# ==========================================================
# STAGE 3: EXISTING EDO RECORD WORKFLOW
# ==========================================================

def extract_edo_tags(
    client,
    product_family,
    product,
    templatename,
    pipeline_config,
    edo_document,
    db: DatabaseHandler
):
    """
    CANONICAL - requested function #1, taken from edo_existing_final.py.
    """
    logging.info("=" * 80)
    logging.info("STAGE 3: EXTRACTING EXISTING EDO TAGS")
    logging.info("=" * 80)

    prompt = get_prompt(
        client,
        product_family,
        product,
        templatename,
        "EDO_Existing_Tag",
        db
    )

    prompt_row = build_prompt_row(
        prompt,
        "Extract all Known Active Design Output Tracking Numbers."
    )

    result = execute_prompt(
        pipeline_config,
        edo_document["edo_proposed"]["collection"],
        prompt_row
    )

    items = deep_extract_records(result)
    logging.info(f"Unconditional Record Extraction count: {len(items)}")

    tags = {}
    for item in items:
        if not isinstance(item, dict):
            continue

        tag = (
            item.get("edo_number")
            or item.get("edo_tag")
            or item.get("EDO Number")
            or item.get("EDO_Tag")
            or item.get("EDO")
            or ""
        )
        tag = normalize_text(tag)

        if tag == "" or tag.lower() == "blank" or tag in tags:
            continue

        tags[tag] = create_empty_edo()
        tags[tag]["edo_type"] = "Existing"
        tags[tag]["edo_tag"] = tag

    return tags


def validate_existing_tags(tags):
    validated = {}
    for key, value in tags.items():
        tag = normalize_text(value.get("edo_tag"))
        if tag == "" or tag.lower() == "blank":
            continue
        validated[tag] = value
    return validated


def extract_ra_fmea_from_text(text):
    """
    Pulls RA_Number / FMEA_Number back out of the column D (dfmea) narrative
    that was already extracted for an existing EDO in extract_edo_details().
    This is the "previous edo dictionary" data used for matching in Stage
    3B - reused instead of asking the LLM for tags a second time.

    The RA&C / FMEA documents identify records with bare tokens like
    "RA-180" and "SYS-147" (sometimes written "FMEA SYS-147") - there is
    no literal word "Number" next to them, so we match the token pattern
    directly rather than looking for a "RA Number:" label.
    """
    text = normalize_text(text)

    ra_match = re.search(r"RA[\s-]?(\d+)", text, re.IGNORECASE)
    fmea_match = re.search(r"SYS[\s-]?(\d+)", text, re.IGNORECASE)

    ra_number = f"RA-{ra_match.group(1)}" if ra_match else ""
    fmea_number = f"SYS-{fmea_match.group(1)}" if fmea_match else ""

    return ra_number, fmea_number


def normalize_id(value):
    """
    Canonicalizes an RA/FMEA identifier for comparison purposes:
    uppercase, single spaces, trimmed. Used so that matching between the
    column-D-derived identifiers and whatever format Stage 3B's
    verification prompt returns doesn't fail on case/whitespace noise.
    """
    value = normalize_text(value).upper()
    return re.sub(r"\s+", " ", value).strip()


def extract_edo_details(
    client,
    product_family,
    product,
    templatename,
    pipeline_config,
    edo_document,
    existing_edos,
    db: DatabaseHandler
):
    """
    CANONICAL version - from edo_existing_final.py.
    Comprehensive attribute hydration: extracts description/reason lists,
    dfmea narrative, location, sysdd, and (crucially) RA_Number /
    FMEA_Number - needed downstream by Stage 3B verification matching.
    """
    logging.info("=" * 80)
    logging.info("STAGE 3: ATTRIBUTE HYDRATION (EXISTING EDO)")
    logging.info("=" * 80)

    prompt = get_prompt(
        client,
        product_family,
        product,
        templatename,
        "EDO_Existing_Generic",
        db
    )

    def normalize_key(key):
        key = normalize_text(key).lower()
        key = key.replace("&", "and")
        key = re.sub(r"[^a-z0-9]+", "_", key)
        key = re.sub(r"_+", "_", key)
        return key.strip("_")

    def first_value(record, keywords):
        """
        Return first non-empty value whose normalized key contains
        any keyword as a substring.
        """
        normalized_record = {normalize_key(k): v for k, v in record.items()}
        for kw in keywords:
            for nk, v in normalized_record.items():
                if kw in nk or nk in kw:
                    if normalize_text(v):
                        return normalize_text(v)
        return ""

    for edo_tag in existing_edos.keys():

        logging.info(f"Hydrating {edo_tag}")

        question = (
            f"Extract the complete row details for the EDO tag requested by the user. Requested EDO Tag: {edo_tag}  Find the row where the EDO Tag exactly matches the requested EDO Ta. Return all available column values for that row, including EDO number/tag, description, reason identified as EDO, RA&C and/or Sys-DFMEA Trace,EDO Location, EDO Description, reason identified as EDO  If the requested EDO tag is not found, return none"
        )

        prompt_row = build_prompt_row(
            prompt,
            question
        )

        result = {}

        # 1. Fetch data
        for attempt in range(3):
            raw_result = execute_prompt(
                pipeline_config,
                edo_document["edo_proposed"]["collection"],
                prompt_row
            )
            records = deep_extract_records(raw_result)

            for record in records:
                tag = normalize_text(
                    record.get("edo_number")
                    or record.get("edo_tag")
                    or record.get("EDO Number")
                    or record.get("EDO_Tag")
                ).lower()

                if tag == edo_tag.lower():
                    result = record
                    break

            # 2. Debug log
            logging.info(f"DEBUG: Hydration data for {edo_tag}: {json.dumps(result, indent=2)}")

            # Always accept whatever we found to avoid "Blank"
            if result:
                break
            time.sleep(INITIAL_RETRY_DELAY)

        normalized = {normalize_key(k): v for k, v in result.items()}
        edo = existing_edos[edo_tag]

        # Description/Reason Parsing
        def extract_list_fields(prefix):
            items = []
            for k, v in normalized.items():
                if prefix in k:
                    if normalize_text(v):
                        items.append(normalize_text(v))
            return sorted(list(set(items)), key=len)

        desc_list = extract_list_fields("description") + extract_list_fields("notes")
        edo["edo_description"] = blank(desc_list[0] if desc_list else "")
        edo["description_2"] = blank(desc_list[-1] if len(desc_list) > 1 else "")

        reason_list = extract_list_fields("reason")
        edo["reason_identified"] = blank(reason_list[0] if reason_list else "")
        edo["reason_2"] = blank(reason_list[-1] if len(reason_list) > 1 else "")

        # Remaining fields
        edo["dfmea"] = blank(first_value(result, ["dfmea", "trace", "ra_and", "sys_dfmea"]))
        edo["location"] = blank(first_value(result, ["location"]))
        edo["sysdd"] = blank(first_value(result, ["sysdd", "hardware", "design_reference"]))

        ra_number = first_value(result, ["ra_number", "ra"])
        fmea_number = first_value(result, ["fmea_number", "fmea"])

        if not ra_number or not fmea_number:
            p_ra, p_fmea = extract_ra_fmea_from_text(edo["dfmea"])
            ra_number = ra_number or p_ra
            fmea_number = fmea_number or p_fmea

        edo["ra_number"] = ra_number or ""
        edo["FMEA_Number"] = fmea_number or ""
        edo["verification_reference"] = ""

    return existing_edos


def extract_edo_details_basic(
    client,
    product_family,
    product,
    templatename,
    pipeline_config,
    edo_document,
    existing_edos,
    db: DatabaseHandler
):
    """
    Original simpler version from generate_EDO_template_copy.py - kept for
    backward compatibility. Not used by the merged pipeline
    (extract_edo_details is used instead, since it also derives
    RA_Number/FMEA_Number needed by Stage 3B).
    """
    logging.info("=" * 80)
    logging.info("STAGE 3: WORKFLOW 1 - ATTRIBUTE HYDRATION (EXISTING) [basic]")
    logging.info("=" * 80)

    prompt = get_prompt(
        client,
        product_family,
        product,
        templatename,
        "EDO_Existing_Generic",
        db
    )

    for edo_tag in existing_edos.keys():
        question = f"Extract descriptions, reason matrices, traces, locations, and references for: {edo_tag}"
        prompt_row = build_prompt_row(prompt, question)
        result = execute_prompt(
            pipeline_config,
            edo_document["edo_proposed"]["collection"],
            prompt_row
        )

        if isinstance(result, list):
            result = {} if len(result) == 0 else result[0]

        edo = existing_edos[edo_tag]
        edo["edo_description"] = blank(result.get("edo_description") or result.get("EDO Description") or result.get("description"))
        edo["reason_identified"] = blank(result.get("reason_identified") or result.get("Reason Identified") or result.get("reason"))
        edo["dfmea"] = blank(result.get("dfmea") or result.get("DFMEA") or result.get("RA&C and/or Sys-DFMEA Trace"))
        edo["verification_reference"] = "Blank"
        edo["location"] = blank(result.get("location") or result.get("EDO Location") or result.get("Location"))
        edo["description_2"] = blank(result.get("description_2") or result.get("EDO Description 2") or result.get("Description 2"))
        edo["reason_2"] = blank(result.get("reason_2") or result.get("Reason 2"))
        edo["sysdd"] = blank(result.get("sysdd") or result.get("SYSDD") or result.get("SYS DD") or result.get("sysdd_reference"))

    return existing_edos


# ==========================================================
# STAGE 3B: EXISTING EDO VERIFICATION REFERENCE (COLUMN E)
# ==========================================================

def extract_existing_edo_verification_details(
    client,
    product_family,
    product,
    templatename,
    pipeline_config,
    edo_document,
    existing_edos,
    db: DatabaseHandler
):
    """
    CANONICAL - requested function #2, taken from edo_existing_final.py.
    Dedicated verification-reference extraction for EXISTING EDOs: its own
    prompt ("EDO_NEW_Verification_details"), queried against the
    edo_fmea collection, matched by RA/FMEA number that was pulled from
    the column D (dfmea) data in extract_edo_details().
    """
    logging.info("=" * 80)
    logging.info("STAGE 3B: EXISTING EDO - VERIFICATION REFERENCE EXTRACTION")
    logging.info("=" * 80)

    if "edo_fmea" not in edo_document:
        raise Exception(
            "EDO_FMEA document/collection not configured - cannot run "
            "verification reference extraction."
        )

    prompt_data = get_prompt(
        client,
        product_family,
        product,
        templatename,
        "Edo_existing_Verification",
        db
    )

    targets = "\n".join(
        f"RA_Number : {edo.get('ra_number')}\nFMEA_Number : {edo.get('FMEA_Number')}"
        for edo in existing_edos.values()
        if edo.get("ra_number") not in (None, "", "Blank") or edo.get("FMEA_Number") not in (None, "", "Blank")
    )

    if not targets:
        logging.warning(
            "No RA/FMEA numbers available on any existing EDO - skipping "
            "verification reference extraction."
        )
        return {}

    prompt_row = {
        "prompt_role": prompt_data["prompt_role"],
        "prompt_text": prompt_data["prompt_text"] + "\nTARGETS:\n" + targets,
        "question": "Fetch verification reference records for the folowing sys number <FMEA_Number> number",
        "fulltext": "Yes",
        "where_filter": "",
        "where_document": "",
        "checkpoint": ""
    }

    _, _, response = execute_llm_retry(
        pipeline_config,
        edo_document["edo_fmea"]["collection"],
        prompt_row
    )

    return parse_json(response)


def apply_existing_edo_verification(existing_edos, verification_details):
    """
    CANONICAL - requested function #3, taken from edo_existing_final.py.
    Couples the verification records back onto existing_edos by RA/FMEA
    number. Matching is format-tolerant (case/whitespace normalized, plus
    substring containment).
    """
    records = deep_extract_records(verification_details)

    for row in records:
        if not isinstance(row, dict):
            continue

        trace_fmea = normalize_id(row.get("FMEA_Number") or row.get("FMEA Number") or "")
        trace_ra = normalize_id(row.get("RA_Number") or row.get("RA Number") or "")

        verification = normalize_text(
            row.get("Verification_Reference")
            or row.get("Verification Reference")
            or row.get("verification_reference")
        )
        if not verification or verification.lower() == "none":
            continue

        for edo in existing_edos.values():
            edo_fmea = normalize_id(edo.get("FMEA_Number") or "")
            edo_ra = normalize_id(edo.get("ra_number") or "")

            fmea_match = trace_fmea and edo_fmea and (trace_fmea == edo_fmea or trace_fmea in edo_fmea or edo_fmea in trace_fmea)
            ra_match = trace_ra and edo_ra and (trace_ra == edo_ra or trace_ra in edo_ra or edo_ra in trace_ra)

            if fmea_match or ra_match:
                edo["verification_reference"] = verification

    return existing_edos


# ==========================================================
# STAGE 4: NEW EDO RECORD WORKFLOW
# ==========================================================
def extract_new_edo_tags(
    client,
    product_family,
    product,
    templatename,
    pipeline_config,
    edo_document,
    db
):
    """
    STEP 1 of 2: scans Section 10 of the FMEA against the RA&C and
    returns ONLY the lightweight identification triplet
    {RA_Number, FMEA_Number, Status} for every qualifying new-EDO row.

    No narrative extraction happens here - keeping the response small
    means this call never truncates. Full-detail extraction happens
    afterwards, in extract_new_edo_summary_details(edo_tags=...),
    scoped to exactly the RA/FMEA pairs returned here.
    """
    logging.info("=" * 80)
    logging.info("STAGE 4a: WORKFLOW 2 - NEW EDO IDENTIFICATION (TAGS ONLY)")
    logging.info("=" * 80)

    prompt_data = db.get_prompt_by_name(
        client,
        product_family,
        product,
        templatename,
        "EDO_NEW_tags"
    )

    prompt_row = {
        "prompt_role": prompt_data["prompt_role"],
        "prompt_text": prompt_data["prompt_text"],
        "question": (
            "Scan Section 10 of the FMEA against the RA&C and return "
            "only the RA_Number / FMEA_Number / Status identification "
            "list for every qualifying new EDO row - no narrative "
            "extraction."
        ),
        "fulltext": "Yes",
        "where_filter": "",
        "where_document": "",
        "checkpoint": ""
    }

    _, _, response = execute_llm_retry(
        pipeline_config,
        edo_document["edo_fmea"]["collection"],
        prompt_row
    )

    return parse_json(response)

def extract_new_edo_summary_details(
    client,
    product_family,
    product,
    templatename,
    pipeline_config,
    edo_document,
    edo_tags,
    db,
    min_batch_size=1
):
    """
    STEP 2 of 2: takes the {RA_Number, FMEA_Number} pairs identified by
    extract_new_edo_tags() and extracts the FULL detail record for
    ONLY those pairs (Product_Feature_Function, Reason_Identified_as_EDO,
    Traceability, Verification_Reference, EDO_Location, EDO_Description,
    Reason_Identified_as_EDO_ColH).

    COST-EFFICIENT LAZY SPLIT: tries the WHOLE list of targets in ONE
    LLM call first. Only if that call comes back with fewer records
    than targets sent (truncation) does it split the list in half and
    retry - and it only re-calls the half(s) that are actually short,
    never the whole list again. Best case (no truncation): 1 LLM call
    total. Worst case: falls back toward smaller calls only as needed.
    """
    logging.info("=" * 80)
    logging.info("STAGE 4b: WORKFLOW 2 - NEW EDO FULL DETAIL EXTRACTION (LAZY SPLIT)")
    logging.info("=" * 80)

    if not edo_tags:
        logging.info("No new EDO tags to extract details for.")
        return []

    prompt_data = db.get_prompt_by_name(
        client,
        product_family,
        product,
        templatename,
        "EDO_NEW_details"
    )

    def _pair_key(item):
        return (
            normalize_text(item.get("RA_Number", "")).lower(),
            normalize_text(item.get("FMEA_Number", "")).lower()
        )

    def _call_llm_for_targets(targets_list):
        """
        One LLM call for the given list of targets. Returns the parsed
        records list (may be shorter than targets_list if truncated).
        """
        targets_text = "\n".join(
            [
                f"RA_Number : {item.get('RA_Number')}\nFMEA_Number : {item.get('FMEA_Number')}"
                for item in targets_list
            ]
        )

        prompt_row = {
            "prompt_role": prompt_data["prompt_role"],
            "prompt_text": prompt_data["prompt_text"] + "\nTARGETS:\n" + targets_text,
            "question": (
                "For the listed RA/FMEA target pairs only, extract the "
                "full EDO detail record for each - one JSON object per "
                "target pair, no others."
            ),
            "fulltext": "Yes",
            "where_filter": "",
            "where_document": "",
            "checkpoint": ""
        }

        try:
            _, _, response = execute_llm_retry(
                pipeline_config,
                edo_document["edo_fmea"]["collection"],
                prompt_row
            )
            records = parse_json(response)
            return records if isinstance(records, list) else []

        except Exception as e:
            logging.error(f"LLM call failed for {len(targets_list)} targets: {e}")
            return []

    def _process(targets_list):
        """
        Tries targets_list in ONE call. Only if that call comes back
        incomplete does it split in half and recurse - so a clean run
        costs exactly 1 call, and only the actually-truncated portion
        ever gets a second call.
        """
        records = _call_llm_for_targets(targets_list)

        if len(records) >= len(targets_list):
            return records

        logging.warning(
            f"Call for {len(targets_list)} targets returned only "
            f"{len(records)} records - likely truncation."
        )

        if len(targets_list) <= min_batch_size:
            # Can't split further - accept what came back; the
            # completeness check below will name the exact gap.
            return records

        mid = len(targets_list) // 2
        logging.info(
            f"Splitting into {mid} + {len(targets_list) - mid} and retrying "
            f"ONLY those two halves (not the whole list again)."
        )
        return _process(targets_list[:mid]) + _process(targets_list[mid:])

    # ---- Single attempt for the WHOLE list first (cheapest path) ----
    all_records = _process(edo_tags)

    # ---- Completeness verification (always runs, costs nothing extra) ----
    returned_keys = {_pair_key(r) for r in all_records}
    missing = [
        item for item in edo_tags
        if _pair_key(item) not in returned_keys
    ]

    if missing:
        logging.error(
            f"INCOMPLETE EXTRACTION: {len(missing)} of {len(edo_tags)} "
            f"targets have no returned record: "
            + ", ".join(f"{m.get('RA_Number')}/{m.get('FMEA_Number')}" for m in missing)
        )
    else:
        logging.info(f"All {len(edo_tags)} new EDO targets extracted successfully.")

    return all_records

def extract_new_edo_traceability_details(
    client,
    product_family,
    product,
    templatename,
    pipeline_config,
    edo_document,
    edo_tags,
    db
):
    """
    CANONICAL - requested function #6, taken from generate_EDO_template_copy.py.
    """
    logging.info("=" * 80)
    logging.info("STAGE 4: WORKFLOW 2 - TRACEABILITY VERIFICATION EXTRACTIONS")
    logging.info("=" * 80)

    prompt_data = db.get_prompt_by_name(
        client,
        product_family,
        product,
        templatename,
        "EDO_NEW_Verification_details"
    )

    targets = "\n".join(
        [f"RA_Number : {item.get('RA_Number')}\nFMEA_Number : {item.get('FMEA_Number')}" for item in edo_tags]
    )

    prompt_row = {
        "prompt_role": prompt_data["prompt_role"],
        "prompt_text": prompt_data["prompt_text"] + "\nTARGETS:\n" + targets,
        "question": "Fetch spreadsheets, reports, locations and pass/fail results",
        "fulltext": "Yes",
        "where_filter": "",
        "where_document": "",
        "checkpoint": ""
    }

    _, _, response = execute_llm_retry(
        pipeline_config,
        edo_document["edo_fmea"]["collection"],
        prompt_row
    )
    print(f"new edo:",response)
    return parse_json(response)


def extract_new_edo_ra_details(
    client,
    product_family,
    product,
    templatename,
    pipeline_config,
    edo_document,
    edo_tags,
    db
):
    """
    NEW - companion to extract_new_edo_summary_details / extract_new_edo_traceability_details
    (both of which are left completely UNMODIFIED, per requirement).

    WHY THIS EXISTS:
    extract_new_edo_summary_details / extract_new_edo_traceability_details
    both query the EDO_FMEA collection, so they only ever return data for
    RA records whose Status is "See FMEA" (i.e. records that actually
    have a matching FMEA entry). RA records with Status == "Medium" have
    NO corresponding FMEA entry, so they never show up in those two
    functions' results and were silently being dropped from the final
    output.

    Per requirement: for "Medium" risk-value RA records, do NOT consult
    the FMEA document at all - the EDO_RA_C document itself already
    carries the same Product_Feature_Function / Reason_Identified_as_EDO
    columns, so this function extracts those values directly from
    EDO_RA_C for exactly the Medium-status subset of edo_tags, ensuring
    every RA Number returned by extract_new_edo_tags gets printed to the
    output Excel - none are excluded.
    """
    logging.info("=" * 80)
    logging.info("STAGE 4: WORKFLOW 2 - MEDIUM RISK RA BACKFILL (FROM EDO_RA_C, NOT FMEA)")
    logging.info("=" * 80)

    medium_tags = [
        item for item in edo_tags
        if "medium" in normalize_text(item.get("Status")).lower()
    ]

    if not medium_tags:
        logging.info("No 'Medium' risk-value RA records found - nothing to backfill.")
        return {}

    if "edo_ra_c" not in edo_document:
        logging.warning(
            "EDO_RA_C document/collection not configured - cannot backfill "
            "Medium risk-value RA records. These will still be printed to "
            "the output, but with empty description/reason fields."
        )
        return {}

    prompt_data = db.get_prompt_by_name(
        client,
        product_family,
        product,
        templatename,
        "EDO_NEW_details_Medium_Risk"
    )

    targets = "\n".join(
        [f"RA_Number : {item.get('RA_Number')}\nFMEA_Number : {item.get('FMEA_Number')}" for item in medium_tags]
    )

    prompt_row = {
        "prompt_role": prompt_data["prompt_role"],
        "prompt_text": prompt_data["prompt_text"] + "\nTARGETS:\n" + targets,
        "question": "Extract core function parameters and identification reasons directly from the RA&C document for the listed RA numbers (Medium risk value - no FMEA reference)",
        "fulltext": "Yes",
        "where_filter": "",
        "where_document": "",
        "checkpoint": ""
    }

    _, _, response = execute_llm_retry(
        pipeline_config,
        edo_document["edo_ra_c"]["collection"],
        prompt_row
    )

    return parse_json(response)


def get_llm_value(row, *keys):
    """
    Returns the first valid, non-empty value found across `keys`.
    Treats None, "", and any case-insensitive "none"/"blank" placeholder
    text (e.g. "None", "NONE", "Blank") as invalid/empty, so literal
    placeholder strings coming back from the LLM never get written to
    the output Excel as if they were real data.
    """
    for key in keys:
        value = row.get(key)
        if value is None:
            continue
        text = str(value).strip()
        if text == "" or text.lower() in ("none", "blank"):
            continue
        return value
    return ""


# Column J (SYSDD / HDD Reference) is a single, fixed document reference
# for this template - it is NOT something that needs to be extracted per
# EDO record via the LLM. Kept as a named constant so it's easy to update
# in one place if the reference document ever changes.
FIXED_SYSDD_REFERENCE = "NPD38119 Titan Hardware Detailed Design"


def get_fixed_sysdd_reference():
    """
    Dedicated function (per requirement) for Column J (SYSDD / HDD
    Reference). Always returns the fixed reference document string,
    regardless of EDO type (Existing or New) or of whatever value an
    upstream extraction step may have found.
    """
    return FIXED_SYSDD_REFERENCE


# ==========================================================
# OUTPUT TEXT NORMALIZATION (drives requirement: capitalize sentences,
# fix "npdxxxx" -> "NPDxxxx" and "edo-29" -> "EDO-29" casing)
# ==========================================================

def _capitalize_sentences(text):
    """
    Capitalizes the first letter of every sentence in `text`, line by
    line (a "sentence" ends at '.', '!', or '?' followed by whitespace).
    Leading whitespace/indentation on each line is preserved.
    """
    lines = text.split("\n")
    result_lines = []

    for line in lines:
        stripped = line.lstrip()
        if not stripped:
            result_lines.append(line)
            continue

        leading_ws = line[:len(line) - len(stripped)]
        sentences = re.split(r'(?<=[.!?])\s+', stripped)
        fixed = []
        for sentence in sentences:
            if sentence:
                fixed.append(sentence[0].upper() + sentence[1:])
            else:
                fixed.append(sentence)
        result_lines.append(leading_ws + " ".join(fixed))

    return "\n".join(result_lines)


def _normalize_document_codes(text):
    """
    Fixes casing on document / tag codes wherever they appear inside a
    sentence:
      - "npd36702" / "Npd36702"     -> "NPD36702"
      - "edo-29" / "Edo-29"         -> "EDO-29"
      - "ra-108" / "Ra-108"         -> "RA-108"
      - "fmea sys-82" / "Fmea sys-82" -> "FMEA Sys-82"
    """
    text = re.sub(r'\bnpd(\d+)', lambda m: f"NPD{m.group(1)}", text, flags=re.IGNORECASE)
    text = re.sub(r'\bedo[\s-]*([a-z0-9]+)', lambda m: f"EDO-{m.group(1).upper()}", text, flags=re.IGNORECASE)
    text = re.sub(r'\bfmea\s+sys[\s-]*(\d+)', lambda m: f"FMEA Sys-{m.group(1)}", text, flags=re.IGNORECASE)
    text = re.sub(r'\bra[\s-]+(\d+)', lambda m: f"RA-{m.group(1)}", text, flags=re.IGNORECASE)
    return text


def format_output_text(value):
    """
    Canonical text formatter applied to every descriptive/narrative cell
    value before it's written to the output Excel:
      - first letter of every sentence is capitalized
      - "npdxxxx" document codes are uppercased to "NPDxxxx"
      - "edo-xx" tag codes are uppercased to "EDO-XX"
    Leaves truly empty values as empty ("") - no placeholder text.
    """
    text = _text(value)
    if not text.strip():
        return text

    text = _capitalize_sentences(text)
    text = _normalize_document_codes(text)
    return text


def format_edo_tag_text(value):
    """
    Same code-casing fix as format_output_text(), but WITHOUT sentence
    capitalization - used specifically for Column A (EDO Tag), which is a
    short code/label rather than a narrative sentence (e.g. "edo-29" ->
    "EDO-29").
    """
    text = _text(value)
    if not text.strip():
        return text
    return _normalize_document_codes(text)


def get_existing_identifier_set(existing_edos):
    """
    Builds a set of every RA Number / FMEA Number already tracked under
    the Existing EDOs (Stage 3), normalized to uppercase for
    case-insensitive comparison. Used to drop any "New EDO" record whose
    RA id or FMEA id is already represented by an Existing EDO, so the
    same underlying issue is never printed twice.
    """
    identifiers = set()
    for edo in existing_edos.values():
        ra = normalize_text(edo.get("ra_number")).upper()
        fmea = normalize_text(edo.get("FMEA_Number")).upper()
        if ra:
            identifiers.add(ra)
        if fmea:
            identifiers.add(fmea)
    return identifiers


# ==========================================================
# FINAL RECORD NORMALIZATION (identical body in both files - kept once)
# ==========================================================

def normalize_edo_record(edo_record):
    default_structure = {
        "edo_type": "Existing",
        "edo_tag": "",
        "RA_Number": "",
        "FMEA_Number": "",
        "edo_description": "",
        "reason_identified": "",
        "dfmea": "",
        "verification_reference": "",
        "location": "",
        "description_2": "",
        "reason_2": "",
        "sysdd": ""
    }
    if not isinstance(edo_record, dict):
        return default_structure

    for key, value in default_structure.items():
        if key not in edo_record:
            edo_record[key] = value
    return edo_record


# ==========================================================
# STAGE 5: UNCONDITIONAL COUPLING / MERGE ENGINE
# ==========================================================

def merge_new_edo_dictionary(
    edo_tags,
    edo_summary_details,
    edo_trace_details
):
    """
    From generate_EDO_template_copy.py - only pipeline that builds NEW EDO
    records, so it is kept as-is and used by the merged pipeline.
    """
    logging.info("=" * 80)
    logging.info("STAGE 5: WORKFLOW MATRIX COUPLING ENGINE")
    logging.info("=" * 80)

    merged = {}

    # Deep Structure Unwrapping
    records = deep_extract_records(edo_summary_details)
    logging.info(f"SUMMARY DETAILS COUNT EXTRICATED : {len(records)}")

    for index, row in enumerate(records):
        ra_number = normalize_text(get_llm_value(row, "RA_Number", "RA Number", "ra_num"))
        fmea_number = normalize_text(get_llm_value(row, "FMEA_Number", "FMEA Number", "fmea_num"))
        status_label = normalize_text(get_llm_value(row, "Status", "status")) or "New EDO"

        key = fmea_number or ra_number
        if not key:
            key = f"NEW-EDO-RECORD-{index}"

        # Safe attribute alignment ensuring literal 'None' strings from the dictionary are preserved
        edo_desc = normalize_text(get_llm_value(row, "Product_Feature_Function", "Product Feature Function", "EDO_Description", "edo_description"))
        reason_id = normalize_text(get_llm_value(row, "Reason_Identified_as_EDO", "Reason Identified as EDO", "reason_identified"))
        dfmea_trace = normalize_text(get_llm_value(row, "Traceability", "dfmea", "traceability"))
        ver_ref = normalize_text(get_llm_value(row, "Verification_Reference", "Verification Reference", "verification_reference"))

        loc_val = normalize_text(get_llm_value(row, "EDO_Location", "location")) or "None"
        desc2_val = normalize_text(get_llm_value(row, "EDO_Description", "Description_2")) or "None"
        reason2_val = normalize_text(get_llm_value(row, "Reason_2")) or "None"
        sys_dd_val = normalize_text(get_llm_value(row, "Reason_Identified_as_EDO_ColH", "sysdd", "SYS_DD")) or "None"

        merged[key] = {
            "edo_type": "New",
            "edo_tag": f"EDO-XX\n{status_label}",
            "edo_description": edo_desc if edo_desc else "Blank",
            "reason_identified": reason_id if reason_id else "Blank",
            "dfmea": dfmea_trace if dfmea_trace else "Blank",
            "verification_reference": ver_ref if ver_ref else "Blank",
            "location": loc_val,
            "description_2": desc2_val,
            "reason_2": reason2_val,
            "sysdd": sys_dd_val,
            "RA_Number": ra_number,
            "FMEA_Number": fmea_number
        }

    # Asymmetric Verification Coupling
    trace_records = deep_extract_records(edo_trace_details)
    for row in trace_records:
        if not isinstance(row, dict):
            continue

        trace_fmea = normalize_text(get_llm_value(row, "FMEA_Number", "FMEA Number"))
        trace_ra = normalize_text(get_llm_value(row, "RA_Number", "RA Number"))
        trace_key = trace_fmea or trace_ra

        verification = normalize_text(get_llm_value(row, "Verification_Reference", "Verification Reference", "verification_reference"))
        if not verification or verification.lower() == "none" or verification == "":
            continue

        if trace_key and trace_key in merged:
            merged[trace_key]["verification_reference"] = verification
        else:
            for main_key, data in merged.items():
                if (trace_ra and data["RA_Number"] == trace_ra) or (trace_fmea and data["FMEA_Number"] == trace_fmea):
                    data["verification_reference"] = verification

    return merged


def merge_new_edo_dictionary_full(
    edo_tags,
    edo_summary_details,
    edo_trace_details,
    edo_ra_details
):
    """
    CANONICAL new-EDO merge - used by generate_edo_template().

    merge_new_edo_dictionary() (above) is left completely UNTOUCHED, but
    it silently drops any RA record whose Status is "Medium", because it
    only ever looks at edo_summary_details / edo_trace_details, both of
    which come exclusively from the EDO_FMEA collection.

    This function fixes that: it builds the output starting from
    edo_tags itself (the full, unfiltered list returned by
    extract_new_edo_tags), so EVERY RA Number is guaranteed a row in the
    final Excel output - none are excluded.

    For each tag:
      - Status contains "Medium"  -> description/reason are pulled ONLY
        from edo_ra_details (EDO_RA_C document) - the FMEA-sourced
        edo_summary_details is deliberately NOT consulted for these,
        exactly as requested.
      - Any other Status (e.g. "See FMEA") -> description/reason/trace/
        verification are pulled from edo_summary_details / edo_trace_details,
        exactly like the original merge_new_edo_dictionary() behaviour.
      - If no matching record is found in the relevant source at all, the
        tag is still written to the output - with empty ("") fields
        rather than a "Blank"/"None" placeholder string, and rather than
        being skipped/excluded.
      - Column A (edo_tag) is always the fixed literal "EDO-XX\\nNew EDO"
        for every New EDO record - the risk Status (Medium / See FMEA) is
        used internally to choose the data source but is never printed.
      - Column D (dfmea) for "Medium" records is always the RA Number
        itself (there is no FMEA trace to show for these), so every RA
        Number is visibly represented in Column D regardless of Status.
    """
    logging.info("=" * 80)
    logging.info("STAGE 5: WORKFLOW MATRIX COUPLING ENGINE (FULL - INCLUDES MEDIUM RA)")
    logging.info("=" * 80)

    def index_records(records):
        index = {}
        for row in records:
            if not isinstance(row, dict):
                continue
            ra = normalize_text(get_llm_value(row, "RA_Number", "RA Number", "ra_num"))
            fmea = normalize_text(get_llm_value(row, "FMEA_Number", "FMEA Number", "fmea_num"))
            if ra and ra not in index:
                index[ra] = row
            if fmea and fmea not in index:
                index[fmea] = row
        return index

    summary_index = index_records(deep_extract_records(edo_summary_details))
    ra_index = index_records(deep_extract_records(edo_ra_details))

    # Verification lookups from the FMEA-sourced traceability extraction -
    # only ever applied to non-Medium ("See FMEA") records.
    trace_index = {}
    for row in deep_extract_records(edo_trace_details):
        if not isinstance(row, dict):
            continue
        trace_ra = normalize_text(get_llm_value(row, "RA_Number", "RA Number"))
        trace_fmea = normalize_text(get_llm_value(row, "FMEA_Number", "FMEA Number"))
        verification = normalize_text(get_llm_value(row, "Verification_Reference", "Verification Reference", "verification_reference"))
        if not verification or verification.lower() == "none":
            continue
        if trace_ra:
            trace_index.setdefault(trace_ra, verification)
        if trace_fmea:
            trace_index.setdefault(trace_fmea, verification)

    merged = {}

    for index, item in enumerate(edo_tags):
        ra_number = normalize_text(item.get("RA_Number"))
        fmea_number = normalize_text(item.get("FMEA_Number"))
        status_label = normalize_text(item.get("Status")) or "New EDO"
        is_medium = "medium" in status_label.lower()

        key = fmea_number or ra_number
        if not key:
            key = f"NEW-EDO-RECORD-{index}"
        # avoid clobbering an existing key on duplicate RA/FMEA numbers
        if key in merged:
            key = f"{key}_{index}"

        if is_medium:
            # Medium risk value -> EDO_RA_C ONLY, never the FMEA document.
            # Keyed by RA Number ONLY - the tag's FMEA_Number is a
            # placeholder ("FMEA-UNKNOWN") for Medium records and is not
            # a reliable lookup key.
            source_row = ra_index.get(ra_number)
        else:
            # See FMEA / anything else -> FMEA-sourced summary, as before
            source_row = summary_index.get(fmea_number) or summary_index.get(ra_number)

        if source_row:
            edo_desc = normalize_text(get_llm_value(source_row, "Product_Feature_Function", "Product Feature Function", "EDO_Description", "edo_description"))
            reason_id = normalize_text(get_llm_value(source_row, "Reason_Identified_as_EDO", "Reason Identified as EDO", "reason_identified"))
            dfmea_trace = normalize_text(get_llm_value(source_row, "Traceability", "dfmea", "traceability"))
            ver_ref = normalize_text(get_llm_value(source_row, "Verification_Reference", "Verification Reference", "verification_reference"))
            loc_val = normalize_text(get_llm_value(source_row, "EDO_Location", "location"))
            desc2_val = normalize_text(get_llm_value(source_row, "EDO_Description", "Description_2"))
            reason2_val = normalize_text(get_llm_value(source_row, "Reason_2"))
        else:
            edo_desc = reason_id = dfmea_trace = ver_ref = ""
            loc_val = desc2_val = reason2_val = ""

        # Per requirement: for "Medium" records, Column D (dfmea) always
        # shows the RA Number itself - there is no FMEA trace document to
        # pull a narrative from, so the RA id is the traceability value.
        if is_medium:
            dfmea_trace = ra_number

        # Verification backfill only applies to non-Medium records, since
        # the trace/verification extraction is itself FMEA-sourced.
        if not is_medium and not ver_ref:
            ver_ref = trace_index.get(fmea_number) or trace_index.get(ra_number) or ver_ref

        # Column J (SYSDD / HDD Reference) is a fixed document reference
        # for this template, not an LLM-extracted value - see
        # get_fixed_sysdd_reference().
        sys_dd_val = get_fixed_sysdd_reference()

        merged[key] = {
            "edo_type": "New",
            # Per requirement: Column A never prints the risk Status
            # (Medium / See FMEA) - always the fixed literal tag text.
            "edo_tag": "EDO-XX\nNew EDO",
            "edo_description": edo_desc,
            "reason_identified": reason_id,
            "dfmea": dfmea_trace,
            "verification_reference": ver_ref,
            "location": loc_val,
            "description_2": desc2_val,
            "reason_2": reason2_val,
            "sysdd": sys_dd_val,
            "RA_Number": ra_number,
            "FMEA_Number": fmea_number
        }

    return merged


def merge_new_edo_records(new_records, existing_edos):
    """
    CANONICAL new-EDO merge - used by generate_edo_template(), replacing
    the old 4-source merge_new_edo_dictionary_full() now that
    extract_new_edo_tags() has been removed and extract_new_edo_summary_details()
    returns complete records directly from the single merged "EDO_NEW_details"
    prompt.

    Fixes the "column mapped wrongly" bug: the previous merge logic looked
    for keys like "Description_2" / "Reason_2" that never actually existed
    in the LLM's JSON output - the real keys are "EDO_Description" and
    "Reason_Identified_as_EDO_ColH". This function maps directly off the
    ACTUAL field names the prompt returns:

        edo_description   <- Product_Feature_Function
        reason_identified <- Reason_Identified_as_EDO
        dfmea             <- Traceability
        verification_reference <- Verification_Reference
        location          <- EDO_Location
        description_2     <- EDO_Description
        reason_2          <- Reason_Identified_as_EDO_ColH
        sysdd             <- fixed constant (get_fixed_sysdd_reference())

    Also applies requirement #2: any new record whose RA Number or FMEA
    Number already belongs to an Existing EDO is dropped from the output,
    so the same issue is never printed twice under both types.

    Column A (edo_tag) is always the fixed literal "EDO-XX\\nNew" - the
    risk Status is never printed there.
    """
    logging.info("=" * 80)
    logging.info("STAGE 5: NEW EDO MERGE (SINGLE-PASS RECORDS, DE-DUPED AGAINST EXISTING)")
    logging.info("=" * 80)

    existing_identifiers = get_existing_identifier_set(existing_edos)

    records = deep_extract_records(new_records)
    merged = {}

    for index, row in enumerate(records):
        if not isinstance(row, dict):
            continue

        ra_number = normalize_text(get_llm_value(row, "RA_Number", "RA Number"))
        fmea_number = normalize_text(get_llm_value(row, "FMEA_Number", "FMEA Number"))

        # Requirement #2: skip any new record already tracked as an
        # Existing EDO (same RA id or FMEA id), so it isn't duplicated.
        ra_key = ra_number.upper()
        fmea_key = fmea_number.upper()
        if (ra_key and ra_key in existing_identifiers) or (fmea_key and fmea_key in existing_identifiers):
            logging.info(
                f"Skipping New EDO record RA={ra_number!r} FMEA={fmea_number!r} - "
                f"already tracked as an Existing EDO."
            )
            continue

        key = fmea_number or ra_number or f"NEW-EDO-RECORD-{index}"
        if key in merged:
            key = f"{key}_{index}"

        merged[key] = {
            "edo_type": "New",
            # Per requirement: Column A never prints the risk Status -
            # always the fixed literal tag text, correctly cased.
            "edo_tag": "EDO-XX\nNew",
            "edo_description": normalize_text(get_llm_value(row, "Product_Feature_Function", "Product Feature Function")),
            "reason_identified": normalize_text(get_llm_value(row, "Reason_Identified_as_EDO", "Reason Identified as EDO")),
            "dfmea": normalize_text(get_llm_value(row, "Traceability", "traceability")),
            "verification_reference": normalize_text(get_llm_value(row, "Verification_Reference", "Verification Reference")),
            "location": normalize_text(get_llm_value(row, "EDO_Location", "location")),
            "description_2": normalize_text(get_llm_value(row, "EDO_Description", "description")),
            "reason_2": normalize_text(get_llm_value(row, "Reason_Identified_as_EDO_ColH", "reason_2")),
            "sysdd": get_fixed_sysdd_reference(),
            "RA_Number": ra_number,
            "FMEA_Number": fmea_number
        }

    return merged


def merge_existing_edo_dictionary(existing_details):
    """
    CANONICAL version - from edo_existing_final.py.
    Richer than the copy-file version: carries edo_type, and correctly
    threads RA_Number/FMEA_Number through (needed since Stage 3B already
    populated ra_number/FMEA_Number on each existing EDO).

    NOTE: missing values default to "" (a truly empty cell), not the
    literal text "Blank". Column J (sysdd) is always the fixed reference
    from get_fixed_sysdd_reference(), per requirement.
    """
    final = {}
    for tag, data in existing_details.items():
        final[tag] = {
            "edo_type": "Existing",
            "edo_tag": data.get("edo_tag", tag),
            "edo_description": data.get("edo_description", ""),
            "reason_identified": data.get("reason_identified", ""),
            "dfmea": data.get("dfmea", ""),
            "location": data.get("location", ""),
            "description_2": data.get("description_2", ""),
            "reason_2": data.get("reason_2", ""),
            "sysdd": get_fixed_sysdd_reference(),
            "verification_reference": data.get("verification_reference", ""),
            "RA_Number": data.get("ra_number", ""),
            "FMEA_Number": data.get("FMEA_Number", "")
        }
    return final


def merge_existing_edo_dictionary_legacy(existing_details):
    """
    Original version from generate_EDO_template_copy.py - kept for
    backward compatibility. Does not carry edo_type/verification_reference
    and hardcodes FMEA_Number to "Blank". Not used by the merged pipeline.
    """
    final = {}
    for tag, data in existing_details.items():
        final[tag] = {
            "edo_tag": data.get("edo_tag", tag),
            "edo_description": data.get("edo_description", "Blank"),
            "reason_identified": data.get("reason_identified", "Blank"),
            "dfmea": data.get("dfmea", "Blank"),
            "location": data.get("location", "Blank"),
            "description_2": data.get("description_2", "Blank"),
            "reason_2": data.get("reason_2", "Blank"),
            "sysdd": data.get("sysdd", "Blank"),
            "RA_Number": data.get("ra_number", "Blank"),
            "FMEA_Number": "Blank"
        }
    return final


def merge_all_edos(existing_edos, new_edos):
    """
    From generate_EDO_template_copy.py - only pipeline that combines
    Existing + New EDO dictionaries into one, so it is kept as-is.
    """
    final_edos = {}

    for key, value in existing_edos.items():
        normalized = normalize_edo_record(value)
        normalized["edo_type"] = "Existing"
        final_edos[key] = normalized

    for key, value in new_edos.items():
        normalized = normalize_edo_record(value)
        normalized["edo_type"] = "New"

        final_key = key
        if final_key in final_edos:
            counter = 1
            while f"{key}_{counter}" in final_edos:
                counter += 1
            final_key = f"{key}_{counter}"

        final_edos[final_key] = normalized

    return final_edos


def validate_final_edos(final_edos):
    """
    CANONICAL version - from generate_EDO_template_copy.py.
    General-purpose: handles both "Existing" and "New" typed records
    (assigns a placeholder tag for blank New EDOs). Needed because the
    merged pipeline's final_edos dictionary contains both types.
    """
    validated = {}
    for key, value in final_edos.items():
        if not isinstance(value, dict):
            continue

        edo_tag = normalize_text(value.get("edo_tag"))
        if edo_tag == "" and value.get("edo_type") == "New":
            value["edo_tag"] = "EDO-XX\nNew EDO"

        normalized = normalize_edo_record(value)
        validated[key] = normalized
    return validated


def validate_final_edos_existing_only(final_edos):
    """
    Original version from edo_existing_final.py - kept for backward
    compatibility. Forces edo_type to "Existing" unconditionally, so it is
    only correct for the existing-EDO-only pipeline
    (generate_and_download_edo1). Not used by the combined pipeline.
    """
    validated = {}
    for key, value in final_edos.items():
        if not isinstance(value, dict):
            continue

        normalized = normalize_edo_record(value)
        normalized["edo_type"] = "Existing"
        validated[key] = normalized
    return validated


# ==========================================================
# STAGE 6: OUTPUT MAPPING AND FILE STORAGE
# ==========================================================

def write_edo_excel(sheet, final_edos, start_row):
    """
    MERGED canonical writer, used by the combined pipeline
    (generate_edo_template).

    Combines:
      - edo_existing_final.py's location-splitting logic (a single
        location cell like "(i) 210119 (metal cyl) (ii) 210120 (bracket)"
        is split into one output row per bracketed item), and Aptos
        Narrow/size-8 formatting.
      - generate_EDO_template_copy.py's per-edo_type Column D logic (New
        EDOs get a constructed "Document Traceability Matrix Info" block
        built from RA_Number/FMEA_Number/dfmea narrative; Existing EDOs
        just show the raw dfmea narrative).

    Writes columns A-J only (1-10).
    """
    logging.info("=" * 80)
    logging.info("STAGE 6: EXCEL OUTPUT FORMATTING RANGE (MERGED)")
    logging.info("=" * 80)

    current_row = start_row

    for key, edo in final_edos.items():
        logging.info(f"Writing row {current_row}: {key}")

        # 1. SPLIT LOGIC for the location field, e.g. turns
        #    "(i) 210119 (metal cyl) (ii) 210120 (bracket)" into
        #    ["210119 (metal cyl)", "210120 (bracket)"]
        raw_location = str(edo.get("location") or "Blank")
        matches = re.findall(r'\d+\s*\([^\)]+\)', raw_location)
        split_items = matches if matches else [raw_location]

        # 2. Column D (Traceability) value depends on edo_type
        if edo.get("edo_type") == "New":
            ra_id = edo.get("RA_Number", "") or "N/A"
            fmea_num = edo.get("FMEA_Number", "") or "N/A"
            trace_narrative = edo.get("dfmea", "")

            trace_value = f"Document Traceability Matrix Info:\nRA Number: {ra_id}\nFMEA Number: {fmea_num}"
            if trace_narrative and trace_narrative.lower() != "none" and trace_narrative.strip() != "":
                trace_value += f"\n\nDetails:\n{trace_narrative}"
        else:
            trace_value = edo.get("dfmea") or "Blank"

        # 3. Determine displayed tag
        if edo.get("edo_type") == "New":
            tag_value = edo.get("edo_tag") or "EDO-XX\nNew EDO"
        else:
            tag_value = edo.get("edo_tag") or key

        # 4. Iterate through the split location items - one output row each
        for i, item in enumerate(split_items):
            # Column A (1): Tag
            sheet.cell(row=current_row, column=1).value = tag_value

            # Column B (2): EDO Description (only shown on the first split row)
            sheet.cell(row=current_row, column=2).value = edo.get("edo_description") if i == 0 else ""

            # Column C (3): Reason Identified
            sheet.cell(row=current_row, column=3).value = edo.get("reason_identified") or "Blank"

            # Column D (4): Traceability (type-aware)
            sheet.cell(row=current_row, column=4).value = trace_value

            # Column E (5): Verification Reference
            sheet.cell(row=current_row, column=5).value = edo.get("verification_reference") or "Blank"

            # Column F (6): Clean Location (the split item)
            sheet.cell(row=current_row, column=7).value = item

            # Column G (7): Location counterpart / secondary field
            #sheet.cell(row=current_row, column=8).value = item

            # Column H (8): Secondary Description
            sheet.cell(row=current_row, column=8).value = edo.get("description_2") or "Blank"

            # Column I (9): Secondary Reason
            sheet.cell(row=current_row, column=9).value = edo.get("reason_2") or "Blank"

            # Column J (10): SYSDD / Design Reference
            sheet.cell(row=current_row, column=10).value = edo.get("sysdd") or "Blank"

            # 5. Apply formatting
            for col in range(1, 11):
                cell = sheet.cell(row=current_row, column=col)
                cell.font = aptos_font
                cell.alignment = cell_alignment
                cell.border = thin_border

            sheet.row_dimensions[current_row].height = 32
            current_row += 1

    return current_row


def write_edo_excel_existing_style(sheet, final_edos, start_row):
    """
    Original version from edo_existing_final.py - kept for backward
    compatibility (used by generate_and_download_edo1, the
    existing-EDO-only pipeline). Writes columns A-K (1-11).
    """
    logging.info("Formatting with Aptos Narrow, Size 8 and structured location splitting.")

    current_row = start_row

    for key, edo in final_edos.items():
        raw_data = str(edo.get("location") or "Blank")

        matches = re.findall(r'\d+\s*\([^\)]+\)', raw_data)
        split_items = matches if matches else [raw_data]

        for i, item in enumerate(split_items):
            sheet.cell(row=current_row, column=1).value = edo.get("edo_tag") or key
            sheet.cell(row=current_row, column=2).value = edo.get("edo_description") if i == 0 else ""
            sheet.cell(row=current_row, column=3).value = edo.get("reason_identified") or "Blank"
            sheet.cell(row=current_row, column=4).value = edo.get("dfmea") or "Blank"
            sheet.cell(row=current_row, column=5).value = edo.get("verification_reference") or "Blank"
            sheet.cell(row=current_row, column=7).value = item
            sheet.cell(row=current_row, column=8).value = edo.get("description_2") or "Blank"
            sheet.cell(row=current_row, column=9).value = edo.get("reason_2") or "Blank"
            sheet.cell(row=current_row, column=10).value = edo.get("sysdd") or "Blank"
            #sheet.cell(row=current_row, column=10).value = "N/A"
            #sheet.cell(row=current_row, column=11).value = "None"

            for col in range(1, 12):
                cell = sheet.cell(row=current_row, column=col)
                cell.font = aptos_font
                cell.alignment = cell_alignment
                cell.border = thin_border

            sheet.row_dimensions[current_row].height = 32
            current_row += 1

    return current_row


def write_edo_excel_new_style(sheet, final_edos, start_row):
    """
    Original version from generate_EDO_template_copy.py - kept for
    backward compatibility. Writes columns A-J (1-10), no location
    splitting, no font override (relies on template's existing font).
    """
    logging.info("=" * 80)
    logging.info("STAGE 6: EXCEL OUTPUT FORMATTING RANGE")
    logging.info("=" * 80)

    current_row = start_row

    for key, edo in final_edos.items():
        logging.info(f"Writing row {current_row}: {key}")

        if edo.get("edo_type") == "New":
            value = edo.get("edo_tag") or "EDO-XX\nNew EDO"
        else:
            value = edo.get("edo_tag") or key
        sheet.cell(row=current_row, column=1).value = value

        sheet.cell(row=current_row, column=2).value = edo.get("edo_description") or "Blank"
        sheet.cell(row=current_row, column=3).value = edo.get("reason_identified") or "Blank"

        if edo.get("edo_type") == "New":
            ra_id = edo.get("RA_Number", "") or "N/A"
            fmea_num = edo.get("FMEA_Number", "") or "N/A"
            trace_narrative = edo.get("dfmea", "")

            trace_value = f"Document Traceability Matrix Info:\nRA Number: {ra_id}\nFMEA Number: {fmea_num}"
            if trace_narrative and trace_narrative.lower() != "none" and trace_narrative.strip() != "":
                trace_value += f"\n\nDetails:\n{trace_narrative}"

            sheet.cell(row=current_row, column=4).value = trace_value
        else:
            sheet.cell(row=current_row, column=4).value = edo.get("dfmea") or "Blank"

        sheet.cell(row=current_row, column=5).value = edo.get("verification_reference") or "Blank"
        #sheet.cell(row=current_row, column=6).value = edo.get("location") or "Blank"
        sheet.cell(row=current_row, column=7).value = edo.get("location") or "Blank"
        sheet.cell(row=current_row, column=8).value = edo.get("description_2") or "Blank"
        sheet.cell(row=current_row, column=9).value = edo.get("reason_2") or "Blank"
        sheet.cell(row=current_row, column=10).value = edo.get("sysdd") or "Blank"

        for col in range(1, 11):
            cell = sheet.cell(row=current_row, column=col)
            cell.alignment = cell_alignment
            cell.border = thin_border

        sheet.row_dimensions[current_row].height = 32
        current_row += 1

    return current_row


def _style_value(value, base_bold, is_new):
    """
    Shared column-coloring rule used by every plain (non rich-text) cell:
      - New EDO row      -> always RED, keeping the column's normal bold flag.
      - Existing EDO row -> BLACK, normal bold flag (an empty cell has no
        text to color, so this only matters when there IS a value).
    """
    if is_new:
        return RED, base_bold

    return BLACK, base_bold



def classify_risk_status(description_text, pipeline_config):
    """
    Risk classification using LLM based on the provided Risk Classification table.
    Compares/evaluates the description_text against the risk term guidelines
    below (same call_llm(prompt, pipeline_config) pattern as
    evaluate_comparison()) and returns High / Medium / Low / None.
    """

    print(f"description text", description_text)

    text = normalize_text(description_text).lower()

    print(f"clasification f risk text", text)

    if not text:
        print(f"risk status", "None")
        return "None"

    prompt = f"""
You are a Risk Classification expert.

Classify the following description into exactly one of these categories:
- High
- Medium
- Low
- None

Risk Classification Guidelines:

High:
- Potential impact to patient safety
- Device performance issues
- Not meeting SOA requirements
- Gaps in EDO lists
- Non-existent EDO
- Missing required V&V
- Missing required risk documents
- Satisfies FA / Meets FA requirements
- Any issue that can significantly affect safety, regulatory compliance, or product functionality

Medium:
- Documentation gaps
- Missing or incomplete documents
- Tracing errors in RTMs
- RAS
- DID
- Updates to existing V&V
- Risk document updates
- Compliance issues
- Moderate documentation or traceability problems

Low:
- Template updates
- Drawing updates or creation
- Minor document updates
- Clarifications
- Typographical corrections
- Missing component qualifications
- Missing control plans
- PCS
- MVP
- Cosmetic or administrative changes with no impact to safety or functionality

None:
- Description does not match any of the above categories.

Description:
{text}

Return ONLY one word:
High
Medium
Low
None
"""

    risk_status = "None"

    try:
        response = call_llm(prompt, pipeline_config)
        response = clean_response(response)

        normalized = str(response).strip().lower()

        if normalized.startswith("high"):
            risk_status = "High"
        elif normalized.startswith("medium"):
            risk_status = "Medium"
        elif normalized.startswith("low"):
            risk_status = "Low"
        else:
            risk_status = "None"

    except Exception as e:
        logging.error(f"Risk classification LLM failed: {e}")
        risk_status = "None"

    print(f"risk status", risk_status)

    return risk_status


def apply_risk_cell_style(cell, risk):
    from openpyxl.styles import PatternFill, Font

    fills = {
        "High": "FF0000",
        "Medium": "FFFF00",
        "Low": "00B050"
    }

    if risk in fills:
        cell.fill = PatternFill("solid", fgColor=fills[risk])
        cell.font = Font(bold=True, color="000000", name="Calibri", size=10)
    else:
        cell.fill = PatternFill(fill_type=None)

def generate_remarks_and_recommendation(edo, tag_value, is_new, risk, pipeline_config):
    """
    Builds Column M ("Remarks and Recommendation") content:
      1. Boilerplate Gap + Verification Status (New EDO vs Existing EDO wording)
      2. Optional "Observation" block - flagged when the FMEA/RA&C trace text
         (Column D) states a DIFFERENT risk level than the assigned Risk
         Classification (Column L)
      3. Optional "Recommendation"/"Design" block - feature-specific safety
         mitigation text (manual warning, drawing/design note, or training),
         generated from the EDO's feature/reason fields when the hazard
         warrants it
    """

    # ---- 1. Base boilerplate (Gap + Verification Status) ----
    if is_new:
        base_text = (
            "Gap:\n"
            "As identified in the Risk Assessment & Control (RA&C) and System "
            "DFMEA, this risk impacts the product’s functions and features. "
            "Therefore, it is classified as a new Essential Design Output "
            "(EDO) and must be incorporated into the existing EDO list.\n\n"
            "Verification Status:\n"
            f"Design verification has been conducted for this EDO ({tag_value}), "
            "and the corresponding reports are traced in the Verification "
            "Reference (Column E)."
        )
    else:
        base_text = (
            "Gap:\n"
            "The design verification reference corresponding to where the EDO "
            "is controlled has not been included in the existing EDO list.\n\n"
            "Verification Status:\n"
            f"Design verification has been conducted for this EDO ({tag_value}), "
            "and the corresponding reports are traced in the Verification "
            "Reference (Column E)."
        )

    # ---- 2. Observation block (risk-trace mismatch check) ----
    observation_text = ""
    dfmea_text = format_output_text(edo.get("dfmea"))

    if dfmea_text and risk:
        obs_prompt = f"""
        You are reviewing an Essential Design Output (EDO) risk record.

        Risk Classification assigned (Column L): {risk}
        System FMEA / RA&C trace text (Column D): {dfmea_text}

        Does the FMEA/RA&C trace text explicitly state a DIFFERENT risk
        evaluation (e.g. 'Low', 'Medium', 'High') than the assigned Risk
        Classification above?

        If yes, reply with EXACTLY this sentence, filling in the FMEA's
        stated level and the assigned classification (lowercase):
        Observation: In Sys-FMEA, the risk evaluation is '<fmea_level>'. recommended to change in the sys-FMEA as <assigned_level_lowercase>.

        If no mismatch is found, reply with exactly: NONE
        """

        try:
            obs_response = call_llm(obs_prompt, pipeline_config)
            obs_response = clean_response(obs_response)
            if obs_response and obs_response.strip().upper() != "NONE":
                observation_text = obs_response.strip()
        except Exception as e:
            logging.error(f"Observation generation failed: {e}")

    # ---- 3. Recommendation / Design block (feature-specific safety mitigation) ----
    recommendation_text = ""
    feature_text = format_output_text(edo.get("edo_description") or edo.get("description_2"))
    reason_text = format_output_text(edo.get("reason_identified") or edo.get("reason_2"))

    if feature_text or reason_text:
        rec_prompt = f"""
        You are drafting a "Recommendation" note for an Essential Design
        Output (EDO) risk record, matching this house style:

        - If the hazard needs a User Manual warning/precaution, write it
          starting with "Recommendation:" (or "Recommendation to add in the
          User Manual:") followed by concrete warning text.
        - If the hazard needs a drawing/design change (e.g. a load rating,
          dimension note, or EDO symbol placement), write it starting with
          "Design:" followed by the specific design note.
        - If training is the appropriate mitigation, write it starting with
          "Recommendation:" followed by the training instruction.
        - If NONE of the above genuinely apply (the base Gap/Verification
          Status boilerplate is already sufficient), reply with exactly:
          NONE

        Product Feature/Function: {feature_text}
        Reason Identified as EDO: {reason_text}

        Only return the recommendation block itself (or NONE) - do not
        repeat the Gap or Verification Status text.
        """

        try:
            rec_response = call_llm(rec_prompt, pipeline_config)
            rec_response = clean_response(rec_response)
            if rec_response and rec_response.strip().upper() != "NONE":
                recommendation_text = rec_response.strip()
        except Exception as e:
            logging.error(f"Recommendation generation failed: {e}")

    # ---- Assemble final Column M text ----
    parts = [base_text]
    if observation_text:
        parts.append(observation_text)
    if recommendation_text:
        parts.append(recommendation_text)

    return "\n\n".join(parts)

def merge_existing_edo_rows(sheet, start_row, end_row):
    """
    Merge A:E for repeated Existing EDO tags only.
    Each EDO tag group is merged independently.
    """
    current = start_row

    while current <= end_row:
        tag = sheet.cell(current, 1).value

        if not tag:
            current += 1
            continue

        last = current
        while last + 1 <= end_row and sheet.cell(last + 1, 1).value == tag:
            last += 1

        if last > current:
            for col in range(1, 6):
                sheet.merge_cells(
                    start_row=current,
                    start_column=col,
                    end_row=last,
                    end_column=col
                )
                sheet.cell(current, col).alignment = cell_alignment

        current = last + 1


def format_edo_worksheet(sheet, final_edos, start_row, pipeline_config):
    """
    Final writer:
    A-E : existing columns
    F-I : new EDO fields
    K   : Risk Classification
    L   : Risk evaluation text / classification trigger
    M   : Gap and Verification Status statement
    """

    current_row = start_row
    existing_ranges = []

    for key, edo in final_edos.items():

        is_new = edo.get("edo_type") == "New"

        if is_new:
            tag_value = "EDO-XX\nNew"
        else:
            tag_value = format_edo_tag_text(edo.get("edo_tag") or key)

        raw_location = format_output_text(edo.get("location"))
        locations = re.findall(r'\d+\s*\([^\)]+\)', raw_location) or [raw_location]

        first = current_row

        for idx, location in enumerate(locations):

            values = {
                1: tag_value,
                2: format_output_text(edo.get("edo_description")) if idx == 0 else "",
                3: format_output_text(edo.get("reason_identified")),
                4: format_output_text(edo.get("dfmea")),
                5: format_output_text(edo.get("verification_reference")),
                7: location,
                8: format_output_text(edo.get("description_2")),
                9: format_output_text(edo.get("reason_2")),
                10: get_fixed_sysdd_reference(),
                11: "None",
            }

            # --- Risk Classification (Column L) computed first, since
            # --- Column M's content depends on it ---
            risk_input = (
                f"{format_output_text(edo.get('reason_identified') or '')} "
                f"{format_output_text(edo.get('dfmea') or '')}"
            )
            risk = classify_risk_status(risk_input, pipeline_config)
            values[12] = risk

            # --- Remarks and Recommendation (Column M) ---
            values[13] = generate_remarks_and_recommendation(
                edo, tag_value, is_new, risk, pipeline_config
            )

            for col, value in values.items():
                cell = sheet.cell(current_row, col)
                cell.value = value
                _apply_border_alignment(cell)

            apply_risk_cell_style(sheet.cell(current_row, 12), risk)

            current_row += 1

        if not is_new:
            existing_ranges.append((first, current_row - 1))

    for first, last in existing_ranges:
        if last > first:
            for col in range(1, 6):
                sheet.merge_cells(
                    start_row=first,
                    start_column=col,
                    end_row=last,
                    end_column=col
                )
                sheet.cell(first, col).alignment = cell_alignment
            for col in (12, 13):
                sheet.merge_cells(
                    start_row=first,
                    start_column=col,
                    end_row=last,
                    end_column=col
                )
                sheet.cell(first, col).alignment = cell_alignment

    return current_row

def clear_existing_rows(sheet, start_row, end_column=10):
    row = start_row
    while row <= sheet.max_row:
        empty = True
        for col in range(1, end_column + 1):
            if sheet.cell(row=row, column=col).value:
                empty = False
                break
        if empty:
            break
        for col in range(1, end_column + 1):
            sheet.cell(row=row, column=col).value = None
        row += 1


def save_edo_workbook(workbook, pipeline_config):
    output_path = pipeline_config["output_file_path"]
    workbook.save(output_path)
    return output_path


# ==========================================================
# MAIN EXECUTION PIPELINE #1 (EXISTING EDO ONLY)
# Preserved as-is from edo_existing_final.py, using the "_existing_style"
# / "_existing_only" variants so its behaviour is unchanged.
# ==========================================================

def generate_and_download_edo1(
    client,
    product_family,
    product,
    templatename,
    pipeline_config,
    db: DatabaseHandler
):
    logging.info("=" * 80)
    logging.info("STARTING EXISTING-EDO TEMPLATE GENERATION PIPELINE")
    logging.info("=" * 80)

    try:
        workbook, sheet = initialize_workbook(pipeline_config)
        start_row = pipeline_config.get("templatestartrow", 4)

        clear_existing_rows(sheet, start_row, end_column=10)

        edo_document = get_edo_document(
            client,
            product_family,
            product,
            templatename,
            db
        )

        existing_edos = extract_edo_tags(
            client,
            product_family,
            product,
            templatename,
            pipeline_config,
            edo_document,
            db
        )
        existing_edos = validate_existing_tags(existing_edos)

        if existing_edos:
            existing_edos = extract_edo_details(
                client,
                product_family,
                product,
                templatename,
                pipeline_config,
                edo_document,
                existing_edos,
                db
            )

            try:
                existing_verification_details = extract_existing_edo_verification_details(
                    client,
                    product_family,
                    product,
                    templatename,
                    pipeline_config,
                    edo_document,
                    existing_edos,
                    db
                )

                existing_edos = apply_existing_edo_verification(
                    existing_edos,
                    existing_verification_details
                )

            except Exception as verification_error:
                logging.warning(
                    "STAGE 3B SKIPPED - verification reference extraction "
                    "failed, leaving column E as 'Blank' for this run. "
                    f"Reason: {verification_error}"
                )

        existing_edo_final = merge_existing_edo_dictionary(existing_edos)
        final_edos = validate_final_edos_existing_only(existing_edo_final)

        if not final_edos:
            raise Exception("No EDO records generated.")

        write_edo_excel_existing_style(
            sheet,
            final_edos,
            start_row
        )

        output_file = save_edo_workbook(
            workbook,
            pipeline_config
        )

        logging.info("=" * 80)
        logging.info("EDO PIPELINE COMPLETED SUCCESSFULLY")
        logging.info(f"OUTPUT FILE PERSISTED AT : {output_file}")
        logging.info("=" * 80)

        return output_file

    except Exception as e:
        logging.error("=" * 80)
        logging.error(f"EDO PIPELINE CRITICAL RUNTIME FAILURE : {str(e)}")
        logging.error("=" * 80)
        raise e


# ==========================================================
# MAIN EXECUTION PIPELINE #2 (EXISTING + NEW EDO - COMBINED / CANONICAL)
# This is the primary merged pipeline: Stage 3 + Stage 3B from
# edo_existing_final.py, Stage 4 from generate_EDO_template_copy.py, and
# Stage 5/6 reconciled to carry both EDO types through to one worksheet.
# ==========================================================

def generate_edo_template(
    client,
    product_family,
    product,
    templatename,
    pipeline_config,
    db: DatabaseHandler
):
    logging.info("=" * 80)
    logging.info("STARTING COMPLETE EDO TEMPLATE GENERATION PIPELINE (MERGED)")
    logging.info("=" * 80)

    try:
        workbook, sheet = initialize_workbook(pipeline_config)
        start_row = pipeline_config.get("templatestartrow", 4)

        # Clear active table grid space exclusively up to Column J
        clear_existing_rows(sheet, start_row, end_column=10)

        edo_document = get_edo_document(
            client,
            product_family,
            product,
            templatename,
            db
        )

        # ---------------------------------------------------
        # STAGE 3: Workflow 1 - Processing Existing EDO Records
        # ---------------------------------------------------
        existing_edos = extract_edo_tags(
            client,
            product_family,
            product,
            templatename,
            pipeline_config,
            edo_document,
            db
        )
        existing_edos = validate_existing_tags(existing_edos)

        if existing_edos:
            existing_edos = extract_edo_details(
                client,
                product_family,
                product,
                templatename,
                pipeline_config,
                edo_document,
                existing_edos,
                db
            )

            # ---------------------------------------------------
            # STAGE 3B: Verification Reference extraction for Existing EDOs
            # ---------------------------------------------------
            try:
                existing_verification_details = extract_existing_edo_verification_details(
                    client,
                    product_family,
                    product,
                    templatename,
                    pipeline_config,
                    edo_document,
                    existing_edos,
                    db
                )

                existing_edos = apply_existing_edo_verification(
                    existing_edos,
                    existing_verification_details
                )

            except Exception as verification_error:
                logging.warning(
                    "STAGE 3B SKIPPED - verification reference extraction "
                    "failed, leaving column E as 'Blank' for this run. "
                    f"Reason: {verification_error}"
                )

        # ---------------------------------------------------
        # STAGE 4: Workflow 2 - Processing New EDO Records
        # ---------------------------------------------------
        edo_tags = extract_new_edo_tags(
            client,
            product_family,
            product,
            templatename,
            pipeline_config,
            edo_document,
            db
        )
        print(f"extract_new_edo_tags: ",edo_tags)
        new_records = extract_new_edo_summary_details(
            client,
            product_family,
            product,
            templatename,
            pipeline_config,
            edo_document,
            edo_tags,
            db
        )
        print(f"extract_new_edo_summary_details: ",new_records)
        # ---------------------------------------------------
        # STAGE 5: New EDO Merge (de-duplicated against Existing EDOs)
        # ---------------------------------------------------
        new_final = merge_new_edo_records(
            new_records,
            existing_edos
        )

        existing_edo_final = merge_existing_edo_dictionary(existing_edos)
        final_edos = merge_all_edos(existing_edo_final, new_final)
        final_edos = validate_final_edos(final_edos)

        if not final_edos:
            raise Exception("No EDO records generated.")

        # ---------------------------------------------------
        # STAGE 6: Output Mapping, Formatting and File Storage
        # ---------------------------------------------------
        format_edo_worksheet(
            sheet,
            final_edos,
            start_row,
            pipeline_config
        )

        output_file = save_edo_workbook(
            workbook,
            pipeline_config
        )

        logging.info("=" * 80)
        logging.info("EDO PIPELINE COMPLETED SUCCESSFULLY")
        logging.info(f"OUTPUT FILE PERSISTED AT : {output_file}")
        logging.info("=" * 80)

        return output_file

    except Exception as e:
        logging.error("=" * 80)
        logging.error(f"EDO PIPELINE CRITICAL RUNTIME FAILURE : {str(e)}")
        logging.error("=" * 80)
        raise e