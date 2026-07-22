"""
================================================================================
MERGED EDO TEMPLATE GENERATION PIPELINE
================================================================================
This file merges:
  - edo_existing_final.py            (EXISTING-EDO-ONLY pipeline)
  - generate_EDO_template_copy.py    (EXISTING + NEW EDO combined pipeline)

Per requirement, NO function from either source file has been deleted.
Where both files defined a function with the SAME NAME but a DIFFERENT BODY,
both versions are kept:
    - the version chosen to run in the merged pipeline keeps its original
      name (the "canonical" version), and
    - the other version is kept under a suffixed name (e.g. "_legacy",
      "_basic", "_lenient", "_minimal", "_existing_style", "_new_style",
      "_existing_only") purely for reference / backward compatibility.

Canonical selection for the 6 functions originally requested by the user:
    1. extract_edo_tags                          -> from edo_existing_final.py
    2. extract_existing_edo_verification_details  -> from edo_existing_final.py
    3. apply_existing_edo_verification            -> from edo_existing_final.py
    4. extract_new_edo_tags                       -> restored; builds edo_new_data
       (see "NEW-EDO WORKFLOW RESTRUCTURE" below)
    5. extract_new_edo_summary_details            -> enriches edo_new_data,
       one LLM call per RA/FMEA pair, looped (see restructure notes below)
    6. extract_new_edo_traceability_details        -> enriches edo_new_data,
       same looped, one-call-per-pair pattern (see restructure notes below)

Everything else (normalize_text, blank, deep_extract_records, merge
functions, write_edo_excel, etc.) has been merged / reconciled so that the
combined pipeline (generate_edo_template) can run BOTH the existing-EDO
workflow and the new-EDO workflow end to end, exactly the way
generate_EDO_template_copy.py's main pipeline did - but now using the
richer Stage 3 / Stage 3B logic that only existed in edo_existing_final.py.

The original standalone "existing-EDO-only" pipeline
(generate_and_download_edo1) is also preserved and still works on its own.

ADDITIONALLY MERGED: the dedicated Excel rich-text formatting module
(edo_excel_formatting.py) is folded in as format_edo_worksheet() - the
Stage 6 writer now used by generate_edo_template(). It reproduces the
reference-screenshot styling (black-bold tag for Existing EDOs, red-bold
tag + red-bold "None" fields for New EDOs, per-line red/black rich text
for the RA&C Trace and Verification Reference columns). The older plain
writers (write_edo_excel, write_edo_excel_existing_style,
write_edo_excel_new_style) are kept, unused, for backward compatibility -
nothing was deleted.

================================================================================
NEW-EDO WORKFLOW RESTRUCTURE (latest revision)
================================================================================
Per explicit requirement, the New-EDO workflow (Stage 4) now runs as THREE
steps that all read/write ONE shared dictionary - edo_new_data - keyed by
RA id (RA_Number, falling back to FMEA_Number or a positional key):

  4a. extract_new_edo_tags()
      Scans Section 10 of the FMEA against the RA&C, pulls out every
      qualifying {RA_Number, FMEA_Number, Status} row, and CREATES
      edo_new_data from them - one dictionary entry per RA id.

  4b. extract_new_edo_summary_details(edo_new_data)
      Iterates every entry already in edo_new_data and, inside that loop,
      calls the LLM ONCE PER RA/FMEA pair (the "EDO_NEW_details" prompt,
      scoped to that single pair) to pull its full detail record
      (Product_Feature_Function, Reason_Identified_as_EDO, Traceability,
      Verification_Reference, EDO_Location, EDO_Description,
      Reason_Identified_as_EDO_ColH) and writes it straight back into
      that SAME edo_new_data[key] entry.

  4c. extract_new_edo_traceability_details(edo_new_data)
      Identical loop-per-RA/FMEA-pair pattern to 4b, using the
      "EDO_NEW_Verification_details" prompt to fetch traceability /
      verification info (spreadsheets, reports, locations, pass/fail
      results), again enriching the SAME edo_new_data[key] entry.

By the end of Stage 4, edo_new_data holds every RA id's tag info plus its
full summary and traceability details in one place. generate_edo_template()
converts edo_new_data.values() to a list before handing it to Stage 5.

extract_new_edo_ra_details() (the Medium-risk RA&C backfill) is kept,
defined, but is not part of this three-step chain - it remains available
for backward compatibility under its own older edo_tags-list contract.

merge_new_edo_records() is the canonical Stage 5 merge for New EDOs -
it maps directly off the ACTUAL JSON keys the merged prompt returns
(Product_Feature_Function, Reason_Identified_as_EDO, Traceability,
Verification_Reference, EDO_Location, EDO_Description,
Reason_Identified_as_EDO_ColH), fixing a prior "column mapped wrongly" bug
where the merge logic looked for keys that didn't actually exist in the
LLM output. It also de-duplicates: any New EDO record whose RA Number or
FMEA Number already belongs to an Existing EDO is dropped, so the same
issue is never printed twice. merge_new_edo_dictionary_full() is kept,
unused, for backward compatibility.

format_edo_worksheet() now also runs every descriptive/narrative cell
through format_output_text() (capitalizes the first letter of every
sentence; uppercases "npdxxxx"/"edo-xx" document codes to
"NPDxxxx"/"EDO-XX"), and the New EDO tag is now the fixed literal
"EDO-XX\nNew" (previously "EDO-XX\nNew EDO").
================================================================================
"""

import os
import re
import json
import time
import logging
import sys
import unicodedata
import zipfile
import tempfile
import hashlib
import openpyxl
import fitz

from openpyxl.styles import (
    Alignment,
    Border,
    Side,
    Font
)
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU

from Files.database import DatabaseHandler
from retrieval.retrieve_content_prompt import retrieve_content_for_prompt


# ==========================================================
# GLOBALS
# ==========================================================

CURRENT_FOLDER = os.path.dirname(os.path.abspath(__file__))

MAX_LLM_RETRIES = 3
INITIAL_RETRY_DELAY = 1

# ---- Image pipeline (from edo_image.py) ----
IMAGE_WIDTH = 180
IMAGE_HEIGHT = 110

# ---- PDF image extraction (ported from generate_ICU_Template_final_working_code.py) ----
PDF_GLOBAL_OVERRIDE = os.path.join(CURRENT_FOLDER, "80369-6.pdf")
IMAGE_OUTPUT_DIR = "extracted_images"
IMAGE_TEXT_OFFSET_PX = 45


# ==========================================================
# ==========================================================
# IMAGE EXTRACTION PIPELINE (from edo_image.py)
# ==========================================================
# Replaces the previous db-lookup-based image resolution
# (resolve_edo_document_file / CANDIDATE_DB_FILE_METHODS /
# CANDIDATE_DOCUMENT_PATH_KEYS) with edo_image.py's simpler, more direct
# approach: get_edo_proposed_file() reads the local file path straight
# off the edo_proposed document metadata (or falls back to
# find_file_by_name() searching the working directory), and
# extract_docx_images() now preserves the images' in-document order
# (via word/document.xml's embed sequence) and de-duplicates identical
# images by content hash. The previous db-based versions are kept below,
# unused, under "_legacy" names for backward compatibility - nothing was
# deleted.
#
# NOTE: the legacy .doc -> .docx conversion step (convert_doc_to_docx(),
# which shelled out to LibreOffice) has been removed - source documents
# here are always already .docx, so it was dead weight (and a Docker
# dependency on the "libreoffice" binary that was never actually needed).
# ==========================================================

def extract_docx_images(docx_file):
    """
    CANONICAL version - from edo_image.py. Reads word/document.xml's
    embed-id sequence (falling back to a plain sorted-namelist scan if
    that XML can't be parsed) so images come back in the same order they
    appear in the document, and de-duplicates identical images by
    content hash. See extract_docx_images_legacy() below for the
    original simpler version (kept for backward compatibility, not used
    by the merged pipeline).
    """
    logging.info("READING WORD MEDIA IMAGES")
    images_by_id = {}
    sequence_order = []
    with zipfile.ZipFile(docx_file, "r") as archive:
        try:
            doc_xml = archive.read("word/document.xml").decode("utf-8")
            embed_ids = re.findall(r'embed="([^"]+)"', doc_xml) or re.findall(r'r:id="([^"]+)"', doc_xml)
            rel_xml = archive.read("word/_rels/document.xml.rels").decode("utf-8")
            rel_map = {}
            for match in re.finditer(r'Id="([^"]+)"\s+Type="[^"]+/image"\s+Target="([^"]+)"', rel_xml):
                rel_id, target = match.groups()
                rel_map[rel_id] = os.path.basename(target)
            for r_id in embed_ids:
                if r_id in rel_map:
                    sequence_order.append(rel_map[r_id])
        except Exception as e:
            logging.warning(f"XML layout sequence failed, falling back to sort: {e}")

        for file in sorted(archive.namelist()):
            if file.startswith("word/media/"):
                img_name = os.path.basename(file)
                images_by_id[img_name] = {
                    "name": img_name,
                    "extension": os.path.splitext(file)[1],
                    "bytes": archive.read(file)
                }
                logging.info(f"IMAGE FOUND : {file}")

    ordered_images = []
    seen = set()
    for name in sequence_order:
        if name in images_by_id and name not in seen:
            ordered_images.append(images_by_id[name])
            seen.add(name)
    for name, img_data in images_by_id.items():
        if name not in seen:
            ordered_images.append(img_data)

    # NOTE: previously de-duplicated images by content hash here, which
    # silently dropped any image whose bytes were identical to one
    # already seen - including legitimate repeat placements of the same
    # figure/diagram at more than one point in the document. Per
    # requirement, ALL images extracted from word/media/ are returned,
    # so nothing embedded in the source document is ever discarded.
    logging.info(f"TOTAL IMAGES : {len(ordered_images)}")
    return ordered_images


def extract_docx_images_legacy(docx_file):
    """
    Original version - kept for backward compatibility. Does not
    preserve in-document order or de-duplicate. Not used by the merged
    pipeline.
    """
    logging.info("READING WORD MEDIA IMAGES")
    images = []
    with zipfile.ZipFile(docx_file, "r") as archive:
        for file in archive.namelist():
            if file.startswith("word/media/"):
                images.append({
                    "name": os.path.basename(file),
                    "extension": os.path.splitext(file)[1],
                    "bytes": archive.read(file)
                })
                logging.info(f"IMAGE FOUND : {file}")
    logging.info(f"TOTAL IMAGES : {len(images)}")
    return images


# ==========================================================
# FIGURE-ONLY IMAGE EXTRACTION (captions, duplicates, exclusions)
# ==========================================================
# CANONICAL image extractor - supersedes extract_docx_images() above for
# the purposes of extract_edo_proposed_images(). Per requirement:
#
#   1. Duplicates are ALLOWED and ALL are retrieved. extract_docx_images()
#      above still silently drops repeat placements of the same media
#      file via its `seen` set (each embed_id occurrence is walked here
#      independently instead, so a figure placed twice in the document
#      comes back as two separate entries).
#
#   2. Only genuine FIGURES are extracted - a docx image is kept only if
#      the paragraph containing it (or the paragraph immediately before/
#      after it) has caption text starting with "Figure N". Anything
#      captioned "Table N" is excluded, anything with no figure caption
#      at all is excluded (treated as decorative/text image, not
#      content), and anything whose nearby text contains "Hillrom" is
#      excluded outright (the company logo/letterhead image).
# ==========================================================

CAPTION_FIGURE_PATTERN = re.compile(r'as\s+below\s*:', re.IGNORECASE)
CAPTION_TABLE_PATTERN = re.compile(r'^\s*table\s+\d+', re.IGNORECASE)

# Case-insensitive substrings that mark an image as a logo/letterhead to
# always exclude, regardless of any figure caption nearby. Add more
# brand/watermark keywords here if other logos need excluding too.
EXCLUDED_LOGO_KEYWORDS = ["hillrom"]


def _extract_paragraph_texts_and_images(doc_xml):
    """
    Splits word/document.xml into paragraphs (<w:p>...</w:p> blocks),
    extracting each paragraph's plain text and any image relationship
    IDs (r:embed / r:id) referenced within it, in document order.
    Returns a list of {"text": str, "embed_ids": [str, ...]} dicts.
    """
    paragraphs = re.findall(r'<w:p[ >].*?</w:p>', doc_xml, flags=re.DOTALL)
    result = []
    for p in paragraphs:
        texts = re.findall(r'<w:t[^>]*>(.*?)</w:t>', p, flags=re.DOTALL)
        text = "".join(texts).strip()
        embed_ids = re.findall(r'embed="([^"]+)"', p) or re.findall(r'r:id="([^"]+)"', p)
        result.append({"text": text, "embed_ids": embed_ids})
    return result


def _resolve_caption_for_image(paragraphs, para_index):
    """
    Looks at the paragraph containing the image, then the paragraph
    immediately before it, then immediately after it, and returns the
    first one with any text - that's treated as the image's caption/
    nearby-text context.
    """
    candidates = []
    if 0 <= para_index < len(paragraphs):
        candidates.append(paragraphs[para_index]["text"])
    if para_index - 1 >= 0:
        candidates.append(paragraphs[para_index - 1]["text"])
    if para_index + 1 < len(paragraphs):
        candidates.append(paragraphs[para_index + 1]["text"])

    for text in candidates:
        if text:
            return text
    return ""


def is_figure_caption(caption_text):
    return bool(CAPTION_FIGURE_PATTERN.search(caption_text or ""))


def is_table_caption(caption_text):
    return bool(CAPTION_TABLE_PATTERN.search(caption_text or ""))


def is_excluded_logo(caption_text):
    lowered = (caption_text or "").lower()
    return any(keyword in lowered for keyword in EXCLUDED_LOGO_KEYWORDS)


def extract_docx_figures_only(docx_file):
    """
    CANONICAL image extractor used by extract_edo_proposed_images().
    See the module-level "FIGURE-ONLY IMAGE EXTRACTION" comment above for
    the full rule set. Every image placement is evaluated independently
    (duplicates allowed/retrieved), and only those with a "Figure N"
    caption nearby are kept - "Table N" captions, uncaptioned images, and
    anything mentioning "Hillrom" are all excluded.
    """
    logging.info("READING WORD MEDIA IMAGES (FIGURES ONLY)")

    with zipfile.ZipFile(docx_file, "r") as archive:
        try:
            doc_xml = archive.read("word/document.xml").decode("utf-8")
        except KeyError:
            logging.warning("word/document.xml not found - cannot extract figures.")
            return []

        rel_map = {}
        try:
            rel_xml = archive.read("word/_rels/document.xml.rels").decode("utf-8")
            for match in re.finditer(r'Id="([^"]+)"\s+Type="[^"]+/image"\s+Target="([^"]+)"', rel_xml):
                rel_id, target = match.groups()
                rel_map[rel_id] = os.path.basename(target)
        except KeyError:
            logging.warning("word/_rels/document.xml.rels not found - cannot resolve image relationships.")
            return []

        media_bytes = {}
        for file in archive.namelist():
            if file.startswith("word/media/"):
                media_bytes[os.path.basename(file)] = {
                    "extension": os.path.splitext(file)[1],
                    "bytes": archive.read(file)
                }

        paragraphs = _extract_paragraph_texts_and_images(doc_xml)

        figures = []
        skipped_table = 0
        skipped_logo = 0
        skipped_uncaptioned = 0

        for para_index, para in enumerate(paragraphs):
            for embed_id in para["embed_ids"]:
                filename = rel_map.get(embed_id)
                if not filename or filename not in media_bytes:
                    continue

                caption = _resolve_caption_for_image(paragraphs, para_index)

                if is_excluded_logo(caption):
                    skipped_logo += 1
                    logging.info(f"SKIPPED (Hillrom logo/letterhead) : {filename}")
                    continue

                if is_table_caption(caption):
                    skipped_table += 1
                    logging.info(f"SKIPPED (Table image, not a Figure) : {filename} - caption: {caption!r}")
                    continue

                if not is_figure_caption(caption):
                    skipped_uncaptioned += 1
                    logging.info(f"SKIPPED (no Figure caption found nearby) : {filename} - nearby text: {caption!r}")
                    continue

                media = media_bytes[filename]
                figures.append({
                    "name": filename,
                    "extension": media["extension"],
                    "bytes": media["bytes"],
                    "caption": caption
                })
                logging.info(f"FIGURE FOUND : {filename} - caption: {caption!r}")

    logging.info(
        f"TOTAL FIGURES EXTRACTED : {len(figures)}  "
        f"(skipped {skipped_table} table image(s), {skipped_logo} Hillrom logo/letterhead "
        f"image(s), {skipped_uncaptioned} uncaptioned image(s))"
    )
    return figures


def extract_images_from_file(file_path):
    if file_path.lower().endswith(".docx"):
        return extract_docx_images(file_path)

    raise ValueError(f"Unsupported file type: {file_path}")


def save_images_to_folder(images, output_folder):
    os.makedirs(output_folder, exist_ok=True)
    saved_paths = []
    for img in images:
        ext = img["extension"] if img["extension"] else ".png"
        out_path = os.path.join(output_folder, img["name"])
        # avoid collisions if names repeat
        base, counter = out_path, 1
        while os.path.exists(out_path):
            root, e = os.path.splitext(base)
            out_path = f"{root}_{counter}{e}"
            counter += 1
        with open(out_path, "wb") as f:
            f.write(img["bytes"])
        saved_paths.append(out_path)
        logging.info(f"Saved : {out_path}")
    return saved_paths


# ==========================================================
# DOCUMENT FILE RESOLUTION (edo_proposed -> local .docx path)
# ==========================================================
# CANONICAL version - from edo_image.py. Reads the local file path
# straight off the edo_proposed document metadata dict (checking a
# handful of common key names), falling back to a working-directory
# search by document name via find_file_by_name(). This replaces the
# previous db-method-guessing approach (resolve_edo_document_file() /
# CANDIDATE_DB_FILE_METHODS / CANDIDATE_DOCUMENT_PATH_KEYS), which is
# kept below, unused, for backward compatibility.
# ==========================================================

def find_file_by_name(filename, search_dir="."):
    if not search_dir or not os.path.isdir(search_dir):
        return None
    for root, _, files in os.walk(search_dir):
        if filename in files:
            return os.path.join(root, filename)
    return None


# ==========================================================
# DOCUMENT SEARCH ROOTS (Docker fix)
# ==========================================================
# find_file_by_name() used to default to search_dir="." - the process's
# current working directory. That is almost never meaningful inside a
# Docker container: "." is just whatever WORKDIR the image sets, and
# nothing lands there unless it was explicitly COPY'd into the image or
# mounted as a volume. If the source .docx files live on a mounted
# volume, network share, or anywhere else in the container's filesystem,
# "." alone will silently never find them - get_edo_proposed_file()
# would return None and image extraction is skipped with no error.
#
# Fix: search a configurable list of roots instead of just ".":
#   1. pipeline_config["documents_root"], if set (highest priority -
#      lets each pipeline run point at the correct mount explicitly)
#   2. the EDO_DOCUMENTS_DIR environment variable, if set (set this in
#      your Dockerfile/compose/deployment to the path where the volume
#      containing the source .docx files is mounted)
#   3. a couple of common container mount-point conventions, as a
#      fallback
#   4. "." itself, last, for local/dev runs where the file may already
#      sit next to the process
# ==========================================================

DEFAULT_DOCUMENT_SEARCH_ROOTS = [
    "/mnt/documents",
    "/data/documents",
    ".",
]


def get_document_search_roots(pipeline_config=None):
    """
    Builds the ordered list of directories to search for the actual
    source .docx file, per the priority described above. The
    EDO_DOCUMENTS_DIR environment variable is read here (at call time),
    not baked into a module-level constant, so it's picked up correctly
    regardless of when it was set relative to module import.
    """
    roots = []

    if pipeline_config:
        configured_root = pipeline_config.get("documents_root")
        if configured_root:
            roots.append(configured_root)

    env_root = os.environ.get("EDO_DOCUMENTS_DIR")
    if env_root:
        roots.append(env_root)

    for root in DEFAULT_DOCUMENT_SEARCH_ROOTS:
        if root and root not in roots:
            roots.append(root)

    return roots


def get_edo_proposed_file(edo_document, pipeline_config=None):
    document = edo_document.get("edo_proposed")
    if not document:
        return None

    for key in ["file_path", "document_path", "path", "local_path", "filepath"]:
        val = document.get(key)
        if not val:
            continue
        if os.path.exists(val):
            logging.info(f"Resolved edo_proposed file from document field '{key}' : {val}")
            return val
        logging.warning(
            f"document['{key}'] = {val!r} was set but does not exist on "
            "disk in this container - falling back to filename search."
        )

    doc_name = document.get("document_name") or document.get("name")
    if not doc_name:
        logging.warning(
            "edo_proposed document has no document_name/name field either - "
            "nothing to search for."
        )
        return None

    search_roots = get_document_search_roots(pipeline_config)
    logging.info(f"Searching for '{doc_name}' under: {search_roots}")

    for root in search_roots:
        found = find_file_by_name(doc_name, search_dir=root)
        if found:
            logging.info(f"Found '{doc_name}' at : {found}")
            return found

    logging.error(
        f"Could not find '{doc_name}' under any of {search_roots}. "
        "If this is running in Docker, that directory needs to actually "
        "be mounted into the container (a volume/bind-mount to wherever "
        "the source .docx files live) - set the EDO_DOCUMENTS_DIR "
        "environment variable, or pipeline_config['documents_root'], to "
        "the correct in-container mount path."
    )
    return None


def extract_edo_proposed_images(edo_document, pipeline_config=None):
    """
    CANONICAL version. Uses the SAME "edo_proposed" document reference
    that extract_edo_tags() already resolves (edo_document["edo_proposed"])
    - no separate document lookup. Resolves its local .docx path via
    get_edo_proposed_file(), then extracts its embedded FIGURE images
    (duplicates included, tables/logo excluded) via
    extract_docx_figures_only(). Source documents are always .docx here,
    so no .doc -> .docx conversion step is needed.

    `pipeline_config` is optional and passed straight through to
    get_edo_proposed_file() so pipeline_config["documents_root"] can be
    used to fix Docker deployments where the source .docx files live on
    a mounted volume rather than the container's default CWD - see the
    "DOCUMENT SEARCH ROOTS (Docker fix)" section above.

    See extract_edo_proposed_images_legacy() below for the previous
    db-based version (kept for backward compatibility, not used by the
    merged pipeline).
    """
    try:
        file_path = get_edo_proposed_file(edo_document, pipeline_config)
        if not file_path:
            # get_edo_proposed_file() / find_file_by_name() already log
            # exactly which directories were searched and why nothing
            # was found - this just makes the end result unambiguous.
            logging.warning(
                "IMAGE EXTRACTION SKIPPED - the edo_proposed .docx file "
                "itself could not be located (see the search log above)."
            )
            return []

        images = extract_docx_figures_only(file_path)

        if not images:
            # The file WAS found and read - so if this is empty, it's
            # almost certainly the "Figure N" caption filter in
            # extract_docx_figures_only() not matching this document's
            # actual caption style (see its per-image SKIPPED log lines
            # just above this one for exactly why each image was
            # excluded), NOT a missing-file problem.
            logging.warning(
                f"IMAGE EXTRACTION RETURNED 0 FIGURES from a file that WAS "
                f"found and read successfully ({file_path}). This means "
                "every image in the document was excluded by the Figure/"
                "Table/Hillrom caption filter - check the 'SKIPPED (...)' "
                "log lines just above for the reason each one was "
                "excluded, and compare against this document's actual "
                "caption wording (CAPTION_FIGURE_PATTERN currently "
                "requires text starting with 'Figure <number>')."
            )
        else:
            logging.info(
                f"Extracted {len(images)} figure(s) from edo_proposed "
                f"({edo_document.get('edo_proposed', {}).get('document_name')})."
            )

        return images
    except Exception as e:
        logging.error(f"Image extraction failed : {e}")
        return []


def insert_image_below_text(sheet, image, row, column=8, text_offset_px=IMAGE_TEXT_OFFSET_PX):
    """
    CANONICAL - from edo_image.py. Anchors `image` into `sheet` at
    (row, column) - Column H (column=8) by default - positioned
    `text_offset_px` pixels below the top of the cell, so it renders
    underneath whatever text is already in that cell rather than
    overlapping it. Also grows the row height / column width as needed
    so the image isn't clipped.
    """
    if not image:
        return False
    try:
        suffix = image.get("extension") or ".png"
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as temp:
            temp.write(image["bytes"])
            temp_name = temp.name

        excel_image = XLImage(temp_name)
        excel_image.width = IMAGE_WIDTH
        excel_image.height = IMAGE_HEIGHT

        marker = AnchorMarker(
            col=column - 1, colOff=pixels_to_EMU(2),
            row=row - 1, rowOff=pixels_to_EMU(text_offset_px)
        )
        size = XDRPositiveSize2D(cx=pixels_to_EMU(IMAGE_WIDTH), cy=pixels_to_EMU(IMAGE_HEIGHT))
        excel_image.anchor = OneCellAnchor(_from=marker, ext=size)
        sheet.add_image(excel_image)

        required_height = (text_offset_px + IMAGE_HEIGHT) * 0.75
        current_height = sheet.row_dimensions[row].height or 0
        if required_height > current_height:
            sheet.row_dimensions[row].height = required_height

        required_col_width = (IMAGE_WIDTH + 15) * 0.14
        current_width = sheet.column_dimensions['H'].width or 0
        if required_col_width > current_width:
            sheet.column_dimensions['H'].width = required_col_width
        return True
    except Exception as e:
        logging.error(f"Image insert failed at row {row} : {e}")
        return False


# ---- Legacy (db-based) resolution - kept, unused, for backward
# ---- compatibility. Not called by the merged pipeline. ----

# Method names this will try calling on `db`, in order, each with the
# document's identity/name as the argument. Add your real method name
# here first if you already know it, to skip the guessing.
CANDIDATE_DB_FILE_METHODS = [
    "get_document_file",
    "download_document",
    "get_document_bytes",
    "get_document_content",
    "fetch_document_file",
    "get_file",
    "get_document_blob",
]

# Keys this will check directly on the document metadata dict itself, in
# case the file path/URL is already sitting right there.
CANDIDATE_DOCUMENT_PATH_KEYS = [
    "file_path",
    "path",
    "local_path",
    "document_path",
    "url",
    "file_url",
    "blob_url",
    "s3_key",
    "s3_url",
]


def _write_bytes_to_temp_docx(data, suffix=".docx"):
    fd, temp_path = tempfile.mkstemp(suffix=suffix)
    with os.fdopen(fd, "wb") as f:
        f.write(data)
    return temp_path


def resolve_edo_document_file(document, db):
    """
    Legacy - kept for backward compatibility, not used by the merged
    pipeline (see extract_edo_proposed_images() above for the canonical
    version). Resolves an actual local .docx file path for a document
    metadata dict, trying several common DatabaseHandler patterns:

      1. A file path / URL already present directly on the document dict
         (see CANDIDATE_DOCUMENT_PATH_KEYS). A local path is returned
         as-is; an http(s) URL is downloaded to a temp file.
      2. A db method (see CANDIDATE_DB_FILE_METHODS) called with the
         document's identity/name, which may return either raw bytes or
         a local file path.

    Returns a local filesystem path to a .docx file, or raises a clear
    RuntimeError listing everything it tried, so the real gap is obvious
    instead of silently returning no images.
    """
    identity = document.get("document_identity")
    name = document.get("document_name")
    tried = []

    # ---- 1. Direct path/URL already on the metadata dict ----
    for key in CANDIDATE_DOCUMENT_PATH_KEYS:
        value = document.get(key)
        if not value:
            continue
        tried.append(f"document[{key!r}] = {value!r}")

        if isinstance(value, str) and value.startswith(("http://", "https://")):
            logging.info(f"Downloading document from URL field '{key}': {value}")
            import urllib.request
            temp_path = tempfile.mktemp(suffix=".docx")
            urllib.request.urlretrieve(value, temp_path)
            return temp_path

        if isinstance(value, str) and os.path.exists(value):
            logging.info(f"Using local file path from document field '{key}': {value}")
            return value

    # ---- 2. A DatabaseHandler method that fetches the file ----
    for method_name in CANDIDATE_DB_FILE_METHODS:
        method = getattr(db, method_name, None)
        if not callable(method):
            continue

        for arg in (identity, name):
            if not arg:
                continue
            tried.append(f"db.{method_name}({arg!r})")
            try:
                result = method(arg)
            except Exception as e:
                logging.warning(f"db.{method_name}({arg!r}) raised {e!r}, trying next option.")
                continue

            if not result:
                continue

            if isinstance(result, (bytes, bytearray)):
                logging.info(f"Got raw file bytes from db.{method_name}({arg!r}) - writing to temp file.")
                return _write_bytes_to_temp_docx(result)

            if isinstance(result, str) and os.path.exists(result):
                logging.info(f"Got local file path from db.{method_name}({arg!r}): {result}")
                return result

    # Self-diagnosing failure: instead of just saying nothing matched,
    # dump what's ACTUALLY available on `db` and on the document dict
    # directly into the log, right here, so the real answer shows up in
    # this run's log output without needing a separate inspection script.
    actual_document_keys = list(document.keys())
    actual_db_methods = [
        name for name in dir(db)
        if not name.startswith("_") and callable(getattr(db, name, None))
    ]

    logging.error("=" * 80)
    logging.error("IMAGE FILE RESOLUTION FAILED - DIAGNOSTIC DUMP")
    logging.error("=" * 80)
    logging.error(f"Document dict actual keys : {actual_document_keys}")
    logging.error(f"Document dict full contents: {document}")
    logging.error(f"db object actual methods   : {actual_db_methods}")
    logging.error(
        "None of these matched CANDIDATE_DOCUMENT_PATH_KEYS "
        f"({CANDIDATE_DOCUMENT_PATH_KEYS}) or CANDIDATE_DB_FILE_METHODS "
        f"({CANDIDATE_DB_FILE_METHODS})."
    )
    logging.error(
        "Look at 'db object actual methods' above for the real "
        "file/document/blob/download-fetching method name, then add it "
        "to CANDIDATE_DB_FILE_METHODS at the top of this section - or "
        "look at 'Document dict full contents' for the real field name "
        "and add it to CANDIDATE_DOCUMENT_PATH_KEYS."
    )
    logging.error("=" * 80)

    raise RuntimeError(
        f"Could not resolve an actual .docx file for document "
        f"identity={identity!r} name={name!r}. Tried:\n  - "
        + "\n  - ".join(tried or ["(nothing - no candidate keys/methods matched)"])
        + "\n\nSee the 'IMAGE FILE RESOLUTION FAILED - DIAGNOSTIC DUMP' "
          "block just above in this run's log for the document's actual "
          "keys and db's actual methods, then add the correct one to "
          "CANDIDATE_DB_FILE_METHODS (or CANDIDATE_DOCUMENT_PATH_KEYS) at "
          "the top of this section."
    )


def extract_edo_proposed_images_legacy(edo_document, db):
    """
    Legacy - kept for backward compatibility, not used by the merged
    pipeline. See extract_edo_proposed_images() above for the canonical
    version (from edo_image.py), which no longer needs the `db`
    argument.
    """
    if "edo_proposed" not in edo_document:
        raise Exception("EDO_Proposed document is not configured - cannot extract images.")

    document = edo_document["edo_proposed"]
    docx_path = resolve_edo_document_file(document, db)

    images = extract_images_from_file(docx_path)
    logging.info(f"Extracted {len(images)} image(s) from edo_proposed ({document.get('document_name')}).")
    return images


# ==========================================================
# PDF IMAGE/TABLE EXTRACTION - ported as-is from
# generate_ICU_Template_final_working_code.py. Standalone: does not
# touch or replace the DOCX-based image extraction above
# (extract_docx_figures_only / extract_edo_proposed_images), and is not
# wired into any pipeline stage here - call extract_section_images()
# directly wherever PDF-sourced figures/tables are needed.
# ==========================================================

def resolve_pdf_path(pipeline_config=None):
    """
    Resolves the PDF file to extract images/tables from: prefers
    pipeline_config["pdf_file_path"] if provided, otherwise falls back
    to PDF_GLOBAL_OVERRIDE.
    """
    pipeline_pdf_path = None
    if pipeline_config:
        pipeline_pdf_path = pipeline_config.get("pdf_file_path")
    pdf_path = pipeline_pdf_path or PDF_GLOBAL_OVERRIDE
    pdf_path = normalize_text(pdf_path)
    if not pdf_path:
        raise ValueError("No PDF path provided.")
    return os.path.abspath(pdf_path)


# ==========================================================
# DB-DRIVEN PDF RESOLUTION (EDO_pdf_new) - New EDO diagrams
# ==========================================================
# Mirrors get_edo_proposed_file()'s approach for the docx source:
# reads the local file path straight off the "edo_pdf_new" document
# metadata dict (edo_document["edo_pdf_new"], resolved in
# get_edo_document() from the db document whose document_identity is
# "EDO_pdf_new"), falling back to a filename search across the same
# search roots as the docx flow. This REPLACES relying on
# PDF_GLOBAL_OVERRIDE / pipeline_config["pdf_file_path"] for the New
# EDO diagram workflow - resolve_pdf_path() above is left as-is/unused
# for backward compatibility with any other caller.
# ==========================================================

def get_edo_pdf_new_file(edo_document, pipeline_config=None):
    """
    Resolves a local filesystem path to the "EDO_pdf_new" PDF, using the
    SAME field-then-filename-search pattern as get_edo_proposed_file():

      1. edo_document["edo_pdf_new"] must be present (set by
         get_edo_document() from the db document whose document_identity
         is "EDO_pdf_new").
      2. Check common path fields directly on that document's metadata.
      3. Otherwise, search for its document_name/name across
         get_document_search_roots() (pipeline_config["documents_root"],
         EDO_DOCUMENTS_DIR, then the default mount-point fallbacks).

    Every branch that fails to resolve a path logs exactly why, so an
    empty result downstream is traceable back to a specific cause
    instead of just silently producing no diagrams.
    """
    document = edo_document.get("edo_pdf_new") if edo_document else None
    if not document:
        logging.warning(
            "EDO_pdf_new RESOLUTION FAILED - no document with "
            "document_identity == 'EDO_pdf_new' was configured for this "
            "template (edo_document has no 'edo_pdf_new' key). New EDO "
            "diagrams (Column H) will be skipped."
        )
        return None

    for key in ["file_path", "document_path", "path", "local_path", "filepath"]:
        val = document.get(key)
        if not val:
            continue
        if os.path.exists(val):
            logging.info(f"Resolved edo_pdf_new file from document field '{key}' : {val}")
            return val
        logging.warning(
            f"EDO_pdf_new document['{key}'] = {val!r} was set but does "
            "not exist on disk in this container - falling back to "
            "filename search."
        )

    doc_name = document.get("document_name") or document.get("name")
    if not doc_name:
        logging.error(
            "EDO_pdf_new RESOLUTION FAILED - the document has no usable "
            "path field (checked file_path/document_path/path/"
            "local_path/filepath) AND no document_name/name field to "
            f"search by. Full document dict: {document}"
        )
        return None

    search_roots = get_document_search_roots(pipeline_config)
    logging.info(f"Searching for EDO_pdf_new '{doc_name}' under: {search_roots}")

    for root in search_roots:
        found = find_file_by_name(doc_name, search_dir=root)
        if found:
            logging.info(f"Found EDO_pdf_new '{doc_name}' at : {found}")
            return found

    logging.error(
        f"EDO_pdf_new RESOLUTION FAILED - could not find '{doc_name}' "
        f"under any of {search_roots}. If this is running in Docker, "
        "that directory needs to actually be mounted into the container "
        "- set EDO_DOCUMENTS_DIR or pipeline_config['documents_root'] to "
        "the correct in-container mount path. New EDO diagrams (Column "
        "H) will be skipped for this run."
    )
    return None


def is_valid_rect(rect):
    """True if a PyMuPDF rect is non-empty, finite, and has real size."""
    if rect is None:
        return False
    if rect.is_empty or rect.is_infinite:
        return False
    if rect.width <= 1 or rect.height <= 1:
        return False
    return True


def rect_distance(a, b):
    """Euclidean distance between the centers of two PyMuPDF rects."""
    ax = (a.x0 + a.x1) / 2
    ay = (a.y0 + a.y1) / 2
    bx = (b.x0 + b.x1) / 2
    by = (b.y0 + b.y1) / 2
    return ((ax - bx) ** 2 + (ay - by) ** 2) ** 0.5


def normalize_asset_key(value):
    """
    Normalizes a raw caption fragment (e.g. "FIGURE 3", "Fig. 3") into a
    canonical "Figure 3" / "Table 2" style key.
    """
    value = normalize_text(value)
    value = re.sub(r"\bFIGURE\b|\bFig\.?\b", "Figure", value, flags=re.IGNORECASE)
    value = re.sub(r"\bTABLE\b", "Table", value, flags=re.IGNORECASE)
    match = re.search(r"\b(Figure|Table)\s*[A-Za-z0-9][A-Za-z0-9\.\-]*", value, re.IGNORECASE)
    return match.group(0).strip() if match else ""


def extract_assets_from_text(text):
    """
    Finds every "Figure N" / "Fig. N" / "Table N" style reference inside
    a block of text and returns the normalized, de-duplicated list of
    keys (in order of first appearance).
    """
    text = normalize_text(text)
    pattern = r"\b(?:Figure|Fig\.?|Table)\s*[A-Za-z0-9][A-Za-z0-9\.\-]*"
    assets = []
    for m in re.finditer(pattern, text, re.IGNORECASE):
        key = normalize_asset_key(m.group(0))
        if key:
            assets.append(key)
    return list(dict.fromkeys(assets))


def extract_section_images(pdf_path, output_dir=IMAGE_OUTPUT_DIR):
    """
    Scans every page of a PDF for "Figure N" and "Table N" captions,
    locates the image/drawing content (or table content) associated
    with each caption, rasterizes just that region to a PNG, and
    returns a manifest mapping the normalized caption key (e.g.
    "Figure 3") to the saved PNG's filepath.

    For each page:
      1. Collects every embedded image's bounding box, plus vector-
         drawing bounding boxes clustered into connected groups (so a
         multi-shape diagram is treated as one block, not many).
      2. Searches for "Figure"/"FIGURE"/"Fig." caption text; for each
         caption, finds the nearest image/drawing cluster
         (find_best_content_bbox), decides whether that content sits
         above or below the caption, and rasterizes a padded clip
         region covering caption + content to a PNG.
      3. Separately searches for "Table" captions; for each, estimates
         where the table ends by looking at horizontal ruling lines
         (or, failing that, text-block gaps) below the caption, and
         rasterizes that region to a PNG.

    Returns {} if the PDF can't be found or opened.
    """
    logging.info("--- IMPROVED IMAGE & TABLE EXTRACTOR STARTED ---")
    if not os.path.isfile(pdf_path):
        logging.error(f"PDF file not found: {pdf_path}")
        return {}
    os.makedirs(output_dir, exist_ok=True)
    try:
        doc = fitz.open(pdf_path)
    except Exception as e:
        logging.error(f"Failed to open PDF: {e}")
        return {}

    image_manifest = {}

    try:
        for page_num in range(len(doc)):
            page = doc[page_num]
            page_no = page_num + 1

            images = []
            for img in page.get_images(full=True):
                try:
                    xref = img[0]
                    rects = page.get_image_rects(xref)
                    for bbox in rects:
                        if is_valid_rect(bbox):
                            images.append((xref, bbox))
                except Exception as e:
                    logging.warning(f"Image rect resolve failed on page {page_no}: {e}")

            drawing_bboxes = []
            try:
                drawings = page.get_drawings()
                raw_rects = []
                for d in drawings:
                    r = fitz.Rect(d["rect"])
                    if is_valid_rect(r) and r.width > 5 and r.height > 5:
                        raw_rects.append(r)
                used = [False] * len(raw_rects)
                for i, r in enumerate(raw_rects):
                    if used[i]:
                        continue
                    cluster = fitz.Rect(r)
                    changed = True
                    while changed:
                        changed = False
                        for j, r2 in enumerate(raw_rects):
                            if used[j]:
                                continue
                            expanded = fitz.Rect(
                                cluster.x0 - 50, cluster.y0 - 50,
                                cluster.x1 + 50, cluster.y1 + 50
                            )
                            if expanded.intersects(r2):
                                cluster |= r2
                                used[j] = True
                                changed = True
                    used[i] = True
                    if cluster.width > 40 and cluster.height > 40:
                        drawing_bboxes.append(cluster)
            except Exception:
                pass

            def find_best_content_bbox(caption_rect, proximity=300):
                candidates = [bbox for _, bbox in images] + drawing_bboxes
                if not candidates:
                    return None, None
                cap_cx = (caption_rect.x0 + caption_rect.x1) / 2
                cap_cy = (caption_rect.y0 + caption_rect.y1) / 2
                nearby = [
                    b for b in candidates
                    if abs((b.x0 + b.x1) / 2 - cap_cx) < proximity
                    and abs((b.y0 + b.y1) / 2 - cap_cy) < proximity
                ]
                if not nearby:
                    nearby = [min(candidates, key=lambda b: rect_distance(caption_rect, b))]
                union = fitz.Rect(nearby[0])
                for b in nearby[1:]:
                    union |= b
                content_is_above = union.y1 <= caption_rect.y0 + 5
                return union, content_is_above

            search_terms = ["Figure", "FIGURE", "Fig."]
            caption_matches = []
            for term in search_terms:
                try:
                    caption_matches.extend(page.search_for(term))
                except Exception as e:
                    logging.warning(f"Search failed for '{term}' on page {page_no}: {e}")

            deduped_caption_matches = []
            seen_fig = set()
            for rect in caption_matches:
                key = (round(rect.x0, 2), round(rect.y0, 2), round(rect.x1, 2), round(rect.y1, 2))
                if key not in seen_fig:
                    seen_fig.add(key)
                    deduped_caption_matches.append(rect)

            for rect in deduped_caption_matches:
                try:
                    caption_clip = fitz.Rect(
                        max(0, rect.x0 - 5),
                        max(0, rect.y0 - 5),
                        page.rect.x1,
                        min(page.rect.y1, rect.y1 + 40)
                    )
                    raw_caption = page.get_text("text", clip=caption_clip)
                    caption_text = " ".join(raw_caption.split("\n")[:3])
                    caption_text = normalize_text(caption_text)
                    asset_candidates = extract_assets_from_text(caption_text)
                    if not asset_candidates:
                        continue
                    clean_key = asset_candidates[0]
                    content_bbox, content_is_above = find_best_content_bbox(rect)
                    if content_bbox is None:
                        logging.warning(f"No content found for caption [{clean_key}] on page {page_no}; skipping.")
                        continue
                    PAD = 12
                    if content_is_above:
                        clip_rect = fitz.Rect(
                            max(0, content_bbox.x0 - PAD),
                            max(0, content_bbox.y0 - PAD),
                            min(page.rect.x1, content_bbox.x1 + PAD),
                            min(page.rect.y1, rect.y1 + PAD)
                        )
                    else:
                        clip_rect = fitz.Rect(
                            max(0, content_bbox.x0 - PAD),
                            max(0, rect.y0 - PAD),
                            min(page.rect.x1, content_bbox.x1 + PAD),
                            min(page.rect.y1, content_bbox.y1 + PAD)
                        )
                    if clip_rect.width < 20 or clip_rect.height < 20:
                        logging.warning(f"Clip rect too small for [{clean_key}] on page {page_no}; skipping.")
                        continue
                    pix = page.get_pixmap(matrix=fitz.Matrix(3, 3), clip=clip_rect, alpha=False)
                    safe_name = re.sub(r'[^A-Za-z0-9._-]+', '_', clean_key)
                    filename = f"{safe_name}_p{page_no}.png"
                    filepath = os.path.join(output_dir, filename)
                    pix.save(filepath)
                    image_manifest[clean_key] = filepath
                    logging.info(f"Extracted Figure [{clean_key}] -> {filepath}")
                except Exception as e:
                    logging.warning(f"Failed figure mapping on page {page_no}: {e}")

            try:
                table_caption_hits = []
                for term in ["Table"]:
                    try:
                        table_caption_hits.extend(page.search_for(term))
                    except Exception as e:
                        logging.warning(f"Table search failed on page {page_no}: {e}")

                seen_table = set()
                deduped_table_hits = []
                for rect in table_caption_hits:
                    key = (round(rect.x0, 1), round(rect.y0, 1))
                    if key not in seen_table:
                        seen_table.add(key)
                        deduped_table_hits.append(rect)

                for t_rect in deduped_table_hits:
                    try:
                        caption_clip = fitz.Rect(
                            max(0, t_rect.x0 - 5),
                            max(0, t_rect.y0 - 5),
                            page.rect.x1,
                            min(page.rect.y1, t_rect.y1 + 30)
                        )
                        raw_caption = page.get_text("text", clip=caption_clip)
                        caption_line = " ".join(raw_caption.split("\n")[:2])
                        caption_line = normalize_text(caption_line)
                        asset_candidates = extract_assets_from_text(caption_line)
                        if not asset_candidates:
                            continue
                        clean_key = asset_candidates[0]
                        caption_bottom = t_rect.y1
                        table_bottom = None
                        try:
                            drawings = page.get_drawings()
                            h_lines = []
                            for draw in drawings:
                                r = fitz.Rect(draw["rect"])
                                if r.y0 >= caption_bottom and r.x0 < (t_rect.x0 + 200):
                                    h_lines.append(r.y1)
                            if h_lines:
                                candidates = [y for y in h_lines if y - caption_bottom <= 600]
                                if candidates:
                                    table_bottom = max(candidates)
                        except Exception:
                            pass
                        if table_bottom is None:
                            scan_clip = fitz.Rect(
                                max(0, t_rect.x0 - 20),
                                caption_bottom,
                                min(page.rect.x1, t_rect.x0 + 500),
                                min(page.rect.y1, caption_bottom + 600)
                            )
                            blocks = page.get_text("blocks", clip=scan_clip)
                            if blocks:
                                sorted_blocks = sorted(blocks, key=lambda b: b[1])
                                prev_y1 = caption_bottom
                                table_bottom = caption_bottom
                                for blk in sorted_blocks:
                                    blk_y0 = blk[1]
                                    blk_y1 = blk[3]
                                    gap = blk_y0 - prev_y1
                                    if gap > 22 and table_bottom > caption_bottom:
                                        break
                                    table_bottom = blk_y1
                                    prev_y1 = blk_y1
                            else:
                                table_bottom = min(page.rect.y1, caption_bottom + 300)
                        if table_bottom - caption_bottom < 20:
                            table_bottom = min(page.rect.y1, caption_bottom + 200)
                        padding = 8
                        clip_rect = fitz.Rect(
                            max(0, t_rect.x0 - padding),
                            max(0, t_rect.y0 - padding),
                            min(page.rect.x1, t_rect.x0 + 520),
                            min(page.rect.y1, table_bottom + padding)
                        )
                        pix = page.get_pixmap(matrix=fitz.Matrix(2, 2), clip=clip_rect, alpha=False)
                        safe_name = re.sub(r'[^A-Za-z0-9._-]+', '_', clean_key)
                        filename = f"{safe_name}_p{page_no}.png"
                        filepath = os.path.join(output_dir, filename)
                        pix.save(filepath)
                        image_manifest[clean_key] = filepath
                        logging.info(f"Extracted Table [{clean_key}] -> {filepath}")
                    except Exception as e:
                        logging.warning(f"Failed table caption processing on page {page_no}: {e}")
            except Exception as e:
                logging.warning(f"Table extraction failed on page {page_no}: {e}")

    finally:
        doc.close()

    logging.info(f"Total extracted assets: {len(image_manifest)}")
    return image_manifest


# ==========================================================
# NEW EDO DIAGRAM LOOKUP (EDO_pdf_new, per RA/FMEA record)
# ==========================================================
# For a New EDO record (e.g. RA-141 / FMEA Sys-152), find and return its
# full diagram from the EDO_pdf_new PDF, for insertion into Column H of
# that specific record's row - as opposed to the generic FIFO image
# queue that format_edo_worksheet() otherwise pulls edo_proposed
# figures from.
# ==========================================================

def _normalize_match_text(value):
    """Loose match key: strip everything but letters/digits, uppercase."""
    return re.sub(r"[^A-Z0-9]", "", normalize_text(value).upper())


def extract_edo_pdf_new_diagram(pdf_path, ra_number, fmea_number, output_dir=IMAGE_OUTPUT_DIR):
    """
    Returns the full diagram associated with a given RA_Number /
    FMEA_Number pair from the EDO_pdf_new PDF, as an image dict
    ({"name", "bytes", "extension"}) ready for insert_image_below_text().

    Matching strategy:
      1. Open the PDF and scan every page's text for an occurrence of
         ra_number and/or fmea_number (loose match: punctuation/spacing
         differences ignored).
      2. On the first matching page, look for a "Figure N" / "Fig. N"
         caption reference in that page's text and pull the FULL
         extracted figure for that caption via extract_section_images()
         (this is "the full diagram", not a cropped snippet).
      3. If the matched page has no Figure caption, fall back to
         rasterizing the ENTIRE matched page itself as the diagram, so
         a record with a real page match never comes back empty just
         because it lacks a captioned figure.

    On ANY failure to produce an image, logs the precise reason (this is
    what the caller / Excel writer should surface when Column H prints
    empty) and returns None - it never raises.
    """
    ra_number = normalize_text(ra_number)
    fmea_number = normalize_text(fmea_number)

    if not ra_number and not fmea_number:
        logging.error(
            "NEW EDO DIAGRAM LOOKUP FAILED - both RA_Number and "
            "FMEA_Number are empty, nothing to search the PDF for."
        )
        return None

    if not pdf_path or not os.path.isfile(pdf_path):
        logging.error(
            f"NEW EDO DIAGRAM LOOKUP FAILED for RA={ra_number!r} "
            f"FMEA={fmea_number!r} - PDF file not found at {pdf_path!r}. "
            "Column H will print empty for this record because the "
            "source PDF itself could not be opened."
        )
        return None

    ra_key = _normalize_match_text(ra_number)
    fmea_key = _normalize_match_text(fmea_number)

    try:
        doc = fitz.open(pdf_path)
    except Exception as e:
        logging.error(
            f"NEW EDO DIAGRAM LOOKUP FAILED for RA={ra_number!r} "
            f"FMEA={fmea_number!r} - fitz.open() raised {e!r} on "
            f"{pdf_path!r}. Column H will print empty for this record "
            "because the PDF could not be parsed."
        )
        return None

    matched_page_no = None
    matched_page_text = ""
    try:
        for page_num in range(len(doc)):
            page = doc[page_num]
            page_text = page.get_text("text") or ""
            page_key = _normalize_match_text(page_text)
            if (ra_key and ra_key in page_key) or (fmea_key and fmea_key in page_key):
                matched_page_no = page_num + 1
                matched_page_text = page_text
                break

        if matched_page_no is None:
            logging.error(
                f"NEW EDO DIAGRAM LOOKUP FAILED for RA={ra_number!r} "
                f"FMEA={fmea_number!r} - neither value was found on any "
                f"page's text across all {len(doc)} page(s) of "
                f"{pdf_path!r}. Column H will print empty for this "
                "record because nothing in EDO_pdf_new references this "
                "RA/FMEA pair (check spelling/format against the PDF, "
                "e.g. 'RA-141' vs 'RA141' vs 'RA 141')."
            )
            return None

        figure_keys = extract_assets_from_text(matched_page_text)
        figure_keys = [k for k in figure_keys if k.lower().startswith("figure") or k.lower().startswith("fig")]

        if figure_keys:
            manifest = extract_section_images(pdf_path, output_dir=output_dir)
            for fig_key in figure_keys:
                image_path = manifest.get(fig_key)
                if image_path and os.path.isfile(image_path):
                    with open(image_path, "rb") as f:
                        data = f.read()
                    logging.info(
                        f"NEW EDO DIAGRAM FOUND for RA={ra_number!r} "
                        f"FMEA={fmea_number!r} on page {matched_page_no} "
                        f"-> [{fig_key}] -> {image_path}"
                    )
                    return {
                        "name": os.path.basename(image_path),
                        "bytes": data,
                        "extension": os.path.splitext(image_path)[1] or ".png"
                    }
            logging.warning(
                f"NEW EDO DIAGRAM: page {matched_page_no} matched RA="
                f"{ra_number!r} FMEA={fmea_number!r} and references "
                f"{figure_keys}, but extract_section_images() did not "
                "produce a saved file for any of those captions (see its "
                "own log lines above for why that figure was skipped) - "
                "falling back to a full-page rasterization instead."
            )

        # Fallback: no Figure caption on the matched page (or its figure
        # image couldn't be extracted) - rasterize the whole page so the
        # record still gets its "full diagram" rather than nothing.
        try:
            page = doc[matched_page_no - 1]
            pix = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
            os.makedirs(output_dir, exist_ok=True)
            safe_name = f"{_normalize_match_text(ra_number) or _normalize_match_text(fmea_number)}_p{matched_page_no}_fullpage.png"
            filepath = os.path.join(output_dir, safe_name)
            pix.save(filepath)
            with open(filepath, "rb") as f:
                data = f.read()
            logging.info(
                f"NEW EDO DIAGRAM (full-page fallback) FOUND for RA="
                f"{ra_number!r} FMEA={fmea_number!r} on page "
                f"{matched_page_no} -> {filepath}"
            )
            return {"name": safe_name, "bytes": data, "extension": ".png"}
        except Exception as e:
            logging.error(
                f"NEW EDO DIAGRAM LOOKUP FAILED for RA={ra_number!r} "
                f"FMEA={fmea_number!r} - page {matched_page_no} matched "
                f"but full-page rasterization raised {e!r}. Column H "
                "will print empty for this record."
            )
            return None

    finally:
        doc.close()


def extract_new_edo_diagrams(edo_document, new_records, pipeline_config=None, output_dir=IMAGE_OUTPUT_DIR):
    """
    Batch wrapper: resolves the EDO_pdf_new PDF ONCE (db-driven, same
    pattern as extract_edo_proposed_images() uses for the docx source),
    then for every New EDO record in `new_records` (each expected to
    carry RA_Number / FMEA_Number, as produced by extract_new_edo_tags()
    / merge_new_edo_records()) looks up its full diagram.

    Returns a dict keyed by (RA_Number.upper(), FMEA_Number.upper()) ->
    image dict, for format_edo_worksheet()'s `ra_fmea_images` param.
    Records with no diagram are simply absent from the dict - every
    "not found" case is already logged in detail by
    extract_edo_pdf_new_diagram() above, so the caller doesn't need to
    duplicate that diagnosis, only note the aggregate outcome.
    """
    diagrams = {}

    pdf_path = get_edo_pdf_new_file(edo_document, pipeline_config)
    if not pdf_path:
        logging.warning(
            "NEW EDO DIAGRAM EXTRACTION SKIPPED FOR ALL RECORDS - the "
            "EDO_pdf_new PDF could not be resolved (see the "
            "get_edo_pdf_new_file() log lines directly above for the "
            "exact reason). Column H will print empty for every New EDO "
            "record this run."
        )
        return diagrams

    total = 0
    found = 0
    for record in new_records:
        if not isinstance(record, dict):
            continue
        ra_number = normalize_text(record.get("RA_Number"))
        fmea_number = normalize_text(record.get("FMEA_Number"))
        if not ra_number and not fmea_number:
            continue

        total += 1
        image = extract_edo_pdf_new_diagram(pdf_path, ra_number, fmea_number, output_dir=output_dir)
        key = (ra_number.upper(), fmea_number.upper())
        if image:
            diagrams[key] = image
            found += 1
        # else: extract_edo_pdf_new_diagram() already logged the exact
        # cause - nothing further to log here per-record.

    logging.info(
        f"NEW EDO DIAGRAM EXTRACTION SUMMARY: {found}/{total} New EDO "
        f"record(s) matched to a diagram in EDO_pdf_new."
    )
    if total and not found:
        logging.warning(
            "NEW EDO DIAGRAM EXTRACTION found ZERO diagrams for ANY New "
            "EDO record. Column H will be empty for all of them. Most "
            "likely causes: (1) EDO_pdf_new resolved to the wrong PDF "
            "file, (2) the RA/FMEA identifiers in this PDF are formatted "
            "differently than in edo_new_data (e.g. 'RA141' vs 'RA-141'), "
            "or (3) this PDF simply doesn't reference these RA/FMEA "
            "pairs at all - check the per-record 'NEW EDO DIAGRAM LOOKUP "
            "FAILED' log lines above for the specific reason on each."
        )

    return diagrams


def main_image_extraction():
    if len(sys.argv) != 3:
        print("Usage: python extract_images_standalone.py <input_docx> <output_folder>")
        sys.exit(1)

    input_file = sys.argv[1]
    output_folder = sys.argv[2]

    if not os.path.exists(input_file):
        logging.error(f"Input file not found: {input_file}")
        sys.exit(1)

    logging.info("=" * 80)
    logging.info(f"EXTRACTING IMAGES FROM : {input_file}")
    logging.info("=" * 80)

    images = extract_images_from_file(input_file)

    if not images:
        logging.warning("No images found in document.")
        sys.exit(0)

    saved = save_images_to_folder(images, output_folder)

    logging.info("=" * 80)
    logging.info(f"DONE - {len(saved)} image(s) saved to {output_folder}")
    logging.info("=" * 80)



# ==========================================================
# HARDCODED PROMPT - EDO_Existing_Trace (Stage 3C trace extraction)
# Used directly by extract_existing_edo_trace_details() instead of being
# fetched from the database via get_prompt().
# ==========================================================

EDO_EXISTING_TRACE_PROMPT_TEXT = """Task:

Search Section 10.3 Safety Hazard DFMEA.


Step 1:
Locate the row where "System DFMEA #" exactly equals the requested value.


Step 2:
From the same row extract ONLY:
- System DFMEA #
- Trace To RAC#



Step 3:
All extracted fields must come from the same DFMEA row.


Do not use information from neighboring rows.
Do not combine values from different DFMEA records.


Within "Trace(s) to Module DFMEA risk controls: Document#, Tag#; Cause":


Extract only entries where Tag# starts with:
- MRS CU FMEA-
- MRS Software FMEA -
- MRS ACC FMEA-




Ignore and do not return:



- PFMEA
- Any other tag types


If no MRS CU FMEA or MRS Software FMEA or MRS ACC FMEA entries exist in the matched DFMEA row, return an empty array.


Within "Additional Risk Control Measures from Module and/or Component DFMEAs":

Extract only entries under:
- Inherent Safety by Design and Manufacture

Return only identifiers that start with:
- MS CU Mod-
- MS ACC Mod-
- SRS-CTRL-

Do NOT return descriptions, requirement text, explanations, or any text after the identifier.

Ignore:
- Protective Measures
- Information for Safety
- Any other identifiers

If no matching identifiers exist, return an empty array.

Example:
[
  "MS ACC Mod-536",
  "MS ACC Mod-456",
  "SRS-CTRL-39",
  "SRS-CTRL-53",
  "MS CU Mod-380"
]
Within the matched DFMEA row:

A. From "Risk Control Measures at System Level":
- Inherent Safety by Design and Manufacture
- Protective Measures
- Information for Safety

Extract only identifiers that start with:
- DRS-

B. From "Additional Risk Control Measures from Module and/or Component DFMEAs":
- Inherent Safety by Design and Manufacture

Extract only identifiers that start with:
- MS CU Mod-
- MS ACC Mod-
- SRS-CTRL-

Do NOT require DRS entries to exist.

Always extract matching identifiers from section B even when no DRS identifiers are found in section A.

If section A contains no DRS identifiers, return an empty array for DRS and still return all matching identifiers from section B.

Return only identifiers.
Do not return descriptions or requirement text.
``
output:
{ "System DFMEA #": "FMEA Sys-734", "Trace To RAC#": "RA-141", "Traces to Module DFMEA risk controls": [ { "document_number": "NPD37819", "tag_number": "MRS CU FMEA-522", "cause": "Magnet failure at the air hose connector" }, { "document_number": "NPD40091", "tag_number": "MRS ACC FMEA-380", "cause": "Damage of connector side" } ], "Additional Risk Control Measures from Module and/or Component DFMEAs": [ "MS CU Mod-536", "MS CU Mod-538", "MS ACC Mod-220" ], "DRS Identifiers": [] }"""

EDO_EXISTING_TRACE_PROMPT = {
    "prompt_role": "system",
    "prompt_text": EDO_EXISTING_TRACE_PROMPT_TEXT
}


# ==========================================================
# HARDCODED PROMPT - EDO_Existing_Generic (Stage 3 attribute hydration)
# Used directly by extract_edo_details() instead of being fetched from
# the database via get_prompt(). Instructs the LLM to return the
# existing-EDO record as the EDO_Table / existing_edo_data / design_elements
# JSON structure below.
# ==========================================================

EDO_EXISTING_GENERIC_PROMPT_TEXT = """Task:

For the requested EDO Tag, extract the complete Existing EDO record from
the EDO Proposed document and return it as a single JSON object in the
EXACT structure below - no extra commentary, no text outside the JSON
object.

{
  "EDO_Table": {
    "Total_Rows": "<row count as a string>",
    "existing_edo_data": [
      {
        "edo_no": "<EDO tag, e.g. EDO-29>",
        "edo_description": "<top-level product feature/function description>",
        "reason identified as edo": "<top-level reason identified as EDO>",
        "type": "Old",
        "RA and FMEA no": "<RA&C document/number + System FMEA document/number>",
        "Trace": "<RA&C and/or Sys-DFMEA trace narrative, or \\"Blank\\">",
        "Verification Reference": [],
        "design_elements": [
          {
            "edo_location": "<part number (part name)>",
            "edo_description": "<sentence describing what is controlled at this location, ending in 'as below:'>",
            "image": "Blank",
            "reason identified as edo": "<reason this dimension/feature is controlled>",
            "SysDD or HDD Reference": "<hardware/system detailed design document reference>"
          }
        ],
        "component specification": "Blank",
        "Risk Status": "Blank",
        "remarks and recommendation": "Blank"
      }
    ]
  }
}

Rules:
- Return one entry in "existing_edo_data" per requested EDO Tag, and one
  entry in "design_elements" per distinct location/part associated with
  that EDO.
- Every "design_elements[].edo_description" must end with "as below:"
  (never "Figure N" or "Fig.").
- Use the literal string "Blank" for any field with no value - never
  leave a field empty or omit it.
- "Verification Reference" is always returned as an array (empty array
  if none is found).
- "Total_Rows" is the count of entries in "existing_edo_data", returned
  as a string.

Example output:
{ "EDO_Table": { "Total_Rows": "1", "existing_edo_data": [ { "edo_no": "EDO-29", "edo_description": "Airway clearance therapy", "reason identified as edo": "If the product airway clearance therapy does not function properly, it results in a medium safety risk identified in the RA&C.", "type": "Old", "RA and FMEA no": "NPD36702 Vest APX Risk Assessment and Control RA-180 NPD36706 Vest APX System FMEA FMEA Sys-147", "Trace": "Blank", "Verification Reference": [], "design_elements": [ { "edo_location": "210119 (metal cyl)", "edo_description": "For 210119, the metal cylinder ID is controlled as below:", "image": "Blank", "reason identified as edo": "For this ID of metal cylinder 210119, it is part of critical dimension for controlling O-ring compression % in APG system.", "SysDD or HDD Reference": "NPD38119 Titan Hardware Detailed Design" } ], "component specification": "Blank", "Risk Status": "Blank", "remarks and recommendation": "Blank" } ] } }"""

EDO_EXISTING_GENERIC_PROMPT = {
    "prompt_role": "system",
    "prompt_text": EDO_EXISTING_GENERIC_PROMPT_TEXT
}


logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    stream=sys.stdout
)


# ==========================================================
# EXCEL FORMATTING
# ==========================================================

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

cell_alignment = Alignment(
    horizontal="left",
    vertical="top",
    wrap_text=True
)

aptos_font = Font(name="Aptos Narrow", size=8)


# ==========================================================
# RICH-TEXT / COLOR FORMATTING (drives format_edo_worksheet)
# ==========================================================
# Tweak these to match your exact template look-and-feel.

FONT_NAME = "Calibri"
FONT_SIZE = 10

BLACK = "FF000000"
RED = "FFFF0000"

# Lines matching these patterns (inside Column D / Column E multi-line
# text) are treated as "reference / traceability ID" lines -> RED + BOLD.
RED_BOLD_PATTERNS = [
    r'^\(DRS-\d+\)',        # (DRS-95) NPD44078 Rev 1 ...
    r'^\(MS\s*CU',          # (MS CU Mod-448) NPD45862 Rev 3 ...
    r'^RA-\d+',             # RA-133
    r'^FMEA\s*Sys-\d+',     # FMEA Sys-725
    r'^RRAA',               # RRAA - NPD35987
]

# Lines matching these patterns are treated as "organisation / document
# name" lines -> BLACK + BOLD.
BLACK_BOLD_PATTERNS = [
    r'^\(MS\s*ACC',         # (MS ACC Mod-413, ...) - NPD44001
    r'^NPD\d+',             # NPD36702 Vest APX Risk Assessment and Control
]


def _font(color=BLACK, bold=False, size=FONT_SIZE, name=FONT_NAME):
    return Font(name=name, size=size, color=color, bold=bold)


def _inline_font(color=BLACK, bold=False, size=FONT_SIZE, name=FONT_NAME):
    return InlineFont(rFont=name, sz=size, color=color, b=bold)


def _text(value):
    return "" if value is None else str(value)


def _apply_border_alignment(cell):
    cell.alignment = cell_alignment
    cell.border = thin_border


def classify_line(line):
    """
    Returns (color, bold) for a single line of a multi-line Column D /
    Column E cell value, based on simple regex rules. Anything that
    doesn't match a rule falls back to plain BLACK, regular text
    (continuation / description lines, e.g. "Rev 1 - Vest APX ... TDR").
    """
    stripped = line.strip()

    if not stripped:
        return BLACK, False

    for pattern in RED_BOLD_PATTERNS:
        if re.search(pattern, stripped, re.IGNORECASE):
            return RED, True

    for pattern in BLACK_BOLD_PATTERNS:
        if re.search(pattern, stripped, re.IGNORECASE):
            return BLACK, True

    return BLACK, False


def build_rich_text(value):
    """
    Splits a multi-line cell value on "\\n" and builds a CellRichText
    object, giving each line its own color/bold via classify_line().
    Falls back to a plain string when there's nothing meaningful to
    style, so single-line / blank cells don't get needless rich-text
    wrapping.
    """
    text = _text(value)

    if not text.strip() or "\n" not in text:
        color, bold = classify_line(text)
        if bold or color == RED:
            return CellRichText(TextBlock(_inline_font(color=color, bold=bold), text))
        return text

    lines = text.split("\n")
    blocks = []

    for idx, line in enumerate(lines):
        color, bold = classify_line(line)
        blocks.append(TextBlock(_inline_font(color=color, bold=bold), line))
        if idx != len(lines) - 1:
            blocks.append("\n")

    return CellRichText(*blocks)


# ==========================================================
# TEXT HELPERS (identical in both source files - kept once)
# ==========================================================

def normalize_text(value):
    if value is None:
        return ""

    value = str(value)

    value = value.replace("\u00A0", " ")
    value = value.replace("\u2007", " ")
    value = value.replace("\u202F", " ")

    value = unicodedata.normalize(
        "NFKC",
        value
    )

    return value.strip()


def clean_llm_response(response):

    response = normalize_text(response)

    return re.sub(
        r"^```json\s*|\s*```$",
        "",
        response,
        flags=re.DOTALL
    ).strip()


# ==========================================================
# RED-COLOURED CONSOLE OUTPUT (for extracted text coming back
# from any function that queries the edo_document["edo_fmea"]["collection"])
# ==========================================================

RED_CONSOLE = "\033[91m"
RESET_CONSOLE = "\033[0m"


def print_red(label, text):
    """
    Prints `text` to the console in red (ANSI escape codes), prefixed
    with `label`. Used specifically wherever the LLM is queried against
    edo_document["edo_fmea"]["collection"], so that the raw extracted
    text stands out clearly in the console/log output.
    """
    print(f"{RED_CONSOLE}{label}:\n{text}{RESET_CONSOLE}")


def parse_json(response):
    try:

        cleaned = clean_llm_response(response)

        logging.info("========== CLEANED JSON ==========")
        logging.info(cleaned)

        return json.loads(cleaned)

    except Exception as e:

        logging.error(f"JSON Parse Error : {e}")

        # FALLBACK - the direct parse failed, most likely because the LLM
        # wrapped the JSON in extra prose/text beyond a plain ```json fence
        # (clean_llm_response only strips a leading/trailing fence, not
        # surrounding text). Try to salvage the first {...} or [...] block
        # in the response before giving up, so a well-formed JSON payload
        # isn't silently discarded just because of text around it.
        try:
            match = re.search(r"(\{.*\}|\[.*\])", normalize_text(response), re.DOTALL)
            if match:
                salvaged = json.loads(match.group(1))
                logging.warning(
                    "JSON Parse Error recovered via fallback extraction - "
                    "the LLM response had extra text around the JSON block."
                )
                return salvaged
        except Exception as fallback_error:
            logging.error(f"JSON fallback extraction also failed: {fallback_error}")

        logging.error(response)

        return {}


def blank(value):
    """
    NOTE: previously substituted the literal text "Blank" for empty
    values. Per requirement, no placeholder text should ever be written
    to the output Excel - a missing value should simply be an empty
    cell. This now just normalizes the text and returns "" for anything
    empty, instead of inserting "Blank".
    """
    return normalize_text(value)


# ==========================================================
# LLM EXECUTION (identical in both source files - kept once)
# ==========================================================

# NOTE: classify_risk_status() needs a plain "send this single prompt
# string, get a text answer back" call - it isn't tied to any document
# collection, so it can't go through execute_llm/execute_llm_retry
# (those require a RAG "collection" argument). call_llm() below matches
# the same pipeline_config["llm"].generate(...) pattern already used by
# evaluate_comparison() elsewhere in this app.
def call_llm(prompt, pipeline_config):
    try:
        llm = pipeline_config["llm"]
        return llm.generate(
            prompt,
            context="",
            question="risk classification",
            temperature=pipeline_config["temperature"],
            max_tokens=pipeline_config["max_tokens"]
        )
    except Exception as e:
        logging.error(f"LLM risk classification call failed: {e}")
        return "No"


def clean_response(response):
    """
    Strips whitespace/code-fences from a plain-text LLM response
    (used by classify_risk_status - it only ever expects one bare word
    back, e.g. "High").
    """
    return clean_llm_response(response)


def execute_llm(
    pipeline_config,
    collection,
    prompt_row
):
    """
    Generic LLM wrapper used by every prompt passed via positional arguments.
    """
    return retrieve_content_for_prompt(
        pipeline_config,
        collection,
        prompt_row["question"],
        prompt_row["prompt_role"],
        prompt_row["prompt_text"],
        prompt_row["fulltext"],
        prompt_row["where_filter"],
        prompt_row["where_document"],
        prompt_row.get("checkpoint", "")
    )


def execute_llm_retry(
    pipeline_config,
    collection,
    prompt_row
):
    """
    Executes the LLM with retry logic.
    """

    last_exception = None

    for attempt in range(MAX_LLM_RETRIES):

        try:

            docs, metadata, response = execute_llm(
                pipeline_config,
                collection,
                prompt_row
            )

            if response:
                return docs, metadata, response

        except Exception as ex:

            last_exception = ex

            logging.warning(
                f"LLM Retry {attempt+1}/{MAX_LLM_RETRIES}"
            )

            time.sleep(
                INITIAL_RETRY_DELAY * (attempt + 1)
            )

    raise Exception(
        f"LLM failed after retries : {last_exception}"
    )


# ==========================================================
# WORKBOOK (identical in both source files - kept once)
# ==========================================================

def initialize_workbook(
    pipeline_config
):
    """
    Opens Excel template.
    """

    workbook = openpyxl.load_workbook(
        pipeline_config["input_file_path"]
    )

    sheet = workbook.active

    return workbook, sheet


# ==========================================================
# COMMON EMPTY EDO STRUCTURE
# ==========================================================

def create_empty_edo():
    """
    CANONICAL version - from edo_existing_final.py.
    Superset of the copy-file version: also carries FMEA_Number so that
    Stage 3B (verification-reference matching by RA/FMEA number) has
    something to match against.

    NOTE: defaults are empty strings, not the literal text "Blank" -
    an unfound value should render as a truly empty cell in the output
    Excel, per requirement.
    """
    return {
        "edo_type": "Existing",
        "edo_tag": "",
        "ra_number": "",
        "FMEA_Number": "",
        "edo_description": "",
        "reason_identified": "",
        "dfmea": "",
        "verification_reference": "",
        "existing_trace": "",
        "location": "",
        "description_2": "",
        "reason_2": "",
        "sysdd": ""
    }


def create_empty_edo_minimal():
    """
    Original version from generate_EDO_template_copy.py - kept for
    backward compatibility. Does NOT carry FMEA_Number. Not used by the
    merged pipeline (create_empty_edo is used instead).
    """
    return {
        "edo_type": "Existing",
        "edo_tag": "Blank",
        "ra_number": "Blank",
        "edo_description": "Blank",
        "reason_identified": "Blank",
        "dfmea": "Blank",
        "verification_reference": "Blank",
        "location": "Blank",
        "description_2": "Blank",
        "reason_2": "Blank",
        "sysdd": "Blank"
    }


# ==========================================================
# LOGGING (identical in both source files - kept once)
# ==========================================================

def log_dictionary(title, dictionary):
    logging.info("=" * 80)
    logging.info(title)

    for key, value in dictionary.items():
        logging.info(f"{key}")
        logging.info(json.dumps(value, indent=4))

    logging.info("=" * 80)


# ==========================================================
# DOCUMENT HELPERS
# ==========================================================

def get_template_documents(
    client,
    product_family,
    product,
    templatename,
    db: DatabaseHandler
):
    docs = db.get_template_documents(
        client,
        product_family,
        product,
        templatename
    )

    if not docs:
        raise Exception("No template documents were found.")

    logging.info("=" * 80)
    logging.info("AVAILABLE TEMPLATE DOCUMENTS")
    logging.info("=" * 80)

    for doc in docs:
        logging.info(
            f"{doc.get('document_identity')}  -->  "
            f"{doc.get('document_name')}"
        )

    logging.info("=" * 80)

    return docs


def find_document(documents, identity):
    for document in documents:
        if normalize_text(document.get("document_identity")) == normalize_text(identity):
            return document
    return None


def get_edo_document(
    client,
    product_family,
    product,
    templatename,
    db: DatabaseHandler
):
    """
    CANONICAL version - from edo_existing_final.py.
    Still resolves EDO_Proposed, EDO_RA_C and EDO_FMEA. Raises if
    EDO_Proposed (required by every stage) is missing.
    """
    documents = get_template_documents(
        client,
        product_family,
        product,
        templatename,
        db
    )

    if not documents:
        raise Exception("No EDO template documents found.")

    edo_documents = {}

    for document in documents:
        identity = normalize_text(document.get("document_identity")).lower()
        logging.info(f"AVAILABLE DOCUMENT : {identity}")

        if identity == "edo_proposed":
            edo_documents["edo_proposed"] = document
        elif identity in ["edo_ra_c", "edo_ra&c", "edo_rac", "edo_ra"]:
            edo_documents["edo_ra_c"] = document
        elif identity in ["edo_fmea", "fmea", "system_fmea"]:
            edo_documents["edo_fmea"] = document
        elif identity in ["edo_pdf_new", "edo_pdf-new", "edo pdf new"]:
            # Same db-driven resolution pattern as edo_proposed (docx),
            # just for the PDF source used to pull New EDO diagrams -
            # see get_edo_pdf_new_file() / extract_edo_pdf_new_diagram().
            edo_documents["edo_pdf_new"] = document

    logging.info("=" * 80)
    logging.info("EDO DOCUMENT CONFIGURATION")
    logging.info("=" * 80)

    for key, doc in edo_documents.items():
        logging.info(f"{key}")
        logging.info(f"Identity   : {doc.get('document_identity')}")
        logging.info(f"Name       : {doc.get('document_name')}")
        logging.info(f"Collection : {doc.get('collection')}")

    logging.info("=" * 80)

    if "edo_proposed" not in edo_documents:
        raise Exception("Required document EDO_Proposed was not found.")

    if "edo_pdf_new" not in edo_documents:
        # Not fatal - New EDO diagram insertion is best-effort, same as
        # image extraction from edo_proposed. Logged so it's obvious in
        # the run log why Column H comes back with no diagrams.
        logging.warning(
            "Document with identity 'EDO_pdf_new' was not found in the "
            "template's configured documents - New EDO diagram lookup "
            "(Column H) will be skipped for this run."
        )

    return edo_documents


def get_edo_document_lenient(
    client,
    product_family,
    product,
    templatename,
    db: DatabaseHandler
):
    """
    Original version from generate_EDO_template_copy.py - kept for
    backward compatibility. Does NOT raise if EDO_Proposed is missing.
    Not used by the merged pipeline (get_edo_document is used instead).
    """
    documents = get_template_documents(
        client,
        product_family,
        product,
        templatename,
        db
    )

    if not documents:
        raise Exception("No EDO template documents found.")

    edo_documents = {}

    for document in documents:
        identity = normalize_text(document.get("document_identity")).lower()
        logging.info(f"AVAILABLE DOCUMENT : {identity}")

        if identity == "edo_proposed":
            edo_documents["edo_proposed"] = document
        elif identity in ["edo_ra_c", "edo_ra&c", "edo_rac", "edo_ra"]:
            edo_documents["edo_ra_c"] = document
        elif identity in ["edo_fmea", "fmea", "system_fmea"]:
            edo_documents["edo_fmea"] = document
        elif identity in ["edo_pdf_new", "edo_pdf-new", "edo pdf new"]:
            edo_documents["edo_pdf_new"] = document

    logging.info("=" * 80)
    logging.info("EDO DOCUMENT CONFIGURATION")
    logging.info("=" * 80)

    for key, doc in edo_documents.items():
        logging.info(f"{key}")
        logging.info(f"Identity   : {doc.get('document_identity')}")
        logging.info(f"Name       : {doc.get('document_name')}")
        logging.info(f"Collection : {doc.get('collection')}")

    logging.info("=" * 80)

    return edo_documents


def get_prompt(
    client,
    product_family,
    product,
    templatename,
    prompt_name,
    db: DatabaseHandler
):
    prompt = db.get_prompt_by_name(
        client,
        product_family,
        product,
        templatename,
        prompt_name
    )

    if not prompt:
        raise Exception(f"Prompt '{prompt_name}' was not found.")

    logging.info(f"Loaded Prompt : {prompt_name}")
    return prompt


def build_prompt_row(
    prompt,
    question,
    fulltext="Yes",
    where_filter="",
    where_document="",
    checkpoint=""
):
    return {
        "prompt_role": prompt["prompt_role"],
        "prompt_text": prompt["prompt_text"],
        "question": question,
        "fulltext": fulltext,
        "where_filter": where_filter,
        "where_document": where_document,
        "checkpoint": checkpoint
    }


def execute_prompt(
    pipeline_config,
    collection,
    prompt
):
    _, _, response = execute_llm_retry(
        pipeline_config,
        collection,
        prompt
    )

    logging.info("=" * 80)
    logging.info("RAW LLM RESPONSE")
    logging.info("=" * 80)
    logging.info(response)

    return parse_json(response)


# ==========================================================
# UNCONDITIONAL RECORD DEEP EXTRACTION
# ==========================================================

def deep_extract_records(data):
    """
    Recursive lookup utility to unwrap fluctuating JSON parent containers.
    Merged version: union of the container keys recognised by both source
    files (edo_existing_final.py's set plus generate_EDO_template_copy.py's
    extra "New_EDOs" / "New_Edos" / "new_edos" keys), so this single
    function correctly unwraps both existing-EDO and new-EDO responses.
    """
    if isinstance(data, list):
        return [item for item in data if isinstance(item, dict)]

    if isinstance(data, dict):
        for key in [
            "Records", "records",
            "New_EDOs", "New_Edos", "new_edos",
            "EDO_Table", "EDO_Tag_Values", "Verification_Details"
        ]:
            if key in data and isinstance(data[key], list):
                return [item for item in data[key] if isinstance(item, dict)]

        if any(k in data for k in ["Product_Feature_Function", "RA_Number", "FMEA_Number", "Traceability", "System DFMEA #", "Trace To RAC#"]):
            return [data]

        for val in data.values():
            if isinstance(val, list) and len(val) > 0 and isinstance(val[0], dict):
                return val
            elif isinstance(val, dict):
                res = deep_extract_records(val)
                if res:
                    return res
    return []


# ==========================================================
# STAGE 3: EXISTING EDO RECORD WORKFLOW
# ==========================================================

def extract_edo_tags(
    client,
    product_family,
    product,
    templatename,
    pipeline_config,
    edo_document,
    db: DatabaseHandler
):
    """
    CANONICAL - requested function #1, taken from edo_existing_final.py.
    """
    logging.info("=" * 80)
    logging.info("STAGE 3: EXTRACTING EXISTING EDO TAGS")
    logging.info("=" * 80)

    prompt = get_prompt(
        client,
        product_family,
        product,
        templatename,
        "EDO_Existing_Tag",
        db
    )

    prompt_row = build_prompt_row(
        prompt,
        "Extract all Known Active Design Output Tracking Numbers."
    )

    result = execute_prompt(
        pipeline_config,
        edo_document["edo_proposed"]["collection"],
        prompt_row
    )

    items = deep_extract_records(result)
    logging.info(f"Unconditional Record Extraction count: {len(items)}")

    tags = {}
    for item in items:
        if not isinstance(item, dict):
            continue

        tag = (
            item.get("edo_number")
            or item.get("edo_tag")
            or item.get("EDO Number")
            or item.get("EDO_Tag")
            or item.get("EDO")
            or ""
        )
        tag = normalize_text(tag)

        if tag == "" or tag.lower() == "blank" or tag in tags:
            continue

        tags[tag] = create_empty_edo()
        tags[tag]["edo_type"] = "Existing"
        tags[tag]["edo_tag"] = tag

    return tags


def validate_existing_tags(tags):
    validated = {}
    for key, value in tags.items():
        tag = normalize_text(value.get("edo_tag"))
        if tag == "" or tag.lower() == "blank":
            continue
        validated[tag] = value
    return validated


def extract_ra_fmea_from_text(text):
    """
    Pulls RA_Number / FMEA_Number back out of the column D (dfmea) narrative
    that was already extracted for an existing EDO in extract_edo_details().
    This is the "previous edo dictionary" data used for matching in Stage
    3B - reused instead of asking the LLM for tags a second time.

    The RA&C / FMEA documents identify records with bare tokens like
    "RA-180" and "SYS-147" (sometimes written "FMEA SYS-147") - there is
    no literal word "Number" next to them, so we match the token pattern
    directly rather than looking for a "RA Number:" label.
    """
    text = normalize_text(text)

    ra_match = re.search(r"RA[\s-]?(\d+)", text, re.IGNORECASE)
    fmea_match = re.search(r"SYS[\s-]?(\d+)", text, re.IGNORECASE)

    ra_number = f"RA-{ra_match.group(1)}" if ra_match else ""
    fmea_number = f"SYS-{fmea_match.group(1)}" if fmea_match else ""

    return ra_number, fmea_number


def normalize_id(value):
    """
    Canonicalizes an RA/FMEA identifier for comparison purposes:
    uppercase, single spaces, trimmed. Used so that matching between the
    column-D-derived identifiers and whatever format Stage 3B's
    verification prompt returns doesn't fail on case/whitespace noise.
    """
    value = normalize_text(value).upper()
    return re.sub(r"\s+", " ", value).strip()


def extract_edo_details(
    client,
    product_family,
    product,
    templatename,
    pipeline_config,
    edo_document,
    existing_edos,
    db: DatabaseHandler
):
    """
    CANONICAL version - from edo_existing_final.py.
    Comprehensive attribute hydration: extracts description/reason lists,
    dfmea narrative, location, sysdd, and (crucially) RA_Number /
    FMEA_Number - needed downstream by Stage 3B verification matching.

    UPDATED to match the EDO_Existing_Generic prompt's nested JSON shape
    (EDO_Table -> existing_edo_data[] -> design_elements[]):
      - Tag matching now also recognises "edo_no" - the key this prompt
        actually returns. The old code only checked edo_number/edo_tag/
        "EDO Number"/"EDO_Tag", so `result` was NEVER populated for this
        prompt's output and every field silently stayed blank - this is
        why nothing was printing in Excel.
      - "design_elements" (a LIST of per-location dicts) is walked and
        normalized into {"location", "description", "reason", "sysdd",
        "image"} entries, stored as a list on edo["design_elements"] for
        any downstream code (Excel writer / image placement) that wants
        the full per-location breakdown.
      - The single-value fields the rest of the pipeline reads today
        (edo_description / reason_identified / location / sysdd /
        description_2 / reason_2) are populated from the new top-level
        fields, backfilled from the design_elements list where the old
        schema had no top-level equivalent (e.g. location/sysdd only
        ever existed per design element, never at the top level).
    """
    logging.info("=" * 80)
    logging.info("STAGE 3: ATTRIBUTE HYDRATION (EXISTING EDO)")
    logging.info("=" * 80)

    prompt = EDO_EXISTING_GENERIC_PROMPT

    def normalize_key(key):
        key = normalize_text(key).lower()
        key = key.replace("&", "and")
        key = re.sub(r"[^a-z0-9]+", "_", key)
        key = re.sub(r"_+", "_", key)
        return key.strip("_")

    def first_value(record, keywords):
        """
        Return first non-empty value whose normalized key contains
        any keyword as a substring.
        """
        normalized_record = {normalize_key(k): v for k, v in record.items()}
        for kw in keywords:
            for nk, v in normalized_record.items():
                if kw in nk or nk in kw:
                    if normalize_text(v):
                        return normalize_text(v)
        return ""

    for edo_tag in existing_edos.keys():

        logging.info(f"Hydrating {edo_tag}")

        question = (
            f"Extract the complete row details for the EDO tag requested by the user. Requested EDO Tag: {edo_tag}  Find the row where the EDO Tag exactly matches the requested EDO Ta. Return all available column values for that row, including EDO number/tag, description, reason identified as EDO, RA&C and/or Sys-DFMEA Trace,EDO Location, EDO Description, reason identified as EDO  If the requested EDO tag is not found, return none"
        )

        prompt_row = build_prompt_row(
            prompt,
            question
        )

        result = {}

        # 1. Fetch data
        for attempt in range(3):
            raw_result = execute_prompt(
                pipeline_config,
                edo_document["edo_proposed"]["collection"],
                prompt_row
            )
            records = deep_extract_records(raw_result)

            for record in records:
                tag = normalize_text(
                    record.get("edo_no")
                    or record.get("edo_number")
                    or record.get("edo_tag")
                    or record.get("EDO Number")
                    or record.get("EDO_Tag")
                ).lower()

                if tag == edo_tag.lower():
                    result = record
                    break

            # 2. Debug log
            logging.info(f"DEBUG: Hydration data for {edo_tag}: {json.dumps(result, indent=2)}")

            # Always accept whatever we found to avoid "Blank"
            if result:
                break
            time.sleep(INITIAL_RETRY_DELAY)

        normalized = {normalize_key(k): v for k, v in result.items()}
        edo = existing_edos[edo_tag]

        # ---- Design elements (new nested structure) ----
        # "design_elements" is a LIST of per-location dicts under the new
        # EDO_Existing_Generic prompt shape - normalize every entry so
        # downstream code has a clean, predictable structure to work
        # from, instead of only ever seeing one location's worth of data.
        raw_design_elements = result.get("design_elements") or result.get("Design_Elements") or []
        if not isinstance(raw_design_elements, list):
            raw_design_elements = []

        design_elements = []
        for element in raw_design_elements:
            if not isinstance(element, dict):
                continue
            design_elements.append({
                "location": blank(
                    element.get("edo_location") or element.get("EDO_Location") or element.get("location")
                ),
                "description": blank(
                    element.get("edo_description") or element.get("EDO_Description") or element.get("description")
                ),
                "reason": blank(
                    element.get("reason identified as edo")
                    or element.get("reason_identified_as_edo")
                    or element.get("reason")
                ),
                "sysdd": blank(
                    element.get("SysDD or HDD Reference") or element.get("sysdd") or element.get("SysDD")
                ),
                "image": blank(element.get("image")),
            })

        edo["design_elements"] = design_elements

        # ---- Top-level description / reason ----
        # The new prompt returns ONE top-level "edo_description" /
        # "reason identified as edo" pair for the EDO itself - fall back
        # to the first design element only if the top-level fields are
        # somehow missing.
        top_description = first_value(result, ["edo_description"])
        top_reason = first_value(result, ["reason_identified_as_edo", "reason"])

        edo["edo_description"] = top_description or (design_elements[0]["description"] if design_elements else "")
        edo["reason_identified"] = top_reason or (design_elements[0]["reason"] if design_elements else "")

        # description_2 / reason_2 - backfilled from the SECOND design
        # element (if any), for backward compatibility with any code
        # still reading these two single-value fields.
        edo["description_2"] = design_elements[1]["description"] if len(design_elements) > 1 else ""
        edo["reason_2"] = design_elements[1]["reason"] if len(design_elements) > 1 else ""

        # ---- RA&C / Sys-DFMEA trace (Column D) ----
        # "RA and FMEA no" holds the real RA&C/System FMEA reference text
        # under the new schema - "Trace" is usually just "Blank"/empty,
        # so prefer the former and only fall back to the latter.
        ra_and_fmea = first_value(result, ["ra_and_fmea_no", "ra_and_fmea", "ra_and"])
        trace_narrative = first_value(result, ["trace"])
        edo["dfmea"] = ra_and_fmea or trace_narrative

        # ---- Location / SysDD ----
        # Neither ever existed at the top level under the new schema -
        # both only ever live per design element - so pull them from the
        # first design element.
        edo["location"] = design_elements[0]["location"] if design_elements else first_value(result, ["location"])
        edo["sysdd"] = design_elements[0]["sysdd"] if design_elements else first_value(result, ["sysdd", "hardware", "design_reference"])

        ra_number = first_value(result, ["ra_number"])
        fmea_number = first_value(result, ["fmea_number"])

        if not ra_number or not fmea_number:
            p_ra, p_fmea = extract_ra_fmea_from_text(edo["dfmea"])
            ra_number = ra_number or p_ra
            fmea_number = fmea_number or p_fmea

        edo["ra_number"] = ra_number or ""
        edo["FMEA_Number"] = fmea_number or ""
        edo["verification_reference"] = ""

    return existing_edos


def extract_edo_details_basic(
    client,
    product_family,
    product,
    templatename,
    pipeline_config,
    edo_document,
    existing_edos,
    db: DatabaseHandler
):
    """
    Original simpler version from generate_EDO_template_copy.py - kept for
    backward compatibility. Not used by the merged pipeline
    (extract_edo_details is used instead, since it also derives
    RA_Number/FMEA_Number needed by Stage 3B).
    """
    logging.info("=" * 80)
    logging.info("STAGE 3: WORKFLOW 1 - ATTRIBUTE HYDRATION (EXISTING) [basic]")
    logging.info("=" * 80)

    prompt = get_prompt(
        client,
        product_family,
        product,
        templatename,
        "EDO_Existing_Generic",
        db
    )

    for edo_tag in existing_edos.keys():
        question = f"Extract descriptions, reason matrices, traces, locations, and references for: {edo_tag}"
        prompt_row = build_prompt_row(prompt, question)
        result = execute_prompt(
            pipeline_config,
            edo_document["edo_proposed"]["collection"],
            prompt_row
        )

        if isinstance(result, list):
            result = {} if len(result) == 0 else result[0]

        edo = existing_edos[edo_tag]
        edo["edo_description"] = blank(result.get("edo_description") or result.get("EDO Description") or result.get("description"))
        edo["reason_identified"] = blank(result.get("reason_identified") or result.get("Reason Identified") or result.get("reason"))
        edo["dfmea"] = blank(result.get("dfmea") or result.get("DFMEA") or result.get("RA&C and/or Sys-DFMEA Trace"))
        edo["verification_reference"] = "Blank"
        edo["location"] = blank(result.get("location") or result.get("EDO Location") or result.get("Location"))
        edo["description_2"] = blank(result.get("description_2") or result.get("EDO Description 2") or result.get("Description 2"))
        edo["reason_2"] = blank(result.get("reason_2") or result.get("Reason 2"))
        edo["sysdd"] = blank(result.get("sysdd") or result.get("SYSDD") or result.get("SYS DD") or result.get("sysdd_reference"))

    return existing_edos


# ==========================================================
# STAGE 3B: EXISTING EDO VERIFICATION REFERENCE (COLUMN E)
# ==========================================================

def extract_existing_edo_verification_details(
    client,
    product_family,
    product,
    templatename,
    pipeline_config,
    edo_document,
    existing_edos,
    db: DatabaseHandler
):
    """
    CANONICAL - requested function #2, taken from edo_existing_final.py.
    Dedicated verification-reference extraction for EXISTING EDOs: its own
    prompt ("EDO_NEW_Verification_details"), queried against the
    edo_fmea collection, matched by RA/FMEA number that was pulled from
    the column D (dfmea) data in extract_edo_details().
    """
    logging.info("=" * 80)
    logging.info("STAGE 3B: EXISTING EDO - VERIFICATION REFERENCE EXTRACTION")
    logging.info("=" * 80)

    if "edo_fmea" not in edo_document:
        raise Exception(
            "EDO_FMEA document/collection not configured - cannot run "
            "verification reference extraction."
        )

    prompt_data = get_prompt(
        client,
        product_family,
        product,
        templatename,
        "EDO_NEW_Verification_details",
        db
    )

    targets = "\n".join(
        f"RA_Number : {edo.get('ra_number')}\nFMEA_Number : {edo.get('FMEA_Number')}"
        for edo in existing_edos.values()
        if edo.get("ra_number") not in (None, "", "Blank") or edo.get("FMEA_Number") not in (None, "", "Blank")
    )

    if not targets:
        logging.warning(
            "No RA/FMEA numbers available on any existing EDO - skipping "
            "verification reference extraction."
        )
        return {}

    prompt_row = {
        "prompt_role": prompt_data["prompt_role"],
        "prompt_text": prompt_data["prompt_text"] + "\nTARGETS:\n" + targets,
        "question": "Fetch verification reference records for the folowing sys number <FMEA_Number> number",
        "fulltext": "Yes",
        "where_filter": "",
        "where_document": "",
        "checkpoint": ""
    }

    _, _, response = execute_llm_retry(
        pipeline_config,
        edo_document["edo_fmea"]["collection"],
        prompt_row
    )

    return parse_json(response)


def apply_existing_edo_verification(existing_edos, verification_details):
    """
    CANONICAL - requested function #3, taken from edo_existing_final.py.
    Couples the verification records back onto existing_edos by RA/FMEA
    number. Matching is format-tolerant (case/whitespace normalized, plus
    substring containment).
    """
    records = deep_extract_records(verification_details)

    for row in records:
        if not isinstance(row, dict):
            continue

        trace_fmea = normalize_id(row.get("FMEA_Number") or row.get("FMEA Number") or "")
        trace_ra = normalize_id(row.get("RA_Number") or row.get("RA Number") or "")

        verification = normalize_text(
            row.get("Verification_Reference")
            or row.get("Verification Reference")
            or row.get("verification_reference")
        )
        if not verification or verification.lower() == "none":
            continue

        for edo in existing_edos.values():
            edo_fmea = normalize_id(edo.get("FMEA_Number") or "")
            edo_ra = normalize_id(edo.get("ra_number") or "")

            fmea_match = trace_fmea and edo_fmea and (trace_fmea == edo_fmea or trace_fmea in edo_fmea or edo_fmea in trace_fmea)
            ra_match = trace_ra and edo_ra and (trace_ra == edo_ra or trace_ra in edo_ra or edo_ra in trace_ra)

            if fmea_match or ra_match:
                edo["verification_reference"] = verification

    return existing_edos


def extract_existing_edo_trace_details(
    client,
    product_family,
    product,
    templatename,
    pipeline_config,
    edo_document,
    existing_edos,
    db: DatabaseHandler
):
    """
    ADDED - dedicated trace extraction for EXISTING EDOs, mirroring
    extract_existing_edo_verification_details() exactly: its own prompt
    ("EDO_Existing_Trace"), queried against the edo_fmea collection,
    matched by RA/FMEA number that was pulled from the column D (dfmea)
    data in extract_edo_details(). The result is written to Column D,
    below the existing RA/FMEA data, in the same row (see
    apply_existing_edo_trace() and the Column D value construction in
    write_edo_excel_existing_style() / format_edo_worksheet()).
    """
    logging.info("=" * 80)
    logging.info("STAGE 3C: EXISTING EDO - TRACE EXTRACTION")
    logging.info("=" * 80)

    if "edo_fmea" not in edo_document:
        raise Exception(
            "EDO_FMEA document/collection not configured - cannot run "
            "trace extraction."
        )

    prompt_data = EDO_EXISTING_TRACE_PROMPT

    targets = "\n".join(
        f"RA_Number : {edo.get('ra_number')}\nFMEA_Number : {edo.get('FMEA_Number')}"
        for edo in existing_edos.values()
        if edo.get("ra_number") not in (None, "", "Blank") or edo.get("FMEA_Number") not in (None, "", "Blank")
    )

    if not targets:
        logging.warning(
            "No RA/FMEA numbers available on any existing EDO - skipping "
            "trace extraction."
        )
        return {}

    prompt_row = {
        "prompt_role": prompt_data["prompt_role"],
        "prompt_text": prompt_data["prompt_text"] + "\nTARGETS:\n" + targets,
        "question": "Fetch trace records for the folowing sys number <FMEA_Number> number",
        "fulltext": "Yes",
        "where_filter": "",
        "where_document": "",
        "checkpoint": ""
    }

    _, _, response = execute_llm_retry(
        pipeline_config,
        edo_document["edo_fmea"]["collection"],
        prompt_row
    )

    parsed = parse_json(response)
    if deep_extract_records(parsed):
        return parsed

    code_block_records = extract_json_code_blocks(response)
    if code_block_records:
        logging.warning(
            "extract_existing_edo_trace_details: LLM response wasn't a "
            f"single valid JSON document - recovered {len(code_block_records)} "
            "record(s) from individual ```json fenced blocks."
        )
        return code_block_records

    markdown_records = parse_markdown_trace_response(response)
    if markdown_records:
        logging.warning(
            "extract_existing_edo_trace_details: LLM response wasn't "
            "valid JSON (or contained no records) - recovered "
            f"{len(markdown_records)} record(s) via Markdown fallback parser."
        )
        return markdown_records

    return parsed


def extract_json_code_blocks(text):
    """
    Fallback parser for extract_existing_edo_trace_details(). The LLM
    sometimes wraps EACH FMEA number's answer in its own separate
    ```json ... ``` fenced code block (e.g. one block per
    "### **For FMEA_Number: SYS-XXX**" section) instead of returning one
    single JSON document for the whole response. A single json.loads()
    call on the whole response fails in that case, because multiple
    top-level JSON objects side-by-side with Markdown headers/separators
    in between them isn't valid JSON as a whole - even though each
    individual fenced block IS perfectly valid JSON on its own. This
    extracts every ```json fenced block and parses each independently,
    returning the combined list of records.
    """
    text = normalize_text(text)
    blocks = re.findall(r"```(?:json)?\s*(.*?)```", text, re.DOTALL | re.IGNORECASE)

    records = []
    for block in blocks:
        try:
            parsed_block = json.loads(block.strip())
        except Exception:
            continue

        if isinstance(parsed_block, list):
            records.extend(item for item in parsed_block if isinstance(item, dict))
        elif isinstance(parsed_block, dict):
            records.append(parsed_block)

    return records


def parse_markdown_trace_response(response):
    """
    Fallback parser for extract_existing_edo_trace_details(). The
    EDO_Existing_Trace prompt is supposed to return JSON, but the LLM
    sometimes replies with Markdown instead - one "### **<FMEA id>**"
    block per FMEA number, with **bold** labels and "-" bullet lists,
    e.g.:

        ### **FMEA Sys-147**
        **Trace To RAC#:** RA-180
        **Traces to Module DFMEA risk controls:**
        - **Empty Array** (No matching entries found)
        **Additional Risk Control Measures from Module and/or Component DFMEAs:**
        - **Empty Array** (No matching identifiers found)
        **DRS Identifiers:**
        - **DRS-570**

    parse_json() can't recover this (there is no {}/[] JSON anywhere in
    it). This converts each block into the same record shape
    apply_existing_edo_trace() expects from a proper JSON response:
    {"System DFMEA #", "Trace To RAC#", "Traces to Module DFMEA risk
    controls", "Additional Risk Control Measures from Module and/or
    Component DFMEAs", "DRS Identifiers"}.
    """
    text = normalize_text(response)

    def extract_bullets(section_text):
        items = []
        for line in section_text.splitlines():
            line = line.strip()
            if not line.startswith("-"):
                continue
            line = line.lstrip("-").strip()
            line = re.sub(r"^\*+|\*+$", "", line).strip()
            if not line or "empty array" in line.lower() or line.lower() in ("none", "n/a"):
                continue
            items.append(line)
        return items

    def extract_section(block, label, next_labels):
        stop_pattern = "|".join(re.escape(nl) for nl in next_labels) if next_labels else None
        pattern = re.escape(label) + r"[:\*\s]*\n?(.*?)" + (
            f"(?=\\*\\*(?:{stop_pattern})|$)" if stop_pattern else "$"
        )
        match = re.search(pattern, block, re.DOTALL | re.IGNORECASE)
        return match.group(1) if match else ""

    records = []

    # Split on "### **<header>**" markers into (header, body) pairs.
    blocks = re.split(r"^#{1,4}\s*\*\*(.+?)\*\*\s*$", text, flags=re.MULTILINE)

    if len(blocks) > 1:
        for i in range(1, len(blocks), 2):
            header = normalize_text(blocks[i])
            body = blocks[i + 1] if i + 1 < len(blocks) else ""

            ra_match = re.search(r"Trace To RAC#\s*[:\*]*\s*([^\n]+)", body, re.IGNORECASE)
            ra_number = re.sub(r"\*+", "", normalize_text(ra_match.group(1))).strip() if ra_match else ""

            module_section = extract_section(
                body, "Traces to Module DFMEA risk controls",
                ["Additional Risk Control Measures", "DRS Identifiers"]
            )
            additional_section = extract_section(
                body, "Additional Risk Control Measures from Module and/or Component DFMEAs",
                ["DRS Identifiers"]
            )
            drs_section = extract_section(body, "DRS Identifiers", [])

            records.append({
                "System DFMEA #": header,
                "Trace To RAC#": ra_number,
                "Traces to Module DFMEA risk controls": extract_bullets(module_section),
                "Additional Risk Control Measures from Module and/or Component DFMEAs": extract_bullets(additional_section),
                "DRS Identifiers": extract_bullets(drs_section),
            })

    return records


def apply_existing_edo_trace(existing_edos, trace_details):
    """
    ADDED - couples the trace records back onto existing_edos by RA/FMEA
    number, mirroring apply_existing_edo_verification() exactly. Matching
    is format-tolerant (case/whitespace normalized, plus substring
    containment). Populates edo["existing_trace"], which is then appended
    below the RA/FMEA data already shown in Column D (dfmea), same row -
    see write_edo_excel_existing_style() / format_edo_worksheet().
    """
    records = deep_extract_records(trace_details)
    logging.info(f"apply_existing_edo_trace: {len(records)} raw trace record(s) extracted from LLM response.")

    matched_count = 0

    def build_trace_text(row):
        """
        Builds a readable trace string from the EDO_Existing_Trace prompt's
        actual response shape:
          - "Traces to Module DFMEA risk controls": list of
            {document_number, tag_number, cause}
          - "Additional Risk Control Measures from Module and/or Component
            DFMEAs": list of identifier strings
          - "DRS Identifiers": list of identifier strings
        Falls back to the older plain-string key names in case the prompt
        output format varies.
        """
        parts = []

        module_controls = row.get("Traces to Module DFMEA risk controls")
        if isinstance(module_controls, list):
            for item in module_controls:
                if isinstance(item, dict):
                    doc = normalize_text(item.get("document_number"))
                    tag = normalize_text(item.get("tag_number"))
                    cause = normalize_text(item.get("cause"))
                    line = " ".join(p for p in [doc, tag] if p)
                    if cause:
                        line = f"{line} - {cause}" if line else cause
                    if line:
                        parts.append(line)
                elif normalize_text(item):
                    parts.append(normalize_text(item))

        additional_measures = row.get("Additional Risk Control Measures from Module and/or Component DFMEAs")
        if isinstance(additional_measures, list):
            parts.extend(normalize_text(m) for m in additional_measures if normalize_text(m))

        drs_identifiers = row.get("DRS Identifiers")
        if isinstance(drs_identifiers, list):
            parts.extend(normalize_text(d) for d in drs_identifiers if normalize_text(d))

        if parts:
            return "\n".join(parts)

        # Fallback - older plain-string key names, kept for compatibility.
        return normalize_text(
            row.get("Trace")
            or row.get("Trace_Reference")
            or row.get("Trace Reference")
            or row.get("trace_reference")
            or row.get("RA_and_FMEA_Trace")
            or row.get("Traceability")
            or row.get("Trace_Text")
            or row.get("Trace Text")
            or row.get("Trace_Details")
            or row.get("Trace Details")
            or row.get("Trace_Value")
            or row.get("trace")
            or row.get("RA_FMEA_Trace")
            or row.get("Sys_DFMEA_Trace")
            or row.get("DFMEA_Trace")
        )

    for row in records:
        if not isinstance(row, dict):
            continue

        trace_fmea = normalize_id(
            row.get("System DFMEA #")
            or row.get("FMEA_Number")
            or row.get("FMEA Number")
            or ""
        )
        trace_ra = normalize_id(
            row.get("Trace To RAC#")
            or row.get("RA_Number")
            or row.get("RA Number")
            or ""
        )

        trace_value = build_trace_text(row)
        if not trace_value or trace_value.lower() == "none":
            # DIAGNOSTIC: if this fires for every record, the LLM's JSON key
            # name for the trace text doesn't match any key checked above -
            # this log line shows the actual keys so the list can be updated.
            logging.warning(
                f"apply_existing_edo_trace: no recognized trace-text key on "
                f"record RA={trace_ra!r} FMEA={trace_fmea!r}. Raw keys present: "
                f"{list(row.keys())}"
            )
            continue

        row_matched = False
        for edo in existing_edos.values():
            edo_fmea = normalize_id(edo.get("FMEA_Number") or "")
            edo_ra = normalize_id(edo.get("ra_number") or "")

            fmea_match = trace_fmea and edo_fmea and (trace_fmea == edo_fmea or trace_fmea in edo_fmea or edo_fmea in trace_fmea)
            ra_match = trace_ra and edo_ra and (trace_ra == edo_ra or trace_ra in edo_ra or edo_ra in trace_ra)

            if fmea_match or ra_match:
                edo["existing_trace"] = trace_value
                row_matched = True
                matched_count += 1

        if not row_matched:
            # DIAGNOSTIC: trace text was found, but couldn't be matched to
            # any existing EDO's RA/FMEA number - shows the mismatch so the
            # normalize_id() comparison can be adjusted if needed.
            logging.warning(
                f"apply_existing_edo_trace: trace text found for "
                f"RA={trace_ra!r} FMEA={trace_fmea!r} but it matched no "
                f"existing EDO (available RA/FMEA on existing_edos: "
                f"{[(e.get('ra_number'), e.get('FMEA_Number')) for e in existing_edos.values()]})."
            )

    logging.info(f"apply_existing_edo_trace: populated existing_trace on {matched_count} existing-EDO match(es).")

    return existing_edos


# ==========================================================
# STAGE 4: NEW EDO RECORD WORKFLOW
# ==========================================================
def extract_new_edo_tags(
    client,
    product_family,
    product,
    templatename,
    pipeline_config,
    edo_document,
    db
):
    """
    STEP 1 of 3: scans Section 10 of the FMEA against the RA&C and
    identifies every qualifying new-EDO row's {RA_Number, FMEA_Number,
    Status}.

    Per requirement, every RA id found is consolidated into ONE
    dictionary - edo_new_data - keyed by RA_Number (falling back to
    FMEA_Number, then to a positional key, only when RA_Number itself
    is missing). This SAME dictionary is what gets passed into and
    progressively enriched by extract_new_edo_summary_details() and
    extract_new_edo_traceability_details() afterwards, so every
    RA id's data lives in one place from identification all the way
    through to the final merge.
    """
    logging.info("=" * 80)
    logging.info("STAGE 4a: WORKFLOW 2 - NEW EDO IDENTIFICATION (TAGS ONLY)")
    logging.info("=" * 80)

    prompt_data = db.get_prompt_by_name(
        client,
        product_family,
        product,
        templatename,
        "EDO_NEW_tags"
    )

    prompt_row = {
        "prompt_role": prompt_data["prompt_role"],
        "prompt_text": prompt_data["prompt_text"],  
        "question": (
    'Extract all records only from the actual Risk Assessment and Control Table under Section 12 and its subsections where Column F (Risk Evaluation) contains exactly "See FMEA". For each matching row, extract the corresponding values from Column A (RA-#), Column F (Risk Evaluation), and the FMEA reference from Column G (Trace(s) to other Risk Documents and Tag #).'
),

        "fulltext": "Yes",
        "where_filter": "",
        "where_document": "",
        "checkpoint": ""
    }

    _, _, response = execute_llm_retry(
        pipeline_config,
        edo_document["edo_ra_c"]["collection"],
        prompt_row
    )

    tags = parse_json(response)
    tag_records = deep_extract_records(tags)

    logging.info(f"extract_new_edo_tags: {len(tag_records)} raw tag row(s) returned by the LLM.")

    # ---- Consolidate every RA id found into ONE dictionary ----
    edo_new_data = {}

    for index, row in enumerate(tag_records):
        if not isinstance(row, dict):
            continue

        ra_number = normalize_text(row.get("RA_Number") or row.get("RA Number") or "")
        fmea_number = normalize_text(row.get("FMEA_Number") or row.get("FMEA Number") or "")
        status = normalize_text(row.get("Status") or row.get("status") or "")

        key = ra_number or fmea_number or f"NEW-EDO-TAG-{index}"
        if key in edo_new_data:
            key = f"{key}_{index}"

        edo_new_data[key] = {
            "RA_Number": ra_number,
            "FMEA_Number": fmea_number,
            "Status": status
        }

    logging.info(
        f"extract_new_edo_tags: consolidated {len(edo_new_data)} RA id(s) "
        f"into edo_new_data."
    )

    return edo_new_data


def extract_new_edo_summary_details(
    client,
    product_family,
    product,
    templatename,
    pipeline_config,
    edo_document,
    edo_new_data,
    db
):
    """
    STEP 2 of 3: takes the SAME edo_new_data dictionary built by
    extract_new_edo_tags() (keyed by RA id) and, for EVERY entry in it,
    calls the LLM ONE TIME - inside the loop, one RA_Number/FMEA_Number
    pair per call - to pull that single row's full detail record
    (Product_Feature_Function, Reason_Identified_as_EDO, Traceability,
    Verification_Reference, EDO_Location, EDO_Description,
    Reason_Identified_as_EDO_ColH).

    Each result is written straight back into edo_new_data[key], so the
    dictionary ends up holding the tag info AND the full detail info
    for every RA id in one place.
    """
    logging.info("=" * 80)
    logging.info("STAGE 4b: WORKFLOW 2 - NEW EDO FULL DETAIL EXTRACTION (PER RA/FMEA, LOOPED)")
    logging.info("=" * 80)

    if not edo_new_data:
        logging.info("No new EDO tags to extract details for.")
        return edo_new_data

    prompt_data = db.get_prompt_by_name(
        client,
        product_family,
        product,
        templatename,
        "EDO_NEW_details"
    )

    for key, entry in edo_new_data.items():
        ra_number = entry.get("RA_Number", "")
        fmea_number = entry.get("FMEA_Number", "")

        target_text = f"RA_Number : {ra_number}\nFMEA_Number : {fmea_number}"

        prompt_row = {
            "prompt_role": prompt_data["prompt_role"],
            "prompt_text": prompt_data["prompt_text"] + "\nTARGETS:\n" + target_text,
            "question": (
                f"For RA_Number {ra_number} / FMEA_Number {fmea_number} "
                "only, extract the full EDO detail record as a single "
                "JSON object - no other RA/FMEA pairs."
            ),
            "fulltext": "Yes",
            "where_filter": "",
            "where_document": "",
            "checkpoint": ""
        }

        detail_row = {}
        try:
            _, _, response = execute_llm_retry(
                pipeline_config,
                edo_document["edo_fmea"]["collection"],
                prompt_row
            )
            parsed = parse_json(response)
            records = deep_extract_records(parsed)
            if records:
                detail_row = records[0]
            elif isinstance(parsed, dict):
                detail_row = parsed

        except Exception as e:
            logging.error(
                f"LLM call failed for RA={ra_number!r} FMEA={fmea_number!r}: {e}"
            )
            detail_row = {}

        if not detail_row:
            logging.warning(
                f"No detail record returned for RA={ra_number!r} FMEA={fmea_number!r}."
            )

        entry["Product_Feature_Function"] = get_llm_value(
            detail_row, "Product_Feature_Function", "Product Feature Function"
        )
        entry["Reason_Identified_as_EDO"] = get_llm_value(
            detail_row, "Reason_Identified_as_EDO", "Reason Identified as EDO"
        )
        entry["Traceability"] = get_llm_value(
            detail_row, "Traceability", "traceability"
        )
        entry["Verification_Reference"] = get_llm_value(
            detail_row, "Verification_Reference", "Verification Reference"
        )
        entry["EDO_Location"] = get_llm_value(
            detail_row, "EDO_Location", "location"
        )
        entry["EDO_Description"] = get_llm_value(
            detail_row, "EDO_Description", "description"
        )
        entry["Reason_Identified_as_EDO_ColH"] = get_llm_value(
            detail_row, "Reason_Identified_as_EDO_ColH", "reason_2"
        )

    logging.info(
        f"extract_new_edo_summary_details: enriched {len(edo_new_data)} "
        f"entries in edo_new_data with full detail records."
    )

    return edo_new_data



# ==========================================================
# TRACE CODE / VALUE MATCHING (used by
# extract_new_edo_traceability_details) - these exist because the
# per-row LLM call doesn't always come back with the exact same JSON
# key names (e.g. one row's code sits under "DRS_Number", the next
# under "Trace_Number", the next isn't labelled at all) - matching by a
# fixed key list alone is what was causing rows to silently come back
# empty. Instead: (1) recognize the CODE itself by pattern - covers
# DRS, MS CU Mod, MS ACC Mod, SRS-CTRL, MRS CU/Software/ACC FMEA, RRAA,
# RA, FMEA Sys, regardless of which key it was returned under; (2) for
# filename/filenumber/location/status, fall back to a key-substring
# scan when the exact key names don't match.
# ==========================================================

TRACE_CODE_PATTERNS = [
    r'DRS[-\s]*\d+',
    r'MS\s*CU\s*Mod[-\s]*\d+',
    r'MS\s*ACC\s*Mod[-\s]*\d+',
    r'SRS-CTRL[-\s]*\d+',
    r'MRS\s*CU\s*FMEA[-\s]*\S+',
    r'MRS\s*Software\s*FMEA[-\s]*\S+',
    r'MRS\s*ACC\s*FMEA[-\s]*\S+',
    r'RRAA[-\s]?\S*',
    r'RA[-\s]*\d+',
    r'FMEA\s*Sys[-\s]*\d+',
]

# key-name -> type label, used ONLY when a value is a BARE number (e.g.
# "570") with no prefix of its own - the type then comes from whichever
# field it was returned under, e.g. {"DRS": 570} or
# {"DRS_Number": "570"} -> "DRS 570", not just "570". Order matters:
# more specific tokens (ms cu mod / ms acc mod / srs ctrl / mrs ... fmea)
# are checked before the plain "drs" fallback.
TRACE_KEY_TYPE_MAP = [
    (["mscumod"], "MS CU Mod"),
    (["msaccmod"], "MS ACC Mod"),
    (["srsctrl"], "SRS-CTRL"),
    (["mrscufmea"], "MRS CU FMEA"),
    (["mrssoftwarefmea"], "MRS Software FMEA"),
    (["mrsaccfmea"], "MRS ACC FMEA"),
    (["rraa"], "RRAA"),
    (["drs"], "DRS"),
]


def _find_trace_code(value):
    """Returns the first recognizable code (DRS 570, MS CU Mod 448,
    SRS-CTRL 39, MRS ACC FMEA-380, RRAA, RA-141, FMEA Sys-152, ...)
    found inside `value`, or "" if none. Accepts either "DRS-570" or
    "DRS 570" in the source text and normalizes the separator between
    the prefix and the number to a single space, so the output is
    always e.g. "DRS 570" - never just the bare "570". The code itself
    already shows which type/state it is - no extra label needed."""
    text = normalize_text(value)
    if not text:
        return ""
    for pattern in TRACE_CODE_PATTERNS:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            code = re.sub(r'[-\s]+', ' ', match.group(0)).strip()
            return code
    return ""


def _find_typed_trace_number(record):
    """Walks a (possibly nested) record looking for a trace code. Unlike
    _find_trace_code(), this also catches a BARE number that has no
    prefix of its own in the text (e.g. {"DRS": 570} or
    {"DRS_Number": "570"}) and prefixes it with the type implied by ITS
    OWN KEY (via TRACE_KEY_TYPE_MAP) - so the result is "DRS 570", never
    a naked "570" that doesn't say what it is."""
    if isinstance(record, dict):
        for k, v in record.items():
            if isinstance(v, (str, int, float)):
                text = normalize_text(v)
                code = _find_trace_code(text)
                if code:
                    return code
                if re.fullmatch(r'\d+', text):
                    key_norm = re.sub(r'[^a-z]', '', str(k).lower())
                    for tokens, label in TRACE_KEY_TYPE_MAP:
                        if any(t in key_norm for t in tokens):
                            return f"{label} {text}"
            else:
                found = _find_typed_trace_number(v)
                if found:
                    return found
    elif isinstance(record, list):
        for item in record:
            found = _find_typed_trace_number(item)
            if found:
                return found
    return ""


def _flatten_record_values(record):
    """Yields every string value found anywhere in a (possibly nested)
    dict/list record, so a code can be found regardless of which key
    the LLM put it under."""
    if isinstance(record, dict):
        for v in record.values():
            yield from _flatten_record_values(v)
    elif isinstance(record, list):
        for v in record:
            yield from _flatten_record_values(v)
    elif isinstance(record, (str, int, float)):
        text = normalize_text(record)
        if text:
            yield text


def _find_value_by_key_substring(record, substrings):
    """Fallback lookup when the exact expected key name isn't present:
    scans the record's own keys for any that CONTAIN one of the given
    substrings (key compared letters-only, lowercased)."""
    if not isinstance(record, dict):
        return ""
    for k, v in record.items():
        key_norm = re.sub(r'[^a-z]', '', str(k).lower())
        if any(s in key_norm for s in substrings) and isinstance(v, (str, int, float)):
            text = normalize_text(v)
            if text:
                return text
    return ""


def extract_new_edo_traceability_details(
    client,
    product_family,
    product,
    templatename,
    pipeline_config,
    edo_document,
    edo_new_data,
    db
):
    """
    STEP 3 of 3: identical pattern to extract_new_edo_summary_details() -
    for EVERY entry already in the SAME edo_new_data dictionary, calls
    the LLM ONE TIME per RA_Number/FMEA_Number pair (inside the loop),
    this time using the "EDO_NEW_Verification_details" prompt to fetch
    traceability / verification info (spreadsheets, reports, locations,
    pass/fail results), and writes the result straight back into that
    same edo_new_data[key] entry.

    Unlike before, ALL traceability records returned for a given
    RA_Number/FMEA_Number pair are kept (not just the first one) - each
    one is normalized into a small dict:

        { "trace_number": ..., "filename": ..., "filenumber": ...,
          "status": ..., "location": ... }

    and the full list is stored on entry["trace_Id"]. The plain-text
    Verification_Reference written to the Excel output is built purely
    from those values (no "Location:"/"Result:"/etc. labels) - one
    traceability record per line.
    """
    logging.info("=" * 80)
    logging.info("STAGE 4c: WORKFLOW 2 - TRACEABILITY VERIFICATION EXTRACTION (PER RA/FMEA, LOOPED)")
    logging.info("=" * 80)

    if not edo_new_data:
        logging.info("No new EDO tags to extract traceability details for.")
        return edo_new_data

    prompt_data = db.get_prompt_by_name(
        client,
        product_family,
        product,
        templatename,
        "EDO_NEW_Verification_details"
    )

    def _format_trace_line(item):
        """
        Builds one plain-value line for a single trace item - no field
        labels, just the values themselves: "(trace_number) filenumber -
        filename "location"".
        """
        line = ""
        if item["trace_number"]:
            line = f"({item['trace_number']})"
        if item["filenumber"]:
            line = f"{line} {item['filenumber']}".strip()
        if item["filename"]:
            line = f"{line} - {item['filename']}" if line else item["filename"]
        if item["location"]:
            line = f'{line} "{item["location"]}"' if line else f'"{item["location"]}"'
        return line.strip()

    for key, entry in edo_new_data.items():
        ra_number = entry.get("RA_Number", "")
        fmea_number = entry.get("FMEA_Number", "")

        target_text = f"RA_Number : {ra_number}\nFMEA_Number : {fmea_number}"

        prompt_row = {
            "prompt_role": prompt_data["prompt_role"],
            "prompt_text": prompt_data["prompt_text"] + "\nTARGETS:\n" + target_text,
            "question": (
                f"For RA_Number {ra_number} / FMEA_Number {fmea_number} "
                "only, fetch spreadsheets, reports, locations and "
                "pass/fail results as a single JSON object."
            ),
            "fulltext": "Yes",
            "where_filter": "",
            "where_document": "",
            "checkpoint": ""
        }

        records = []
        try:
            _, _, response = execute_llm_retry(
                pipeline_config,
                edo_document["edo_fmea"]["collection"],
                prompt_row
            )
            parsed = parse_json(response)
            records = deep_extract_records(parsed)
            if not records and isinstance(parsed, dict):
                records = [parsed]

        except Exception as e:
            logging.error(
                f"Traceability LLM call failed for RA={ra_number!r} FMEA={fmea_number!r}: {e}"
            )
            records = []

        if not records:
            logging.warning(
                f"No traceability record returned for RA={ra_number!r} FMEA={fmea_number!r}."
            )

        # Normalize EVERY record returned (not just the first) into the
        # requested trace_Id structure.
        trace_items = []
        for record in records:
            if not isinstance(record, dict):
                continue
            trace_items.append({
                "trace_number": get_llm_value(
                    record, "Trace_Number", "Trace Number", "DRS_Number", "DRS", "trace_number"
                ),
                "filename": get_llm_value(
                    record, "Report", "report", "File_Name", "filename"
                ),
                "filenumber": get_llm_value(
                    record, "File_Number", "Document_Number", "filenumber"
                ),
                "status": get_llm_value(
                    record, "Result", "Pass_Fail", "pass_fail", "status"
                ),
                "location": get_llm_value(
                    record, "Location", "location"
                ),
            })

        entry["trace_Id"] = trace_items

        # Only overwrite Verification_Reference if this pass actually
        # found traceability records - don't clobber a value already set
        # by extract_new_edo_summary_details() with an empty result.
        ver_lines = [line for line in (_format_trace_line(item) for item in trace_items) if line]
        if ver_lines:
            entry["Verification_Reference"] = "\n".join(ver_lines)

        entry["Traceability_Location"] = trace_items[0]["location"] if trace_items else ""
        entry["Traceability_Report"] = trace_items[0]["filename"] if trace_items else ""
        entry["Traceability_Result"] = trace_items[0]["status"] if trace_items else ""

    logging.info(
        f"extract_new_edo_traceability_details: enriched {len(edo_new_data)} "
        f"entries in edo_new_data with traceability/verification records."
    )

    return edo_new_data

def get_llm_value(row, *keys):
    """
    Returns the first valid, non-empty value found across `keys`.
    Treats None, "", and any case-insensitive "none"/"blank" placeholder
    text (e.g. "None", "NONE", "Blank") as invalid/empty, so literal
    placeholder strings coming back from the LLM never get written to
    the output Excel as if they were real data.
    """
    for key in keys:
        value = row.get(key)
        if value is None:
            continue
        text = str(value).strip()
        if text == "" or text.lower() in ("none", "blank"):
            continue
        return value
    return ""


# Column J (SYSDD / HDD Reference) is a single, fixed document reference
# for this template - it is NOT something that needs to be extracted per
# EDO record via the LLM. Kept as a named constant so it's easy to update
# in one place if the reference document ever changes.
FIXED_SYSDD_REFERENCE = "NPD38119 Titan Hardware Detailed Design"


def get_fixed_sysdd_reference():
    """
    Dedicated function (per requirement) for Column J (SYSDD / HDD
    Reference). Always returns the fixed reference document string,
    regardless of EDO type (Existing or New) or of whatever value an
    upstream extraction step may have found.
    """
    return FIXED_SYSDD_REFERENCE


# ==========================================================
# OUTPUT TEXT NORMALIZATION (drives requirement: capitalize sentences,
# fix "npdxxxx" -> "NPDxxxx" and "edo-29" -> "EDO-29" casing)
# ==========================================================

def _capitalize_sentences(text):
    """
    Capitalizes the first letter of every sentence in `text`, line by
    line (a "sentence" ends at '.', '!', or '?' followed by whitespace).
    Leading whitespace/indentation on each line is preserved.
    """
    lines = text.split("\n")
    result_lines = []

    for line in lines:
        stripped = line.lstrip()
        if not stripped:
            result_lines.append(line)
            continue

        leading_ws = line[:len(line) - len(stripped)]
        sentences = re.split(r'(?<=[.!?])\s+', stripped)
        fixed = []
        for sentence in sentences:
            if sentence:
                fixed.append(sentence[0].upper() + sentence[1:])
            else:
                fixed.append(sentence)
        result_lines.append(leading_ws + " ".join(fixed))

    return "\n".join(result_lines)


def _normalize_document_codes(text):
    """
    Fixes casing on document / tag codes wherever they appear inside a
    sentence:
      - "npd36702" / "Npd36702"     -> "NPD36702"
      - "edo-29" / "Edo-29"         -> "EDO-29"
      - "ra-108" / "Ra-108"         -> "RA-108"
      - "fmea sys-82" / "Fmea sys-82" -> "FMEA Sys-82"
    """
    text = re.sub(r'\bnpd(\d+)', lambda m: f"NPD{m.group(1)}", text, flags=re.IGNORECASE)
    text = re.sub(r'\bedo[\s-]*([a-z0-9]+)', lambda m: f"EDO-{m.group(1).upper()}", text, flags=re.IGNORECASE)
    text = re.sub(r'\bfmea\s+sys[\s-]*(\d+)', lambda m: f"FMEA Sys-{m.group(1)}", text, flags=re.IGNORECASE)
    text = re.sub(r'\bra[\s-]+(\d+)', lambda m: f"RA-{m.group(1)}", text, flags=re.IGNORECASE)
    return text


def format_output_text(value):
    """
    Canonical text formatter applied to every descriptive/narrative cell
    value before it's written to the output Excel:
      - first letter of every sentence is capitalized
      - "npdxxxx" document codes are uppercased to "NPDxxxx"
      - "edo-xx" tag codes are uppercased to "EDO-XX"
    Leaves truly empty values as empty ("") - no placeholder text.
    """
    text = _text(value)
    if not text.strip():
        return text

    text = _capitalize_sentences(text)
    text = _normalize_document_codes(text)
    return text


def format_edo_tag_text(value):
    """
    Same code-casing fix as format_output_text(), but WITHOUT sentence
    capitalization - used specifically for Column A (EDO Tag), which is a
    short code/label rather than a narrative sentence (e.g. "edo-29" ->
    "EDO-29").
    """
    text = _text(value)
    if not text.strip():
        return text
    return _normalize_document_codes(text)


def get_existing_identifier_set(existing_edos):
    """
    Builds a set of every RA Number / FMEA Number already tracked under
    the Existing EDOs (Stage 3), normalized to uppercase for
    case-insensitive comparison. Used to drop any "New EDO" record whose
    RA id or FMEA id is already represented by an Existing EDO, so the
    same underlying issue is never printed twice.
    """
    identifiers = set()
    for edo in existing_edos.values():
        ra = normalize_text(edo.get("ra_number")).upper()
        fmea = normalize_text(edo.get("FMEA_Number")).upper()
        if ra:
            identifiers.add(ra)
        if fmea:
            identifiers.add(fmea)
    return identifiers


# ==========================================================
# FINAL RECORD NORMALIZATION (identical body in both files - kept once)
# ==========================================================

def normalize_edo_record(edo_record):
    default_structure = {
        "edo_type": "Existing",
        "edo_tag": "",
        "RA_Number": "",
        "FMEA_Number": "",
        "edo_description": "",
        "reason_identified": "",
        "dfmea": "",
        "verification_reference": "",
        "existing_trace": "",
        "location": "",
        "description_2": "",
        "reason_2": "",
        "sysdd": ""
    }
    if not isinstance(edo_record, dict):
        return default_structure

    for key, value in default_structure.items():
        if key not in edo_record:
            edo_record[key] = value
    return edo_record


# ==========================================================
# STAGE 5: UNCONDITIONAL COUPLING / MERGE ENGINE
# ==========================================================

def merge_new_edo_dictionary(
    edo_tags,
    edo_summary_details,
    edo_trace_details
):
    """
    From generate_EDO_template_copy.py - only pipeline that builds NEW EDO
    records, so it is kept as-is and used by the merged pipeline.
    """
    logging.info("=" * 80)
    logging.info("STAGE 5: WORKFLOW MATRIX COUPLING ENGINE")
    logging.info("=" * 80)

    merged = {}

    # Deep Structure Unwrapping
    records = deep_extract_records(edo_summary_details)
    logging.info(f"SUMMARY DETAILS COUNT EXTRICATED : {len(records)}")

    for index, row in enumerate(records):
        ra_number = normalize_text(get_llm_value(row, "RA_Number", "RA Number", "ra_num"))
        fmea_number = normalize_text(get_llm_value(row, "FMEA_Number", "FMEA Number", "fmea_num"))
        status_label = normalize_text(get_llm_value(row, "Status", "status")) or "New EDO"

        key = fmea_number or ra_number
        if not key:
            key = f"NEW-EDO-RECORD-{index}"

        # Safe attribute alignment ensuring literal 'None' strings from the dictionary are preserved
        edo_desc = normalize_text(get_llm_value(row, "Product_Feature_Function", "Product Feature Function", "EDO_Description", "edo_description"))
        reason_id = normalize_text(get_llm_value(row, "Reason_Identified_as_EDO", "Reason Identified as EDO", "reason_identified"))
        dfmea_trace = normalize_text(get_llm_value(row, "Traceability", "dfmea", "traceability"))
        ver_ref = normalize_text(get_llm_value(row, "Verification_Reference", "Verification Reference", "verification_reference"))

        loc_val = normalize_text(get_llm_value(row, "EDO_Location", "location")) or "None"
        desc2_val = normalize_text(get_llm_value(row, "EDO_Description", "Description_2")) or "None"
        reason2_val = normalize_text(get_llm_value(row, "Reason_2")) or "None"
        sys_dd_val = normalize_text(get_llm_value(row, "Reason_Identified_as_EDO_ColH", "sysdd", "SYS_DD")) or "None"

        merged[key] = {
            "edo_type": "New",
            "edo_tag": f"EDO-XX\n{status_label}",
            "edo_description": edo_desc if edo_desc else "Blank",
            "reason_identified": reason_id if reason_id else "Blank",
            "dfmea": dfmea_trace if dfmea_trace else "Blank",
            "verification_reference": ver_ref if ver_ref else "Blank",
            "location": loc_val,
            "description_2": desc2_val,
            "reason_2": reason2_val,
            "sysdd": sys_dd_val,
            "RA_Number": ra_number,
            "FMEA_Number": fmea_number
        }

    # Asymmetric Verification Coupling
    trace_records = deep_extract_records(edo_trace_details)
    for row in trace_records:
        if not isinstance(row, dict):
            continue

        trace_fmea = normalize_text(get_llm_value(row, "FMEA_Number", "FMEA Number"))
        trace_ra = normalize_text(get_llm_value(row, "RA_Number", "RA Number"))
        trace_key = trace_fmea or trace_ra

        verification = normalize_text(get_llm_value(row, "Verification_Reference", "Verification Reference", "verification_reference"))
        if not verification or verification.lower() == "none" or verification == "":
            continue

        if trace_key and trace_key in merged:
            merged[trace_key]["verification_reference"] = verification
        else:
            for main_key, data in merged.items():
                if (trace_ra and data["RA_Number"] == trace_ra) or (trace_fmea and data["FMEA_Number"] == trace_fmea):
                    data["verification_reference"] = verification

    return merged


def merge_new_edo_dictionary_full(
    edo_tags,
    edo_summary_details,
    edo_trace_details,
    edo_ra_details
):
    """
    CANONICAL new-EDO merge - used by generate_edo_template().

    merge_new_edo_dictionary() (above) is left completely UNTOUCHED, but
    it silently drops any RA record whose Status is "Medium", because it
    only ever looks at edo_summary_details / edo_trace_details, both of
    which come exclusively from the EDO_FMEA collection.

    This function fixes that: it builds the output starting from
    edo_tags itself (the full, unfiltered list returned by
    extract_new_edo_tags), so EVERY RA Number is guaranteed a row in the
    final Excel output - none are excluded.

    For each tag:
      - Status contains "Medium"  -> description/reason are pulled ONLY
        from edo_ra_details (EDO_RA_C document) - the FMEA-sourced
        edo_summary_details is deliberately NOT consulted for these,
        exactly as requested.
      - Any other Status (e.g. "See FMEA") -> description/reason/trace/
        verification are pulled from edo_summary_details / edo_trace_details,
        exactly like the original merge_new_edo_dictionary() behaviour.
      - If no matching record is found in the relevant source at all, the
        tag is still written to the output - with empty ("") fields
        rather than a "Blank"/"None" placeholder string, and rather than
        being skipped/excluded.
      - Column A (edo_tag) is always the fixed literal "EDO-XX\\nNew EDO"
        for every New EDO record - the risk Status (Medium / See FMEA) is
        used internally to choose the data source but is never printed.
      - Column D (dfmea) for "Medium" records is always the RA Number
        itself (there is no FMEA trace to show for these), so every RA
        Number is visibly represented in Column D regardless of Status.
    """
    logging.info("=" * 80)
    logging.info("STAGE 5: WORKFLOW MATRIX COUPLING ENGINE (FULL - INCLUDES MEDIUM RA)")
    logging.info("=" * 80)

    def index_records(records):
        index = {}
        for row in records:
            if not isinstance(row, dict):
                continue
            ra = normalize_text(get_llm_value(row, "RA_Number", "RA Number", "ra_num"))
            fmea = normalize_text(get_llm_value(row, "FMEA_Number", "FMEA Number", "fmea_num"))
            if ra and ra not in index:
                index[ra] = row
            if fmea and fmea not in index:
                index[fmea] = row
        return index

    summary_index = index_records(deep_extract_records(edo_summary_details))
    ra_index = index_records(deep_extract_records(edo_ra_details))

    # Verification lookups from the FMEA-sourced traceability extraction -
    # only ever applied to non-Medium ("See FMEA") records.
    trace_index = {}
    for row in deep_extract_records(edo_trace_details):
        if not isinstance(row, dict):
            continue
        trace_ra = normalize_text(get_llm_value(row, "RA_Number", "RA Number"))
        trace_fmea = normalize_text(get_llm_value(row, "FMEA_Number", "FMEA Number"))
        verification = normalize_text(get_llm_value(row, "Verification_Reference", "Verification Reference", "verification_reference"))
        if not verification or verification.lower() == "none":
            continue
        if trace_ra:
            trace_index.setdefault(trace_ra, verification)
        if trace_fmea:
            trace_index.setdefault(trace_fmea, verification)

    merged = {}

    for index, item in enumerate(edo_tags):
        ra_number = normalize_text(item.get("RA_Number"))
        fmea_number = normalize_text(item.get("FMEA_Number"))
        status_label = normalize_text(item.get("Status")) or "New EDO"
        is_medium = "medium" in status_label.lower()

        key = fmea_number or ra_number
        if not key:
            key = f"NEW-EDO-RECORD-{index}"
        # avoid clobbering an existing key on duplicate RA/FMEA numbers
        if key in merged:
            key = f"{key}_{index}"

        if is_medium:
            # Medium risk value -> EDO_RA_C ONLY, never the FMEA document.
            # Keyed by RA Number ONLY - the tag's FMEA_Number is a
            # placeholder ("FMEA-UNKNOWN") for Medium records and is not
            # a reliable lookup key.
            source_row = ra_index.get(ra_number)
        else:
            # See FMEA / anything else -> FMEA-sourced summary, as before
            source_row = summary_index.get(fmea_number) or summary_index.get(ra_number)

        if source_row:
            edo_desc = normalize_text(get_llm_value(source_row, "Product_Feature_Function", "Product Feature Function", "EDO_Description", "edo_description"))
            reason_id = normalize_text(get_llm_value(source_row, "Reason_Identified_as_EDO", "Reason Identified as EDO", "reason_identified"))
            dfmea_trace = normalize_text(get_llm_value(source_row, "Traceability", "dfmea", "traceability"))
            ver_ref = normalize_text(get_llm_value(source_row, "Verification_Reference", "Verification Reference", "verification_reference"))
            loc_val = normalize_text(get_llm_value(source_row, "EDO_Location", "location"))
            desc2_val = normalize_text(get_llm_value(source_row, "EDO_Description", "Description_2"))
            reason2_val = normalize_text(get_llm_value(source_row, "Reason_2"))
        else:
            edo_desc = reason_id = dfmea_trace = ver_ref = ""
            loc_val = desc2_val = reason2_val = ""

        # Per requirement: for "Medium" records, Column D (dfmea) always
        # shows the RA Number itself - there is no FMEA trace document to
        # pull a narrative from, so the RA id is the traceability value.
        if is_medium:
            dfmea_trace = ra_number

        # Verification backfill only applies to non-Medium records, since
        # the trace/verification extraction is itself FMEA-sourced.
        if not is_medium and not ver_ref:
            ver_ref = trace_index.get(fmea_number) or trace_index.get(ra_number) or ver_ref

        # Column J (SYSDD / HDD Reference) is a fixed document reference
        # for this template, not an LLM-extracted value - see
        # get_fixed_sysdd_reference().
        sys_dd_val = get_fixed_sysdd_reference()

        merged[key] = {
            "edo_type": "New",
            # Per requirement: Column A never prints the risk Status
            # (Medium / See FMEA) - always the fixed literal tag text.
            "edo_tag": "EDO-XX\nNew EDO",
            "edo_description": edo_desc,
            "reason_identified": reason_id,
            "dfmea": dfmea_trace,
            "verification_reference": ver_ref,
            "location": loc_val,
            "description_2": desc2_val,
            "reason_2": reason2_val,
            "sysdd": sys_dd_val,
            "RA_Number": ra_number,
            "FMEA_Number": fmea_number
        }

    return merged


def merge_new_edo_records(new_records, existing_edos):
    """
    CANONICAL new-EDO merge - used by generate_edo_template(), replacing
    the old 4-source merge_new_edo_dictionary_full() now that
    extract_new_edo_tags() has been removed and extract_new_edo_summary_details()
    returns complete records directly from the single merged "EDO_NEW_details"
    prompt.

    Fixes the "column mapped wrongly" bug: the previous merge logic looked
    for keys like "Description_2" / "Reason_2" that never actually existed
    in the LLM's JSON output - the real keys are "EDO_Description" and
    "Reason_Identified_as_EDO_ColH". This function maps directly off the
    ACTUAL field names the prompt returns:

        edo_description   <- Product_Feature_Function
        reason_identified <- Reason_Identified_as_EDO
        dfmea             <- Traceability
        verification_reference <- Verification_Reference
        location          <- EDO_Location
        description_2     <- EDO_Description
        reason_2          <- Reason_Identified_as_EDO_ColH
        sysdd             <- fixed constant (get_fixed_sysdd_reference())

    Also applies requirement #2: any new record whose RA Number or FMEA
    Number already belongs to an Existing EDO is dropped from the output,
    so the same issue is never printed twice under both types.

    Column A (edo_tag) is always the fixed literal "EDO-XX\\nNew" - the
    risk Status is never printed there.
    """
    logging.info("=" * 80)
    logging.info("STAGE 5: NEW EDO MERGE (SINGLE-PASS RECORDS, DE-DUPED AGAINST EXISTING)")
    logging.info("=" * 80)

    existing_identifiers = get_existing_identifier_set(existing_edos)

    records = deep_extract_records(new_records)
    merged = {}

    for index, row in enumerate(records):
        if not isinstance(row, dict):
            continue

        ra_number = normalize_text(get_llm_value(row, "RA_Number", "RA Number"))
        fmea_number = normalize_text(get_llm_value(row, "FMEA_Number", "FMEA Number"))

        # Requirement #2: skip any new record already tracked as an
        # Existing EDO (same RA id or FMEA id), so it isn't duplicated.
        ra_key = ra_number.upper()
        fmea_key = fmea_number.upper()
        if (ra_key and ra_key in existing_identifiers) and (fmea_key and fmea_key in existing_identifiers):
            logging.info(
                f"Skipping New EDO record RA={ra_number!r} FMEA={fmea_number!r} - "
                f"already tracked as an Existing EDO."
            )
            continue

        key = fmea_number or ra_number or f"NEW-EDO-RECORD-{index}"
        if key in merged:
            key = f"{key}_{index}"

        merged[key] = {
            "edo_type": "New",
            # Per requirement: Column A never prints the risk Status -
            # always the fixed literal tag text, correctly cased.
            "edo_tag": "EDO-XX\nNew",
            "edo_description": normalize_text(get_llm_value(row, "Product_Feature_Function", "Product Feature Function")),
            "reason_identified": normalize_text(get_llm_value(row, "Reason_Identified_as_EDO", "Reason Identified as EDO")),
            "dfmea": normalize_text(get_llm_value(row, "Traceability", "traceability")),
            "verification_reference": normalize_text(get_llm_value(row, "Verification_Reference", "Verification Reference")),
            "location": normalize_text(get_llm_value(row, "EDO_Location", "location")),
            "description_2": normalize_text(get_llm_value(row, "EDO_Description", "description")),
            "reason_2": normalize_text(get_llm_value(row, "Reason_Identified_as_EDO_ColH", "reason_2")),
            "sysdd": get_fixed_sysdd_reference(),
            "RA_Number": ra_number,
            "FMEA_Number": fmea_number
        }

    return merged


def merge_existing_edo_dictionary(existing_details):
    """
    CANONICAL version - from edo_existing_final.py.
    Richer than the copy-file version: carries edo_type, and correctly
    threads RA_Number/FMEA_Number through (needed since Stage 3B already
    populated ra_number/FMEA_Number on each existing EDO).

    NOTE: missing values default to "" (a truly empty cell), not the
    literal text "Blank". Column J (sysdd) is always the fixed reference
    from get_fixed_sysdd_reference(), per requirement.

    Also carries "design_elements" (the full per-location list built by
    extract_edo_details() from the EDO_Existing_Generic prompt's nested
    design_elements[] array) straight through - previously this dict was
    rebuilt with only a fixed set of keys, so design_elements was
    silently dropped here and format_edo_worksheet() only ever saw a
    single backfilled location instead of every location.
    """
    final = {}
    for tag, data in existing_details.items():
        final[tag] = {
            "edo_type": "Existing",
            "edo_tag": data.get("edo_tag", tag),
            "edo_description": data.get("edo_description", ""),
            "reason_identified": data.get("reason_identified", ""),
            "dfmea": data.get("dfmea", ""),
            "location": data.get("location", ""),
            "description_2": data.get("description_2", ""),
            "reason_2": data.get("reason_2", ""),
            "sysdd": get_fixed_sysdd_reference(),
            "verification_reference": data.get("verification_reference", ""),
            "existing_trace": data.get("existing_trace", ""),
            "RA_Number": data.get("ra_number", ""),
            "FMEA_Number": data.get("FMEA_Number", ""),
            "design_elements": data.get("design_elements", [])
        }
    return final


def merge_existing_edo_dictionary_legacy(existing_details):
    """
    Original version from generate_EDO_template_copy.py - kept for
    backward compatibility. Does not carry edo_type/verification_reference
    and hardcodes FMEA_Number to "Blank". Not used by the merged pipeline.
    """
    final = {}
    for tag, data in existing_details.items():
        final[tag] = {
            "edo_tag": data.get("edo_tag", tag),
            "edo_description": data.get("edo_description", "Blank"),
            "reason_identified": data.get("reason_identified", "Blank"),
            "dfmea": data.get("dfmea", "Blank"),
            "location": data.get("location", "Blank"),
            "description_2": data.get("description_2", "Blank"),
            "reason_2": data.get("reason_2", "Blank"),
            "sysdd": data.get("sysdd", "Blank"),
            "RA_Number": data.get("ra_number", "Blank"),
            "FMEA_Number": "Blank"
        }
    return final


def merge_all_edos(existing_edos, new_edos):
    """
    From generate_EDO_template_copy.py - only pipeline that combines
    Existing + New EDO dictionaries into one, so it is kept as-is.
    """
    final_edos = {}

    for key, value in existing_edos.items():
        normalized = normalize_edo_record(value)
        normalized["edo_type"] = "Existing"
        final_edos[key] = normalized

    for key, value in new_edos.items():
        normalized = normalize_edo_record(value)
        normalized["edo_type"] = "New"

        final_key = key
        if final_key in final_edos:
            counter = 1
            while f"{key}_{counter}" in final_edos:
                counter += 1
            final_key = f"{key}_{counter}"

        final_edos[final_key] = normalized

    return final_edos


def validate_final_edos(final_edos):
    """
    CANONICAL version - from generate_EDO_template_copy.py.
    General-purpose: handles both "Existing" and "New" typed records
    (assigns a placeholder tag for blank New EDOs). Needed because the
    merged pipeline's final_edos dictionary contains both types.
    """
    validated = {}
    for key, value in final_edos.items():
        if not isinstance(value, dict):
            continue

        edo_tag = normalize_text(value.get("edo_tag"))
        if edo_tag == "" and value.get("edo_type") == "New":
            value["edo_tag"] = "EDO-XX\nNew EDO"

        normalized = normalize_edo_record(value)
        validated[key] = normalized
        print(f"final values:", final_edos)
    return validated


def validate_final_edos_existing_only(final_edos):
    """
    Original version from edo_existing_final.py - kept for backward
    compatibility. Forces edo_type to "Existing" unconditionally, so it is
    only correct for the existing-EDO-only pipeline
    (generate_and_download_edo1). Not used by the combined pipeline.
    """
    validated = {}
    for key, value in final_edos.items():
        if not isinstance(value, dict):
            continue

        normalized = normalize_edo_record(value)
        normalized["edo_type"] = "Existing"
        validated[key] = normalized
    return validated


# ==========================================================
# STAGE 6: OUTPUT MAPPING AND FILE STORAGE
# ==========================================================

def write_edo_excel(sheet, final_edos, start_row):
    """
    MERGED canonical writer, used by the combined pipeline
    (generate_edo_template).

    Combines:
      - edo_existing_final.py's location-splitting logic (a single
        location cell like "(i) 210119 (metal cyl) (ii) 210120 (bracket)"
        is split into one output row per bracketed item), and Aptos
        Narrow/size-8 formatting.
      - generate_EDO_template_copy.py's per-edo_type Column D logic (New
        EDOs get a constructed "Document Traceability Matrix Info" block
        built from RA_Number/FMEA_Number/dfmea narrative; Existing EDOs
        just show the raw dfmea narrative).

    Writes columns A-J only (1-10).
    """
    logging.info("=" * 80)
    logging.info("STAGE 6: EXCEL OUTPUT FORMATTING RANGE (MERGED)")
    logging.info("=" * 80)

    current_row = start_row

    for key, edo in final_edos.items():
        logging.info(f"Writing row {current_row}: {key}")

        # 1. SPLIT LOGIC for the location field, e.g. turns
        #    "(i) 210119 (metal cyl) (ii) 210120 (bracket)" into
        #    ["210119 (metal cyl)", "210120 (bracket)"]
        raw_location = str(edo.get("location") or "Blank")
        matches = re.findall(r'\d+\s*\([^\)]+\)', raw_location)
        split_items = matches if matches else [raw_location]

        # 2. Column D (Traceability) value depends on edo_type
        if edo.get("edo_type") == "New":
            ra_id = edo.get("RA_Number", "") or "N/A"
            fmea_num = edo.get("FMEA_Number", "") or "N/A"
            trace_narrative = edo.get("dfmea", "")

            trace_value = f"Document Traceability Matrix Info:\nRA Number: {ra_id}\nFMEA Number: {fmea_num}"
            if trace_narrative and trace_narrative.lower() != "none" and trace_narrative.strip() != "":
                trace_value += f"\n\nDetails:\n{trace_narrative}"
        else:
            trace_value = edo.get("dfmea") or "Blank"

        # 3. Determine displayed tag
        if edo.get("edo_type") == "New":
            tag_value = edo.get("edo_tag") or "EDO-XX\nNew EDO"
        else:
            tag_value = edo.get("edo_tag") or key

        # 4. Iterate through the split location items - one output row each
        for i, item in enumerate(split_items):
            # Column A (1): Tag
            sheet.cell(row=current_row, column=1).value = tag_value

            # Column B (2): EDO Description (only shown on the first split row)
            sheet.cell(row=current_row, column=2).value = edo.get("edo_description") if i == 0 else ""

            # Column C (3): Reason Identified
            sheet.cell(row=current_row, column=3).value = edo.get("reason_identified") or "Blank"

            # Column D (4): Traceability (type-aware)
            sheet.cell(row=current_row, column=4).value = trace_value

            # Column E (5): Verification Reference
            sheet.cell(row=current_row, column=5).value = edo.get("verification_reference") or "Blank"

            # Column F (6): Clean Location (the split item)
            sheet.cell(row=current_row, column=7).value = item

            # Column G (7): Location counterpart / secondary field
            #sheet.cell(row=current_row, column=8).value = item

            # Column H (8): Secondary Description
            sheet.cell(row=current_row, column=8).value = edo.get("description_2") or "Blank"

            # Column I (9): Secondary Reason
            sheet.cell(row=current_row, column=9).value = edo.get("reason_2") or "Blank"

            # Column J (10): SYSDD / Design Reference
            sheet.cell(row=current_row, column=10).value = edo.get("sysdd") or "Blank"

            # 5. Apply formatting
            for col in range(1, 11):
                cell = sheet.cell(row=current_row, column=col)
                cell.font = aptos_font
                cell.alignment = cell_alignment
                cell.border = thin_border

            sheet.row_dimensions[current_row].height = 32
            current_row += 1

    return current_row


def write_edo_excel_existing_style(sheet, final_edos, start_row):
    """
    Original version from edo_existing_final.py - kept for backward
    compatibility (used by generate_and_download_edo1, the
    existing-EDO-only pipeline). Writes columns A-K (1-11).
    """
    logging.info("Formatting with Aptos Narrow, Size 8 and structured location splitting.")

    current_row = start_row

    for key, edo in final_edos.items():
        raw_data = str(edo.get("location") or "Blank")

        matches = re.findall(r'\d+\s*\([^\)]+\)', raw_data)
        split_items = matches if matches else [raw_data]

        for i, item in enumerate(split_items):
            dfmea_value = edo.get("dfmea") or "Blank"
            trace_value = normalize_text(edo.get("existing_trace"))
            if trace_value:
                dfmea_value = f"{dfmea_value}\n{trace_value}"

            sheet.cell(row=current_row, column=1).value = edo.get("edo_tag") or key
            sheet.cell(row=current_row, column=2).value = edo.get("edo_description") if i == 0 else ""
            sheet.cell(row=current_row, column=3).value = edo.get("reason_identified") or "Blank"
            sheet.cell(row=current_row, column=4).value = dfmea_value
            logging.info(f"{key} dfmea raw: {edo.get('dfmea')}")
            logging.info(f"{key} existing_trace raw: {edo.get('existing_trace')}")
            logging.info(f"{key} final column D value: {dfmea_value}")
            sheet.cell(row=current_row, column=5).value = edo.get("verification_reference") or "Blank"
            sheet.cell(row=current_row, column=7).value = item
            sheet.cell(row=current_row, column=8).value = edo.get("description_2") or "Blank"
            sheet.cell(row=current_row, column=9).value = edo.get("reason_2") or "Blank"
            sheet.cell(row=current_row, column=10).value = edo.get("sysdd") or "Blank"
            #sheet.cell(row=current_row, column=10).value = "N/A"
            #sheet.cell(row=current_row, column=11).value = "None"

            for col in range(1, 12):
                cell = sheet.cell(row=current_row, column=col)
                cell.font = aptos_font
                cell.alignment = cell_alignment
                cell.border = thin_border

            sheet.row_dimensions[current_row].height = 32
            current_row += 1

    return current_row


def write_edo_excel_new_style(sheet, final_edos, start_row):
    """
    Original version from generate_EDO_template_copy.py - kept for
    backward compatibility. Writes columns A-J (1-10), no location
    splitting, no font override (relies on template's existing font).
    """
    logging.info("=" * 80)
    logging.info("STAGE 6: EXCEL OUTPUT FORMATTING RANGE")
    logging.info("=" * 80)

    current_row = start_row

    for key, edo in final_edos.items():
        logging.info(f"Writing row {current_row}: {key}")

        if edo.get("edo_type") == "New":
            value = edo.get("edo_tag") or "EDO-XX\nNew EDO"
        else:
            value = edo.get("edo_tag") or key
        sheet.cell(row=current_row, column=1).value = value

        sheet.cell(row=current_row, column=2).value = edo.get("edo_description") or "Blank"
        sheet.cell(row=current_row, column=3).value = edo.get("reason_identified") or "Blank"

        if edo.get("edo_type") == "New":
            ra_id = edo.get("RA_Number", "") or "N/A"
            fmea_num = edo.get("FMEA_Number", "") or "N/A"
            trace_narrative = edo.get("dfmea", "")

            trace_value = f"Document Traceability Matrix Info:\nRA Number: {ra_id}\nFMEA Number: {fmea_num}"
            if trace_narrative and trace_narrative.lower() != "none" and trace_narrative.strip() != "":
                trace_value += f"\n\nDetails:\n{trace_narrative}"

            sheet.cell(row=current_row, column=4).value = trace_value
        else:
            sheet.cell(row=current_row, column=4).value = edo.get("dfmea") or "Blank"

        sheet.cell(row=current_row, column=5).value = edo.get("verification_reference") or "Blank"
        #sheet.cell(row=current_row, column=6).value = edo.get("location") or "Blank"
        sheet.cell(row=current_row, column=7).value = edo.get("location") or "Blank"
        sheet.cell(row=current_row, column=8).value = edo.get("description_2") or "Blank"
        sheet.cell(row=current_row, column=9).value = edo.get("reason_2") or "Blank"
        sheet.cell(row=current_row, column=10).value = edo.get("sysdd") or "Blank"

        for col in range(1, 11):
            cell = sheet.cell(row=current_row, column=col)
            cell.alignment = cell_alignment
            cell.border = thin_border

        sheet.row_dimensions[current_row].height = 32
        current_row += 1

    return current_row


def _style_value(value, base_bold, is_new):
    """
    Shared column-coloring rule used by every plain (non rich-text) cell:
      - New EDO row      -> always RED, keeping the column's normal bold flag.
      - Existing EDO row -> BLACK, normal bold flag (an empty cell has no
        text to color, so this only matters when there IS a value).
    """
    if is_new:
        return RED, base_bold
    
    print(f"formatted color")
    return BLACK, base_bold



def classify_risk_status(description_text, pipeline_config):
    """
    Risk classification using LLM based on the provided Risk Classification table.
    Compares/evaluates the description_text against the risk term guidelines
    below (same call_llm(prompt, pipeline_config) pattern as
    evaluate_comparison()) and returns High / Medium / Low / None.
    """

    print(f"description text", description_text)

    text = normalize_text(description_text).lower()

    print(f"clasification f risk text", text)

    if not text:
        print(f"risk status", "None")
        return "None"

    prompt = f"""
You are a Risk Classification expert.

Classify the following description into exactly one of these categories:
- High
- Medium
- Low
- None

Risk Classification Guidelines:

High:
- Potential impact to patient safety
- Device performance issues
- Not meeting SOA requirements
- Gaps in EDO lists
- Non-existent EDO
- Missing required V&V
- Missing required risk documents
- Satisfies FA / Meets FA requirements
- Any issue that can significantly affect safety, regulatory compliance, or product functionality

Medium:
- Documentation gaps
- Missing or incomplete documents
- Tracing errors in RTMs
- RAS
- DID
- Updates to existing V&V
- Risk document updates
- Compliance issues
- Moderate documentation or traceability problems

Low:
- Template updates
- Drawing updates or creation
- Minor document updates
- Clarifications
- Typographical corrections
- Missing component qualifications
- Missing control plans
- PCS
- MVP
- Cosmetic or administrative changes with no impact to safety or functionality

None:
- Description does not match any of the above categories.

Description:
{text}

Return ONLY one word:
High
Medium
Low
None
"""

    risk_status = "None"

    try:
        response = call_llm(prompt, pipeline_config)
        response = clean_response(response)

        normalized = str(response).strip().lower()

        if normalized.startswith("high"):
            risk_status = "High"
        elif normalized.startswith("medium"):
            risk_status = "Medium"
        elif normalized.startswith("low"):
            risk_status = "Low"
        else:
            risk_status = "None"

    except Exception as e:
        logging.error(f"Risk classification LLM failed: {e}")
        risk_status = "None"

    print(f"risk status", risk_status)

    return risk_status


def apply_risk_cell_style(cell, risk):
    from openpyxl.styles import PatternFill, Font

    fills = {
        "High": "FF0000",
        "Medium": "FFFF00",
        "Low": "00B050"
    }

    if risk in fills:
        cell.fill = PatternFill("solid", fgColor=fills[risk])
        cell.font = Font(bold=True, color="000000", name="Calibri", size=10)
    else:
        cell.fill = PatternFill(fill_type=None)

def generate_remarks_and_recommendation(edo, tag_value, is_new, risk, pipeline_config):
    """
    Builds Column M ("Remarks and Recommendation") content:
      1. Boilerplate Gap + Verification Status (New EDO vs Existing EDO wording)
      2. Optional "Observation" block - flagged when the FMEA/RA&C trace text
         (Column D) states a DIFFERENT risk level than the assigned Risk
         Classification (Column L)
      3. Optional "Recommendation"/"Design" block - feature-specific safety
         mitigation text (manual warning, drawing/design note, or training),
         generated from the EDO's feature/reason fields when the hazard
         warrants it
    """

    # ---- 1. Base boilerplate (Gap + Verification Status) ----
    if is_new:
        base_text = (
            "Gap:\n"
            "As identified in the Risk Assessment & Control (RA&C) and System "
            "DFMEA, this risk impacts the product’s functions and features. "
            "Therefore, it is classified as a new Essential Design Output "
            "(EDO) and must be incorporated into the existing EDO list.\n\n"
            "Verification Status:\n"
            f"Design verification has been conducted for this EDO ({tag_value}), "
            "and the corresponding reports are traced in the Verification "
            "Reference (Column E)."
        )
    else:
        base_text = (
            "Gap:\n"
            "The design verification reference corresponding to where the EDO "
            "is controlled has not been included in the existing EDO list.\n\n"
            "Verification Status:\n"
            f"Design verification has been conducted for this EDO ({tag_value}), "
            "and the corresponding reports are traced in the Verification "
            "Reference (Column E)."
        )

    # ---- 2. Observation block (risk-trace mismatch check) ----
    observation_text = ""
    dfmea_text = format_output_text(edo.get("dfmea"))

    if dfmea_text and risk:
        obs_prompt = f"""
        You are reviewing an Essential Design Output (EDO) risk record.

        Risk Classification assigned (Column L): {risk}
        System FMEA / RA&C trace text (Column D): {dfmea_text}

        Does the FMEA/RA&C trace text explicitly state a DIFFERENT risk
        evaluation (e.g. 'Low', 'Medium', 'High') than the assigned Risk
        Classification above?

        If yes, reply with EXACTLY this sentence, filling in the FMEA's
        stated level and the assigned classification (lowercase):
        Observation: In Sys-FMEA, the risk evaluation is '<fmea_level>'. recommended to change in the sys-FMEA as <assigned_level_lowercase>.

        If no mismatch is found, reply with exactly: NONE
        """

        try:
            obs_response = call_llm(obs_prompt, pipeline_config)
            obs_response = clean_response(obs_response)
            if obs_response and obs_response.strip().upper() != "NONE":
                observation_text = obs_response.strip()
        except Exception as e:
            logging.error(f"Observation generation failed: {e}")

    # ---- 3. Recommendation / Design block (feature-specific safety mitigation) ----
    recommendation_text = ""
    feature_text = format_output_text(edo.get("edo_description") or edo.get("description_2"))
    reason_text = format_output_text(edo.get("reason_identified") or edo.get("reason_2"))

    if feature_text or reason_text:
        rec_prompt = f"""
        You are drafting a "Recommendation" note for an Essential Design
        Output (EDO) risk record, matching this house style:

        - If the hazard needs a User Manual warning/precaution, write it
          starting with "Recommendation:" (or "Recommendation to add in the
          User Manual:") followed by concrete warning text.
        - If the hazard needs a drawing/design change (e.g. a load rating,
          dimension note, or EDO symbol placement), write it starting with
          "Design:" followed by the specific design note.
        - If training is the appropriate mitigation, write it starting with
          "Recommendation:" followed by the training instruction.
        - If NONE of the above genuinely apply (the base Gap/Verification
          Status boilerplate is already sufficient), reply with exactly:
          NONE

        Product Feature/Function: {feature_text}
        Reason Identified as EDO: {reason_text}

        Only return the recommendation block itself (or NONE) - do not
        repeat the Gap or Verification Status text.
        """

        try:
            rec_response = call_llm(rec_prompt, pipeline_config)
            rec_response = clean_response(rec_response)
            if rec_response and rec_response.strip().upper() != "NONE":
                recommendation_text = rec_response.strip()
        except Exception as e:
            logging.error(f"Recommendation generation failed: {e}")

    # ---- Assemble final Column M text ----
    parts = [base_text]
    if observation_text:
        parts.append(observation_text)
    if recommendation_text:
        parts.append(recommendation_text)

    return "\n\n".join(parts)

def merge_existing_edo_rows(sheet, start_row, end_row):
    """
    Merge A:E for repeated Existing EDO tags only.
    Each EDO tag group is merged independently.
    """
    current = start_row

    while current <= end_row:
        tag = sheet.cell(current, 1).value

        if not tag:
            current += 1
            continue

        last = current
        while last + 1 <= end_row and sheet.cell(last + 1, 1).value == tag:
            last += 1

        if last > current:
            for col in range(1, 6):
                sheet.merge_cells(
                    start_row=current,
                    start_column=col,
                    end_row=last,
                    end_column=col
                )
                sheet.cell(current, col).alignment = cell_alignment

        current = last + 1


def format_edo_worksheet(sheet, final_edos, start_row, pipeline_config, images=None, ra_fmea_images=None):
    """
    Final writer:
    A-E : existing columns
    F-I : new EDO fields
    H   : also carries any images extracted from edo_proposed (see
          extract_edo_proposed_images() / insert_image_below_text()),
          stacked below the description_2 text. One image is placed per
          split row within each EDO tag's merged block (Column H is
          never merged across split rows, unlike A-E), pulling from the
          shared queue in document order; once the queue runs dry, the
          last image already used for that SAME EDO tag is duplicated
          across its remaining split rows.

          For New EDO records specifically, `ra_fmea_images` (built by
          extract_new_edo_diagrams(), keyed by
          (RA_Number.upper(), FMEA_Number.upper())) is checked FIRST -
          if that record's own RA/FMEA pair (Column D) has a matched
          diagram, that exact image is placed instead of the generic
          FIFO queue. Only when there's no per-record match does the
          generic queue apply. Every case where Column H ends up with
          no image at all is logged with the specific reason.
    K   : Risk Classification
    L   : Risk evaluation text / classification trigger
    M   : Gap and Verification Status statement
    """

    current_row = start_row
    existing_ranges = []
    image_queue = list(images) if images else []
    ra_fmea_images = ra_fmea_images or {}
    last_image_row = start_row

    for key, edo in final_edos.items():

        is_new = edo.get("edo_type") == "New"

        if is_new:
            tag_value = "EDO-XX\nNew"
        else:
            tag_value = format_edo_tag_text(edo.get("edo_tag") or key)

        raw_location = format_output_text(edo.get("location"))
        design_elements = edo.get("design_elements") or []

        if design_elements:
            # Existing EDOs using the new nested design_elements[]
            # structure - one split row PER DESIGN ELEMENT, each with
            # its own location/description/reason, instead of only ever
            # showing the single location that got backfilled onto
            # edo["location"].
            split_rows = [
                {
                    "location": element.get("location", ""),
                    "description_2": element.get("description", ""),
                    "reason_2": element.get("reason", ""),
                }
                for element in design_elements
            ]
        else:
            # New EDOs (and any Existing EDO with no design_elements) -
            # original behaviour: regex-split a single concatenated
            # "(i) loc1 (ii) loc2 ..." location string, reusing the same
            # description_2/reason_2 value on every split row.
            locations = re.findall(r'\d+\s*\([^\)]+\)', raw_location) or [raw_location]
            split_rows = [
                {
                    "location": location,
                    "description_2": edo.get("description_2"),
                    "reason_2": edo.get("reason_2"),
                }
                for location in locations
            ]

        first = current_row
        last_image_for_this_edo = None

        for idx, row_data in enumerate(split_rows):

            # Per requirement: for New EDO rows, Columns G (location),
            # H (description_2), and I (reason_2) should never be left
            # blank when the LLM didn't return a value - they must show
            # the literal text "None" instead of an empty cell.
            col_g_value = row_data["location"]
            col_h_value = format_output_text(row_data["description_2"])
            col_i_value = format_output_text(row_data["reason_2"])

            if is_new:
                if not normalize_text(col_g_value):
                    col_g_value = "None"
                if not normalize_text(col_h_value):
                    col_h_value = "None"
                if not normalize_text(col_i_value):
                    col_i_value = "None"

            # Column D: for Existing EDOs only, append the trace text
            # (existing_trace, from extract_existing_edo_trace_details /
            # apply_existing_edo_trace) below the RA/FMEA data already in
            # dfmea, in the same cell/row. New EDOs are left untouched.
            col_d_value = format_output_text(edo.get("dfmea"))
            if not is_new:
                trace_text = format_output_text(edo.get("existing_trace"))
                if trace_text:
                    col_d_value = f"{col_d_value}\n Trace:{trace_text}" if col_d_value else trace_text

            logging.info(f"{key} dfmea raw: {edo.get('dfmea')}")
            logging.info(f"{key} existing_trace raw: {edo.get('existing_trace')}")
            logging.info(f"{key} final column D value: {col_d_value}")

            values = {
                1: tag_value,
                2: format_output_text(edo.get("edo_description")) if idx == 0 else "",
                3: format_output_text(edo.get("reason_identified")),
                4: col_d_value,
                5: format_output_text(edo.get("verification_reference")),
                7: col_g_value,
                8: col_h_value,
                9: col_i_value,
                10: get_fixed_sysdd_reference(),
                11: "None",
            }

            # --- Risk Classification (Column L) computed first, since
            # --- Column M's content depends on it ---
            risk_input = (
                f"{format_output_text(edo.get('reason_identified') or '')} "
                f"{format_output_text(edo.get('dfmea') or '')}"
            )
            risk = classify_risk_status(risk_input, pipeline_config)
            values[12] = risk

            # --- Remarks and Recommendation (Column M) ---
            values[13] = generate_remarks_and_recommendation(
                edo, tag_value, is_new, risk, pipeline_config
            )

            for col, value in values.items():
                cell = sheet.cell(current_row, col)
                cell.value = value
                _apply_border_alignment(cell)

            apply_risk_cell_style(sheet.cell(current_row, 12), risk)

            # PLACE AN IMAGE ON THIS SPLIT ROW, WITHIN THIS EDO TAG'S
            # MERGED BLOCK (Column H/8 is never merged across split rows,
            # so each split row keeps its own independent image slot).
            # Pull the next image from the shared queue while one is
            # available; once the queue is exhausted, duplicating the
            # last image already used for THIS SAME EDO tag across its
            # remaining split rows is allowed, so every split row under
            # one Column A tag still shows an image.
            row_image = None
            matched_by_ra_fmea = False

            if is_new and ra_fmea_images:
                match_key = (
                    normalize_text(edo.get("RA_Number") or "").upper(),
                    normalize_text(edo.get("FMEA_Number") or "").upper()
                )
                row_image = ra_fmea_images.get(match_key)
                if row_image:
                    matched_by_ra_fmea = True
                else:
                    logging.warning(
                        f"COLUMN H (row {current_row}, key {key!r}): no "
                        f"per-record diagram matched for RA="
                        f"{edo.get('RA_Number')!r} FMEA="
                        f"{edo.get('FMEA_Number')!r} in ra_fmea_images - "
                        "see extract_edo_pdf_new_diagram()'s log lines "
                        "for this RA/FMEA pair for the exact cause. "
                        "Falling back to the generic image queue."
                    )

            if not row_image:
                if image_queue:
                    row_image = image_queue.pop(0)
                    last_image_for_this_edo = row_image
                else:
                    row_image = last_image_for_this_edo

            if row_image:
                insert_image_below_text(sheet, row_image, row=current_row, column=8, text_offset_px=IMAGE_TEXT_OFFSET_PX)
                last_image_row = current_row
                if not matched_by_ra_fmea:
                    last_image_for_this_edo = row_image
            else:
                # Nothing at all to place - explicit log so an empty
                # Column H cell is never silently unexplained: either no
                # per-record match, an exhausted generic queue, or both.
                logging.error(
                    f"COLUMN H WILL BE EMPTY at row {current_row} for "
                    f"key {key!r} (RA={edo.get('RA_Number')!r} FMEA="
                    f"{edo.get('FMEA_Number')!r}, is_new={is_new}) - no "
                    "ra_fmea_images match AND the generic image queue is "
                    "exhausted with no prior image for this EDO tag to "
                    "duplicate."
                )

            current_row += 1

        if not is_new:
            existing_ranges.append((first, current_row - 1))

    # TRUE OVERFLOW: only reachable if there were literally more images
    # than physical rows across the ENTIRE sheet (every split row of
    # every EDO tag already received one, including duplicates within a
    # tag's own block). Stack whatever's left below the last row that
    # received an image, so no extracted image is ever silently unplaced.
    if image_queue:
        running_offset = IMAGE_TEXT_OFFSET_PX + IMAGE_HEIGHT + 15
        while image_queue:
            overflow_img = image_queue.pop(0)
            insert_image_below_text(sheet, overflow_img, row=last_image_row, column=8, text_offset_px=running_offset)
            running_offset += IMAGE_HEIGHT + 15

    for first, last in existing_ranges:
        if last > first:
            for col in range(1, 6):
                sheet.merge_cells(
                    start_row=first,
                    start_column=col,
                    end_row=last,
                    end_column=col
                )
                sheet.cell(first, col).alignment = cell_alignment
            for col in (12, 13):
                sheet.merge_cells(
                    start_row=first,
                    start_column=col,
                    end_row=last,
                    end_column=col
                )
                sheet.cell(first, col).alignment = cell_alignment
    print(f"current _value:")
    return current_row

def clear_existing_rows(sheet, start_row, end_column=10):
    row = start_row
    while row <= sheet.max_row:
        empty = True
        for col in range(1, end_column + 1):
            if sheet.cell(row=row, column=col).value:
                empty = False
                break
        if empty:
            break
        for col in range(1, end_column + 1):
            sheet.cell(row=row, column=col).value = None
        row += 1


def save_edo_workbook(workbook, pipeline_config):
    output_path = pipeline_config["output_file_path"]
    workbook.save(output_path)
    return output_path


# ==========================================================
# MAIN EXECUTION PIPELINE #1 (EXISTING EDO ONLY)
# Preserved as-is from edo_existing_final.py, using the "_existing_style"
# / "_existing_only" variants so its behaviour is unchanged.
# ==========================================================

def generate_and_download_edo1(
    client,
    product_family,
    product,
    templatename,
    pipeline_config,
    db: DatabaseHandler
):
    logging.info("=" * 80)
    logging.info("STARTING EXISTING-EDO TEMPLATE GENERATION PIPELINE")
    logging.info("=" * 80)

    try:
        workbook, sheet = initialize_workbook(pipeline_config)
        start_row = pipeline_config.get("templatestartrow", 4)

        clear_existing_rows(sheet, start_row, end_column=10)

        edo_document = get_edo_document(
            client,
            product_family,
            product,
            templatename,
            db
        )

        existing_edos = extract_edo_tags(
            client,
            product_family,
            product,
            templatename,
            pipeline_config,
            edo_document,
            db
        )
        existing_edos = validate_existing_tags(existing_edos)

        if existing_edos:
            existing_edos = extract_edo_details(
                client,
                product_family,
                product,
                templatename,
                pipeline_config,
                edo_document,
                existing_edos,
                db
            )

            try:
                existing_verification_details = extract_existing_edo_verification_details(
                    client,
                    product_family,
                    product,
                    templatename,
                    pipeline_config,
                    edo_document,
                    existing_edos,
                    db
                )

                existing_edos = apply_existing_edo_verification(
                    existing_edos,
                    existing_verification_details
                )

            except Exception as verification_error:
                logging.warning(
                    "STAGE 3B SKIPPED - verification reference extraction "
                    "failed, leaving column E as 'Blank' for this run. "
                    f"Reason: {verification_error}"
                )

            try:
                existing_trace_details = extract_existing_edo_trace_details(
                    client,
                    product_family,
                    product,
                    templatename,
                    pipeline_config,
                    edo_document,
                    existing_edos,
                    db
                )

                existing_edos = apply_existing_edo_trace(
                    existing_edos,
                    existing_trace_details
                )

            except Exception as trace_error:
                logging.warning(
                    "STAGE 3C SKIPPED - trace extraction failed, leaving "
                    f"the trace part of column D blank for this run. "
                    f"Reason: {trace_error}"
                )

        existing_edo_final = merge_existing_edo_dictionary(existing_edos)
        final_edos = validate_final_edos_existing_only(existing_edo_final)

        if not final_edos:
            raise Exception("No EDO records generated.")

        write_edo_excel_existing_style(
            sheet,
            final_edos,
            start_row
        )

        output_file = save_edo_workbook(
            workbook,
            pipeline_config
        )

        logging.info("=" * 80)
        logging.info("EDO PIPELINE COMPLETED SUCCESSFULLY")
        logging.info(f"OUTPUT FILE PERSISTED AT : {output_file}")
        logging.info("=" * 80)

        return output_file

    except Exception as e:
        logging.error("=" * 80)
        logging.error(f"EDO PIPELINE CRITICAL RUNTIME FAILURE : {str(e)}")
        logging.error("=" * 80)
        raise e


# ==========================================================
# MAIN EXECUTION PIPELINE #2 (EXISTING + NEW EDO - COMBINED / CANONICAL)
# This is the primary merged pipeline: Stage 3 + Stage 3B from
# edo_existing_final.py, Stage 4 from generate_EDO_template_copy.py, and
# Stage 5/6 reconciled to carry both EDO types through to one worksheet.
# ==========================================================

def generate_edo_template(
    client,
    product_family,
    product,
    templatename,
    pipeline_config,
    db: DatabaseHandler
):
    logging.info("=" * 80)
    logging.info("STARTING COMPLETE EDO TEMPLATE GENERATION PIPELINE (MERGED)")
    logging.info("=" * 80)

    try:
        workbook, sheet = initialize_workbook(pipeline_config)
        start_row = pipeline_config.get("templatestartrow", 4)

        # Clear active table grid space exclusively up to Column J
        clear_existing_rows(sheet, start_row, end_column=10)

        edo_document = get_edo_document(
            client,
            product_family,
            product,
            templatename,
            db
        )

        # ---------------------------------------------------
        # STAGE 3 (image extraction): pull embedded images out of the
        # SAME "edo_proposed" document that extract_edo_tags() below
        # queries for Existing EDO tags - no separate document lookup.
        # Non-fatal: if the DatabaseHandler doesn't yet expose a way to
        # fetch the raw file (see resolve_edo_document_file()), this is
        # logged and the pipeline continues without images rather than
        # failing the whole run.
        # ---------------------------------------------------
        try:
            edo_proposed_images = extract_edo_proposed_images(edo_document, pipeline_config)

            images_output_dir = os.path.join(
                os.path.dirname(pipeline_config.get("output_file_path", ".")) or ".",
                "edo_proposed_images"
            )
            if edo_proposed_images:
                save_images_to_folder(edo_proposed_images, images_output_dir)

        except Exception as image_error:
            edo_proposed_images = []
            logging.warning(
                "IMAGE EXTRACTION SKIPPED - could not extract images from "
                f"edo_proposed for this run. Reason: {image_error}"
            )

        # ---------------------------------------------------
        # STAGE 3: Workflow 1 - Processing Existing EDO Records
        # ---------------------------------------------------
        existing_edos = extract_edo_tags(
            client,
            product_family,
            product,
            templatename,
            pipeline_config,
            edo_document,
            db
        )
        existing_edos = validate_existing_tags(existing_edos)

        if existing_edos:
            existing_edos = extract_edo_details(
                client,
                product_family,
                product,
                templatename,
                pipeline_config,
                edo_document,
                existing_edos,
                db
            )

            # ---------------------------------------------------
            # STAGE 3B: Verification Reference extraction for Existing EDOs
            # ---------------------------------------------------
            try:
                existing_verification_details = extract_existing_edo_verification_details(
                    client,
                    product_family,
                    product,
                    templatename,
                    pipeline_config,
                    edo_document,
                    existing_edos,
                    db
                )

                existing_edos = apply_existing_edo_verification(
                    existing_edos,
                    existing_verification_details
                )

            except Exception as verification_error:
                logging.warning(
                    "STAGE 3B SKIPPED - verification reference extraction "
                    "failed, leaving column E as 'Blank' for this run. "
                    f"Reason: {verification_error}"
                )

            # ---------------------------------------------------
            # STAGE 3C: Trace extraction for Existing EDOs (Column D,
            # appended below the RA/FMEA data, same row)
            # ---------------------------------------------------
            try:
                existing_trace_details = extract_existing_edo_trace_details(
                    client,
                    product_family,
                    product,
                    templatename,
                    pipeline_config,
                    edo_document,
                    existing_edos,
                    db
                )

                existing_edos = apply_existing_edo_trace(
                    existing_edos,
                    existing_trace_details
                )

            except Exception as trace_error:
                logging.warning(
                    "STAGE 3C SKIPPED - trace extraction failed, leaving "
                    "the trace part of column D blank for this run. "
                    f"Reason: {trace_error}"
                )

        # ---------------------------------------------------
        # STAGE 4: Workflow 2 - Processing New EDO Records
        #
        # All three sub-stages now share ONE dictionary (edo_new_data),
        # keyed by RA id:
        #   4a. extract_new_edo_tags()                 -> creates it
        #   4b. extract_new_edo_summary_details()       -> enriches it
        #       (loops over every entry, one LLM call per RA/FMEA pair)
        #   4c. extract_new_edo_traceability_details()  -> enriches it
        #       further (same loop-per-pair pattern)
        # ---------------------------------------------------
        edo_new_data = extract_new_edo_tags(
            client,
            product_family,
            product,
            templatename,
            pipeline_config,
            edo_document,
            db
        )
        print(f"extract_new_edo_tags: ", edo_new_data)

        edo_new_data = extract_new_edo_summary_details(
            client,
            product_family,
            product,
            templatename,
            pipeline_config,
            edo_document,
            edo_new_data,
            db
        )
        print(f"extract_new_edo_summary_details: ", edo_new_data)

        edo_new_data = extract_new_edo_traceability_details(
            client,
            product_family,
            product,
            templatename,
            pipeline_config,
            edo_document,
            edo_new_data,
            db
        )
        print(f"extract_new_edo_traceability_details: ", edo_new_data)

        new_records = list(edo_new_data.values())

        # ---------------------------------------------------
        # STAGE 4D: New EDO Diagram Lookup (Column H, EDO_pdf_new)
        #
        # DB-driven, same pattern as the edo_proposed/docx image
        # extraction above (Stage before Stage 3): resolves the
        # "EDO_pdf_new" document via get_edo_pdf_new_file() and, for
        # every New EDO record (e.g. RA-141 / FMEA Sys-152), finds and
        # extracts its full diagram. Never fatal - logged and the
        # pipeline continues with no per-record diagrams (falling back
        # to the generic edo_proposed image queue) rather than failing
        # the whole run.
        # ---------------------------------------------------
        try:
            new_edo_diagrams = extract_new_edo_diagrams(
                edo_document,
                new_records,
                pipeline_config
            )
        except Exception as new_diagram_error:
            new_edo_diagrams = {}
            logging.warning(
                "NEW EDO DIAGRAM EXTRACTION SKIPPED - could not extract "
                f"diagrams from EDO_pdf_new for this run. Reason: {new_diagram_error}"
            )

        # ---------------------------------------------------
        # STAGE 5: New EDO Merge (de-duplicated against Existing EDOs)
        # ---------------------------------------------------
        new_final = merge_new_edo_records(
            new_records,
            existing_edos
        )

        existing_edo_final = merge_existing_edo_dictionary(existing_edos)
        final_edos = merge_all_edos(existing_edo_final, new_final)
        final_edos = validate_final_edos(final_edos)

        if not final_edos:
            raise Exception("No EDO records generated.")

        # ---------------------------------------------------
        # STAGE 6: Output Mapping, Formatting and File Storage
        # ---------------------------------------------------
        format_edo_worksheet(
            sheet,
            final_edos,
            start_row,
            pipeline_config,
            images=edo_proposed_images,
            ra_fmea_images=new_edo_diagrams
        )

        output_file = save_edo_workbook(
            workbook,
            pipeline_config
        )

        logging.info("=" * 80)
        logging.info("EDO PIPELINE COMPLETED SUCCESSFULLY")
        logging.info(f"OUTPUT FILE PERSISTED AT : {output_file}")
        logging.info("=" * 80)

        return output_file

    except Exception as e:
        logging.error("=" * 80)
        logging.error(f"EDO PIPELINE CRITICAL RUNTIME FAILURE : {str(e)}")
        logging.error("=" * 80)
        raise e